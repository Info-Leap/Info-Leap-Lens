"""
mapping_workbook.py — Rich Excel mapping workbook generator for InfoLeap Pulse ingestion.

Two modes:
  1. Pre-ingest (classify → download) — produces workbook from schema_doc + raw DataFrame
  2. Post-ingest (after DB written) — adds sampled fact/dim tables from the live SQLite DB

Output workbook structure:
  MASTER_MAPPING    — one row per survey question, AI mapping + human override columns with
                      dropdown DataValidation on bucket column (all 22 TARGET_BUCKETS)
  RAW_DATA          — exact copy of raw survey DataFrame, zero transformation
  dim.brand         — brand lookup table
  dim.bq3_attribute — attribute lookup table
  dim.city / dim.zone
  fact.brand_awareness / fact.brand_imagery / fact.brand_nps
  fact.satisfaction / fact.need_importance / fact.attitudes
  fact.purchase_journey / fact.price_paid / fact.portfolio_awareness
  INSTRUCTIONS      — usage guide

Read-back:
  read_master_mapping(path) → override dict {question_code: {bucket, source_column, notes}}
  This feeds straight into generic_loader.assignment_from_schema() to re-ingest with corrections.
"""

from __future__ import annotations

import io
import sqlite3
from pathlib import Path
from typing import Optional

import pandas as pd

# ── Bucket list (must stay in sync with codebook_parser.TARGET_BUCKETS) ──────
BUCKET_NAMES = [
    "TOM", "SPONT", "AIDED", "EVER_USED", "CURRENT_USER", "CONSIDERATION",
    "PREFERRED", "LAST_PURCHASED", "CSAT", "NPS", "BRAND_IMAGERY", "IMPORTANCE",
    "ATTITUDE", "PORTFOLIO_AWARENESS", "PRICE_PAID", "PURCHASE_JOURNEY",
    "GENDER", "AGE", "CITY", "ZONE", "DEMOGRAPHIC", "SKIP",
]

# ── Column headers ─────────────────────────────────────────────────────────────
MASTER_COLS = [
    "question_code",
    "question_text",
    "ai_detected_bucket",
    "ai_detected_shape",
    "ai_source_columns",
    "ai_confidence",
    "ai_reasoning",
    "value_labels",              # code→label pairs (read-only reference)
    "OVERRIDE_bucket",           # ← human fills this; dropdown validated
    "OVERRIDE_source_column",    # ← human fills this if different from ai_source_columns
    "OVERRIDE_notes",            # ← free text
]

# Column indices (1-based) for DataValidation
_COL_OVERRIDE_BUCKET = 9   # column I
_COL_OVERRIDE_SOURCE = 10  # column J
_COL_OVERRIDE_NOTES  = 11  # column K


# ─────────────────────────────────────────────────────────────────────────────
# Styling helpers
# ─────────────────────────────────────────────────────────────────────────────

def _apply_header_style(ws, n_cols: int, fill_hex: str = "1F4E79"):
    """Bold white header row with colored fill."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    fill   = PatternFill("solid", fgColor=fill_hex)
    font   = Font(bold=True, color="FFFFFF", size=10)
    align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(bottom=thin)
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill   = fill
        cell.font   = font
        cell.alignment = align
        cell.border = border


def _apply_override_col_style(ws, n_rows: int):
    """Light green tint on OVERRIDE columns so they stand out."""
    from openpyxl.styles import PatternFill, Font, Alignment
    fill  = PatternFill("solid", fgColor="E2EFDA")
    font  = Font(size=9, italic=True, color="375623")
    align = Alignment(wrap_text=False, vertical="top")
    for col in (_COL_OVERRIDE_BUCKET, _COL_OVERRIDE_SOURCE, _COL_OVERRIDE_NOTES):
        for row in range(2, n_rows + 2):
            cell = ws.cell(row=row, column=col)
            cell.fill      = fill
            cell.font      = font
            cell.alignment = align


def _auto_col_widths(ws, max_width: int = 60):
    """Best-effort auto column widths."""
    from openpyxl.utils import get_column_letter
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), max_width)


def _style_data_rows(ws, n_rows: int, n_cols: int):
    """Alternating row shading for readability."""
    from openpyxl.styles import PatternFill, Alignment
    fill_even = PatternFill("solid", fgColor="F5F9FF")
    align = Alignment(wrap_text=True, vertical="top")
    for row in range(2, n_rows + 2):
        shade = row % 2 == 0
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = align
            if shade and col not in (_COL_OVERRIDE_BUCKET, _COL_OVERRIDE_SOURCE, _COL_OVERRIDE_NOTES):
                cell.fill = fill_even


def _freeze_and_filter(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


import re as _re

def _sanitize_question_text(text: str, question_code: str, bucket: str) -> str:
    """Replace AI-generated placeholder question_text with a useful label."""
    if not text:
        return text
    t = text.strip()
    # Patterns that indicate AI filled with a dummy/column label instead of real text:
    # "Brand 1", "Brand_1", "q15", "q15_1", column codes like "mq10f", etc.
    placeholder_patterns = [
        r"^Brand\s*\d+$",         # "Brand 1", "Brand 2"
        r"^[a-z]+\d+[a-z]*$",     # "q15", "q15a", "mq10f"
        r"^[a-z]+\d+_\d+$",       # "q15_1", "q17_3"
        r"^Brand_\d+$",            # "Brand_1"
    ]
    for pat in placeholder_patterns:
        if _re.match(pat, t, _re.IGNORECASE):
            return f"[{bucket}] {question_code}"
    return t


# ─────────────────────────────────────────────────────────────────────────────
# MASTER_MAPPING sheet
# ─────────────────────────────────────────────────────────────────────────────

def _build_master_mapping_sheet(wb, schema_doc: dict):
    """
    schema_doc format (from schema_ingest.classify_all_questions):
      {"project_id": "...", "questions": [
        {"question_code": "q17_1", "question_text": "...", "bucket": "AIDED",
         "shape": "brand_battery", "source_columns": ["q17_1","q17_2",...],
         "confidence": 0.92, "reasoning": "...", "value_labels": {"1": "Amul", ...}},
        ...
      ]}
    """
    from openpyxl.worksheet.datavalidation import DataValidation

    ws = wb.create_sheet("MASTER_MAPPING", 0)

    questions = schema_doc.get("questions", [])

    # Header row
    for col, header in enumerate(MASTER_COLS, 1):
        ws.cell(row=1, column=col, value=header)

    # Data rows
    for row_i, q in enumerate(questions, 2):
        # Handle multiple field-name conventions across schema versions
        src_cols = (q.get("source_columns") or q.get("data_columns")
                    or ([q["source_column"]] if q.get("source_column") else []))
        # Also include dummy_columns (brand battery sub-columns)
        dummy = q.get("dummy_columns") or []
        all_cols = list(src_cols) + [c for c in dummy if c not in src_cols]
        src_str  = ", ".join(str(c) for c in all_cols) if all_cols else ""
        vl       = q.get("value_labels") or q.get("code_to_label") or {}
        vl_str   = "; ".join(f"{k}={v}" for k, v in list(vl.items())[:12]) if vl else ""
        _qcode = q.get("question_code", "")
        _qtext = _sanitize_question_text(q.get("question_text", ""), _qcode, q.get("bucket", ""))
        ws.cell(row=row_i, column=1,  value=_qcode)
        ws.cell(row=row_i, column=2,  value=_qtext)
        ws.cell(row=row_i, column=3,  value=q.get("bucket", ""))
        ws.cell(row=row_i, column=4,  value=q.get("shape", ""))
        ws.cell(row=row_i, column=5,  value=src_str)
        ws.cell(row=row_i, column=6,  value=round(float(q.get("confidence", 0)), 2))
        ws.cell(row=row_i, column=7,  value=q.get("reasoning", ""))
        ws.cell(row=row_i, column=8,  value=vl_str)
        # OVERRIDE cols left blank for human
        ws.cell(row=row_i, column=9,  value="")
        ws.cell(row=row_i, column=10, value="")
        ws.cell(row=row_i, column=11, value="")

    n_rows = len(questions)

    # DataValidation dropdown on OVERRIDE_bucket column (col I)
    bucket_formula = '"' + ",".join(BUCKET_NAMES) + '"'
    dv = DataValidation(
        type="list",
        formula1=bucket_formula,
        allow_blank=True,
        showDropDown=False,   # False = show the dropdown arrow (confusingly named)
        showErrorMessage=True,
        errorTitle="Invalid bucket",
        error="Choose from the dropdown list or leave blank to keep AI mapping.",
    )
    dv.sqref = f"I2:I{n_rows + 1}"
    ws.add_data_validation(dv)

    # Styling
    _apply_header_style(ws, len(MASTER_COLS), fill_hex="1F4E79")
    _apply_override_col_style(ws, n_rows)
    _style_data_rows(ws, n_rows, len(MASTER_COLS))
    _freeze_and_filter(ws)

    # Fixed widths for key cols
    from openpyxl.utils import get_column_letter
    widths = {1: 18, 2: 45, 3: 20, 4: 18, 5: 35, 6: 12, 7: 50, 8: 40, 9: 22, 10: 30, 11: 30}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Row height for readability
    ws.row_dimensions[1].height = 30
    for r in range(2, n_rows + 2):
        ws.row_dimensions[r].height = 18


# ─────────────────────────────────────────────────────────────────────────────
# RAW_DATA sheet
# ─────────────────────────────────────────────────────────────────────────────

def _build_raw_data_sheet(wb, df_raw: pd.DataFrame):
    """Exact copy of raw survey data — zero transformation, zero column renaming."""
    ws = wb.create_sheet("RAW_DATA")
    n_cols = len(df_raw.columns)
    n_rows = len(df_raw)

    import numpy as np

    # Headers
    ws.append([str(c) for c in df_raw.columns])

    # Data — use append() for speed (avoids cell-by-cell overhead on wide frames)
    for _, row in df_raw.iterrows():
        clean_row = []
        for v in row:
            if hasattr(v, "item"):
                v = v.item()
            # Excel can't handle NaN/inf — convert to None
            if isinstance(v, float) and (v != v or v == float("inf") or v == float("-inf")):
                v = None
            clean_row.append(v)
        ws.append(clean_row)

    _apply_header_style(ws, n_cols, fill_hex="264653")
    _freeze_and_filter(ws)

    # Auto widths (cap at 25 for raw data — can be very wide)
    from openpyxl.utils import get_column_letter
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        vals = [str(c.value or "") for c in col_cells[:20]]
        ws.column_dimensions[col_letter].width = min(max(len(max(vals, key=len)) + 2, 8), 25)

    ws.row_dimensions[1].height = 22
    ws.sheet_properties.tabColor = "E76F51"


# ─────────────────────────────────────────────────────────────────────────────
# DB fact/dim sheets (post-ingest)
# ─────────────────────────────────────────────────────────────────────────────

_FACT_TABLES = [
    ("fact.brand_awareness",   "fact_brand_awareness",   300),
    ("fact.brand_imagery",     "fact_brand_imagery",     300),
    ("fact.brand_nps",         "fact_brand_nps",         300),
    ("fact.satisfaction",      "fact_satisfaction",      300),
    ("fact.need_importance",   "fact_need_importance",   300),
    ("fact.attitudes",         "fact_attitudes",         200),
    ("fact.purchase_journey",  "fact_purchase_journey",  300),
    ("fact.price_paid",        "fact_price_paid",        300),
    ("fact.portfolio_aware",   "fact_portfolio_awareness", 200),
]

_DIM_TABLES = [
    ("dim.brand",          "dim_brand"),
    ("dim.bq3_attribute",  "dim_bq3_attribute"),
    ("dim.city",           "dim_city"),
    ("dim.zone",           "dim_zone"),
]

_DIM_COLORS = {
    "dim.brand": "2A9D8F", "dim.bq3_attribute": "2A9D8F",
    "dim.city": "2A9D8F",  "dim.zone": "2A9D8F",
}

_FACT_COLOR = "457B9D"


def _add_db_sheets(wb, db_path: str):
    """Add sampled fact tables and full dim tables from the project SQLite DB."""
    conn = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True)
    existing = {r[0] for r in conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table'")}

    # Dim tables — full data
    for sheet_name, tbl in _DIM_TABLES:
        if tbl not in existing:
            continue
        try:
            df = pd.read_sql(f"SELECT * FROM {tbl}", conn)
        except Exception:
            continue
        if df.empty:
            continue
        ws = wb.create_sheet(sheet_name)
        for col_i, col in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_i, value=col)
        for row_i, (_, row) in enumerate(df.iterrows(), 2):
            for col_i, val in enumerate(row, 1):
                if hasattr(val, "item"):
                    val = val.item()
                ws.cell(row=row_i, column=col_i, value=val)
        _apply_header_style(ws, len(df.columns), fill_hex=_DIM_COLORS.get(sheet_name, "2A9D8F"))
        _freeze_and_filter(ws)
        _auto_col_widths(ws)
        ws.sheet_properties.tabColor = "2A9D8F"

    # Fact tables — sampled
    for sheet_name, tbl, limit in _FACT_TABLES:
        if tbl not in existing:
            continue
        try:
            total = conn.execute(f"SELECT COUNT(*) FROM {tbl}").fetchone()[0]
            df = pd.read_sql(f"SELECT * FROM {tbl} LIMIT {limit}", conn)
        except Exception:
            continue
        if df.empty:
            continue
        ws = wb.create_sheet(sheet_name)
        # Caption row with total count
        ws.cell(row=1, column=1, value=f"[Sample — {limit} of {total:,} rows. Full table in {tbl} in project DB.]")
        from openpyxl.styles import Font, PatternFill
        ws.cell(row=1, column=1).font = Font(italic=True, color="666666", size=9)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))

        # Headers in row 2
        for col_i, col in enumerate(df.columns, 1):
            ws.cell(row=2, column=col_i, value=col)
        for row_i, (_, row) in enumerate(df.iterrows(), 3):
            for col_i, val in enumerate(row, 1):
                if hasattr(val, "item"):
                    val = val.item()
                ws.cell(row=row_i, column=col_i, value=val)

        _apply_header_style(ws, len(df.columns), fill_hex=_FACT_COLOR)
        # Re-style the caption row (was overwritten by loop above — re-apply)
        ws.cell(row=2, column=1)  # ensure row 2 stays as header
        ws.freeze_panes = "A3"
        _auto_col_widths(ws)
        ws.sheet_properties.tabColor = "457B9D"

    conn.close()


# ─────────────────────────────────────────────────────────────────────────────
# INSTRUCTIONS sheet
# ─────────────────────────────────────────────────────────────────────────────

def _build_instructions_sheet(wb, project_id: str = ""):
    from openpyxl.styles import Font, PatternFill, Alignment
    ws = wb.create_sheet("INSTRUCTIONS")
    ws.sheet_properties.tabColor = "F4A261"

    lines = [
        ("InfoLeap Pulse — Mapping Workbook", True, 14),
        (f"Project: {project_id or '(not set)'}", False, 10),
        ("", False, 10),
        ("HOW TO USE THIS WORKBOOK", True, 11),
        ("", False, 10),
        ("1. MASTER_MAPPING sheet", True, 10),
        ("   • Columns A–H: AI-detected mapping (read-only reference — do not edit)", False, 10),
        ("   • Column I — OVERRIDE_bucket: Use the dropdown to change the bucket if AI got it wrong.", False, 10),
        ("     Leave blank to accept the AI mapping.", False, 10),
        ("   • Column J — OVERRIDE_source_column: If the source column name differs, type it here.", False, 10),
        ("   • Column K — OVERRIDE_notes: Any notes for the data team.", False, 10),
        ("", False, 10),
        ("2. RAW_DATA sheet", True, 10),
        ("   • Exact copy of your uploaded data. Use to visually verify column contents vs mappings.", False, 10),
        ("   • Do NOT edit — it is reference only.", False, 10),
        ("", False, 10),
        ("3. fact.* sheets", True, 10),
        ("   • Sampled rows (up to 300) from the ingested fact tables in the project database.", False, 10),
        ("   • Use to verify that ingested data looks correct.", False, 10),
        ("", False, 10),
        ("4. dim.* sheets", True, 10),
        ("   • Full dimension tables (brands, attributes, cities, zones).", False, 10),
        ("   • Verify brand names and attribute labels are correct.", False, 10),
        ("", False, 10),
        ("BUCKET REFERENCE", True, 11),
        ("", False, 10),
    ]

    bucket_desc = {
        "TOM": "Top-of-mind (first brand named, unprompted)",
        "SPONT": "Spontaneous recall (unprompted, after TOM)",
        "AIDED": "Aided awareness (from a shown list)",
        "EVER_USED": "Ever used the brand",
        "CURRENT_USER": "Currently uses the brand",
        "CONSIDERATION": "Would consider purchasing",
        "PREFERRED": "Most preferred brand",
        "LAST_PURCHASED": "Most recently purchased brand",
        "CSAT": "Satisfaction score (0–10)",
        "NPS": "Net Promoter Score (likelihood to recommend)",
        "BRAND_IMAGERY": "Brand attribute association battery",
        "IMPORTANCE": "Attribute importance ratings",
        "ATTITUDE": "Category attitude/agreement statements",
        "PORTFOLIO_AWARENESS": "Brand-category portfolio association",
        "PRICE_PAID": "Price paid / price tier",
        "PURCHASE_JOURNEY": "Purchase journey / channel / decision",
        "GENDER": "Respondent gender",
        "AGE": "Respondent age",
        "CITY": "Respondent city",
        "ZONE": "Respondent zone/region",
        "DEMOGRAPHIC": "Other demographic (income, occupation, etc.)",
        "SKIP": "Not used — leave unmapped",
    }

    for b, desc in bucket_desc.items():
        lines.append((f"   {b}: {desc}", False, 9))

    row = 1
    for text, bold, size in lines:
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(bold=bold, size=size,
                         color="1F4E79" if bold else "333333")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 16 if not bold else 20
        row += 1

    ws.column_dimensions["A"].width = 90
    ws.freeze_panes = "A1"


# ─────────────────────────────────────────────────────────────────────────────
# COLUMN_PICKER sheet — every raw column with bucket dropdown
# ─────────────────────────────────────────────────────────────────────────────

def _build_column_picker_sheet(wb, df_raw: pd.DataFrame, schema_doc: dict):
    """
    One row per raw data column. Columns:
      A: raw_column_name
      B: detected_bucket (from schema_doc, or blank)
      C: question_code
      D: confidence
      E: shape
      F: sample_values (first 5 unique non-null values)
      G: OVERRIDE_bucket (dropdown)
      H: OVERRIDE_notes
    Reviewers can verify every column is accounted for and correct bucket.
    """
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import PatternFill, Font, Alignment

    ws = wb.create_sheet("COLUMN_PICKER")
    ws.sheet_properties.tabColor = "7B2D8B"

    headers = ["raw_column_name", "detected_bucket", "question_code",
               "confidence", "shape", "sample_values",
               "OVERRIDE_bucket", "OVERRIDE_notes"]
    for col_i, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_i, value=h)

    # Build column → question lookup from schema_doc
    col_to_q: dict[str, dict] = {}
    for q in schema_doc.get("questions", []):
        src_cols = (q.get("source_columns") or q.get("data_columns")
                    or ([q["source_column"]] if q.get("source_column") else []))
        dummy = q.get("dummy_columns") or []
        for c in list(src_cols) + dummy:
            col_to_q[str(c)] = q

    for row_i, col_name in enumerate(df_raw.columns, 2):
        col_str = str(col_name)
        q = col_to_q.get(col_str, {})

        # Sample values — first 5 unique non-null
        try:
            samples = df_raw[col_name].dropna().unique()[:5]
            sample_str = " | ".join(str(v) for v in samples)
        except Exception:
            sample_str = ""

        ws.cell(row=row_i, column=1, value=col_str)
        ws.cell(row=row_i, column=2, value=q.get("bucket", ""))
        ws.cell(row=row_i, column=3, value=q.get("question_code", ""))
        conf = q.get("confidence")
        ws.cell(row=row_i, column=4, value=round(float(conf), 2) if conf is not None else "")
        ws.cell(row=row_i, column=5, value=q.get("shape", ""))
        ws.cell(row=row_i, column=6, value=sample_str[:200])
        ws.cell(row=row_i, column=7, value="")   # OVERRIDE_bucket
        ws.cell(row=row_i, column=8, value="")   # OVERRIDE_notes

        # Highlight unmapped columns
        if not q:
            fill = PatternFill("solid", fgColor="FFF3CD")
            for c in range(1, 7):
                ws.cell(row=row_i, column=c).fill = fill

    n_rows = len(df_raw.columns)

    # Dropdown on OVERRIDE_bucket (col G)
    bucket_formula = '"' + ",".join(BUCKET_NAMES) + '"'
    dv = DataValidation(
        type="list", formula1=bucket_formula,
        allow_blank=True, showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid bucket",
        error="Choose from dropdown or leave blank.",
    )
    dv.sqref = f"G2:G{n_rows + 1}"
    ws.add_data_validation(dv)

    _apply_header_style(ws, len(headers), fill_hex="7B2D8B")
    _freeze_and_filter(ws)
    _style_data_rows(ws, n_rows, len(headers))

    from openpyxl.utils import get_column_letter
    widths = {1: 28, 2: 20, 3: 18, 4: 12, 5: 18, 6: 50, 7: 22, 8: 30}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 28
    # Highlight override cols
    _apply_override_col_style(ws, n_rows)


# ─────────────────────────────────────────────────────────────────────────────
# PER-BUCKET sheets — one sheet per bucket with all mapped questions
# ─────────────────────────────────────────────────────────────────────────────

_BUCKET_TAB_COLORS = {
    "TOM": "C0392B", "SPONT": "E67E22", "AIDED": "F1C40F",
    "EVER_USED": "27AE60", "CURRENT_USER": "1ABC9C",
    "CONSIDERATION": "2980B9", "PREFERRED": "8E44AD",
    "LAST_PURCHASED": "2C3E50",
    "NPS": "E74C3C", "CSAT": "E91E63",
    "BRAND_IMAGERY": "3498DB", "IMPORTANCE": "16A085",
    "ATTITUDE": "8E44AD", "PORTFOLIO_AWARENESS": "27AE60",
    "PRICE_PAID": "795548", "PURCHASE_JOURNEY": "FF5722",
    "GENDER": "607D8B", "AGE": "78909C", "CITY": "546E7A", "ZONE": "455A64",
    "DEMOGRAPHIC": "90A4AE", "SKIP": "BDBDBD",
}

_BUCKET_SHEET_COLS = [
    "question_code", "question_text", "shape",
    "source_columns", "confidence", "reasoning", "value_labels",
]


def _build_bucket_sheets(wb, schema_doc: dict):
    """One sheet per bucket — all questions mapped to that bucket."""
    from collections import defaultdict

    questions = schema_doc.get("questions", [])
    by_bucket: dict[str, list] = defaultdict(list)
    for q in questions:
        bucket = q.get("bucket", "SKIP") or "SKIP"
        by_bucket[bucket].append(q)

    # Order: important buckets first
    order = [b for b in BUCKET_NAMES if b in by_bucket]
    for bucket in order:
        qs = by_bucket[bucket]
        sheet_name = f"b.{bucket}"[:31]  # Excel max 31 chars
        ws = wb.create_sheet(sheet_name)
        tab_color = _BUCKET_TAB_COLORS.get(bucket, "607D8B")
        ws.sheet_properties.tabColor = tab_color

        # Header
        for col_i, h in enumerate(_BUCKET_SHEET_COLS, 1):
            ws.cell(row=1, column=col_i, value=h)

        for row_i, q in enumerate(qs, 2):
            src_cols = (q.get("source_columns") or q.get("data_columns")
                        or ([q["source_column"]] if q.get("source_column") else []))
            dummy = q.get("dummy_columns") or []
            all_cols = list(src_cols) + [c for c in dummy if c not in src_cols]
            vl = q.get("value_labels") or q.get("code_to_label") or {}
            _bqcode = q.get("question_code", "")
            _bqtext = _sanitize_question_text(q.get("question_text", ""), _bqcode, q.get("bucket", ""))
            ws.cell(row=row_i, column=1, value=_bqcode)
            ws.cell(row=row_i, column=2, value=_bqtext)
            ws.cell(row=row_i, column=3, value=q.get("shape", ""))
            ws.cell(row=row_i, column=4, value=", ".join(str(c) for c in all_cols))
            conf = q.get("confidence")
            ws.cell(row=row_i, column=5, value=round(float(conf), 2) if conf is not None else "")
            ws.cell(row=row_i, column=6, value=q.get("reasoning", ""))
            ws.cell(row=row_i, column=7, value="; ".join(f"{k}={v}" for k, v in list(vl.items())[:15]))

        n_rows = len(qs)
        _apply_header_style(ws, len(_BUCKET_SHEET_COLS), fill_hex=tab_color)
        _freeze_and_filter(ws)
        _style_data_rows(ws, n_rows, len(_BUCKET_SHEET_COLS))
        _auto_col_widths(ws)
        ws.row_dimensions[1].height = 26


# ─────────────────────────────────────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────────────────────────────────────

def build_mapping_workbook(
    schema_doc: dict,
    df_raw: Optional[pd.DataFrame] = None,
    db_path: Optional[str] = None,
    project_id: str = "",
) -> bytes:
    """
    Build and return the full mapping workbook as bytes.

    Args:
        schema_doc: Output of classify_all_questions — {"project_id": ..., "questions": [...]}
        df_raw:     Raw survey DataFrame (copied as-is into RAW_DATA sheet). Optional.
        db_path:    Path to project SQLite DB (for fact/dim table sheets). Optional.
        project_id: Used in INSTRUCTIONS sheet title and filename suggestion.

    Returns:
        bytes — Excel workbook (.xlsx)
    """
    import openpyxl
    wb = openpyxl.Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    _pid = project_id or schema_doc.get("project_id", "")

    # Sheet 1: MASTER_MAPPING (primary override UI)
    _build_master_mapping_sheet(wb, schema_doc)

    # Sheet 2: COLUMN_PICKER (all raw cols with bucket dropdowns)
    if df_raw is not None and not df_raw.empty:
        _build_column_picker_sheet(wb, df_raw, schema_doc)

    # Sheet 3+: Per-bucket breakdown sheets
    _build_bucket_sheets(wb, schema_doc)

    # INSTRUCTIONS (last — easy to find but not in the way)
    _build_instructions_sheet(wb, project_id=_pid)

    # RAW_DATA — full survey data
    if df_raw is not None and not df_raw.empty:
        _build_raw_data_sheet(wb, df_raw)

    if db_path and Path(db_path).exists():
        _add_db_sheets(wb, db_path)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# Read-back: parse corrected workbook → override dict
# ─────────────────────────────────────────────────────────────────────────────

def read_master_mapping(workbook_bytes: bytes) -> dict[str, dict]:
    """
    Parse a corrected mapping workbook and return override dict.

    Returns:
        {question_code: {"bucket": str, "source_column": str, "notes": str}}
        Only rows where at least one OVERRIDE column is filled are included.
    """
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
    if "MASTER_MAPPING" not in wb.sheetnames:
        raise ValueError("Workbook has no MASTER_MAPPING sheet")

    ws = wb["MASTER_MAPPING"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    headers = [str(h).strip() if h else "" for h in rows[0]]
    try:
        idx_code   = headers.index("question_code")
        idx_bucket = headers.index("OVERRIDE_bucket")
        idx_src    = headers.index("OVERRIDE_source_column")
        idx_notes  = headers.index("OVERRIDE_notes")
    except ValueError as e:
        raise ValueError(f"MASTER_MAPPING missing expected column: {e}")

    overrides = {}
    for row in rows[1:]:
        code    = str(row[idx_code]).strip() if row[idx_code] else ""
        bucket  = str(row[idx_bucket]).strip() if row[idx_bucket] else ""
        src_col = str(row[idx_src]).strip() if row[idx_src] else ""
        notes   = str(row[idx_notes]).strip() if row[idx_notes] else ""
        if code and (bucket or src_col or notes):
            overrides[code] = {
                "bucket": bucket or None,
                "source_column": src_col or None,
                "notes": notes or None,
            }

    wb.close()
    return overrides


def apply_overrides_to_schema(schema_doc: dict, overrides: dict[str, dict]) -> dict:
    """
    Merge human overrides from MASTER_MAPPING into schema_doc questions.
    Returns new schema_doc with overridden bucket/source_column values.
    """
    import copy
    new_doc = copy.deepcopy(schema_doc)
    override_map = {k: v for k, v in overrides.items()}

    for q in new_doc.get("questions", []):
        code = q.get("question_code", "")
        if code in override_map:
            ov = override_map[code]
            if ov.get("bucket"):
                q["bucket"] = ov["bucket"]
            if ov.get("source_column"):
                # Update whichever key this schema version uses
                if "source_columns" in q:
                    q["source_columns"] = [ov["source_column"]]
                else:
                    q["source_column"] = ov["source_column"]
            if ov.get("notes"):
                q["_override_notes"] = ov["notes"]

    return new_doc
