"""
master_excel.py — Authoritative per-project master workbook.

Sheet layout (6 sheets, intuitive):
  META           — project metadata (auto, read-only)
  BUCKET_CONFIG  — PRIMARY EDIT SURFACE: one row per bucket (TOM/SPONT/AIDED/NPS/etc.)
                   Non-technical user edits source_column, brand_names here.
                   Color-coded by bucket type. OVERRIDE columns highlighted green.
  BRANDS         — canonical brand list: code → name, junk flag, in-analysis flag
  RAW_DATA       — all respondent rows (editable for data corrections)
  IMAGERY_TAGGED — attribute × brand × assoc% (auto-computed, reference only)
  VALIDATION     — awareness funnel % per brand, green/red flags (auto, reference)

Workflow:
  1. Open BUCKET_CONFIG — see every bucket, its source column, fill rate, brand count
  2. Change source_column if wrong column detected
  3. Fix brand names in BRANDS sheet (change brand_name col)
  4. Upload back to app → re-ingest → numbers update instantly
"""

from __future__ import annotations

import re
from datetime import datetime
from typing import Optional

import pandas as pd
import numpy as np

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


AWARENESS_STAGES = [
    "TOM", "SPONT", "AIDED", "EVER_USED",
    "CONSIDERATION", "CURRENT_USER", "PREFERRED", "LAST_PURCHASED",
]
JUNK_PATTERNS = [r"^\$\{", r"^NoneOfThese$", r"^None$", r"^Other$", r"^Raw/Loose"]

# Bucket type → colour (so user can visually group)
BUCKET_COLORS = {
    # Awareness funnel — blue family
    "TOM":           "BDD7EE", "SPONT":         "9DC3E6", "AIDED":         "2E75B6",
    "EVER_USED":     "1F4E79", "CONSIDERATION": "2F5597", "CURRENT_USER":  "4472C4",
    "PREFERRED":     "5B9BD5", "LAST_PURCHASED":"70AD47",
    # NPS / CSAT — purple
    "NPS":           "D9B3FF", "CSAT":          "C490E4",
    # Imagery / Importance — orange
    "BRAND_IMAGERY": "FCE4D6", "IMPORTANCE":    "FFD966",
    # Demographics — grey
    "GENDER":        "F2F2F2", "AGE":           "F2F2F2", "CITY":          "F2F2F2",
    "ZONE":          "F2F2F2", "DEMOGRAPHIC":   "F2F2F2", "INCOME":        "F2F2F2",
    "OCCUPATION":    "F2F2F2", "NCCS":          "F2F2F2",
    # Other
    "PRICE_PAID":    "E2EFDA", "PURCHASE_JOURNEY": "E2EFDA", "ATTITUDE":  "FFF2CC",
    "SKIP":          "D9D9D9",
}

# Bucket → plain-English description
BUCKET_DESC = {
    "TOM":              "Top-of-mind brand (unaided, first mention)",
    "SPONT":            "Spontaneous brand recall (unaided, all mentions)",
    "AIDED":            "Aided brand awareness (shown a list)",
    "EVER_USED":        "Ever used / owned this brand at home",
    "CONSIDERATION":    "Would consider buying next",
    "CURRENT_USER":     "Currently own / use this brand",
    "PREFERRED":        "Most preferred brand for next purchase",
    "LAST_PURCHASED":   "Most recently purchased brand",
    "NPS":              "Net Promoter Score (0–10 likelihood to recommend)",
    "CSAT":             "Customer satisfaction score (0–10)",
    "BRAND_IMAGERY":    "Brand attribute associations (multi-select per attribute)",
    "IMPORTANCE":       "Attribute importance rating (1–7 scale)",
    "GENDER":           "Respondent gender",
    "AGE":              "Respondent age / age band",
    "CITY":             "Respondent city",
    "ZONE":             "Respondent zone / region",
    "DEMOGRAPHIC":      "Other demographic variable",
    "INCOME":           "Household income",
    "OCCUPATION":       "Respondent occupation",
    "NCCS":             "NCCS socio-economic category",
    "PRICE_PAID":       "Price / price tier paid for last purchase",
    "PURCHASE_JOURNEY": "Purchase journey question (channel, reason, etc.)",
    "ATTITUDE":         "Category attitude statement rating",
    "SKIP":             "Not ingested (ignored)",
}

_C   = lambda hex: PatternFill("solid", fgColor=hex) if HAS_OPENPYXL else None
GREEN_FILL    = _C("C6EFCE")
RED_FILL      = _C("FFC7CE")
YELLOW_FILL   = _C("FFEB9C")
HEADER_FILL   = _C("1F3864")
OVERRIDE_FILL = _C("E2EFDA")   # light green = editable by user
HEADER_FONT   = Font(color="FFFFFF", bold=True) if HAS_OPENPYXL else None
BOLD_FONT     = Font(bold=True) if HAS_OPENPYXL else None
ITALIC_FONT   = Font(italic=True, color="595959") if HAS_OPENPYXL else None
THIN_BORDER   = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
) if HAS_OPENPYXL else None


def _is_junk(name: str) -> bool:
    for pat in JUNK_PATTERNS:
        if re.search(pat, str(name), re.IGNORECASE):
            return True
    return False


def _safe_int(val):
    try:
        return int(val)
    except (ValueError, TypeError):
        return val


def _safe_val(val):
    if val is None:
        return None
    try:
        if pd.isna(val):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(val, (np.integer,)):
        return int(val)
    if isinstance(val, (np.floating,)):
        return float(val)
    if isinstance(val, (np.bool_,)):
        return bool(val)
    return val


def _header_row(ws, values: list[str], row: int = 1):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        if HAS_OPENPYXL:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _freeze_filter(ws, freeze_at="A2"):
    ws.freeze_panes = freeze_at
    ws.auto_filter.ref = ws.dimensions


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def write_master_excel(
    output_path: str,
    schema_doc: dict,
    raw_data_path: str,
    project_id: str,
    raw_df: Optional[pd.DataFrame] = None,
):
    """Write master_mapping.xlsx — 6-sheet intuitive workbook."""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl required: pip install openpyxl")

    wb = Workbook()
    wb.remove(wb.active)

    questions  = schema_doc.get("questions", [])
    n_resp     = _count_respondents(raw_df)

    _write_meta(wb, schema_doc, project_id, raw_data_path, n_resp)
    _write_bucket_config(wb, questions, raw_df, n_resp)
    _write_brands(wb, questions, schema_doc.get("_ap_brand_universe", []))
    if raw_df is not None:
        _write_raw_data(wb, raw_df)
        _write_imagery_tagged(wb, questions, raw_df, n_resp)
        _write_validation(wb, schema_doc, raw_df, n_resp)

    wb.save(output_path)
    print(f"[OK] master_mapping.xlsx → {output_path}")


def _count_respondents(raw_df: Optional[pd.DataFrame]) -> int:
    if raw_df is None:
        return 0
    for col in ("caseid", "respondent_id", "id", "CASEID", "resp_id"):
        if col in raw_df.columns and raw_df[col].nunique() > 1:
            return raw_df[col].nunique()
    return len(raw_df)


_NON_BRAND_BUCKETS = {
    "CITY", "ZONE", "DEMOGRAPHIC", "AGE", "GENDER", "INCOME",
    "OCCUPATION", "NCCS", "PRICE_PAID", "PURCHASE_JOURNEY", "ATTITUDE",
    "IMPORTANCE", "SKIP",
}

# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------

def _write_meta(wb, schema_doc, project_id, raw_data_path, n_resp):
    ws = wb.create_sheet("META")
    rows = [
        ("project_id",     project_id),
        ("raw_data_path",  raw_data_path),
        ("n_respondents",  n_resp),
        ("generated_at",   datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("schema_version", "3.1"),
        ("how_to_edit",    "Open BUCKET_CONFIG — change source_column or OVERRIDE_source_column. Open BRANDS — fix brand names. Re-upload to app to re-ingest."),
        ("sheets",         "META | BUCKET_CONFIG | BRANDS | RAW_DATA | IMAGERY_TAGGED | VALIDATION"),
    ]
    _header_row(ws, ["field", "value"])
    for i, (k, v) in enumerate(rows, 2):
        c = ws.cell(row=i, column=1, value=k)
        if HAS_OPENPYXL:
            c.font = BOLD_FONT
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 90


def _write_bucket_config(wb, questions, raw_df, n_resp):
    """
    PRIMARY EDIT SURFACE — one row per unique bucket assignment.

    Columns:
      [auto] bucket | bucket_description | question_code | question_text
      [auto] source_column | dummy_columns | n_columns_parsed
      [auto] n_filled | fill_rate_% | brand_count | example_values
      [EDIT] OVERRIDE_source_column — paste correct column name here if wrong
      [EDIT] notes — any comments
    """
    ws = wb.create_sheet("BUCKET_CONFIG")

    # Build headers with two annotation rows
    header_main = [
        "bucket", "bucket_description", "question_code", "question_text",
        "source_column", "dummy_columns", "n_columns_parsed",
        "n_filled", "fill_rate_%", "brand_count", "example_values",
        "OVERRIDE_source_column", "notes",
    ]
    # Row 1: group labels
    groups = [
        ("AUTO — do not edit", 11),  # cols 1-11
        ("EDITABLE — change these", 2),  # cols 12-13
    ]
    col_idx = 1
    for label, span in groups:
        cell = ws.cell(row=1, column=col_idx, value=label)
        if HAS_OPENPYXL:
            fill = HEADER_FILL if label.startswith("AUTO") else _C("375623")
            cell.fill = fill
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            if span > 1:
                ws.merge_cells(
                    start_row=1, start_column=col_idx,
                    end_row=1, end_column=col_idx + span - 1
                )
        col_idx += span

    # Row 2: actual column headers
    for c, val in enumerate(header_main, 1):
        cell = ws.cell(row=2, column=c, value=val)
        if HAS_OPENPYXL:
            is_edit = c >= 12
            cell.fill = _C("375623") if is_edit else HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data rows start at row 3
    # Deduplicate by source_column (not question_code) so multi-attribute batteries
    # like q34_1_1 / q34_2_1 / ... each get their own row, not collapsed to one.
    seen_src: set[str] = set()
    data_row = 3
    for q in questions:
        src_key = q.get("source_column") or q.get("question_code", "")
        if src_key in seen_src:
            continue
        seen_src.add(src_key)

        qcode    = q.get("question_code", "")
        bucket   = str(q.get("bucket", "SKIP")).strip().upper()
        src_col  = q.get("source_column") or qcode
        dummies  = q.get("dummy_columns", [])
        ctl      = {k: v for k, v in q.get("code_to_label", {}).items() if not _is_junk(str(v))}
        brand_ct = len(set(ctl.values()))
        dummy_str = "; ".join(dummies)

        # Compute fill stats from raw_df
        n_filled, fill_pct, examples = 0, 0.0, ""
        check_cols = [src_col] + dummies[:3]
        if raw_df is not None:
            if dummies:
                # For multi-select: count rows where ANY dummy = 1
                avail = [c for c in dummies if c in raw_df.columns]
                if avail:
                    n_filled = int(raw_df[avail].any(axis=1).sum())
            elif src_col in raw_df.columns:
                n_filled = int(raw_df[src_col].notna().sum())
            fill_pct = round(n_filled / n_resp * 100, 1) if n_resp > 0 else 0.0

            # Example values: first 5 unique non-null values from source col
            if src_col in (raw_df.columns if raw_df is not None else []):
                ex = raw_df[src_col].dropna().unique()[:5]
                # Show code → brand if CTL exists
                ex_parts = []
                for v in ex:
                    code_s = str(_safe_int(v))
                    label  = ctl.get(code_s) or ctl.get(f"{code_s}.0")
                    ex_parts.append(f"{code_s}={label}" if label else code_s)
                examples = ", ".join(ex_parts)

        n_cols_parsed = len(dummies) if dummies else (1 if src_col else 0)
        bucket_color = BUCKET_COLORS.get(bucket, "FFFFFF")
        bucket_fill  = _C(bucket_color) if bucket_color != "FFFFFF" else None

        row_vals = [
            bucket,
            BUCKET_DESC.get(bucket, ""),
            qcode,
            q.get("question_text", ""),
            src_col,
            dummy_str,
            n_cols_parsed,
            n_filled,
            fill_pct,
            brand_ct if brand_ct else "",
            examples,
            "",   # OVERRIDE_source_column (user edits)
            "",   # notes
        ]
        for c, val in enumerate(row_vals, 1):
            cell = ws.cell(row=data_row, column=c, value=val)
            if HAS_OPENPYXL:
                if c <= 2 and bucket_fill:
                    cell.fill = bucket_fill
                    cell.font = Font(bold=True, color="000000")
                elif c in (12, 13):
                    cell.fill = OVERRIDE_FILL
                if fill_pct < 20 and c == 9:
                    cell.fill = RED_FILL
                elif fill_pct < 60 and c == 9:
                    cell.fill = YELLOW_FILL

        data_row += 1

    # Column widths
    col_widths = [18, 55, 16, 55, 18, 60, 16, 10, 12, 12, 45, 28, 30]
    for c, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(header_main))}2"
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 30
    print(f"  [BUCKET_CONFIG] {data_row - 3} bucket rows")


def _collect_all_brands(questions):
    """All unique brand code→name from all brand-related questions."""
    # qcode → {code_str: brand_name}
    by_q: dict[str, dict[str, str]] = {}
    for q in questions:
        bucket = str(q.get("bucket", "")).upper()
        if bucket in _NON_BRAND_BUCKETS or bucket.startswith("DEMOGRAPHIC"):
            continue
        qcode = q.get("question_code", "")
        ctl   = q.get("code_to_label", {})
        if not ctl:
            continue
        clean: dict[str, str] = {}
        for k, v in ctl.items():
            try:
                ki = str(int(float(k)))
            except (ValueError, TypeError):
                ki = str(k)
            if ki not in clean:
                clean[ki] = str(v).strip()
        by_q[qcode] = clean
    return by_q


def _write_brands(wb, questions, ap_universe):
    """
    Consolidated brand code table — one row per (question_code, brand_code, brand_name).
    User edits brand_name column to fix spellings.
    is_junk = TRUE → excluded from all analytics.
    in_analysis = TRUE → included in AP brand universe.
    """
    ws = wb.create_sheet("BRANDS")

    # Instruction row
    ws.cell(row=1, column=1,
            value="Edit brand_name to fix spellings. Set is_junk=TRUE to exclude. in_analysis=TRUE = counted in funnel.")
    if HAS_OPENPYXL:
        ws.cell(row=1, column=1).font = ITALIC_FONT
    ws.merge_cells("A1:G1")

    headers = ["question_code", "bucket", "brand_code", "brand_name", "is_junk", "in_analysis", "notes"]
    _header_row(ws, headers, row=2)

    ap_set  = set(ap_universe)
    by_q    = _collect_all_brands(questions)
    qcode_to_bucket = {q.get("question_code",""): str(q.get("bucket","")).upper() for q in questions}

    row_idx = 3
    for qcode, codes in by_q.items():
        bucket = qcode_to_bucket.get(qcode, "")
        for code, name in sorted(codes.items(), key=lambda x: _safe_int(x[0])):
            is_junk    = _is_junk(name)
            in_analysis = name in ap_set
            ws.cell(row=row_idx, column=1, value=qcode)
            ws.cell(row=row_idx, column=2, value=bucket)
            ws.cell(row=row_idx, column=3, value=_safe_int(code))
            ws.cell(row=row_idx, column=4, value=name)
            ws.cell(row=row_idx, column=5, value="TRUE" if is_junk else "FALSE")
            ws.cell(row=row_idx, column=6, value="TRUE" if in_analysis else "FALSE")
            ws.cell(row=row_idx, column=7, value="")
            if is_junk and HAS_OPENPYXL:
                for c in range(1, 8):
                    ws.cell(row=row_idx, column=c).fill = YELLOW_FILL
            row_idx += 1

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(7)}2"
    for col, w in zip("ABCDEFG", [16, 22, 12, 35, 10, 12, 30]):
        ws.column_dimensions[col].width = w
    print(f"  [BRANDS] {row_idx - 3} brand-code rows")


def _write_raw_data(wb, raw_df: pd.DataFrame):
    ws = wb.create_sheet("RAW_DATA")
    cols = list(raw_df.columns)

    # Instruction row
    ws.cell(row=1, column=1,
            value="All respondent rows. Edit to correct data. Re-ingest to apply changes.")
    if HAS_OPENPYXL:
        ws.cell(row=1, column=1).font = ITALIC_FONT
    ws.merge_cells(f"A1:{get_column_letter(min(len(cols), 10))}1")

    _header_row(ws, cols, row=2)
    for r_idx, row in enumerate(raw_df.itertuples(index=False), 3):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=_safe_val(val))

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}2"
    for c_idx, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max(len(str(col)) + 2, 8), 20)
    print(f"  [RAW_DATA] {len(raw_df)} rows × {len(cols)} cols")


def _write_imagery_tagged(wb, questions, raw_df: pd.DataFrame, n_resp: int):
    """
    Pre-compute imagery association % for each attribute × brand.
    Handles both dummy-column format (bq3b/1, bq3b/2...) and single-column format.
    """
    ws = wb.create_sheet("IMAGERY_TAGGED")
    ws.cell(row=1, column=1,
            value="Auto-computed from BRAND_IMAGERY questions. Reference only — do not edit.")
    if HAS_OPENPYXL:
        ws.cell(row=1, column=1).font = ITALIC_FONT
    ws.merge_cells("A1:G1")

    headers = ["attribute_code", "attribute_text", "brand_name", "assoc_%", "n_assoc", "n_base", "is_junk"]
    _header_row(ws, headers, row=2)

    imagery_qs = [q for q in questions if str(q.get("bucket","")).upper() == "BRAND_IMAGERY"]
    if not imagery_qs:
        ws.cell(row=3, column=1, value="No BRAND_IMAGERY questions found in schema")
        return

    row_idx = 3
    for q in imagery_qs:
        attr_code = q.get("question_code", "")
        attr_text = q.get("question_text", "")
        dummies   = q.get("dummy_columns", [])
        ctl       = q.get("code_to_label", {})

        code_to_brand: dict[str, str] = {}
        for k, v in ctl.items():
            try:
                ki = str(int(float(k)))
            except (ValueError, TypeError):
                ki = str(k)
            code_to_brand[ki] = str(v)

        if dummies:
            # Dummy column per brand: attr_code/{brand_code}
            for dummy_col in dummies:
                if dummy_col not in raw_df.columns:
                    continue
                parts      = dummy_col.rsplit("/", 1)
                brand_code = parts[-1] if len(parts) == 2 else dummy_col
                brand_name = code_to_brand.get(brand_code, f"code_{brand_code}")
                col_data   = pd.to_numeric(raw_df[dummy_col], errors="coerce").fillna(0)
                n_assoc    = int(col_data.sum())
                assoc_pct  = round(n_assoc / n_resp * 100, 1) if n_resp > 0 else 0.0
                is_junk    = _is_junk(brand_name)
                _write_imagery_row(ws, row_idx, attr_code, attr_text, brand_name,
                                   assoc_pct, n_assoc, n_resp, is_junk)
                row_idx += 1
        else:
            # Single column with brand code values — count each code occurrence
            src_col = q.get("source_column") or attr_code
            if src_col in raw_df.columns:
                for code, brand_name in code_to_brand.items():
                    if _is_junk(brand_name):
                        continue
                    mask    = raw_df[src_col].astype(str).str.strip() == str(code)
                    try:
                        mask2 = raw_df[src_col] == int(code)
                        mask  = mask | mask2
                    except (ValueError, TypeError):
                        pass
                    n_assoc   = int(mask.sum())
                    assoc_pct = round(n_assoc / n_resp * 100, 1) if n_resp > 0 else 0.0
                    is_junk   = _is_junk(brand_name)
                    _write_imagery_row(ws, row_idx, attr_code, attr_text, brand_name,
                                       assoc_pct, n_assoc, n_resp, is_junk)
                    row_idx += 1

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(7)}2"
    for col, w in zip("ABCDEFG", [18, 55, 35, 12, 10, 10, 10]):
        ws.column_dimensions[col].width = w
    print(f"  [IMAGERY_TAGGED] {row_idx - 3} rows")


def _write_imagery_row(ws, row_idx, attr_code, attr_text, brand_name,
                       assoc_pct, n_assoc, n_resp, is_junk):
    ws.cell(row=row_idx, column=1, value=attr_code)
    ws.cell(row=row_idx, column=2, value=attr_text)
    ws.cell(row=row_idx, column=3, value=brand_name)
    ws.cell(row=row_idx, column=4, value=assoc_pct)
    ws.cell(row=row_idx, column=5, value=n_assoc)
    ws.cell(row=row_idx, column=6, value=n_resp)
    ws.cell(row=row_idx, column=7, value="TRUE" if is_junk else "FALSE")
    if is_junk and HAS_OPENPYXL:
        for c in range(1, 8):
            ws.cell(row=row_idx, column=c).fill = YELLOW_FILL


def _write_validation(wb, schema_doc, raw_df, n_resp):
    """Awareness funnel % per brand + violation flags."""
    from infoleap.data_layer import compute_brand_awareness

    try:
        aw_df = compute_brand_awareness(raw_df, schema_doc)
    except Exception as e:
        ws = wb.create_sheet("VALIDATION")
        ws.cell(row=1, column=1, value=f"Could not compute: {e}")
        return

    ws = wb.create_sheet("VALIDATION")
    ws.cell(row=1, column=1,
            value="Auto-computed awareness funnel. Reference only — green=ok, red=monotonicity violation.")
    if HAS_OPENPYXL:
        ws.cell(row=1, column=1).font = ITALIC_FONT
    ws.merge_cells("A1:I1")

    avail_stages = [s for s in AWARENESS_STAGES if s in aw_df["stage"].unique()] if len(aw_df) else []
    headers = ["brand"] + [f"{s} %" for s in avail_stages] + ["violations"]
    _header_row(ws, headers, row=2)

    pivot = aw_df.groupby(["brand_name", "stage"])["respondent_id"].nunique().unstack(fill_value=0)
    rules = [
        ("TOM", "SPONT"), ("SPONT", "AIDED"), ("EVER_USED", "AIDED"),
        ("CURRENT_USER", "EVER_USED"), ("PREFERRED", "CURRENT_USER"), ("CONSIDERATION", "AIDED"),
    ]
    for i, brand in enumerate(pivot.index, 3):
        row = pivot.loc[brand]
        pct  = {s: round(row.get(s, 0) / n_resp * 100, 1) for s in avail_stages}
        viols = [
            f"{sm}({pct.get(sm,0)}%)>{lg}({pct.get(lg,0)}%)"
            for sm, lg in rules if pct.get(sm, 0) > pct.get(lg, 0) + 2
        ]
        ws.cell(row=i, column=1, value=brand)
        for j, stage in enumerate(avail_stages, 2):
            cell = ws.cell(row=i, column=j, value=pct.get(stage, 0))
            if HAS_OPENPYXL:
                cell.fill = RED_FILL if viols else GREEN_FILL
        ws.cell(row=i, column=len(avail_stages) + 2, value="; ".join(viols) if viols else "ok")

    ws.freeze_panes = "B3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"
    ws.column_dimensions["A"].width = 28
    for c in range(2, len(avail_stages) + 2):
        ws.column_dimensions[get_column_letter(c)].width = 14
    ws.column_dimensions[get_column_letter(len(avail_stages) + 2)].width = 50


# ---------------------------------------------------------------------------
# Read master Excel → schema_doc (backwards-compatible)
# ---------------------------------------------------------------------------

def read_master_excel(excel_path: str) -> dict:
    """
    Read master_mapping.xlsx → schema_doc.
    Reads BUCKET_CONFIG (new) OR legacy MAPPING sheet.
    OVERRIDE_source_column takes precedence over source_column if filled.
    Brand codes read from BRANDS sheet (new) or legacy BRAND_CODES / CTL_* sheets.
    """
    xl         = pd.ExcelFile(excel_path)
    sheets     = xl.sheet_names

    # META
    meta_df = pd.read_excel(excel_path, sheet_name="META")
    meta    = dict(zip(meta_df.iloc[:, 0].astype(str), meta_df.iloc[:, 1].astype(str)))

    # Build brand code lookup from BRANDS sheet (new) or BRAND_CODES (legacy)
    # qcode → {code_str: brand_name}
    brand_lookup: dict[str, dict[str, str]] = {}

    if "BRANDS" in sheets:
        # New format has instruction row at row 1, header at row 2 → header=1
        # Old format has header at row 1 → header=0. Detect by checking first cell.
        _b_raw = pd.read_excel(excel_path, sheet_name="BRANDS", header=0)
        _first_col = str(_b_raw.columns[0]).lower() if len(_b_raw.columns) else ""
        _b_header  = 1 if "question_code" not in _first_col else 0
        b_df = pd.read_excel(excel_path, sheet_name="BRANDS", header=_b_header)
        for _, row in b_df.iterrows():
            qcode = str(row.get("question_code", "")).strip()
            code  = row.get("brand_code")
            name  = row.get("brand_name")
            if not qcode or pd.isna(code) or pd.isna(name):
                continue
            code_s = str(int(float(code))) if isinstance(code, (int, float)) else str(code)
            name_s = str(name).strip()
            if qcode not in brand_lookup:
                brand_lookup[qcode] = {}
            brand_lookup[qcode][code_s]       = name_s
            brand_lookup[qcode][f"{code_s}.0"] = name_s

    elif "BRAND_CODES" in sheets:
        bc_df = pd.read_excel(excel_path, sheet_name="BRAND_CODES")
        for _, row in bc_df.iterrows():
            qcode = str(row.get("question_code", "")).strip()
            code  = row.get("code")
            name  = row.get("brand_name")
            if not qcode or pd.isna(code) or pd.isna(name):
                continue
            code_s = str(int(float(code))) if isinstance(code, (int, float)) else str(code)
            name_s = str(name).strip()
            if qcode not in brand_lookup:
                brand_lookup[qcode] = {}
            brand_lookup[qcode][code_s]       = name_s
            brand_lookup[qcode][f"{code_s}.0"] = name_s

    # Choose mapping sheet: prefer BUCKET_CONFIG, fall back to MAPPING
    map_sheet = "BUCKET_CONFIG" if "BUCKET_CONFIG" in sheets else "MAPPING"
    # BUCKET_CONFIG has instruction row 1 + header row 2 → header=1
    # MAPPING (legacy) has header at row 1 (Excel row 1, pandas index 0) → header=0
    header_row = 0 if map_sheet == "MAPPING" else 1
    map_df = pd.read_excel(excel_path, sheet_name=map_sheet, header=header_row)

    # Normalise column names (BUCKET_CONFIG / MAPPING have different col sets)
    map_df.columns = [str(c).strip() for c in map_df.columns]

    questions = []
    seen: set[str] = set()
    for _, row in map_df.iterrows():
        qcode = str(row.get("question_code", "")).strip()
        if not qcode or qcode == "nan":
            continue
        # Deduplicate by source_column (not question_code) to preserve per-attribute rows
        # e.g. q34_1_1 / q34_2_1 are both question_code=q34 but different source columns
        def _clean(v):
            s = str(v).strip() if v is not None else ""
            return "" if s.lower() == "nan" else s
        src_for_dedup = (
            _clean(row.get("OVERRIDE_source_column")) or
            _clean(row.get("source_column")) or
            qcode
        )
        if src_for_dedup in seen:
            continue
        seen.add(src_for_dedup)

        # Bucket — override wins
        bucket = _clean(row.get("OVERRIDE_bucket")) or _clean(row.get("bucket")) or "SKIP"

        # Source column — override wins; use _clean() to strip NaN floats correctly
        src_col = (
            _clean(row.get("OVERRIDE_source_column")) or
            _clean(row.get("source_column")) or
            qcode
        )

        # Dummy columns
        dummy_raw = _clean(row.get("dummy_columns", ""))
        dummy_cols = [d.strip() for d in dummy_raw.split(";") if d.strip() and d.strip() != "nan"]

        # CTL: new BRANDS sheet > legacy BRAND_CODES > legacy CTL_* > empty
        ctl = brand_lookup.get(qcode, {})
        if not ctl:
            ctl_sheet = str(row.get("ctl_sheet", "") or "").strip()
            if ctl_sheet and ctl_sheet != "nan" and ctl_sheet in sheets:
                ctl_df = pd.read_excel(excel_path, sheet_name=ctl_sheet)
                for _, crow in ctl_df.iterrows():
                    code = crow.get("code")
                    name = crow.get("brand_name")
                    if pd.notna(code) and pd.notna(name):
                        cs = str(int(float(code))) if isinstance(code, (int, float)) else str(code)
                        ns = str(name).strip()
                        ctl[cs]       = ns
                        ctl[f"{cs}.0"] = ns

        conf_raw = row.get("confidence")
        try:
            confidence = float(conf_raw) if pd.notna(conf_raw) else 1.0
        except (ValueError, TypeError):
            confidence = 1.0

        questions.append({
            "question_code":  qcode,
            "question_text":  _clean(row.get("question_text", "")),
            "bucket":         bucket,
            "shape":          _clean(row.get("shape", "")) or "single_value",
            "source_column":  src_col,
            "dummy_columns":  dummy_cols,
            "code_to_label":  ctl,
            "confidence":     confidence,
            "notes":          _clean(row.get("notes", "")),
        })

    return {
        "project_id":      meta.get("project_id", "unknown"),
        "questions":       questions,
        "_source":         "master_excel",
        "_raw_data_path":  meta.get("raw_data_path", ""),
    }
