import json
import pathlib
import jsonschema
import pytest
import pandas as pd
from infoleap.ingestion.schema_ingest import build_context_packet

SCHEMA_PATH = pathlib.Path(__file__).parent.parent / "schemas" / "ingestion_mapping.schema.json"


def _load_schema():
    return json.loads(SCHEMA_PATH.read_text())


def test_valid_multivalent_question_passes():
    doc = {
        "project_id": "akshayakalpa",
        "questions": [
            {
                "question_code": "q17", "question_text": "Which brands are you aware of?",
                "bucket": "AIDED", "shape": "multivalent_source", "source_column": "q17",
                "delimiter": " ", "dummy_columns": [f"q17_{i}" for i in range(1, 18)],
                "code_to_label": {"1": "Amul", "2": "Nandini"},
                "confidence": 0.92, "reasoning": "q17 holds space-delimited numeric codes matching the Values sheet's brand list; q17_1..17 are one-hot dummies of the same codes.",
            }
        ],
    }
    jsonschema.validate(doc, _load_schema())


def test_missing_source_column_fails():
    doc = {
        "project_id": "akshayakalpa",
        "questions": [
            {
                "question_code": "q17", "question_text": "x", "bucket": "AIDED",
                "shape": "multivalent_source", "confidence": 0.9, "reasoning": "x",
            }
        ],
    }
    with pytest.raises(jsonschema.ValidationError):
        jsonschema.validate(doc, _load_schema())


def test_multivalent_with_null_delimiter_fails():
    doc = {
        "project_id": "akshayakalpa",
        "questions": [
            {
                "question_code": "q17", "question_text": "Which brands are you aware of?",
                "bucket": "AIDED", "shape": "multivalent_source", "source_column": "q17",
                "delimiter": None, "dummy_columns": [f"q17_{i}" for i in range(1, 18)],
                "code_to_label": {"1": "Amul", "2": "Nandini"},
                "confidence": 0.92, "reasoning": "x",
            }
        ],
    }
    with pytest.raises(jsonschema.ValidationError):
        jsonschema.validate(doc, _load_schema())


def test_brand_bucket_missing_code_to_label_warns_not_raises():
    """2026-08-03: schema relaxed â€” code_to_label is now advisory (not schema-enforced) for
    brand-identity buckets, because all 4 fallback models returned brand questions with missing
    code_to_label and the strict minProperties:1 rule hard-failed the whole batch.
    The defensive repair in _classify_with_single_model now injects {} and adds a warning
    instead. Schema should NOT raise for a brand bucket missing code_to_label."""
    doc = {
        "project_id": "akshayakalpa",
        "questions": [
            {
                "question_code": "q17", "question_text": "Which brands are you aware of?",
                "bucket": "AIDED", "shape": "single_value", "source_column": "q17",
                "delimiter": None, "dummy_columns": [],
                "confidence": 0.92, "reasoning": "x",
            }
        ],
    }
    jsonschema.validate(doc, _load_schema())  # must NOT raise


def test_multivalent_with_delimiter_omitted_fails():
    doc = {
        "project_id": "akshayakalpa",
        "questions": [
            {
                "question_code": "q17", "question_text": "Which brands are you aware of?",
                "bucket": "AIDED", "shape": "multivalent_source", "source_column": "q17",
                "dummy_columns": [f"q17_{i}" for i in range(1, 18)],
                "code_to_label": {"1": "Amul", "2": "Nandini"},
                "confidence": 0.92, "reasoning": "x",
            }
        ],
    }
    with pytest.raises(jsonschema.ValidationError):
        jsonschema.validate(doc, _load_schema())


def test_build_context_packet_includes_all_columns_but_sampled_rows():
    df = pd.DataFrame({
        "q17": ["2 5", "1 3", "2"] * 40,   # 120 rows
        "q17_1": [0, 1, 0] * 40,
        "q17_2": [1, 0, 1] * 40,
    })
    packet = build_context_packet(
        ap_text="AP TABPLAN CONTENT HERE", datamap_text="DATAMAP CONTENT HERE",
        data_df=df, sample_rows=15,
    )
    assert "AP TABPLAN CONTENT HERE" in packet["ap_text"]
    assert "DATAMAP CONTENT HERE" in packet["datamap_text"]
    assert packet["raw_sample_columns"] == ["q17", "q17_1", "q17_2"]
    assert len(packet["raw_sample_rows"]) == 15


def test_build_context_packet_raw_sample_rows_is_list_of_lists_paired_positionally():
    df = pd.DataFrame({
        "q17": ["2 5", "1 3"],
        "q17_1": [0, 1],
        "q17_2": [1, 0],
    })
    packet = build_context_packet(ap_text="x", datamap_text="y", data_df=df, sample_rows=2)
    assert isinstance(packet["raw_sample_rows"], list)
    for row in packet["raw_sample_rows"]:
        assert isinstance(row, list)
        assert not isinstance(row, dict)
        assert len(row) == len(packet["raw_sample_columns"])
    assert packet["raw_sample_columns"] == ["q17", "q17_1", "q17_2"]
    assert packet["raw_sample_rows"][0] == ["2 5", 0, 1]
    assert packet["raw_sample_rows"][1] == ["1 3", 1, 0]


from infoleap.ingestion.schema_ingest import flatten_datamap_to_text


def test_flatten_datamap_to_text_includes_variable_and_value_rows(tmp_path):
    import openpyxl
    wb = openpyxl.Workbook()
    var_sheet = wb.active
    var_sheet.title = "Variables"
    var_sheet.append(["Variable", "Label"])
    var_sheet.append(["q17_1", "Amul aware"])
    val_sheet = wb.create_sheet("Values")
    val_sheet.append(["Variable", "Value", "Label"])
    val_sheet.append(["q17_1", 1, "Yes"])
    path = tmp_path / "datamap.xlsx"
    wb.save(path)

    text = flatten_datamap_to_text(str(path), "Variables", "Values")
    assert "q17_1" in text
    assert "Amul aware" in text
    assert "Yes" in text


def test_flatten_datamap_to_text_ffills_blank_variable_rows(tmp_path):
    import openpyxl
    wb = openpyxl.Workbook()
    var_sheet = wb.active
    var_sheet.title = "Variables"
    var_sheet.append(["Variable", "Label"])
    var_sheet.append(["q17", "Which brands are you aware of?"])
    val_sheet = wb.create_sheet("Values")
    val_sheet.append(["Variable", "Value", "Label"])
    val_sheet.append(["q17", 1, "Amul"])
    val_sheet.append([None, 2, "Nandini"])
    val_sheet.append([None, 3, "Nestle"])
    path = tmp_path / "datamap.xlsx"
    wb.save(path)

    text = flatten_datamap_to_text(str(path), "Variables", "Values")
    lines = [ln for ln in text.splitlines() if "Nandini" in ln or "Nestle" in ln]
    assert len(lines) == 2
    for ln in lines:
        assert "q17" in ln


def test_flatten_datamap_to_text_missing_sheet_raises(tmp_path):
    import openpyxl
    wb = openpyxl.Workbook()
    wb.active.title = "SomeOtherSheet"
    path = tmp_path / "datamap.xlsx"
    wb.save(path)
    with pytest.raises(ValueError, match="have:"):
        flatten_datamap_to_text(str(path), "Variables", "Values")


def test_flatten_datamap_to_text_missing_file_raises_valueerror():
    with pytest.raises(ValueError):
        flatten_datamap_to_text("nonexistent_datamap_file_xyz.xlsx", "Variables", "Values")


from unittest.mock import patch
from infoleap.ingestion.schema_ingest import classify_all_questions


def test_classify_all_questions_returns_validated_dict():
    fake_response_body = json.dumps({
        "project_id": "akshayakalpa",
        "questions": [
            {
                "question_code": "q17", "question_text": "Aware of which brands?",
                "bucket": "AIDED", "shape": "multivalent_source", "source_column": "q17",
                "delimiter": " ", "dummy_columns": ["q17_1", "q17_2"],
                "code_to_label": {"1": "Amul", "2": "Nandini"},
                "confidence": 0.9, "reasoning": "test reasoning",
            }
        ],
    })

    class FakeResp:
        def read(self):
            return json.dumps({"choices": [{"message": {"content": fake_response_body}}]}).encode()

        def __enter__(self):
            return self

        def __exit__(self, *a):
            return False

    with patch("lens.ingestion.schema_ingest.urlopen", return_value=FakeResp()):
        result = classify_all_questions(
            packet={"ap_text": "x", "datamap_text": "y", "raw_sample_columns": ["q17"], "raw_sample_rows": []},
            project_id="akshayakalpa", api_key="fake-key", model="fake-model",
        )
    assert result["questions"][0]["bucket"] == "AIDED"
    assert result["questions"][0]["source_column"] == "q17"


def test_classify_all_questions_malformed_response_raises_runtimeerror():
    class FakeResp:
        def read(self):
            return b"not json at all {{{"

        def __enter__(self):
            return self

        def __exit__(self, *a):
            return False

    with patch("lens.ingestion.schema_ingest.urlopen", return_value=FakeResp()):
        with pytest.raises(RuntimeError):
            classify_all_questions(
                packet={"ap_text": "x", "datamap_text": "y", "raw_sample_columns": ["q17"], "raw_sample_rows": []},
                project_id="akshayakalpa", api_key="fake-key", model="fake-model",
            )


def test_build_context_packet_handles_datetime_columns():
    df = pd.DataFrame({
        "q1": ["a", "b"],
        "survey_date": pd.to_datetime(["2026-01-01", "2026-01-02"]),
    })
    packet = build_context_packet(ap_text="x", datamap_text="y", data_df=df, sample_rows=2)
    json.dumps(packet["raw_sample_rows"])  # must not raise


def test_classify_all_questions_raises_valueerror_before_network_call_when_oversized():
    huge_packet = {
        "ap_text": "x",
        "datamap_text": "y" * 1_000_000,  # ~250K+ estimated tokens, way over default budget
        "raw_sample_columns": ["q17"],
        "raw_sample_rows": [],
    }
    with patch("lens.ingestion.schema_ingest.urlopen") as mock_urlopen:
        with pytest.raises(ValueError):
            classify_all_questions(
                packet=huge_packet, project_id="akshayakalpa",
                api_key="fake-key", model="fake-model",
            )
        mock_urlopen.assert_not_called()


def test_classify_all_questions_valueerror_reflects_tokenizer_inflation_correction():
    """The raw chars/4 estimate for this packet is UNDER the 100K budget on its own â€”
    only the corrected (x2.35) estimate exceeds it. This proves the guard is applying
    the inflation correction, not just the old uncorrected chars/4 check.
    """
    # ~200K chars -> raw chars/4 estimate ~= 50K (under 100K budget on its own),
    # but 50K * 2.35 ~= 117.5K (over the 100K budget).
    packet = {
        "ap_text": "x",
        "datamap_text": "y" * 200_000,
        "raw_sample_columns": ["q17"],
        "raw_sample_rows": [],
    }
    raw_estimate = len(packet["datamap_text"]) // 4
    assert raw_estimate < 100_000, "test setup assumption: raw estimate must be under budget"
    assert raw_estimate * 2.35 > 100_000, "test setup assumption: corrected estimate must exceed budget"

    with patch("lens.ingestion.schema_ingest.urlopen") as mock_urlopen:
        with pytest.raises(ValueError):
            classify_all_questions(
                packet=packet, project_id="akshayakalpa",
                api_key="fake-key", model="fake-model",
            )
        mock_urlopen.assert_not_called()


def test_classify_all_questions_raises_runtimeerror_when_response_badly_incomplete():
    """Regression for the live Akshayakalpa gap: gpt-4o-mini returned 10 questions for a ~300-
    question survey with no error. Simulate a large synthetic raw_sample_columns (many distinct
    stems) but a deliberately short LLM response (few questions) and assert the completeness
    guard raises a RuntimeError naming both counts.
    """
    # 50 distinct question stems, well over the completeness-check floor.
    many_columns = [f"q{i}" for i in range(1, 51)]

    fake_response_body = json.dumps({
        "project_id": "akshayakalpa",
        "questions": [
            {
                "question_code": "q1", "question_text": "x", "bucket": "GENDER",
                "shape": "single_value", "source_column": "q1", "confidence": 0.9,
                "reasoning": "test",
            }
        ],  # only 1 question returned out of 50 stems
    })

    class FakeResp:
        def read(self):
            return json.dumps({"choices": [{"message": {"content": fake_response_body}}]}).encode()

        def __enter__(self):
            return self

        def __exit__(self, *a):
            return False

    with patch("lens.ingestion.schema_ingest.urlopen", return_value=FakeResp()):
        with pytest.raises(RuntimeError, match=r"only 1 question.*~50.*stems"):
            classify_all_questions(
                packet={"ap_text": "x", "datamap_text": "y",
                        "raw_sample_columns": many_columns, "raw_sample_rows": []},
                project_id="akshayakalpa", api_key="fake-key", model="fake-model",
            )


def test_classify_all_questions_small_fixture_does_not_trip_completeness_check():
    """A genuinely tiny test fixture (well under the completeness-check floor of 5 distinct
    stems) must NOT trigger the completeness guard even though it only returns 1 question â€”
    this is the existing happy-path shape, kept working."""
    fake_response_body = json.dumps({
        "project_id": "akshayakalpa",
        "questions": [
            {
                "question_code": "q17", "question_text": "Aware of which brands?",
                "bucket": "AIDED", "shape": "multivalent_source", "source_column": "q17",
                "delimiter": " ", "dummy_columns": ["q17_1", "q17_2"],
                "code_to_label": {"1": "Amul", "2": "Nandini"},
                "confidence": 0.9, "reasoning": "test reasoning",
            }
        ],
    })

    class FakeResp:
        def read(self):
            return json.dumps({"choices": [{"message": {"content": fake_response_body}}]}).encode()

        def __enter__(self):
            return self

        def __exit__(self, *a):
            return False

    with patch("lens.ingestion.schema_ingest.urlopen", return_value=FakeResp()):
        result = classify_all_questions(
            packet={"ap_text": "x", "datamap_text": "y",
                    "raw_sample_columns": ["q17", "q17_1", "q17_2"], "raw_sample_rows": []},
            project_id="akshayakalpa", api_key="fake-key", model="fake-model",
        )
    assert result["questions"][0]["bucket"] == "AIDED"


def test_classify_all_questions_downgrades_multi_select_dummies_with_empty_dummy_columns():
    """Regression for the live Akshayakalpa gap: gpt-4o-mini picked shape='multi_select_dummies'
    with an empty dummy_columns array, which the schema correctly rejects via minItems. Rather
    than hard-failing the whole batch, classify_all_questions should downgrade that one question
    to 'single_value' and record a warning, letting the rest of the (potentially large) response
    through.
    """
    fake_response_body = json.dumps({
        "project_id": "akshayakalpa",
        "questions": [
            {
                "question_code": "q67", "question_text": "Reasons for not trying",
                "bucket": "ATTITUDE", "shape": "multi_select_dummies", "source_column": "q67",
                "dummy_columns": [], "confidence": 0.7, "reasoning": "test",
            }
        ],
    })

    class FakeResp:
        def read(self):
            return json.dumps({"choices": [{"message": {"content": fake_response_body}}]}).encode()

        def __enter__(self):
            return self

        def __exit__(self, *a):
            return False

    with patch("lens.ingestion.schema_ingest.urlopen", return_value=FakeResp()):
        result = classify_all_questions(
            packet={"ap_text": "x", "datamap_text": "y", "raw_sample_columns": ["q67"], "raw_sample_rows": []},
            project_id="akshayakalpa", api_key="fake-key", model="fake-model",
        )
    assert result["questions"][0]["shape"] == "single_value"
    assert result["questions"][0]["delimiter"] is None
    assert any("q67" in w for w in result.get("_downgrade_warnings", []))


def test_classify_all_questions_raises_clear_error_on_bucket_shape_field_swap():
    """Regression for the live Akshayakalpa gap: gpt-4o-mini put 'multi_select_dummies' (a shape
    value) into the bucket field. classify_all_questions should raise a clear RuntimeError naming
    the likely cause (field swap) rather than letting the raw jsonschema enum error surface.
    """
    fake_response_body = json.dumps({
        "project_id": "akshayakalpa",
        "questions": [
            {
                "question_code": "q67", "question_text": "x",
                "bucket": "multi_select_dummies", "shape": "single_value",
                "source_column": "q67", "confidence": 0.8, "reasoning": "test",
            }
        ],
    })

    class FakeResp:
        def read(self):
            return json.dumps({"choices": [{"message": {"content": fake_response_body}}]}).encode()

        def __enter__(self):
            return self

        def __exit__(self, *a):
            return False

    with patch("lens.ingestion.schema_ingest.urlopen", return_value=FakeResp()):
        with pytest.raises(RuntimeError, match=r"swapped the bucket/shape"):
            classify_all_questions(
                packet={"ap_text": "x", "datamap_text": "y", "raw_sample_columns": ["q67"], "raw_sample_rows": []},
                project_id="akshayakalpa", api_key="fake-key", model="fake-model",
            )


def test_classify_all_questions_injects_project_id_when_llm_omits_it():
    """Regression for the live Akshayakalpa gap: the LLM omitted the required top-level
    project_id field entirely. classify_all_questions should inject the caller's own project_id
    parameter unconditionally rather than requiring/trusting the LLM's echo.
    """
    fake_response_body = json.dumps({
        # NOTE: no "project_id" key at all here.
        "questions": [
            {
                "question_code": "q17", "question_text": "Aware of which brands?",
                "bucket": "AIDED", "shape": "multivalent_source", "source_column": "q17",
                "delimiter": " ", "dummy_columns": ["q17_1", "q17_2"],
                "code_to_label": {"1": "Amul", "2": "Nandini"},
                "confidence": 0.9, "reasoning": "test reasoning",
            }
        ],
    })

    class FakeResp:
        def read(self):
            return json.dumps({"choices": [{"message": {"content": fake_response_body}}]}).encode()

        def __enter__(self):
            return self

        def __exit__(self, *a):
            return False

    with patch("lens.ingestion.schema_ingest.urlopen", return_value=FakeResp()):
        result = classify_all_questions(
            packet={"ap_text": "x", "datamap_text": "y", "raw_sample_columns": ["q17"], "raw_sample_rows": []},
            project_id="akshayakalpa", api_key="fake-key", model="fake-model",
        )
    assert result["project_id"] == "akshayakalpa"


def test_classify_all_questions_overwrites_llm_supplied_project_id():
    """Even if the LLM DOES echo a project_id back, the caller's own parameter must win â€” never
    trust the LLM's echo over the authoritative value already known to the caller."""
    fake_response_body = json.dumps({
        "project_id": "some_wrong_value_the_llm_made_up",
        "questions": [
            {
                "question_code": "q17", "question_text": "x", "bucket": "GENDER",
                "shape": "single_value", "source_column": "q17",
                "confidence": 0.9, "reasoning": "test",
            }
        ],
    })

    class FakeResp:
        def read(self):
            return json.dumps({"choices": [{"message": {"content": fake_response_body}}]}).encode()

        def __enter__(self):
            return self

        def __exit__(self, *a):
            return False

    with patch("lens.ingestion.schema_ingest.urlopen", return_value=FakeResp()):
        result = classify_all_questions(
            packet={"ap_text": "x", "datamap_text": "y", "raw_sample_columns": ["q17"], "raw_sample_rows": []},
            project_id="akshayakalpa", api_key="fake-key", model="fake-model",
        )
    assert result["project_id"] == "akshayakalpa"


def test_classify_all_questions_missing_questions_key_includes_raw_result_dump():
    """Regression for the live Akshayakalpa gap: the parsed JSON genuinely lacked the top-level
    `questions` key (a clean jsonschema 'required' miss, not a JSONDecodeError). The resulting
    RuntimeError must include a dump of the actual parsed result so the real response shape is
    visible without burning another paid API call to re-diagnose.
    """
    fake_response_body = json.dumps({"project_id": "x"})  # no "questions" key at all

    class FakeResp:
        def read(self):
            return json.dumps({"choices": [{"message": {"content": fake_response_body}}]}).encode()

        def __enter__(self):
            return self

        def __exit__(self, *a):
            return False

    with patch("lens.ingestion.schema_ingest.urlopen", return_value=FakeResp()):
        with pytest.raises(RuntimeError, match=r"'questions' is a required property.*Parsed result"):
            classify_all_questions(
                packet={"ap_text": "x", "datamap_text": "y", "raw_sample_columns": ["q17"], "raw_sample_rows": []},
                project_id="akshayakalpa", api_key="fake-key", model="fake-model",
            )


def test_classify_all_questions_auto_unwraps_schema_echo_response():
    """Regression for the live Akshayakalpa gap: the LLM returned the JSON Schema's own
    structure (title/type/properties) as the top-level object, with the real data instance
    nested one level down inside result["properties"]. classify_all_questions must detect and
    auto-unwrap this exact shape, recovering the real questions/project_id data.
    """
    fake_response_body = json.dumps({
        "title": "IngestionMapping",
        "type": "object",
        "properties": {
            "project_id": "akshayakalpa_uitest8",
            "questions": [
                {
                    "question_code": "q2", "question_text": "current city",
                    "bucket": "CITY", "shape": "single_value", "source_column": "q2",
                    "delimiter": None, "dummy_columns": [],
                    "code_to_label": {"1": "Bangalore", "2": "Chennai"},
                    "confidence": 1, "reasoning": "test",
                }
            ],
        },
    })

    class FakeResp:
        def read(self):
            return json.dumps({"choices": [{"message": {"content": fake_response_body}}]}).encode()

        def __enter__(self):
            return self

        def __exit__(self, *a):
            return False

    with patch("lens.ingestion.schema_ingest.urlopen", return_value=FakeResp()):
        result = classify_all_questions(
            packet={"ap_text": "x", "datamap_text": "y", "raw_sample_columns": ["q2"], "raw_sample_rows": []},
            project_id="akshayakalpa_uitest8", api_key="fake-key", model="fake-model",
        )
    assert "questions" in result
    assert result["questions"][0]["question_code"] == "q2"
    assert result["questions"][0]["bucket"] == "CITY"
    assert result["project_id"] == "akshayakalpa_uitest8"
    assert "title" not in result
    assert "properties" not in result


def test_classify_all_questions_defaults_missing_question_text_to_question_code():
    """Regression for the live Akshayakalpa gap: one question in an otherwise-good response
    omitted question_text entirely. question_text is display-only (never read by
    assignment_from_schema/load_confirmed_assignment) so it's now optional in the schema â€”
    classify_all_questions should still validate successfully and default_fill it from
    question_code so the review table never shows a blank.
    """
    fake_response_body = json.dumps({
        "project_id": "akshayakalpa",
        "questions": [
            {
                "question_code": "q2", "bucket": "CITY", "shape": "single_value",
                "source_column": "q2", "delimiter": None, "dummy_columns": [],
                "code_to_label": {"1": "Bangalore", "2": "Chennai"},
                "confidence": 1, "reasoning": "test",
                # no "question_text" key at all
            }
        ],
    })

    class FakeResp:
        def read(self):
            return json.dumps({"choices": [{"message": {"content": fake_response_body}}]}).encode()

        def __enter__(self):
            return self

        def __exit__(self, *a):
            return False

    with patch("lens.ingestion.schema_ingest.urlopen", return_value=FakeResp()):
        result = classify_all_questions(
            packet={"ap_text": "x", "datamap_text": "y", "raw_sample_columns": ["q2"], "raw_sample_rows": []},
            project_id="akshayakalpa", api_key="fake-key", model="fake-model",
        )
    assert result["questions"][0]["question_text"] == "q2"


def test_classify_all_questions_falls_back_to_next_model_on_failure():
    """New feature: model may be a list of model-ids tried in order. If the first model's
    response fails validation, the function should retry with the next model in the list rather
    than failing the whole classification outright.
    """
    good_body = json.dumps({
        "project_id": "akshayakalpa",
        "questions": [
            {
                "question_code": "q17", "question_text": "Aware of which brands?",
                "bucket": "AIDED", "shape": "multivalent_source", "source_column": "q17",
                "delimiter": " ", "dummy_columns": ["q17_1", "q17_2"],
                "code_to_label": {"1": "Amul", "2": "Nandini"},
                "confidence": 0.9, "reasoning": "test reasoning",
            }
        ],
    })
    bad_body = "not json at all {{{"

    call_log = []

    def fake_urlopen(req, timeout=None):
        import json as _json
        body = _json.loads(req.data)
        model_used = body["model"]
        call_log.append(model_used)

        class FakeResp:
            def read(self_inner):
                if model_used == "model-a-bad":
                    return bad_body.encode()
                return json.dumps({"choices": [{"message": {"content": good_body}}]}).encode()

            def __enter__(self_inner):
                return self_inner

            def __exit__(self_inner, *a):
                return False

        return FakeResp()

    with patch("lens.ingestion.schema_ingest.urlopen", side_effect=fake_urlopen):
        result = classify_all_questions(
            packet={"ap_text": "x", "datamap_text": "y", "raw_sample_columns": ["q17"], "raw_sample_rows": []},
            project_id="akshayakalpa", api_key="fake-key",
            model=["model-a-bad", "model-b-good"],
        )
    assert call_log == ["model-a-bad", "model-b-good"]
    assert result["_model_used"] == "model-b-good"
    assert result["questions"][0]["bucket"] == "AIDED"


def test_classify_all_questions_all_models_fail_raises_aggregated_error():
    """If every model in the fallback list fails, classify_all_questions must raise ONE
    RuntimeError listing every model tried and its specific error."""
    class FakeResp:
        def read(self):
            return b"not json at all {{{"

        def __enter__(self):
            return self

        def __exit__(self, *a):
            return False

    with patch("lens.ingestion.schema_ingest.urlopen", return_value=FakeResp()):
        with pytest.raises(RuntimeError, match=r"ALL 2 model\(s\) failed.*model-a.*model-b"):
            classify_all_questions(
                packet={"ap_text": "x", "datamap_text": "y", "raw_sample_columns": ["q17"], "raw_sample_rows": []},
                project_id="akshayakalpa", api_key="fake-key",
                model=["model-a", "model-b"],
            )


def test_classify_all_questions_size_guard_does_not_trigger_fallback():
    """The size-guard ValueError (context packet too large) is model-independent and must raise
    IMMEDIATELY without trying any models in the fallback list â€” a bigger prompt problem won't be
    fixed by switching models."""
    huge_packet = {
        "ap_text": "x",
        "datamap_text": "y" * 1_000_000,
        "raw_sample_columns": ["q17"],
        "raw_sample_rows": [],
    }
    with patch("lens.ingestion.schema_ingest.urlopen") as mock_urlopen:
        with pytest.raises(ValueError, match="too large"):
            classify_all_questions(
                packet=huge_packet, project_id="akshayakalpa", api_key="fake-key",
                model=["model-a", "model-b"],
            )
        mock_urlopen.assert_not_called()


def test_invalid_bucket_enum_fails():
    doc = {
        "project_id": "akshayakalpa",
        "questions": [
            {
                "question_code": "q17", "question_text": "x",
                "bucket": "NOT_A_BUCKET", "shape": "single_value", "source_column": "q17",
                "confidence": 0.9, "reasoning": "x",
            }
        ],
    }
    with pytest.raises(jsonschema.ValidationError):
        jsonschema.validate(doc, _load_schema())


def test_valid_multi_select_dummies_question_passes():
    doc = {
        "project_id": "akshayakalpa",
        "questions": [
            {
                "question_code": "q67",
                "question_text": "Reasons for not trying the AKSHAYAKALPA milk brand",
                "bucket": "ATTITUDE", "shape": "multi_select_dummies", "source_column": "q67",
                "dummy_columns": [f"q67_{i}" for i in range(1, 14)],
                "code_to_label": {}, "confidence": 1.0,
                "reasoning": "Pure one-hot dummy battery, no combined delimited column found in raw sample.",
            }
        ],
    }
    jsonschema.validate(doc, _load_schema())


def test_multi_select_dummies_missing_dummy_columns_fails():
    doc = {
        "project_id": "akshayakalpa",
        "questions": [
            {
                "question_code": "q67", "question_text": "x",
                "bucket": "ATTITUDE", "shape": "multi_select_dummies", "source_column": "q67",
                "dummy_columns": [], "confidence": 1.0, "reasoning": "x",
            }
        ],
    }
    with pytest.raises(jsonschema.ValidationError):
        jsonschema.validate(doc, _load_schema())


from infoleap.ingestion.generic_loader import assignment_from_schema


def test_assignment_from_schema_maps_multivalent_source_correctly():
    schema_doc = {
        "project_id": "akshayakalpa",
        "questions": [
            {
                "question_code": "q17", "question_text": "Aware of which brands?",
                "bucket": "AIDED", "shape": "multivalent_source", "source_column": "q17",
                "delimiter": " ", "dummy_columns": ["q17_1", "q17_2"],
                "code_to_label": {"1": "Amul", "2": "Nandini"},
                "confidence": 0.9, "reasoning": "test",
            },
            {
                "question_code": "q99", "question_text": "skip me", "bucket": "SKIP",
                "shape": "single_value", "source_column": "q99", "delimiter": None,
                "dummy_columns": [], "code_to_label": {}, "confidence": 0.5, "reasoning": "not needed",
            },
        ],
    }
    result = assignment_from_schema(schema_doc)
    assert "q99" not in result["value_labels_by_code"]   # SKIP bucket excluded entirely
    aided_questions = result["assignment"]["AIDED"]
    assert len(aided_questions) == 1
    q = aided_questions[0]
    assert q["question_code"] == "q17"
    assert q["data_columns"] == ["q17"]                  # NOT the dummy columns
    assert result["value_labels_by_code"]["q17"] == {"1": "Amul", "2": "Nandini"}
    assert result["delimiter_by_code"]["q17"] == " "
    # confirmed via grep of generic_loader.py: shape_by_code values are 'single' /
    # 'multi_select' / 'delimited_multi_select' / 'battery_member' â€” NOT the schema's own
    # enum names. multivalent_source (delimiter-based) maps to 'delimited_multi_select'.
    assert result["shape_by_code"]["q17"] == "delimited_multi_select"


def test_assignment_from_schema_maps_multi_select_dummies_to_dummy_columns():
    schema_doc = {
        "project_id": "akshayakalpa",
        "questions": [
            {
                "question_code": "q67",
                "question_text": "Reasons for not trying the AKSHAYAKALPA milk brand",
                "bucket": "ATTITUDE", "shape": "multi_select_dummies", "source_column": "q67",
                "dummy_columns": [f"q67_{i}" for i in range(1, 14)],
                "code_to_label": {}, "confidence": 1.0, "reasoning": "test",
            },
        ],
    }
    result = assignment_from_schema(schema_doc)
    q = result["assignment"]["ATTITUDE"][0]
    assert q["question_code"] == "q67"
    # data_columns must be the dummy_columns list, NOT [source_column] â€” the real data lives
    # across the independent one-hot columns, source_column ("q67") is nominal-only.
    assert q["data_columns"] == [f"q67_{i}" for i in range(1, 14)]
    assert result["shape_by_code"]["q67"] == "multi_select"
