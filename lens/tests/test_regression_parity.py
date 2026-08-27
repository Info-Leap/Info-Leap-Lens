"""
Regression parity test — verifies oxdata/r_scripts/logistic_regression.R reproduces
XLSTAT's fitted model on real workbook data to 4 decimal places.

Ground truth: .regression_reference/AUDIT.md §2, extracted from
.regression_reference/extracted/blush regression_logistical.xlsm, sheet
"Log(Binary) all" (n=3745, depvar ~ 10 binary imagery attributes).
"""
import sys
import os
sys.path.append(os.getcwd())

import pytest
import openpyxl
import pandas as pd

from oxdata.skills.r_bridge import run_r_stat, r_available

WORKBOOK = os.path.join(
    ".regression_reference", "extracted", "blush regression_logistical.xlsm"
)

# XLSTAT ground truth — .regression_reference/AUDIT.md §2
EXPECTED_FIT = {
    "mcfadden_r2": 0.33192,
    "cox_snell_r2": 0.36877,
    "nagelkerke_r2": 0.49105,
    "aic": 3490.0929,
    "bic": 3558.6028,
    "accuracy": 0.7666,
}

# attribute -> (coef, std_error, odds_ratio) from AUDIT.md §2 / XLSTAT "Model parameters"
EXPECTED_DRIVERS = {
    "Offers.long.lasting.makeup": (0.79150, 0.09542, 2.2067),
}

TOL = 1e-4
# Nagelkerke R2 involves an extra division (Cox-Snell / max-Cox-Snell) that compounds
# rounding — AUDIT.md §2 documents this field as matching to ~3dp, not the exact 4dp
# of every other field. Everything else uses the tighter TOL.
FIELD_TOL = {"nagelkerke_r2": 1e-3}


def _load_blush_df() -> pd.DataFrame:
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True, keep_vba=True)
    ws = wb["tomspontaided"]
    header = [ws.cell(row=1, column=c).value for c in range(1, 17)]
    rows = [
        [ws.cell(row=r, column=c).value for c in range(1, 17)]
        for r in range(2, ws.max_row + 1)
    ]
    df = pd.DataFrame(rows, columns=header).rename(columns={"depvar": "nps_score"})
    predictor_cols = header[6:16]
    df = df[["nps_score"] + predictor_cols].dropna()
    df["nps_score"] = df["nps_score"].astype(int)
    for c in predictor_cols:
        df[c] = df[c].astype(int)
    return df


@pytest.fixture(scope="module")
def blush_result():
    if not r_available():
        pytest.skip("Rscript.exe not available on this host")
    df = _load_blush_df()
    res = run_r_stat("logistic_regression", df)
    if "error" in res:
        pytest.fail(f"R script returned an error: {res['error']}")
    return res


def test_n_matches_workbook(blush_result):
    assert blush_result["n"] == 3745, (
        f"n mismatch: expected 3745 (blush workbook row count), got {blush_result['n']}"
    )


def test_goodness_of_fit_matches_xlstat(blush_result):
    for field, expected in EXPECTED_FIT.items():
        actual = blush_result.get(field)
        tol = FIELD_TOL.get(field, TOL)
        assert actual is not None, f"{field} missing from R result"
        assert actual == pytest.approx(expected, abs=tol), (
            f"{field} mismatch: XLSTAT={expected} InfoLeap={actual} "
            f"(diff={abs(actual - expected):.6f}, tolerance={tol})"
        )


def test_driver_coefficients_match_xlstat(blush_result):
    drivers = {d["attribute"]: d for d in blush_result["significant_drivers"]}
    for attr, (exp_coef, exp_se, exp_or) in EXPECTED_DRIVERS.items():
        assert attr in drivers, f"driver '{attr}' missing from R result"
        d = drivers[attr]
        assert d["coef"] == pytest.approx(exp_coef, abs=TOL), (
            f"{attr} coef mismatch: XLSTAT={exp_coef} InfoLeap={d['coef']}"
        )
        assert d["std_error"] == pytest.approx(exp_se, abs=TOL), (
            f"{attr} std_error mismatch: XLSTAT={exp_se} InfoLeap={d['std_error']}"
        )
        assert d["odds_ratio"] == pytest.approx(exp_or, abs=1e-3), (
            f"{attr} odds_ratio mismatch: XLSTAT={exp_or} InfoLeap={d['odds_ratio']}"
        )


def test_missing_data_safety_net_fields_present(blush_result):
    assert "n_before_na_omit" in blush_result, (
        "n_before_na_omit missing — the na.omit() missing-data safety net "
        "(see .regression_reference/AUDIT.md §3) is not wired into this result"
    )
    assert "pct_dropped_na" in blush_result, "pct_dropped_na missing from R result"
    assert blush_result["pct_dropped_na"] < 5, (
        f"unexpectedly high missing-data drop on a clean workbook: "
        f"{blush_result['pct_dropped_na']}% dropped "
        f"(n_before={blush_result['n_before_na_omit']}, n_after={blush_result['n']})"
    )
