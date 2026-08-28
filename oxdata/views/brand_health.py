"""
Brand Health Dashboard — Deep Dive Edition
==========================================
Data available in current wave:
  - Full awareness+conversion funnel: AIDED/SPONT/TOM/EVER_USED/CONSIDERATION/CURRENT_USER/PREFERRED/LAST_PURCHASED
  - NPS loyalty scores per brand           (v_brand_nps)
  - Geographic breakdown by zone / city    (v_respondents)
  - Brand imagery attribute associations   (fact_brand_imagery, 442K rows)
  - Need attribute importance              (fact_need_importance, 258K rows)
"""

import streamlit as st
import pandas as pd
import os
import sys
from pathlib import Path
import plotly.express as px
import plotly.graph_objects as go

current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(current_dir)  # oxdata/ -> info-leap/
if project_root not in sys.path:
    sys.path.insert(0, project_root)

from lens.analytics.brand_imagery_engine import BrandImageryEngine
from lens.analytics.brand_narrative import generate_brand_narrative
from lens.analytics.mapping_utils import _get_attr_themes, ATTRIBUTE_THEMES
from oxdata.utils.test_data_loader import list_test_datasets, load_test_matrix
from oxdata.utils.ui_styles import (inject_pulse_styles, sidebar_context_block,
                                    section_header as _shared_section_header,
                                    kpi_card as _shared_kpi_card,
                                    empty_state as _shared_empty_state,
                                    COLORS, CHART_LAYOUT)
from oxdata.utils.error_handlers import show_error_card
from oxdata.config.project_1 import BENCHMARKS

# data_layer: raw-Excel compute path (activated when project has raw_data.xlsx + llm_mapping_raw.json)
try:
    from lens.data_layer import get_project_layer, ProjectDataLayer, get_all_brands
    _DATA_LAYER_AVAILABLE = True
except Exception:
    _DATA_LAYER_AVAILABLE = False

def _get_layer() -> "ProjectDataLayer | None":
    """Return data layer for active project, or None (fall back to SQL)."""
    if not _DATA_LAYER_AVAILABLE:
        return None
    try:
        pid = st.session_state.get("active_project_id", "project_1")
        return get_project_layer(pid)
    except Exception:
        return None

# Force reload of analytics engines each time brand_health.py is (re)loaded
# so Streamlit's module cache never serves stale .pyc versions.
import importlib
import importlib.util
import lens.analytics.bip_engine as _bip_mod_ref
import lens.analytics.can_map_engine as _ca_mod_ref
import lens.analytics.brand_imagery_engine as _bie_mod_ref
importlib.reload(_bip_mod_ref)
importlib.reload(_ca_mod_ref)
importlib.reload(_bie_mod_ref)


# ══════════════════════════════════════════════════════════════════════════
# Data Workbench — AI Q&A over the currently-displayed section's data
# ══════════════════════════════════════════════════════════════════════════
from dotenv import load_dotenv as _wb_load_dotenv
_WB_ENV_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
_wb_load_dotenv(_WB_ENV_FILE, override=False)

_WB_OR_KEY = os.getenv("OPENROUTER_API_KEY")
_WB_OR_BASE_URL = os.getenv("OPENROUTER_BASE_URL", "https://openrouter.ai/api/v1")

# Free-tier default (OpenRouter ":free" suffix — no cost, rate-limited ~20 rpm/200 rpd).
# Paid fallback chain used automatically if the free model errors or times out.
_WB_MODEL_FREE = "meta-llama/llama-3.3-70b-instruct:free"
_WB_MODEL_CHOICES = {
    "Free (default)":        _WB_MODEL_FREE,
    "PRO (openai/gpt-4.1-mini)":  os.getenv("OPENROUTER_MODEL_PRO", "openai/gpt-4.1-mini"),
    "MINI (openai/gpt-4o-mini)":  os.getenv("OPENROUTER_MODEL_MINI", "openai/gpt-4o-mini"),
    "LLAMA 70B (paid)":       os.getenv("OPENROUTER_MODEL_LLAMA", "meta-llama/llama-3.3-70b-instruct"),
}
_WB_FALLBACK_MODEL = os.getenv("OPENROUTER_MODEL_MINI", "openai/gpt-4o-mini")


def _wb_json_safe(obj, _depth=0):
    """Coerce numpy/pandas types to plain JSON-serializable Python types."""
    import numpy as _np
    if _depth > 6:
        return str(obj)
    if isinstance(obj, dict):
        return {str(k): _wb_json_safe(v, _depth + 1) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [_wb_json_safe(v, _depth + 1) for v in obj]
    if isinstance(obj, pd.DataFrame):
        return _wb_json_safe(obj.to_dict(orient="records"), _depth + 1)
    if isinstance(obj, pd.Series):
        return _wb_json_safe(obj.tolist(), _depth + 1)
    if isinstance(obj, (_np.integer,)):
        return int(obj)
    if isinstance(obj, (_np.floating,)):
        v = float(obj)
        return None if v != v else v  # NaN -> None
    if isinstance(obj, _np.ndarray):
        return _wb_json_safe(obj.tolist(), _depth + 1)
    if isinstance(obj, float) and obj != obj:  # NaN
        return None
    if isinstance(obj, (str, int, float, bool)) or obj is None:
        return obj
    return str(obj)


def _wb_truncate_rows(data, max_rows=50):
    """Cap any list-of-dicts inside `data` to `max_rows`, noting truncation, so the
    JSON payload sent to the LLM stays bounded regardless of section size."""
    def _walk(v):
        if isinstance(v, list):
            if len(v) > max_rows and v and isinstance(v[0], dict):
                return {"_truncated": True, "shown": max_rows, "of": len(v),
                         "rows": [_walk(x) for x in v[:max_rows]]}
            return [_walk(x) for x in v]
        if isinstance(v, dict):
            return {k: _walk(x) for k, x in v.items()}
        return v
    return _walk(data)


def _workbench_capture(section_label: str, data: dict, note: str = ""):
    """Record the structured data behind the currently-rendered section so the
    Data Workbench AI panel (bottom of page) can answer questions about exactly
    what's on screen. Cheap — just a session_state write, safe to call on every
    rerun of a section-render function."""
    try:
        import datetime as _dt
        st.session_state["_wb_active_section_data"] = {
            "section": section_label,
            "note": note,
            "data": _wb_truncate_rows(_wb_json_safe(data)),
            "captured_at": _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
    except Exception:
        pass  # capture must never break the page render it's piggybacking on


def _call_workbench_llm(model: str, system_prompt: str, user_prompt: str, timeout: int = 45):
    """Synchronous OpenRouter chat call (brand_health.py has no existing asyncio
    infra, so a sync client keeps this self-contained inside Streamlit's
    synchronous render path). Returns (text, model_used, error)."""
    if not _WB_OR_KEY:
        return None, model, "OPENROUTER_API_KEY not configured (oxdata/.env)."
    try:
        from openai import OpenAI
        client = OpenAI(base_url=_WB_OR_BASE_URL, api_key=_WB_OR_KEY, timeout=timeout)
        resp = client.chat.completions.create(
            model=model,
            messages=[{"role": "system", "content": system_prompt},
                      {"role": "user", "content": user_prompt}],
            temperature=0.2,
        )
        text = resp.choices[0].message.content
        return (text.strip() if text else None), model, None
    except Exception as e:
        return None, model, str(e)


def _wb_ask(question: str, model_choice: str):
    """Ask the Data Workbench AI about the currently-captured section data.
    Tries the selected model first; if it fails and wasn't already the paid
    fallback, retries once with the paid fallback model."""
    ctx = st.session_state.get("_wb_active_section_data")
    if not ctx:
        return None, None, "No structured data captured for this section yet — switch tabs/re-run the section, then try again."

    import json as _json
    section = ctx.get("section", "current section")
    payload = _json.dumps(ctx.get("data", {}), default=str)
    if len(payload) > 60000:
        payload = payload[:60000] + '... "(truncated — payload too large)"'

    # Build project-aware description for the system prompt (dynamic, not project_1-only)
    try:
        from oxdata.db_loader import get_project_meta as _get_meta
        _pmeta = _get_meta()
        _proj_desc = _pmeta.get("description") or f"{_pmeta.get('n_respondents', '?')} respondents"
        _proj_industry = _pmeta.get("industry") or "consumer goods"
    except Exception:
        _proj_desc = "~6,631 respondents across 6 electrical-appliance categories in 18 Indian cities"
        _proj_industry = "consumer goods"

    system_prompt = (
        f"You are a market-research data analyst embedded in InfoLeap Pulse, a Streamlit "
        f"dashboard analyzing survey data ({_proj_desc}, {_proj_industry} sector). "
        "You are given the exact structured data currently "
        "shown on screen for one dashboard section (JSON below). Statistical terms follow "
        "XLSTAT convention (p<0.05 = '*', p<0.01 = '**', p<0.001 = '***'; VIF>5 = moderate, "
        ">10 = severe multicollinearity; McFadden R2 0.2+ = adequate fit). "
        "Answer using ONLY the data provided — do not invent numbers. If the user's question "
        "is empty or generic, give a concise general analysis of what this section's data shows: "
        "headline findings, notable strengths/risks, and one actionable recommendation. "
        "Be specific, cite the actual numbers, keep it tight (bullet points over prose)."
    )
    user_prompt = (
        f"Section: {section}\n\n"
        f"Data (JSON):\n{payload}\n\n"
        f"Question: {question if question.strip() else '(none — give a general analysis of this section)'}"
    )

    model = _WB_MODEL_CHOICES.get(model_choice, _WB_MODEL_FREE)
    text, used_model, err = _call_workbench_llm(model, system_prompt, user_prompt)
    if text is None and model != _WB_FALLBACK_MODEL:
        text, used_model, err2 = _call_workbench_llm(_WB_FALLBACK_MODEL, system_prompt, user_prompt)
        if text is not None:
            return text, used_model, f"⚠️ Fell back to {_WB_FALLBACK_MODEL} — {model} failed: {err}"
        return None, used_model, f"Both {model} and fallback {_WB_FALLBACK_MODEL} failed: {err} / {err2}"
    return text, used_model, err


# ── Colour palette ──────────────────────────────────────────────────────────
PALETTE = {
    "tom":   "#1a5d4d",
    "spont": "#30a76a",
    "aided": "#86efac",
    "nps_promoter":  "#22c55e",
    "nps_passive":   "#fbbf24",
    "nps_detractor": "#ef4444",
    "highlight":     "#0ea5e9",
    "neutral":       "#9ca3af",
}
ZONE_COLORS = {"North": "#3b82f6", "South": "#f59e0b",
               "East":  "#8b5cf6", "West":  "#ec4899"}
NPS_INDUSTRY_AVG = BENCHMARKS.get("nps_industry_avg", 45)  # project_1 fallback; overridden per-render below

# ── Chart Theme (customisable — change here to reflect across all charts) ────
# Predefined palettes: "pulse_green" | "corporate_blue" | "warm_amber" | "mono"
CHART_THEME = {
    # Active palette name (swap to change all chart colours at once)
    "palette": "pulse_green",
    # Per-palette colour sets. Every entry in the first 7 slots must stay readable at full
    # opacity on a white background (up to 7 brands can be shown at once — base + 6 compare) —
    # no pastel/near-white colours here. Anything lighter (e.g. an unselected-state tint) is
    # derived at render time via _hex_lighten()/_hex_to_rgba(), never baked into this list.
    "palettes": {
        "pulse_green":    ["#1a5d4d", "#30a76a", "#0ea5e9", "#f59e0b", "#8b5cf6", "#ec4899", "#ef4444", "#166534"],
        "corporate_blue": ["#1e3a5f", "#2563eb", "#0891b2", "#f59e0b", "#6b7280", "#dc2626", "#16a34a", "#7c3aed"],
        "warm_amber":     ["#92400e", "#d97706", "#1e3a5f", "#6b7280", "#16a34a", "#dc2626", "#7c3aed", "#0891b2"],
        "mono":           ["#111827", "#374151", "#6b7280", "#4b5563", "#1a5d4d", "#ef4444", "#0ea5e9", "#f59e0b"],
    },
    # Font family — change to match report template
    "font_family": "Inter, Arial, sans-serif",
    # Base font size for chart labels
    "font_size": 12,
    # Chart height multipliers for PPT export (larger = better PPT quality)
    "ppt_scale": 1.0,   # set to 1.3 to scale all charts up for PPT
}

def _get_csat_scale() -> int:
    """Read CSAT rating scale (5 or 10) from active project's project_meta.json."""
    try:
        import streamlit as st
        from oxdata.db_loader import get_project_meta
        pid = st.session_state.get("active_project_id", "project_1")
        return int(get_project_meta(pid).get("csat_scale", 10))
    except Exception:
        return 10


def _chart_colors() -> list:
    """Return active palette colour list."""
    p = CHART_THEME["palette"]
    return CHART_THEME["palettes"].get(p, CHART_THEME["palettes"]["pulse_green"])

def _chart_layout_base(height: int = 500) -> dict:
    """Common Plotly layout kwargs — apply to all charts for consistency."""
    base = dict(CHART_LAYOUT)  # shared tokens: white bg, Inter font, standard margins
    base.update(
        height=int(height * CHART_THEME["ppt_scale"]),
        font=dict(family=CHART_THEME["font_family"], size=CHART_THEME["font_size"],
                  color="#374151"),
        margin=dict(t=50, b=50, l=60, r=40),
    )
    return base


def _theme_fig(fig, height: int = None):
    """Stamp active CHART_THEME (font, size, scale) onto any Plotly figure.
    Call before every `return fig` so sidebar theme controls actually apply.
    Existing chart-specific settings are preserved; only font family/size and
    paper/plot background are overridden.
    """
    layout_patch = dict(
        font=dict(family=CHART_THEME["font_family"], size=CHART_THEME["font_size"]),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
    )
    if height is not None:
        layout_patch["height"] = int(height * CHART_THEME["ppt_scale"])
    fig.update_layout(**layout_patch)
    return fig


# ── Helpers ──────────────────────────────────────────────────────────────────

def _metric_card(label, value, icon, subtext="", delta=None, color=None):
    """Delegates to shared kpi_card. icon param ignored (legacy compat)."""
    delta_str = ""
    if delta is not None:
        sign = "+" if delta >= 0 else ""
        delta_str = f"  {sign}{delta}"
    _shared_kpi_card(label, f"{value}{delta_str}", color or COLORS["teal"], subtext=subtext)


def _insight_callout(findings, label, icon="💡"):
    if not findings:
        return
    bullets = "".join(f'<div style="font-size:0.85rem;margin-bottom:5px;line-height:1.5;opacity:0.9;">• {f}</div>' for f in findings)
    header = f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:8px;"><span style="font-size:1rem;">{icon}</span><span style="font-size:0.78rem;font-weight:700;text-transform:uppercase;letter-spacing:0.05em;color:#22c55e;">{label}</span></div>'
    st.markdown(f'<div style="border-radius:10px;padding:14px 16px;border:1px solid rgba(34,197,94,0.3);background:rgba(34,197,94,0.07);margin-top:10px;">{header}{bullets}</div>', unsafe_allow_html=True)


def _section_header(title: str, subtitle: str = ""):
    _shared_section_header(title, subtitle)


def _ai_chart_headline(section_key: str, data_summary: str, brand: str = "", project_id: str = "project_1") -> str | None:
    """Generate a data-driven newspaper-style headline for a chart section.
    Session-cached per (section_key, brand, project_id) — never regenerates on rerun.
    Strict guardrails: numbers from data_summary only, no speculation.
    Returns headline string or None on failure/no API key.
    """
    _cache_key = f"_ai_headline_{project_id}_{brand}_{section_key}"
    if _cache_key in st.session_state:
        return st.session_state[_cache_key]
    if not _WB_OR_KEY:
        return None
    try:
        import httpx as _hx, json as _json
        _model = os.getenv("OPENROUTER_MODEL_MINI", "openai/gpt-4o-mini")
        _prompt = (
            "You are a market research analyst writing chart headlines for a brand health report. "
            "Write ONE headline (max 18 words) that captures the most important insight from the data below. "
            "Rules: (1) Use only the numbers given — no invention. "
            "(2) Active voice, declarative sentence. "
            "(3) Mention the brand or top finding explicitly. "
            "(4) No questions, no ellipsis, no hype words. "
            "Output only the headline text — no quotes, no prefix.\n\n"
            f"Brand: {brand or 'Category'}\n"
            f"Data: {data_summary}"
        )
        _r = _hx.post(
            f"{_WB_OR_BASE_URL}/chat/completions",
            headers={"Authorization": f"Bearer {_WB_OR_KEY}", "Content-Type": "application/json"},
            json={"model": _model, "messages": [{"role": "user", "content": _prompt}], "max_tokens": 60},
            timeout=12,
        )
        _hl = _r.json()["choices"][0]["message"]["content"].strip().strip('"').strip("'")
        st.session_state[_cache_key] = _hl
        return _hl
    except Exception:
        return None


def _render_ai_headline(section_key: str, data_summary: str, brand: str = "", project_id: str = "project_1"):
    """Render AI headline as a styled subtitle under a section. No-op if generation fails."""
    _hl = _ai_chart_headline(section_key, data_summary, brand=brand, project_id=project_id)
    if _hl:
        st.markdown(
            f"<div style='font-size:0.97rem;font-weight:600;color:#0f4c75;"
            f"background:linear-gradient(90deg,#e8f4fd,transparent);padding:6px 12px;"
            f"border-left:3px solid #0ea5e9;border-radius:3px;margin:4px 0 12px;'>"
            f"{_hl}</div>",
            unsafe_allow_html=True,
        )


def _sig_letters_proportions(items, alpha: float = 0.05):
    """Column-proportion significance test with letters (XLSTAT / pValue-style).

    items: list of (label, count, base). Pairwise two-proportion pooled z-tests.
    Convention: each column gets a letter (A = highest proportion). A column's
    annotation lists the letters of the OTHER columns it is significantly HIGHER than.

    Returns (letters, beats, pct):
      letters: {label -> 'A'|'B'|...}   beats: {label -> ['B','C',...]}   pct: {label -> %}
    """
    import numpy as _np
    from itertools import combinations as _comb
    from scipy.stats import norm as _norm

    def _col_letter(i: int) -> str:
        # Excel-style labels so >26 brands don't overflow into '[', '\\', ']' …
        s, i = "", i + 1
        while i > 0:
            i, r = divmod(i - 1, 26)
            s = chr(65 + r) + s
        return s

    data = []
    for lbl, cnt, base in items:
        base = float(base) if base else 0.0
        p = (float(cnt) / base) if base > 0 else 0.0
        data.append([lbl, float(cnt), base, p])
    data.sort(key=lambda x: -x[3])  # highest proportion first
    letters = {d[0]: _col_letter(i) for i, d in enumerate(data)}  # A = highest
    beats = {d[0]: [] for d in data}
    for (la, ca, ba, pa), (lb, cb, bb, pb) in _comb(data, 2):
        if ba < 1 or bb < 1:
            continue
        pooled = (ca + cb) / (ba + bb)
        se = _np.sqrt(pooled * (1 - pooled) * (1 / ba + 1 / bb))
        if se == 0:
            continue
        z = (pa - pb) / se
        pval = 2 * (1 - _norm.cdf(abs(z)))
        if pval < alpha:  # pa >= pb by sort → higher beats lower
            beats[la].append(letters[lb])
    pct = {d[0]: d[3] * 100 for d in data}
    return letters, beats, pct


import re as _re

def _apply_to_plain(text: str, pattern, replacement, flags=0) -> str:
    """Apply regex only to text segments that are NOT inside any <mark> element.
    Prevents nested marks when sequential highlighting passes run on already-marked text.
    """
    result = []
    parts = _re.split(r'(</?[a-zA-Z][^>]*>)', text)
    in_mark = 0
    for i, part in enumerate(parts):
        if i % 2 == 1:  # HTML tag token
            if _re.match(r'<mark\b', part, _re.IGNORECASE):
                in_mark += 1
            elif _re.match(r'</mark>', part, _re.IGNORECASE):
                in_mark = max(0, in_mark - 1)
            result.append(part)
        else:  # text content
            if in_mark == 0:
                result.append(_re.sub(pattern, replacement, part, flags=flags))
            else:
                result.append(part)  # inside a mark — leave unchanged
    return "".join(result)


def _highlight_text(text: str, known_names: list = None) -> str:
    """Highlighter-style markup: coloured backgrounds like physical marker pens.
    Each pass runs only on plain-text segments to prevent nested marks.
    """
    if not text:
        return text

    # 1. Escape bare < and > that are NOT part of HTML tags (e.g. p<0.05, x>y)
    #    so the browser doesn't treat them as malformed tags.
    text = _re.sub(r'<(?![a-zA-Z/!])', '&lt;', text)
    text = _re.sub(r'(?<![a-zA-Z"\'=\d])>', '&gt;', text)

    # Dark-theme color palette — dark backgrounds, light text for readability
    # 2. **bold** → gold badge
    text = _re.sub(
        r'\*\*(.+?)\*\*',
        r'<mark style="background:#92400e;color:#fef3c7;font-weight:700;'
        r'border-radius:3px;padding:2px 5px;">\1</mark>',
        text,
    )

    # 3. Known brand/attr names → teal badge (skip if already inside a mark)
    if known_names:
        for name in sorted(known_names, key=len, reverse=True):
            if not name:
                continue
            escaped = _re.escape(name)
            text = _apply_to_plain(
                text,
                rf'\b({escaped})\b',
                r'<mark style="background:#0e7490;color:#e0f2fe;font-weight:600;'
                r'border-radius:3px;padding:2px 5px;">\1</mark>',
                flags=_re.IGNORECASE,
            )

    # 4. Percentages → orange badge (plain segments only)
    text = _apply_to_plain(
        text,
        r'(\b\d+\.?\d*%)',
        r'<mark style="background:#c2410c;color:#fff7ed;font-weight:600;'
        r'border-radius:3px;padding:2px 5px;">\1</mark>',
    )

    # 5. Standalone decimals → orange badge (plain segments only)
    text = _apply_to_plain(
        text,
        r'([+-]?\b\d+\.\d+)\b',
        r'<mark style="background:#c2410c;color:#fff7ed;font-weight:600;'
        r'border-radius:3px;padding:2px 5px;">\1</mark>',
    )

    # 6. χ²/p-value stat markers → blue badge (plain segments only)
    text = _apply_to_plain(
        text,
        r'(χ²=[\d.]+|p=[\d.]+|p-value[\s=]+[\d.]+)',
        r'<mark style="background:#1e40af;color:#dbeafe;font-weight:600;'
        r'border-radius:3px;padding:2px 5px;">\1</mark>',
    )

    # 7. Strategic keywords → green badge (plain segments only)
    _TERMS = (
        r'\b(significant(?:ly)?|over-indexed|under-indexed|contested|exclusive(?:ly)?|'
        r'dominant(?:ly)?|differentiat\w+|outperform\w*|strongest|weakest|battleground|'
        r'table.?stakes|ownership|anchor\w*|outlier\w*|reliable|unreliable|'
        r'defensible|competitive advantage|strategic|critical)\b'
    )
    text = _apply_to_plain(text, _TERMS,
        r'<mark style="background:#166534;color:#dcfce7;font-weight:600;'
        r'border-radius:3px;padding:2px 5px;">\1</mark>',
        flags=_re.IGNORECASE,
    )

    return text


def _structured_ai_card(raw_text: str, label: str, accent: str, known_names: list = None):
    """
    Parse FINDING/DETAIL/ACTION structured LLM output and render as formatted card.
    Falls back to single-block rendering for unstructured text.
    Uses explicit string building (no nested f-strings) for Python 3.12 compatibility.
    """
    if not raw_text:
        return

    _bh_secs: dict = {"FINDING": [], "DETAIL": [], "ACTION": []}
    _bh_cur = None
    for line in raw_text.strip().splitlines():
        ls = line.strip()
        if ls.upper().startswith("FINDING:"):
            _bh_cur = "FINDING"; rest = ls[8:].strip();
            if rest: _bh_secs["FINDING"].append(rest)
        elif ls.upper().startswith("DETAIL:"):
            _bh_cur = "DETAIL"; rest = ls[7:].strip();
            if rest: _bh_secs["DETAIL"].append(rest)
        elif ls.upper().startswith("ACTION:"):
            _bh_cur = "ACTION"; rest = ls[7:].strip();
            if rest: _bh_secs["ACTION"].append(rest)
        elif _bh_cur and ls:
            _bh_secs[_bh_cur].append(ls)
    finding = " ".join(_bh_secs["FINDING"])
    detail  = " ".join(_bh_secs["DETAIL"])
    action  = " ".join(_bh_secs["ACTION"])

    structured = bool(finding or detail or action)
    names = known_names or []

    # Card wrapper styles (reused in both branches)
    outer = (
        '<div style="border-left:4px solid ' + accent + ';border-radius:6px;'
        'background:rgba(255,255,255,0.04);padding:14px 18px;margin:12px 0;">'
    )
    label_div = (
        '<div style="font-size:0.72rem;font-weight:700;letter-spacing:0.08em;'
        'color:' + accent + ';margin-bottom:10px;text-transform:uppercase;">'
        + label + '</div>'
    )

    if structured:
        finding_html = _highlight_text(finding, names) if finding else ""
        detail_html  = _highlight_text(detail,  names) if detail  else ""
        action_html  = _highlight_text(action,  names) if action  else ""

        parts = [outer, label_div]
        if finding_html:
            parts.append(
                '<div style="font-size:0.95rem;font-weight:600;color:#f1f5f9;'
                'margin-bottom:8px;line-height:1.5;">' + finding_html + '</div>'
            )
        if detail_html:
            parts.append(
                '<div style="font-size:0.88rem;color:#9ca3af;line-height:1.65;'
                'margin-bottom:8px;">' + detail_html + '</div>'
            )
        if action_html:
            parts.append(
                '<div style="font-size:0.85rem;background:rgba(255,255,255,0.07);'
                'border-radius:4px;padding:8px 12px;color:#9ca3af;'
                'border-left:3px solid ' + accent + ';line-height:1.5;">'
                '<span style="font-weight:700;color:' + accent + ';">&#9658; Action: </span>'
                + action_html + '</div>'
            )
        parts.append('</div>')
        st.markdown('\n'.join(parts), unsafe_allow_html=True)
    else:
        body = _highlight_text(raw_text, names)
        html = (
            outer + label_div
            + '<div style="font-size:0.88rem;color:#9ca3af;line-height:1.65;">'
            + body + '</div></div>'
        )
        st.markdown(html, unsafe_allow_html=True)


def _ai_card(content: str, label: str, accent: str):
    if not content or content == "Analysis pending...":
        return
    # Convert newlines → paragraph breaks so multi-sentence content renders correctly
    paragraphs = [p.strip() for p in content.strip().split("\n") if p.strip()]
    body_html = "".join(
        f'<p style="margin:0 0 8px 0;">{_highlight_text(p)}</p>'
        for p in paragraphs
    )
    st.markdown(
        f'<div class="bh-ai-card" style="border-left:3px solid {accent};min-height:auto;">'
        f'<div class="bh-ai-label" style="color:{accent};">{label}</div>'
        f'<div class="bh-ai-body">{body_html}</div>'
        f'</div>',
        unsafe_allow_html=True,
    )


def _h_escape(text) -> str:
    import html as _h
    return _h.escape(str(text if text is not None else ""))


def _md_bold(text: str) -> str:
    """Convert **bold** markdown to <strong> HTML, escaping the rest."""
    import html as _h
    import re as _re2
    if not text:
        return ""
    out, last = [], 0
    for m in _re2.finditer(r"\*\*(.+?)\*\*", text):
        out.append(_h.escape(text[last:m.start()]))
        out.append(f'<strong>{_h.escape(m.group(1))}</strong>')
        last = m.end()
    out.append(_h.escape(text[last:]))
    return "".join(out)


def _render_executive_briefing(brief: dict, sel_brand: str, filter_label: str = "All India"):
    """Premium C-suite Executive Command Briefing — bottom-line verdict + 3 role lenses
    (CEO / CMO / Head of Product). brief: dict from generate_executive_briefing()."""
    if not brief or not brief.get("bottom_line"):
        return
    quad = brief.get("quadrant", "")
    tom_rank = brief.get("tom_rank_str", "-")
    nps_rank = brief.get("nps_rank_str", "-")
    quad_colors = {
        "Market Leader":      "#10b981",
        "Loyalty Hidden Gem": "#38bdf8",
        "Awareness Leader":   "#f59e0b",
        "Growth Opportunity": "#a78bfa",
    }
    qc = quad_colors.get(quad, "#38bdf8")

    # ── Hero strip: eyebrow + quadrant badge + bottom line ───────────────────
    st.markdown(
        f'<div style="background:radial-gradient(circle at top right,#1e293b,#0f172a);'
        f'border:1px solid rgba(255,255,255,0.08);border-radius:20px 20px 0 0;'
        f'padding:26px 30px 22px;box-shadow:0 18px 44px rgba(0,0,0,0.28);">'
        f'<div style="display:flex;justify-content:space-between;align-items:center;'
        f'flex-wrap:wrap;gap:10px;margin-bottom:14px;">'
        f'<div style="font-size:0.68rem;font-weight:800;letter-spacing:0.22em;'
        f'text-transform:uppercase;color:#38bdf8;">⬢ Executive Command Briefing</div>'
        f'<div style="display:flex;gap:8px;align-items:center;">'
        f'<span style="background:{qc};color:#0f172a;font-size:0.66rem;font-weight:800;'
        f'padding:4px 14px;border-radius:30px;text-transform:uppercase;letter-spacing:0.05em;'
        f'box-shadow:0 0 18px {qc}55;">{_h_escape(quad)}</span>'
        f'<span style="color:rgba(255,255,255,0.45);font-size:0.66rem;font-weight:600;">'
        f'Salience {_h_escape(tom_rank)} · Advocacy {_h_escape(nps_rank)}</span>'
        f'</div></div>'
        f'<div style="font-size:0.62rem;color:rgba(255,255,255,0.4);text-transform:uppercase;'
        f'letter-spacing:0.14em;margin-bottom:7px;">{_h_escape(sel_brand)} · {_h_escape(filter_label)}</div>'
        f'<div style="font-size:1.32rem;font-weight:800;color:#f8fafc;line-height:1.4;'
        f'letter-spacing:-0.01em;">{_md_bold(brief["bottom_line"])}</div>'
        f'</div>',
        unsafe_allow_html=True,
    )

    # ── Three objective analysis dimensions (neutral — no role framing) ──────
    # .get() with legacy-key fallback so a stale cache never renders empty cards.
    _m = brief.get("market") or brief.get("ceo") or ""
    _d = brief.get("demand") or brief.get("cmo") or ""
    _e = brief.get("experience") or brief.get("product") or ""
    lenses = [
        ("Market Position",       "Salience · Rank · Geography", "◆", "#10b981", _m),
        ("Demand & Funnel",       "Awareness → Preference",      "◆", "#38bdf8", _d),
        ("Experience & Advocacy", "Promoters · Detractors",      "◆", "#a78bfa", _e),
    ]
    cols = st.columns(3)
    for col, (title, sub, icon, color, body) in zip(cols, lenses):
        with col:
            st.markdown(
                f'<div style="background:#0f172a;border:1px solid rgba(255,255,255,0.07);'
                f'border-top:3px solid {color};border-radius:0 0 4px 4px;padding:18px 18px 20px;'
                f'min-height:200px;height:100%;box-sizing:border-box;">'
                f'<div style="display:flex;align-items:center;gap:9px;margin-bottom:12px;">'
                f'<span style="font-size:0.8rem;color:{color};">{icon}</span>'
                f'<div><div style="font-size:0.82rem;font-weight:800;color:{color};'
                f'line-height:1;">{title}</div>'
                f'<div style="font-size:0.58rem;color:rgba(255,255,255,0.4);'
                f'text-transform:uppercase;letter-spacing:0.08em;margin-top:2px;">{sub}</div></div>'
                f'</div>'
                f'<div style="font-size:0.82rem;line-height:1.66;color:#cbd5e1;">'
                f'{_md_bold(body).replace(chr(10)," ")}</div>'
                f'</div>',
                unsafe_allow_html=True,
            )
    st.markdown('<div style="height:14px;"></div>', unsafe_allow_html=True)


def _glass_card(content: str, title: str = None, subtitle: str = None):
    """Wraps content in a glass-card div for the Midnight Glass aesthetic."""
    header_html = ""
    if title:
        sub_html = f'<div style="font-size: 0.8rem; color: rgba(255,255,255,0.6);">{subtitle}</div>' if subtitle else ""
        header_html = f"""<div style="margin-bottom: 15px;"><div style="font-size: 1.1rem; font-weight: 700; color: var(--neon-cyan);">{title}</div>{sub_html}</div>"""
    
    st.markdown(f"""<div class="glass-card">{header_html}{content}</div>""", unsafe_allow_html=True)


def _hero_banner(brand_name: str, brand_data: dict, base_n: int,
                 seg_label: str, brands_list: list):
    """
    Refactored Hero Banner with Midnight Glass aesthetic.
    High contrast, neon cyan accents, and flexbox layout.
    """
    strat = brand_data.get("strat_score", 0)
    nps   = brand_data.get("nps")
    tom   = brand_data.get("tom_pct", 0)

    # NPS rank
    eligible = [b for b in brands_list if b.get("nps") is not None]
    eligible.sort(key=lambda x: x["nps"], reverse=True)
    nps_rank = next((i + 1 for i, b in enumerate(eligible) if b["brand_name"] == brand_name), None)
    nps_rank_str = f"#{nps_rank} of {len(eligible)}" if nps_rank else "—"

    # TOM rank
    tom_sorted = sorted(brands_list, key=lambda x: x.get("tom_pct", 0), reverse=True)
    tom_rank = next((i + 1 for i, b in enumerate(tom_sorted) if b["brand_name"] == brand_name), None)
    tom_rank_str = f"#{tom_rank} of {len(tom_sorted)}" if tom_rank else "—"

    # Positioning quadrant
    median_tom = sum(b.get("tom_pct", 0) for b in brands_list) / max(len(brands_list), 1)
    nps_val = nps or 0
    if tom > median_tom and nps_val >= 45:
        quadrant, q_color = "Market Leader", "#86efac"
    elif tom <= median_tom and nps_val >= 45:
        quadrant, q_color = "Loyalty Hidden Gem", "#93c5fd"
    elif tom > median_tom:
        quadrant, q_color = "Awareness Leader", "#fcd34d"
    else:
        quadrant, q_color = "Growth Opportunity", "#d1d5db"

    # Executive Hero Refactor
    hero_html = f"""<div class="exec-hero"><p class="exec-hero-eyebrow">Strategic Brand Intelligence &middot; {seg_label}</p><h1 class="exec-hero-name">{brand_name}</h1><div style="display: flex; align-items: center; gap: 18px; margin-bottom: 30px;"><span class="exec-badge">{quadrant}</span><span style="color: rgba(255,255,255,0.4); font-size: 0.85rem; font-weight: 500; letter-spacing: 0.02em;">{base_n:,} tracked respondents &middot; Wave 1 (2026)</span></div><div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 20px;"><div class="glass-card" style="border-top: 3px solid var(--neon-cyan); background: rgba(56, 189, 248, 0.03);"><div style="font-size: 2.8rem; font-weight: 900; color: #e5e7eb; line-height: 1; letter-spacing: -0.02em;">{strat}</div><div style="font-size: 0.65rem; text-transform: uppercase; letter-spacing: 0.12em; color: rgba(255,255,255,0.5); margin-top: 10px; display: flex; align-items: center; gap: 6px;">Health Index<span class="hs-tooltip-anchor">&#9432;<span class="hs-tooltip-box"><strong>Pulse Health Index Formula</strong><br><br>Decomposed as:<br>• 40% TOM Awareness<br>• 10% Spontaneous Recall<br>• 50% Normalized Advocacy (NPS)<br><br><em>Normalized to 100-pt scale.</em></span></span></div></div><div class="glass-card"><div style="font-size: 2.2rem; font-weight: 800; color: #e5e7eb; line-height: 1;">{tom_rank_str}</div><div style="font-size: 0.65rem; text-transform: uppercase; letter-spacing: 0.12em; color: rgba(255,255,255,0.5); margin-top: 10px;">Market Salience</div></div><div class="glass-card"><div style="font-size: 2.2rem; font-weight: 800; color: #e5e7eb; line-height: 1;">{nps_rank_str}</div><div style="font-size: 0.65rem; text-transform: uppercase; letter-spacing: 0.12em; color: rgba(255,255,255,0.5); margin-top: 10px;">Advocacy Rank</div></div></div></div>"""
    st.markdown(hero_html, unsafe_allow_html=True)





# ── Chart builders ────────────────────────────────────────────────────────────

def _weighted_mean(items, val_key, wt_key=None, default=0.0):
    """Respondent-weighted mean of a per-brand metric.

        weighted_mean = Σ(valueᵢ · wᵢ) / Σ(wᵢ)

    wt_key selects the weight (e.g. 'nps_base' so brands with more raters count
    more). When wt_key is None/absent or all weights are 0, this reduces to the
    simple arithmetic mean — which is the correct result when every brand shares
    the same base (e.g. awareness %, asked of all respondents). Used study-wide so
    market/category benchmarks are respondent-weighted, not brand-count averages.
    """
    num = den = 0.0
    n = 0
    for b in items:
        v = b.get(val_key)
        if v is None:
            continue
        try:
            v = float(v)
        except (TypeError, ValueError):
            continue
        w = 1.0
        if wt_key is not None:
            try:
                w = float(b.get(wt_key, 0) or 0)
            except (TypeError, ValueError):
                w = 0.0
        num += v * w
        den += w
        n += 1
    if den > 0:
        return num / den
    # all weights zero → fall back to simple mean
    vals = []
    for b in items:
        try:
            vals.append(float(b.get(val_key)))
        except (TypeError, ValueError):
            pass
    return (sum(vals) / len(vals)) if vals else default


def _awareness_landscape_chart(brands_list, base_n: int = 0, top_n: int = 15,
                               all_brands_ref: list = None):
    """PPT-style vertical stacked bar + descending total-awareness line + ALL reference.
    all_brands_ref: full unfiltered brand list used for ALL bar (defaults to brands_list).
    Segments bottom→top: TOM (dark) | Spont-incremental (medium) | Aided-only (light).
    Line connects brands only (NOT ALL). Dashed reference line at ALL aided level.
    Tier bands + black header. Sig arrows vs ALL average @ 95% CI.
    """
    import math

    # Caller already sorted by the user's chosen rank-by key — preserve that order
    top = brands_list[:top_n]
    n_brands = len(top)
    if n_brands == 0:
        return go.Figure()

    # ── ALL benchmark — ALWAYS computed from full reference list, never from filter ──
    # Respondent-weighted mean via _weighted_mean. Awareness %s are all computed
    # over the SAME respondent base (asked of everyone), so the correct weight is
    # equal across brands → this equals the simple mean. (NPS, with per-brand rater
    # bases, IS base-weighted elsewhere where that changes the result.)
    _ref = all_brands_ref if (all_brands_ref is not None and len(all_brands_ref) > 0) else top
    _n_ref = len(_ref)
    def _aware_pct(b):
        return b.get("total_awareness_pct", b.get("aided_pct", 0)) or 0

    all_tom   = round(_weighted_mean(_ref, "tom_pct"),   1)
    all_spont = round(_weighted_mean(_ref, "spont_pct"), 1)
    all_aided = round(sum(_aware_pct(b) for b in _ref) / max(len(_ref), 1), 1)
    all_entry = {
        "brand_name":     "ALL",
        "tom_only_pct":   all_tom,
        "spont_incr_pct": round(all_spont - all_tom,   1),
        "aided_incr_pct": round(all_aided - all_spont, 1),
        "aided_pct": all_aided, "spont_pct": all_spont, "tom_pct": all_tom,
        "total_awareness_pct": all_aided,
    }
    display = [all_entry] + top     # ALL leftmost, then brands desc
    n_disp  = len(display)

    names     = [b["brand_name"]     for b in display]
    tom_seg   = [b["tom_only_pct"]   for b in display]
    spont_seg = [b["spont_incr_pct"] for b in display]
    aided_seg = [b["aided_incr_pct"] for b in display]
    total_all = [_aware_pct(b)   for b in display]
    spont_cum = [b["spont_pct"] for b in display]

    # Line connects brands only — ALL excluded (None causes gap, no upward jump)
    line_y = [None] + [_aware_pct(b) for b in top]

    # ── Tier split by tertile (brands only, index offset +1 for ALL) ─────────
    t1 = max(1, n_brands // 3) + 1
    t2 = max(t1 + 1, 2 * n_brands // 3 + 1)

    # ── Significance vs ALL average (z-test on aided%) ───────────────────────
    sig_arrows = [""]
    for b in top:
        if base_n > 0 and all_aided > 0:
            p0  = all_aided / 100
            p1  = _aware_pct(b) / 100
            se  = math.sqrt(p0 * (1 - p0) / base_n)
            z   = (p1 - p0) / se if se > 0 else 0
            sig_arrows.append("▲" if z > 1.96 else ("▼" if z < -1.96 else ""))
        else:
            sig_arrows.append("")

    # ── Colours — match app PALETTE (green theme) ────────────────────────────
    _c_tom   = PALETTE["tom"]    # "#1a5d4d" dark green (TOM = strongest recall)
    _c_spont = PALETTE["spont"]  # "#30a76a" mid green (spontaneous)
    _c_aided = PALETTE["aided"]  # "#86efac" light green (aided only)
    # ALL bar: 3 distinct greys so segments remain readable
    _c_all_tom   = "#374151"   # dark grey   → TOM segment
    _c_all_spont = "#6b7280"   # mid grey    → Spont segment
    _c_all_aided = "#d1d5db"   # light grey  → Aided segment

    fig = go.Figure()

    # ── Tier background bands (brand columns only) ────────────────────────────
    tier_cfg = [
        (1,  t1 - 1, "rgba(220,252,231,0.45)"),
        (t1, t2 - 1, "rgba(254,249,195,0.45)"),
        (t2, n_disp - 1, "rgba(254,226,226,0.45)"),
    ]
    for xi0, xi1, fill in tier_cfg:
        if xi0 >= n_disp or xi1 < xi0:
            continue
        fig.add_vrect(
            x0=names[xi0], x1=names[min(xi1, n_disp - 1)],
            fillcolor=fill, opacity=1.0, layer="below", line_width=0,
        )

    # ── ALL column: subtle grey fill to distinguish from brand columns ────────
    fig.add_vrect(
        x0="ALL", x1="ALL",
        fillcolor="rgba(229,231,235,0.6)", opacity=1.0, layer="below", line_width=0,
    )

    # ── Stacked bars ─────────────────────────────────────────────────────────
    fig.add_trace(go.Bar(
        name="Top of Mind", x=names, y=tom_seg,
        marker_color=[_c_all_tom if n == "ALL" else _c_tom for n in names],
        text=[f"<b>{int(v)}</b>" if v >= 4 else "" for v in tom_seg],
        textposition="inside", textfont=dict(color="white", size=11),
        hovertemplate="<b>%{x}</b><br>TOM: %{y:.1f}%<extra></extra>",
    ))
    fig.add_trace(go.Bar(
        name="Spontaneous", x=names, y=spont_seg,
        marker_color=[_c_all_spont if n == "ALL" else _c_spont for n in names],
        text=[f"<b>{int(v)}</b>" if v >= 4 else "" for v in spont_seg],
        textposition="inside", textfont=dict(color="white", size=11),
        customdata=spont_cum,
        hovertemplate="<b>%{x}</b><br>Spont (incremental): %{y:.1f}%<br>Cumul. Spont: %{customdata:.1f}%<extra></extra>",
    ))
    fig.add_trace(go.Bar(
        name="Aided only", x=names, y=aided_seg,
        marker_color=[_c_all_aided if n == "ALL" else _c_aided for n in names],
        text=[f"<b>{int(v)}</b>" if v >= 4 else "" for v in aided_seg],
        textposition="inside", textfont=dict(color="#1a5d4d", size=11),
        hovertemplate="<b>%{x}</b><br>Aided only: %{y:.1f}%<extra></extra>",
    ))

    # ── Total awareness line — brands only (ALL is None = no jump) ───────────
    fig.add_trace(go.Scatter(
        name="Total Awareness", x=names, y=line_y,
        mode="lines+markers",
        line=dict(color="#1a5d4d", width=2.5),
        marker=dict(size=7, color="#1a5d4d", symbol="circle"),
        connectgaps=False,
        hovertemplate="<b>%{x}</b><br>Total awareness: %{y:.1f}%<extra></extra>",
    ))

    # ── Dashed reference line at ALL aided level ──────────────────────────────
    fig.add_hline(
        y=all_aided,
        line=dict(dash="dot", color="#6b7280", width=1.5),
        annotation_text=f"Avg {all_aided:.1f}%",
        annotation_position="top right",
        annotation_font=dict(size=9, color="#6b7280"),
    )

    # ── Total % + sig arrows above each bar ──────────────────────────────────
    for name, tot, arr in zip(names, total_all, sig_arrows):
        if name == "ALL":
            # ALL: just show the number, no arrow
            fig.add_annotation(
                x=name, y=tot + 4, text=f"<b>{int(round(tot))}</b>",
                showarrow=False, yref="y", xanchor="center",
                font=dict(size=11, color="#4b5563"),
            )
        else:
            c   = "#16a34a" if arr == "▲" else ("#dc2626" if arr == "▼" else "#374151")
            txt = f"<b>{int(round(tot))}</b>" + (f" {arr}" if arr else "")
            fig.add_annotation(
                x=name, y=tot + 4, text=txt, showarrow=False, yref="y",
                font=dict(size=12, color=c), xanchor="center",
            )

    # ── Teal tier header band above plot (no black bg) ────────────────────────
    fig.add_shape(
        type="rect", xref="paper", yref="paper",
        x0=0, y0=1.10, x1=1.0, y1=1.22,
        fillcolor="rgba(26,93,77,0.90)", line_width=0,
    )
    for tx, tlbl, tc in [
        (0.22, "← High Recall Brands",  "#86efac"),
        (0.55, "Mid Recall Brands",      "#fde68a"),
        (0.83, "Low Recall Brands →",    "#fca5a5"),
    ]:
        fig.add_annotation(
            xref="paper", yref="paper", x=tx, y=1.16,
            text=f"<b>{tlbl}</b>", showarrow=False,
            font=dict(size=10, color=tc), xanchor="center",
        )

    # ── "Total Awareness" badge + "Data in %" (teal bg) ──────────────────────
    fig.add_annotation(
        xref="paper", yref="paper", x=0.0, y=1.085,
        text="<b>Total Awareness  ·  Demographic-filtered</b>", showarrow=False,
        font=dict(size=10, color="#1a5d4d"),
        bgcolor="rgba(220,252,231,0.85)", borderpad=4, xanchor="left",
        bordercolor="#86efac", borderwidth=1,
    )
    fig.add_annotation(
        xref="paper", yref="paper", x=1.0, y=1.085,
        text="<i>Data in %</i>", showarrow=False,
        font=dict(size=10, color="#6b7280"), xanchor="right",
    )

    fig.update_layout(
        barmode="stack",
        height=540,
        margin=dict(t=130, b=70, l=50, r=40),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        legend=dict(
            orientation="h", yanchor="bottom", y=1.0,
            xanchor="center", x=0.5, font=dict(size=12),
            traceorder="normal",
        ),
        yaxis=dict(
            range=[0, 118], ticksuffix="%", dtick=20,
            gridcolor="#e5e7eb", zeroline=False,
            title=dict(text="% of Respondents", font=dict(size=12)),
        ),
        xaxis=dict(
            gridcolor="rgba(0,0,0,0)",
            tickfont=dict(size=11), tickangle=0,
        ),
        bargap=0.22,
        hoverlabel=dict(bgcolor="white", font_size=12),
    )
    return _theme_fig(fig)


def _render_funnel_with_arrows(brand_data: dict, visible_stages: list = None):
    """HTML awareness-depth funnel with content auto-fit widths and stage selection."""
    if visible_stages is None:
        visible_stages = ["Total Awareness", "Spontaneous Recall", "Consideration", "Top of Mind"]

    total_awa_pct  = float(brand_data.get("total_awareness_pct", brand_data.get("aided_pct", 0)) or 0)
    total_awa_n    = int(brand_data.get("total_awareness", brand_data.get("aided", 0)) or 0)

    aided_only_pct = float(brand_data.get("aided_pct", 0) or 0)
    aided_only_n   = int(brand_data.get("aided", 0) or 0)

    spont_pct  = float(brand_data.get("spont_pct", 0) or 0)
    consid_pct = float(brand_data.get("consideration_pct", 0) or 0)
    tom_pct    = float(brand_data.get("tom_pct",   0) or 0)

    spont_n    = int(brand_data.get("spont", 0) or 0)
    consid_n   = int(brand_data.get("consideration", 0) or 0)
    tom_n      = int(brand_data.get("tom",   0) or 0)

    all_stages = [
        ("Total Awareness", total_awa_pct, total_awa_n, PALETTE['aided'], "#166534", "#14532d"),
        ("Spontaneous Recall", spont_pct, spont_n, PALETTE['spont'], "#dcfce7", "#ffffff"),
        ("Consideration", consid_pct, consid_n, "#7c3aed", "#ede9fe", "#ffffff"),
        ("Top of Mind", tom_pct, tom_n, PALETTE['tom'], "#86efac", "#ffffff")
    ]

    filtered = [s for s in all_stages if s[0] in visible_stages]
    if not filtered:
        st.info("Select at least one stage to render the awareness funnel.")
        return

    def _arrow_color(pct: float) -> str:
        return "#22c55e" if pct >= 50 else ("#f59e0b" if pct >= 25 else "#ef4444")

    font_fam = CHART_THEME["font_family"]
    html = f'<div style="max-width:640px; margin:0 auto; padding:15px 0; font-family:{font_fam};">'

    # Width calculated strictly based on percentage of stage value (relative to top stage = 100%)
    top_pct = filtered[0][1] if filtered[0][1] > 0 else 100

    for idx, (label, pct, n, bg, sub_col, main_col) in enumerate(filtered):
        is_first = (idx == 0)
        is_last = (idx == len(filtered) - 1)
        
        # Percentage-proportional width (min 22% so n= and text fit inside)
        rel_w = max(22, min(100, round(pct / top_pct * 100)))

        br = f"{'14px 14px 4px 4px' if is_first else '4px'}"
        br_b = f"{'4px 4px 14px 14px' if is_last else '4px'}"

        if not is_first:
            prev_label = filtered[idx - 1][0]
            curr_label = label
            prev_pct = filtered[idx - 1][1]
            conv_rate = round(pct / prev_pct * 100, 1) if prev_pct > 0 else 0
            c_col = _arrow_color(conv_rate)

            p_short = prev_label.replace("Recall", "").replace("Awareness", "Aware").strip()
            c_short = curr_label.replace("Recall", "").replace("Awareness", "Aware").replace("Top of Mind", "TOM").strip()

            html += (
                f'<div style="display:flex; justify-content:center; padding:6px 0; position:relative;">'
                f'<div style="position:absolute; left:50%; top:0; bottom:0; width:2px;'
                f'background:linear-gradient({filtered[idx-1][3]},{bg}); transform:translateX(-50%); z-index:0;"></div>'
                f'<div style="position:relative; z-index:1; background:white; border:2px solid {c_col};'
                f'border-radius:30px; padding:4px 16px; display:inline-flex; align-items:center; gap:8px;'
                f'box-shadow:0 2px 8px rgba(0,0,0,0.10);">'
                f'<span style="font-size:1rem; color:{c_col}; font-weight:700;">↓</span>'
                f'<div style="text-align:left;">'
                f'<div style="font-size:0.95rem; font-weight:900; color:{c_col}; line-height:1;">{conv_rate:.1f}%</div>'
                f'<div style="font-size:0.45rem; color:#6b7280; text-transform:uppercase; font-weight:600;'
                f'letter-spacing:0.06em;">{p_short} → {c_short}</div></div></div></div>'
            )

        # Dynamically scale font size proportional to rel_w percentage so text auto-adjusts seamlessly
        title_fs = max(0.42, min(0.65, round(0.58 * (rel_w / 100) + 0.15, 2)))
        num_fs   = max(1.1,  min(2.1,  round(1.9  * (rel_w / 100) + 0.4, 2)))
        sub_fs   = max(0.45, min(0.62, round(0.58 * (rel_w / 100) + 0.12, 2)))

        html += (
            f'<div style="width:{rel_w}%; margin:0 auto; background:{bg}; border-radius:{br};'
            f'border-bottom-left-radius:{br_b.split()[2] if is_last else "4px"};'
            f'border-bottom-right-radius:{br_b.split()[3] if is_last else "4px"};'
            f'padding:10px 8px; text-align:center; box-shadow:0 2px 8px rgba(0,0,0,0.08); box-sizing:border-box; overflow:hidden;">'
            f'<div style="font-size:{title_fs}rem; font-weight:800; text-transform:uppercase; color:{sub_col};'
            f'letter-spacing:0.06em; margin-bottom:2px; white-space:nowrap; text-overflow:ellipsis; overflow:hidden;">{label}</div>'
            f'<div style="font-size:{num_fs}rem; font-weight:900; color:{main_col}; line-height:1;">{pct:.1f}%</div>'
            f'<div style="font-size:{sub_fs}rem; color:{sub_col}; margin-top:2px; white-space:nowrap;">n={n:,}</div></div>'
        )





    html += "</div>"
    st.markdown(html, unsafe_allow_html=True)


def _render_conversion_funnel(brand_data: dict, visible_stages: list = None):
    """HTML PPT-Aligned Brand Funnel — 4-stage with auto-fit widths and stage selection."""
    if visible_stages is None:
        visible_stages = ["Total Awareness", "Ever Tried (Trial)", "Current Usage", "Most Often Used Brand (MOUB)"]

    total_awa_pct = float(brand_data.get("total_awareness_pct", brand_data.get("aided_pct", 0)) or 0)
    ever_used_pct = float(brand_data.get("ever_used_pct", 0) or 0)
    current_pct   = float(brand_data.get("current_pct", 0) or 0)
    preferred_pct = float(brand_data.get("preferred_pct", 0) or 0)

    total_awa_n = int(brand_data.get("total_awareness", brand_data.get("aided", 0)) or 0)
    ever_used_n = int(brand_data.get("ever_used", 0) or 0)
    current_n   = int(brand_data.get("current_use", 0) or 0)
    preferred_n = int(brand_data.get("preferred", 0) or 0)

    all_stages = [
        ("Total Awareness", total_awa_pct, total_awa_n, "#047857", "#a7f3d0", "#ffffff"),
        ("Ever Tried (Trial)", ever_used_pct, ever_used_n, "#0e7490", "#a5f3fc", "#ffffff"),
        ("Current Usage", current_pct, current_n, "#1d4ed8", "#bfdbfe", "#ffffff"),
        ("Most Often Used Brand (MOUB)", preferred_pct, preferred_n, "#4338ca", "#c7d2fe", "#ffffff")
    ]

    filtered = [s for s in all_stages if s[0] in visible_stages]
    if not filtered:
        st.info("Select at least one stage to render the conversion funnel.")
        return

    def _ac(pct): return "#22c55e" if pct >= 50 else ("#f59e0b" if pct >= 25 else "#ef4444")

    font_fam = CHART_THEME["font_family"]
    html = f'<div style="max-width:600px; margin:0 auto; padding:15px 0; font-family:{font_fam};">'

    # Width calculated strictly based on percentage of stage value (relative to top stage = 100%)
    top_pct = filtered[0][1] if filtered[0][1] > 0 else 100

    for idx, (label, pct, n, bg, sub_col, main_col) in enumerate(filtered):
        is_first = (idx == 0)
        is_last = (idx == len(filtered) - 1)
        
        # Percentage-proportional width (min 22% so n= and text fit inside)
        rel_w = max(22, min(100, round(pct / top_pct * 100)))

        br = f"{'14px 14px 4px 4px' if is_first else '4px'}"
        br_b = f"{'4px 4px 14px 14px' if is_last else '4px'}"

        if not is_first:
            prev_label = filtered[idx - 1][0]
            curr_label = label
            prev_pct = filtered[idx - 1][1]
            conv_rate = round(pct / prev_pct * 100, 1) if prev_pct > 0 else 0
            c_col = _ac(conv_rate)

            p_short = prev_label.replace(" (Trial)", "").replace(" (MOUB)", "").replace("Awareness", "Aware").strip()
            c_short = curr_label.replace(" (Trial)", "").replace(" (MOUB)", "").replace("Awareness", "Aware").replace("Most Often Used Brand", "Preferred").strip()

            html += (
                f'<div style="display:flex; justify-content:center; padding:6px 0; position:relative;">'
                f'<div style="position:absolute; left:50%; top:0; bottom:0; width:2px;'
                f'background:linear-gradient({filtered[idx-1][3]},{bg}); transform:translateX(-50%); z-index:0;"></div>'
                f'<div style="position:relative; z-index:1; background:white; border:2px solid {c_col};'
                f'border-radius:30px; padding:4px 16px; display:inline-flex; align-items:center; gap:8px;'
                f'box-shadow:0 2px 8px rgba(0,0,0,0.10);">'
                f'<span style="font-size:1rem; color:{c_col}; font-weight:700;">↓</span>'
                f'<div style="text-align:left;">'
                f'<div style="font-size:0.95rem; font-weight:900; color:{c_col}; line-height:1;">{conv_rate:.1f}%</div>'
                f'<div style="font-size:0.45rem; color:#6b7280; text-transform:uppercase; font-weight:600;'
                f'letter-spacing:0.06em;">{p_short} → {c_short}</div></div></div></div>'
            )

        # Dynamically scale font size proportional to rel_w percentage so text auto-adjusts seamlessly
        title_fs = max(0.42, min(0.65, round(0.58 * (rel_w / 100) + 0.15, 2)))
        num_fs   = max(1.1,  min(2.1,  round(1.9  * (rel_w / 100) + 0.4, 2)))
        sub_fs   = max(0.45, min(0.62, round(0.58 * (rel_w / 100) + 0.12, 2)))

        html += (
            f'<div style="width:{rel_w}%; margin:0 auto; background:{bg}; border-radius:{br};'
            f'border-bottom-left-radius:{br_b.split()[2] if is_last else "4px"};'
            f'border-bottom-right-radius:{br_b.split()[3] if is_last else "4px"};'
            f'padding:10px 8px; text-align:center; box-shadow:0 2px 8px rgba(0,0,0,0.08); box-sizing:border-box; overflow:hidden;">'
            f'<div style="font-size:{title_fs}rem; font-weight:800; text-transform:uppercase; color:{sub_col};'
            f'letter-spacing:0.06em; margin-bottom:2px; white-space:nowrap; text-overflow:ellipsis; overflow:hidden;">{label}</div>'
            f'<div style="font-size:{num_fs}rem; font-weight:900; color:{main_col}; line-height:1;">{pct:.1f}%</div>'
            f'<div style="font-size:{sub_fs}rem; color:{sub_col}; margin-top:2px; white-space:nowrap;">n={n:,}</div></div>'
        )





    html += "</div>"
    st.markdown(html, unsafe_allow_html=True)




def _funnel_png_bytes(brand_data) -> bytes:
    """
    Matplotlib awareness funnel PNG — matches HTML funnel visual.
    Uses data coordinates with explicit non-overlapping y positions.
    """
    import io
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches

    total_awa_pct = float(brand_data.get("total_awareness_pct", brand_data.get("aided_pct", 0)) or 0)
    spont_pct     = float(brand_data.get("spont_pct", 0) or 0)
    consid_pct    = float(brand_data.get("consideration_pct", 0) or 0)
    tom_pct       = float(brand_data.get("tom_pct",   0) or 0)
    total_awa_n   = int(brand_data.get("total_awareness", brand_data.get("aided", 0)) or 0)
    spont_n       = int(brand_data.get("spont", 0) or 0)
    consid_n      = int(brand_data.get("consideration", 0) or 0)
    tom_n         = int(brand_data.get("tom",   0) or 0)
    brand_name    = brand_data.get("brand_name", "")

    conv_to_spont = round(spont_pct  / total_awa_pct * 100, 1) if total_awa_pct > 0 else 0
    conv_to_cons  = round(consid_pct / spont_pct  * 100, 1) if spont_pct  > 0 else 0
    conv_to_tom   = round(tom_pct    / consid_pct * 100, 1) if consid_pct > 0 else 0

    # Bar widths (0–1 scale, total awareness always full width)
    spont_w = max(0.30, min(0.95, spont_pct  / total_awa_pct)) if total_awa_pct > 0 else 0.55
    cons_w  = max(0.25, min(0.92, consid_pct / total_awa_pct)) if total_awa_pct > 0 else 0.45
    tom_w   = max(0.20, min(0.85, tom_pct    / total_awa_pct)) if total_awa_pct > 0 else 0.35

    COLOR_AIDED = "#bbf7d0"
    COLOR_SPONT = "#16a34a"
    COLOR_CONS  = "#7c3aed"
    COLOR_TOM   = "#166534"
    TEXT_DARK   = "#14532d"
    TEXT_LIGHT  = "white"

    def conv_color(pct):
        return "#22c55e" if pct >= 50 else ("#f59e0b" if pct >= 25 else "#ef4444")

    # ── Layout in data coords ──────────────────────────────────────────────────
    # matplotlib y = 0 at BOTTOM, increases upward.
    # We work in a 100-unit tall space so values are easy to reason about.
    # Bars are DRAWN with (x0, y_bottom) and height BH. 4 stages stacked bottom-up:
    # TOM (bottom) → Consideration → Spontaneous → Total Awareness (top).
    BH   = 12   # bar height
    GAP  = 9    # gap between bar bottom and next bar top  (badge sits in centre)
    CX   = 50   # centre x
    FW   = 90   # full width (total awareness bar)
    PADX = 4    # left/right canvas padding

    Y_TOM_BOT    = 10
    Y_BADGE3     = Y_TOM_BOT + BH + GAP / 2
    Y_CONS_BOT   = Y_TOM_BOT + BH + GAP
    Y_BADGE2     = Y_CONS_BOT + BH + GAP / 2
    Y_SPONT_BOT  = Y_CONS_BOT + BH + GAP
    Y_BADGE1     = Y_SPONT_BOT + BH + GAP / 2
    Y_AIDED_BOT  = Y_SPONT_BOT + BH + GAP
    Y_TITLE      = Y_AIDED_BOT + BH + 5

    fig, ax = plt.subplots(figsize=(7, 6.5), facecolor="white")
    ax.set_xlim(PADX, 100 - PADX)
    ax.set_ylim(3, Y_TITLE + 6)
    ax.axis("off")
    fig.patch.set_facecolor("white")

    def draw_bar(cx, y_bot, bar_w, color, label, pct, n, tc):
        x0 = cx - bar_w / 2
        rect = mpatches.FancyBboxPatch(
            (x0, y_bot), bar_w, BH,
            boxstyle="round,pad=0.5",
            facecolor=color, edgecolor="none", zorder=3,
        )
        ax.add_patch(rect)
        # label (small caps style)
        ax.text(cx, y_bot + BH * 0.78, label,
                ha="center", va="center", fontsize=7, fontweight="bold",
                color=tc, alpha=0.9, zorder=4, linespacing=1)
        # big percentage
        ax.text(cx, y_bot + BH * 0.45, f"{pct:.1f}%",
                ha="center", va="center", fontsize=20, fontweight="black",
                color=tc, zorder=4)
        # sample size
        ax.text(cx, y_bot + BH * 0.13, f"n={n:,}",
                ha="center", va="center", fontsize=7,
                color=tc, alpha=0.75, zorder=4)

    def draw_badge(cx, cy, conv, label, bc):
        bw, bh = 44, 5.5
        rect = mpatches.FancyBboxPatch(
            (cx - bw / 2, cy - bh / 2), bw, bh,
            boxstyle="round,pad=0.5",
            facecolor="white", edgecolor=bc, linewidth=1.5, zorder=5,
        )
        ax.add_patch(rect)
        ax.text(cx - bw / 2 + 3, cy, f"↓  {conv:.1f}%",
                ha="left", va="center", fontsize=9.5, fontweight="black",
                color=bc, zorder=6)
        ax.text(cx + 3, cy, f"  {label}",
                ha="left", va="center", fontsize=6.8, color="#6b7280", zorder=6)

    # Draw title
    if brand_name:
        ax.text(CX, Y_TITLE, f"{brand_name}  —  Awareness Funnel",
                ha="center", va="bottom", fontsize=11, fontweight="bold",
                color="#111827", zorder=4)

    # Draw bars (bottom-up)
    draw_bar(CX, Y_AIDED_BOT, FW,             COLOR_AIDED, "TOTAL AWARENESS",   total_awa_pct,  total_awa_n,  TEXT_DARK)
    draw_badge(CX, Y_BADGE1,  conv_to_spont, "converted to spontaneous",         conv_color(conv_to_spont))
    draw_bar(CX, Y_SPONT_BOT, spont_w * FW,  COLOR_SPONT, "SPONTANEOUS RECALL", spont_pct,  spont_n,  TEXT_LIGHT)
    draw_badge(CX, Y_BADGE2,  conv_to_cons,  "entered consideration",            conv_color(conv_to_cons))
    draw_bar(CX, Y_CONS_BOT,  cons_w  * FW,  COLOR_CONS,  "CONSIDERATION",       consid_pct, consid_n, TEXT_LIGHT)
    draw_badge(CX, Y_BADGE3,  conv_to_tom,   "converted to top of mind",         conv_color(conv_to_tom))
    draw_bar(CX, Y_TOM_BOT,   tom_w   * FW,  COLOR_TOM,   "TOP OF MIND",        tom_pct,    tom_n,    TEXT_LIGHT)

    buf = io.BytesIO()
    plt.savefig(buf, format="png", dpi=160, bbox_inches="tight",
                facecolor="white", edgecolor="none")
    plt.close(fig)
    buf.seek(0)
    return buf.read()


def _conversion_funnel_png_bytes(brand_data) -> bytes:
    """
    Matplotlib conversion funnel PNG — matches _render_conversion_funnel's HTML visual
    (4-stage: Total Awareness → Ever Tried (Trial) → Current Usage → Most Often Used Brand (MOUB)).
    Separate from _funnel_png_bytes (which renders the awareness-depth funnel).
    """
    import io
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches

    total_awa_pct = float(brand_data.get("total_awareness_pct", brand_data.get("aided_pct", 0)) or 0)
    ever_used_pct = float(brand_data.get("ever_used_pct", 0) or 0)
    current_pct   = float(brand_data.get("current_pct", 0) or 0)
    preferred_pct = float(brand_data.get("preferred_pct", 0) or 0)
    total_awa_n   = int(brand_data.get("total_awareness", brand_data.get("aided", 0)) or 0)
    ever_used_n   = int(brand_data.get("ever_used", 0) or 0)
    current_n     = int(brand_data.get("current_use", 0) or 0)
    preferred_n   = int(brand_data.get("preferred", 0) or 0)
    brand_name    = brand_data.get("brand_name", "")

    stages = [
        ("TOTAL AWARENESS",         total_awa_pct, total_awa_n, "#047857"),
        ("EVER TRIED (TRIAL)",      ever_used_pct, ever_used_n, "#0e7490"),
        ("CURRENT USAGE",           current_pct,   current_n,   "#1d4ed8"),
        ("MOST OFTEN USED (MOUB)",  preferred_pct, preferred_n, "#4338ca"),
    ]

    def conv_color(pct):
        return "#22c55e" if pct >= 50 else ("#f59e0b" if pct >= 25 else "#ef4444")

    def _w(pct):
        return max(0.20, min(0.98, pct / total_awa_pct)) if total_awa_pct > 0 else 0.4

    BH, GAP, CX, FW, PADX = 12, 9, 50, 90, 4
    n = len(stages)
    y_bots = [10 + i * (BH + GAP) for i in range(n)][::-1]  # stage0 (widest) drawn topmost
    y_badges = [y_bots[i] + BH + GAP / 2 for i in range(1, n)][::-1]
    y_title = y_bots[0] + BH + 5

    fig, ax = plt.subplots(figsize=(7, 6.5), facecolor="white")
    ax.set_xlim(PADX, 100 - PADX)
    ax.set_ylim(3, y_title + 6)
    ax.axis("off")
    fig.patch.set_facecolor("white")

    def draw_bar(y_bot, bar_w, color, label, pct, cnt):
        x0 = CX - bar_w / 2
        ax.add_patch(mpatches.FancyBboxPatch((x0, y_bot), bar_w, BH, boxstyle="round,pad=0.5",
                                              facecolor=color, edgecolor="none", zorder=3))
        ax.text(CX, y_bot + BH * 0.78, label, ha="center", va="center", fontsize=7,
                fontweight="bold", color="white", alpha=0.9, zorder=4)
        ax.text(CX, y_bot + BH * 0.45, f"{pct:.1f}%", ha="center", va="center", fontsize=18,
                fontweight="black", color="white", zorder=4)
        ax.text(CX, y_bot + BH * 0.13, f"n={cnt:,}", ha="center", va="center", fontsize=7,
                color="white", alpha=0.75, zorder=4)

    def draw_badge(cy, conv, bc):
        bw, bh = 44, 5.5
        ax.add_patch(mpatches.FancyBboxPatch((CX - bw / 2, cy - bh / 2), bw, bh,
                                              boxstyle="round,pad=0.5", facecolor="white",
                                              edgecolor=bc, linewidth=1.5, zorder=5))
        ax.text(CX - bw / 2 + 3, cy, f"↓  {conv:.1f}%", ha="left", va="center", fontsize=9.5,
                fontweight="black", color=bc, zorder=6)

    if brand_name:
        ax.text(CX, y_title, f"{brand_name}  —  Conversion Funnel",
                ha="center", va="bottom", fontsize=11, fontweight="bold", color="#111827", zorder=4)

    for i, (label, pct, cnt, color) in enumerate(stages):
        bar_w = FW if i == 0 else _w(pct) * FW
        draw_bar(y_bots[i], bar_w, color, label, pct, cnt)
        if i > 0:
            prev_pct = stages[i - 1][1]
            conv = round(pct / prev_pct * 100, 1) if prev_pct > 0 else 0
            draw_badge(y_badges[i - 1], conv, conv_color(conv))

    buf = io.BytesIO()
    plt.savefig(buf, format="png", dpi=160, bbox_inches="tight", facecolor="white", edgecolor="none")
    plt.close(fig)
    buf.seek(0)
    return buf.read()


def _awareness_funnel_chart(brand_data):
    """
    Plotly awareness funnel — kept for comparison/fallback use.
    Primary download now uses _funnel_png_bytes() for visual fidelity.
    """
    total_awa_pct = float(brand_data.get("total_awareness_pct", brand_data.get("aided_pct", 0)) or 0)
    spont_pct = float(brand_data.get("spont_pct", 0) or 0)
    tom_pct   = float(brand_data.get("tom_pct",   0) or 0)
    total_awa_n = int(brand_data.get("total_awareness", brand_data.get("aided", 0)) or 0)
    spont_n   = int(brand_data.get("spont", 0) or 0)
    tom_n     = int(brand_data.get("tom",   0) or 0)

    conv_to_spont = round(spont_pct / total_awa_pct * 100, 1) if total_awa_pct > 0 else 0
    conv_to_tom   = round(tom_pct   / spont_pct * 100, 1) if spont_pct > 0 else 0

    def _conv_color(pct):
        return "#22c55e" if pct >= 50 else ("#f59e0b" if pct >= 25 else "#ef4444")

    stages = ["Total Awareness", "Spontaneous Recall", "Top of Mind"]
    values = [total_awa_pct, spont_pct, tom_pct]
    ns     = [total_awa_n, spont_n, tom_n]
    convs  = [None, conv_to_spont, conv_to_tom]
    conv_labels = [None, "converted from Total Aware", "converted from Spont"]

    inside_text = [
        f"<b>{total_awa_pct:.1f}%</b><br><span style='font-size:11px'>n={total_awa_n:,}</span>",
        f"<b>{spont_pct:.1f}%</b><br><span style='font-size:11px'>↓ {conv_to_spont:.1f}% from Total Aware · n={spont_n:,}</span>",
        f"<b>{tom_pct:.1f}%</b><br><span style='font-size:11px'>↓ {conv_to_tom:.1f}% from Spont · n={tom_n:,}</span>",
    ]

    fig = go.Figure(go.Funnel(
        y=stages,
        x=values,
        text=inside_text,
        textinfo="text",
        textposition="inside",
        connector=dict(
            fillcolor="rgba(180,180,180,0.18)",
            line=dict(color="rgba(150,150,150,0.3)", width=1),
        ),
        marker=dict(
            color=[PALETTE["aided"], PALETTE["spont"], PALETTE["tom"]],
            line=dict(width=0),
        ),
        textfont=dict(size=14, color="white", family="Inter, Arial, sans-serif"),
        hovertemplate="<b>%{y}</b><br>%{x:.1f}% of respondents<br>n=%{customdata:,}<extra></extra>",
        customdata=[aided_n, spont_n, tom_n],
    ))

    # Conversion rate annotations between stages
    for i, (conv, lbl) in enumerate(zip(convs, conv_labels)):
        if conv is None:
            continue
        col = _conv_color(conv)
        fig.add_annotation(
            x=0, y=stages[i],
            xref="paper", yref="y",
            yshift=-48,
            text=f"<b>↓ {conv:.1f}%</b> {lbl}",
            showarrow=False,
            font=dict(size=11, color=col, family="Inter, Arial, sans-serif"),
            xanchor="left",
            bgcolor=f"rgba(255,255,255,0.85)",
            bordercolor=col,
            borderwidth=1,
            borderpad=3,
        )

    fig.update_layout(
        height=420,
        margin=dict(t=20, b=20, l=160, r=60),
        paper_bgcolor="white",
        plot_bgcolor="white",
        font=dict(size=13, family="Inter, Arial, sans-serif"),
        yaxis=dict(tickfont=dict(size=13, color="#374151"), showgrid=False),
        funnelmode="stack",
    )
    return _theme_fig(fig)


def _nps_breakdown_tiles(brand_data) -> bool:
    """Renders robust NPS tiles using a single markdown block."""
    p_pct  = brand_data.get("nps_promoters_pct")
    pa_pct = brand_data.get("nps_passives_pct")
    d_pct  = brand_data.get("nps_detractors_pct")
    nps    = brand_data.get("nps")
    n_total = brand_data.get("nps_base", 0)
    
    if p_pct is None or pa_pct is None or d_pct is None:
        return False

    p_n = int(round(p_pct * n_total / 100))
    pa_n = int(round(pa_pct * n_total / 100))
    d_n = int(round(d_pct * n_total / 100))

    nps_color = "#22c55e" if (nps or 0) >= NPS_INDUSTRY_AVG else \
                "#f59e0b" if (nps or 0) >= 0 else "#ef4444"
    delta = round((nps or 0) - NPS_INDUSTRY_AVG, 1)
    sign  = "+" if delta >= 0 else ""

    html = f"""
    <div style="display:grid; grid-template-columns:1fr 1fr 1fr; gap:10px; margin-bottom:12px;">
        <div style="background:#f0fdf4; border:1px solid #dcfce7; border-radius:10px; padding:12px; text-align:center;">
            <div style="font-size:0.65rem; color:#166534; font-weight:700; text-transform:uppercase;">Promoters</div>
            <div style="font-size:1.6rem; font-weight:800; color:#15803d;">{p_pct:.0f}%</div>
            <div style="font-size:0.6rem; color:#16a34a;">n={p_n:,}</div>
        </div>
        <div style="background:#fffbeb; border:1px solid #fef3c7; border-radius:10px; padding:12px; text-align:center;">
            <div style="font-size:0.65rem; color:#92400e; font-weight:700; text-transform:uppercase;">Passives</div>
            <div style="font-size:1.6rem; font-weight:800; color:#d97706;">{pa_pct:.0f}%</div>
            <div style="font-size:0.6rem; color:#b45309;">n={pa_n:,}</div>
        </div>
        <div style="background:#fef2f2; border:1px solid #fee2e2; border-radius:10px; padding:12px; text-align:center;">
            <div style="font-size:0.65rem; color:#991b1b; font-weight:700; text-transform:uppercase;">Detractors</div>
            <div style="font-size:1.6rem; font-weight:800; color:#dc2626;">{d_pct:.0f}%</div>
            <div style="font-size:0.6rem; color:#b91c1c;">n={d_n:,}</div>
        </div>
    </div>
    <div style="background:#f9fafb; border:1px solid #e5e7eb; border-radius:10px; padding:14px; display:flex; justify-content:space-between; align-items:center;">
        <div>
            <div style="font-size:0.65rem; color:#6b7280; font-weight:700; text-transform:uppercase;">Net Promoter Score</div>
            <div style="font-size:2rem; font-weight:800; color:{nps_color};">{f"{nps:+.0f}" if nps is not None else "N/A"}</div>
        </div>
        <div style="text-align:right;">
            <div style="font-size:0.65rem; color:#9ca3af;">vs Industry +{NPS_INDUSTRY_AVG}</div>
            <div style="font-size:1rem; font-weight:700; color:{nps_color};">{sign}{delta} pts</div>
            <div style="font-size:0.6rem; color:#9ca3af;">n={n_total:,} raters</div>
        </div>
    </div>
    """
    st.markdown(html, unsafe_allow_html=True)
    return True


def _brand_radar_chart(brand_data: dict, brands_list: list, compare_data: dict = None):
    """Spider chart: selected brand vs comparison brand (or Top-5 avg).
    Performance zone rings at 33 and 66 show danger / average / leader bands.
    """
    def norm_nps(v):
        return round((v + 100) / 2, 1) if v is not None else 50.0

    # Rater Depth = each brand's NPS sample size relative to the LARGEST tracked
    # brand's base (self-normalising, 0-100). Dynamic — no hardcoded study size —
    # and consistent across focus / comparison / peer-average (was ×3 vs ×5: a bug).
    _max_base = max((b.get("nps_base", 0) or 0) for b in (brands_list or [])) or 1
    def _rater_depth(d):
        return min(round((d.get("nps_base", 0) or 0) / _max_base * 100, 1), 100)

    b_tom   = brand_data.get("tom_pct", 0)
    b_spont = brand_data.get("spont_pct", 0)
    b_aided = brand_data.get("aided_pct", 0)
    b_loyal = norm_nps(brand_data.get("nps"))
    b_rater = _rater_depth(brand_data)

    if compare_data:
        cmp_name = compare_data.get("brand_name", "Comparison")
        cmp_vals = [
            compare_data.get("tom_pct", 0),
            compare_data.get("spont_pct", 0),
            compare_data.get("aided_pct", 0),
            norm_nps(compare_data.get("nps")),
            _rater_depth(compare_data),
        ]
        cmp_fill = "rgba(234, 179, 8, 0.12)"
        cmp_line = dict(color="#f59e0b", width=2, dash="dot")
    else:
        top5 = sorted(brands_list, key=lambda x: x.get("aided_pct", 0), reverse=True)[:5]
        cmp_name = "Top-5 Avg"
        # Awareness components share a base → equal-weight mean is correct. NPS has
        # per-brand rater bases → respondent-weight it (Σ norm_nps·base / Σ base).
        _t5_nps_w = sum(norm_nps(b.get("nps")) * (b.get("nps_base", 0) or 0) for b in top5)
        _t5_nps_d = sum((b.get("nps_base", 0) or 0) for b in top5)
        _t5_nps = round(_t5_nps_w / _t5_nps_d, 1) if _t5_nps_d > 0 else round(
            sum(norm_nps(b.get("nps")) for b in top5) / len(top5), 1)
        cmp_vals = [
            round(sum(b["tom_pct"]   for b in top5) / len(top5), 1),
            round(sum(b["spont_pct"] for b in top5) / len(top5), 1),
            round(sum(b["aided_pct"] for b in top5) / len(top5), 1),
            _t5_nps,
            round(sum(_rater_depth(b) for b in top5) / len(top5), 1),
        ]
        cmp_fill = "rgba(156, 163, 175, 0.10)"
        cmp_line = dict(color="#9ca3af", width=1.5, dash="dot")

    axes       = ["Salience<br>(TOM %)", "Recall<br>(SPONT %)", "Total<br>Reach %",
                  "Loyalty<br>(NPS norm)", "Rater<br>Depth"]
    brand_vals = [b_tom, b_spont, b_aided, b_loyal, b_rater]
    brand_name = brand_data.get("brand_name", "Brand")

    fig = go.Figure()

    # ── Zone reference rings ──────────────────────────────────────────────────
    # Ring at 66 = "Leader territory" (light green fill), 33 = "Average" (light amber)
    for r_val, r_color, r_label in [
        (66, "rgba(220,252,231,0.35)", "Leader"),
        (33, "rgba(254,249,195,0.35)", "Average"),
    ]:
        ring_r = [r_val] * 5
        fig.add_trace(go.Scatterpolar(
            r=ring_r + [ring_r[0]],
            theta=axes + [axes[0]],
            fill="toself",
            fillcolor=r_color,
            line=dict(color="rgba(0,0,0,0)", width=0),
            showlegend=False,
            hoverinfo="skip",
        ))

    # ── Comparison trace ─────────────────────────────────────────────────────
    fig.add_trace(go.Scatterpolar(
        r=cmp_vals + [cmp_vals[0]],
        theta=axes + [axes[0]],
        fill="toself",
        name=cmp_name,
        fillcolor=cmp_fill,
        line=cmp_line,
        hovertemplate="%{theta}: %{r:.1f}<extra></extra>",
    ))

    # ── Primary brand trace ───────────────────────────────────────────────────
    fig.add_trace(go.Scatterpolar(
        r=brand_vals + [brand_vals[0]],
        theta=axes + [axes[0]],
        fill="toself",
        name=brand_name,
        fillcolor="rgba(26, 93, 77, 0.22)",
        line=dict(color="#1a5d4d", width=2.8),
        marker=dict(size=7, color="#1a5d4d"),
        hovertemplate="%{theta}: %{r:.1f}<extra></extra>",
    ))

    # ── Zone ring legend entries (added as dummy scatter traces, hidden from chart) ─
    # These appear in the legend to explain the reference rings
    fig.add_trace(go.Scatterpolar(
        r=[None], theta=[axes[0]],
        name="─ Leader zone (66+)",
        line=dict(color="#22c55e", width=2, dash="dot"),
        marker=dict(size=0),
        showlegend=True, mode="lines",
    ))
    fig.add_trace(go.Scatterpolar(
        r=[None], theta=[axes[0]],
        name="─ Average zone (33+)",
        line=dict(color="#f59e0b", width=2, dash="dot"),
        marker=dict(size=0),
        showlegend=True, mode="lines",
    ))

    fig.update_layout(
        polar=dict(
            bgcolor="rgba(0,0,0,0)",
            radialaxis=dict(
                visible=True, range=[0, 100],
                gridcolor="#d1d5db", linecolor="#d1d5db",
                tickfont=dict(size=8), showticklabels=True,
                tickvals=[0, 33, 66, 100],
            ),
            angularaxis=dict(
                gridcolor="#e5e7eb", linecolor="#e5e7eb",
                tickfont=dict(size=10),
            ),
        ),
        showlegend=True,
        legend=dict(orientation="h", yanchor="bottom", y=1.10, xanchor="center", x=0.5,
                    font=dict(size=9), itemwidth=40),
        paper_bgcolor="rgba(0,0,0,0)",
        margin=dict(t=75, b=25, l=40, r=40),
        height=460,
        font=dict(size=11),
    )
    return _theme_fig(fig)


def _brand_positioning_chart(brands_list: list, sel_brand: str):
    """
    Bubble chart: TOM% (x) vs NPS (y), bubble size = aided%, colour = NPS intensity.
    Coloured quadrant backgrounds + brand count badges per quadrant.
    Quadrant split: median TOM (x) × NPS industry avg (y).
    """
    eligible = [b for b in brands_list
                if b.get("nps") is not None and b.get("tom_pct", 0) > 0]
    if len(eligible) < 3:
        return None

    names  = [b["brand_name"]               for b in eligible]
    x_vals = [b["tom_pct"]                  for b in eligible]
    y_vals = [float(b["nps"])               for b in eligible]
    sizes  = [max(b["aided_pct"] * 0.85, 8) for b in eligible]

    import statistics as _stats
    x_vals_sorted = sorted(x_vals)
    median_tom = round(_stats.median(x_vals_sorted), 1)
    x_max = max(x_vals) * 1.20
    x_min = -0.5
    y_max = max(y_vals) + 14
    y_min = min(y_vals) - 10

    # Quadrant membership
    def _quadrant(x, y):
        if x >= median_tom and y >= NPS_INDUSTRY_AVG: return "leaders"
        if x < median_tom  and y >= NPS_INDUSTRY_AVG: return "gems"
        if x >= median_tom and y <  NPS_INDUSTRY_AVG: return "awareness"
        return "growth"

    quad_counts = {"leaders": 0, "gems": 0, "awareness": 0, "growth": 0}
    for x, y in zip(x_vals, y_vals):
        quad_counts[_quadrant(x, y)] += 1

    fig = go.Figure()

    # ── Coloured quadrant backgrounds ─────────────────────────────────────────
    fig.add_shape(type="rect", x0=median_tom, x1=x_max,
                  y0=NPS_INDUSTRY_AVG, y1=y_max,
                  fillcolor="rgba(220,252,231,0.45)", line_width=0, layer="below")
    fig.add_shape(type="rect", x0=x_min, x1=median_tom,
                  y0=NPS_INDUSTRY_AVG, y1=y_max,
                  fillcolor="rgba(219,234,254,0.45)", line_width=0, layer="below")
    fig.add_shape(type="rect", x0=median_tom, x1=x_max,
                  y0=y_min, y1=NPS_INDUSTRY_AVG,
                  fillcolor="rgba(254,243,199,0.45)", line_width=0, layer="below")
    fig.add_shape(type="rect", x0=x_min, x1=median_tom,
                  y0=y_min, y1=NPS_INDUSTRY_AVG,
                  fillcolor="rgba(243,244,246,0.45)", line_width=0, layer="below")

    # ── Non-selected brands: colour by NPS value (green→amber→red) ────────────
    other_idx = [i for i, n in enumerate(names) if n != sel_brand]
    sel_idx   = [i for i, n in enumerate(names) if n == sel_brand]

    def _nps_color(nps_val):
        if nps_val >= NPS_INDUSTRY_AVG: return "#22c55e"
        if nps_val >= 0:                return "#f59e0b"
        return "#ef4444"

    other_colors = [_nps_color(y_vals[i]) for i in other_idx]

    fig.add_trace(go.Scatter(
        x=[x_vals[i] for i in other_idx],
        y=[y_vals[i] for i in other_idx],
        mode="markers+text",
        name="Other brands",
        text=[names[i] for i in other_idx],
        textposition="top center",
        textfont=dict(size=9, color="#4b5563"),
        marker=dict(
            size=[sizes[i] for i in other_idx],
            color=other_colors,
            opacity=0.72,
            line=dict(width=1, color="white"),
        ),
        hovertemplate=(
            "<b>%{text}</b><br>"
            "TOM: %{x:.1f}%  |  NPS: %{y:+.0f}<br>"
            "<i>Bubble size = Aided awareness %</i>"
            "<extra></extra>"
        ),
    ))
    if sel_idx:
        i = sel_idx[0]
        fig.add_trace(go.Scatter(
            x=[x_vals[i]], y=[y_vals[i]],
            mode="markers+text",
            name=sel_brand,
            text=[f"★ {sel_brand}"],
            textposition="top center",
            textfont=dict(size=11, color="#1a5d4d", family="Arial Black"),
            marker=dict(
                size=max(sizes[i] * 1.1, 22),
                color="#1a5d4d", opacity=1.0,
                line=dict(width=2.5, color="white"),
                symbol="star",
            ),
            hovertemplate=(
                f"<b>{sel_brand}</b><br>"
                "TOM: %{x:.1f}%  |  NPS: %{y:+.0f}"
                "<extra></extra>"
            ),
        ))

    # ── Quadrant dividers ─────────────────────────────────────────────────────
    fig.add_vline(x=median_tom, line_dash="dot", line_color="#9ca3af", line_width=1.5,
                  annotation_text=f"Median TOM {median_tom:.1f}%",
                  annotation_position="bottom right", annotation_font_size=9)
    fig.add_hline(y=NPS_INDUSTRY_AVG, line_dash="dot", line_color="#9ca3af", line_width=1.5,
                  annotation_text=f"Industry NPS avg +{NPS_INDUSTRY_AVG}",
                  annotation_position="top right", annotation_font_size=9)

    # ── Quadrant labels with brand counts ─────────────────────────────────────
    x_r = x_max * 0.94   # right quadrants
    x_l = x_min + (median_tom - x_min) * 0.15  # left quadrants
    y_t = y_max - 2      # top row
    y_b = y_min + 2      # bottom row

    for lx, ly, label, count, color in [
        (x_r, y_t, "Market Leaders",       quad_counts["leaders"],   "#14532d"),
        (x_l, y_t, "Loyalty Hidden Gems",  quad_counts["gems"],      "#1e3a8a"),
        (x_r, y_b, "Awareness Leaders",    quad_counts["awareness"], "#78350f"),
        (x_l, y_b, "Growth Opportunities", quad_counts["growth"],    "#374151"),
    ]:
        fig.add_annotation(
            x=lx, y=ly,
            text=f"<b>{label}</b>  n={count}",
            showarrow=False, font=dict(size=9, color=color, family="Arial"),
            xanchor="right" if lx > median_tom else "left",
            bgcolor="rgba(255,255,255,0.88)", borderpad=4,
            bordercolor=color, borderwidth=1,
        )

    # Bubble size note
    fig.add_annotation(xref="paper", yref="paper", x=0.0, y=-0.1,
                       text="<i>Bubble size = Aided awareness %</i>",
                       showarrow=False, font=dict(size=9, color="#6b7280"), xanchor="left")

    fig.update_layout(
        height=500,
        margin=dict(t=30, b=55, l=55, r=30),
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        xaxis=dict(title="Top-of-Mind % (Salience)", gridcolor="#f3f4f6",
                   ticksuffix="%", range=[x_min, x_max], zeroline=False),
        yaxis=dict(title="Net Promoter Score (Loyalty)", gridcolor="#f3f4f6",
                   range=[y_min, y_max], zeroline=False),
        showlegend=True,
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="center", x=0.5,
                    font=dict(size=10)),
        font=dict(size=11),
        hoverlabel=dict(bgcolor="white", font_size=12),
    )
    return _theme_fig(fig)


def _nps_rankings_chart(brands_list, min_raters=30, top_n=15, highlight: str = None):
    """
    Horizontal NPS league table.
    Zone backgrounds: red (NPS<0) / amber (0–industry avg) / green (above avg).
    Black header strip with zone tier labels. Rank # in y-axis labels.
    """
    eligible = [b for b in brands_list
                if b.get("nps") is not None and b.get("nps_base", 0) >= min_raters]
    if not eligible:
        return None
    eligible.sort(key=lambda x: x["nps"], reverse=True)
    rank_map = {b["brand_name"]: i + 1 for i, b in enumerate(eligible)}
    top = eligible[:top_n]
    if highlight and not any(b["brand_name"] == highlight for b in top):
        hl_brand = next((b for b in eligible if b["brand_name"] == highlight), None)
        if hl_brand:
            top.append(hl_brand)
            top.sort(key=lambda x: x["nps"], reverse=True)

    top_rev = list(reversed(top))
    names  = [b["brand_name"] for b in top_rev]
    scores = [b["nps"]        for b in top_rev]
    bases  = [b["nps_base"]   for b in top_rev]

    x_min = min(scores) - 22
    x_max = max(scores) + 32

    # Rank-prefixed y-axis labels
    y_labels = [f"#{rank_map.get(n, '?')}  {n}" for n in names]

    # Colour by NPS tier
    colors = []
    for b in top_rev:
        if highlight and b["brand_name"] == highlight:
            colors.append("#1a5d4d")       # selected brand — dark teal
        elif b["nps"] >= NPS_INDUSTRY_AVG:
            colors.append("#16a34a")       # champion — rich green
        elif b["nps"] >= 0:
            colors.append("#f59e0b")       # developing — amber
        else:
            colors.append("#ef4444")       # detractor zone — red

    border_colors = ["#0a2e22" if (highlight and n == highlight) else "rgba(0,0,0,0)" for n in names]
    border_widths = [2        if (highlight and n == highlight) else 0               for n in names]
    text_colors   = ["#1a5d4d" if (highlight and n == highlight) else "#6b7280"       for n in names]
    text_sizes    = [13        if (highlight and n == highlight) else 11               for n in names]

    fig = go.Figure()

    # ── NPS zone background bands — ultra-subtle so bars stay readable ─────────
    fig.add_vrect(x0=x_min, x1=0,
                  fillcolor="rgba(239,68,68,0.07)", opacity=1, layer="below", line_width=0)
    fig.add_vrect(x0=0, x1=NPS_INDUSTRY_AVG,
                  fillcolor="rgba(245,158,11,0.07)", opacity=1, layer="below", line_width=0)
    fig.add_vrect(x0=NPS_INDUSTRY_AVG, x1=x_max,
                  fillcolor="rgba(34,197,94,0.07)", opacity=1, layer="below", line_width=0)

    # ── Zero line ────────────────────────────────────────────────────────────
    fig.add_vline(x=0, line_color="#d1d5db", line_width=1.2)

    # ── Bars — white outline separates bars from subtle backgrounds ───────────
    bar_borders = border_colors[:]
    bar_border_w = border_widths[:]
    for i, (bc, bw) in enumerate(zip(border_colors, border_widths)):
        if bw == 0:
            bar_borders[i] = "rgba(255,255,255,0.8)"
            bar_border_w[i] = 0.6

    fig.add_trace(go.Bar(
        y=y_labels, x=scores, orientation="h",
        marker=dict(color=colors, opacity=0.88,
                    line=dict(color=bar_borders, width=bar_border_w)),
        text=[f"<b>{s:+.0f}</b>" if (highlight and names[i] == highlight) else f"{s:+.0f}"
              for i, s in enumerate(scores)],
        textposition="outside",
        textfont=dict(size=text_sizes, color=text_colors),
        customdata=bases,
        hovertemplate=(
            "<b>%{y}</b><br>"
            "NPS: %{x:+.0f}<br>"
            "Raters: %{customdata:,}"
            "<extra></extra>"
        ),
    ))

    # ── Industry avg reference line ───────────────────────────────────────────
    fig.add_vline(x=NPS_INDUSTRY_AVG, line_dash="dot", line_color="#6b7280", line_width=1.8)
    fig.add_annotation(
        xref="x", yref="paper", x=NPS_INDUSTRY_AVG, y=-0.06,
        text=f"<b>Ind. avg +{NPS_INDUSTRY_AVG}</b>", showarrow=False,
        font=dict(size=9, color="#6b7280"), xanchor="center",
        bgcolor="rgba(255,255,255,0.85)", borderpad=2,
    )

    # ── Light teal header with NPS zone labels (no black) ─────────────────────
    fig.add_shape(type="rect", xref="paper", yref="paper",
                  x0=0, y0=1.06, x1=1.0, y1=1.22,
                  fillcolor="rgba(26,93,77,0.10)", line=dict(color="#d1fae5", width=1))
    for tx, tlbl, tc in [
        (0.12, "← Detractor",   "#dc2626"),
        (0.40, "Developing",    "#d97706"),
        (0.72, "Champion →",    "#16a34a"),
    ]:
        fig.add_annotation(
            xref="paper", yref="paper", x=tx, y=1.14,
            text=f"<b>{tlbl}</b>", showarrow=False,
            font=dict(size=9, color=tc), xanchor="center",
        )

    # ── Subtitle caption ──────────────────────────────────────────────────────
    fig.add_annotation(
        xref="paper", yref="paper", x=1.0, y=1.045,
        text="<i>NPS pts  ·  rank among all eligible brands</i>", showarrow=False,
        font=dict(size=9, color="#9ca3af"), xanchor="right",
    )

    fig.update_layout(
        height=max(440, len(top) * 36),
        margin=dict(t=80, b=55, l=10, r=95),
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        xaxis=dict(title="Net Promoter Score", range=[x_min, x_max],
                   gridcolor="#f3f4f6", zeroline=False),
        yaxis=dict(gridcolor="rgba(0,0,0,0)", tickfont=dict(size=10.5)),
        font=dict(size=11),
        bargap=0.26,
        hoverlabel=dict(bgcolor="white", font_size=12),
    )
    return _theme_fig(fig)


def _zone_comparison_chart(zone_data: dict, brand_name: str):
    """
    Grouped bar: TOM / SPONT / AIDED per zone.
    Zone-colored group headers. All-zone avg reference lines per metric.
    """
    zones = list(zone_data.keys())
    if not zones:
        return go.Figure()

    tom_vals   = [zone_data[z].get("tom_pct",   0) for z in zones]
    spont_vals = [zone_data[z].get("spont_pct", 0) for z in zones]
    aided_vals = [zone_data[z].get("aided_pct", 0) for z in zones]

    # All-zone averages for reference lines
    avg_tom   = round(sum(tom_vals)   / len(zones), 1)
    avg_spont = round(sum(spont_vals) / len(zones), 1)
    avg_aided = round(sum(aided_vals) / len(zones), 1)

    fig = go.Figure()

    fig.add_trace(go.Bar(
        name="Top of Mind", x=zones, y=tom_vals,
        marker_color=PALETTE["tom"],
        text=[f"<b>{v}</b>%" for v in tom_vals],
        textposition="outside",
        textfont=dict(size=11),
        hovertemplate="<b>%{x}</b>  TOM: %{y:.1f}%<extra></extra>",
    ))
    fig.add_trace(go.Bar(
        name="Spontaneous", x=zones, y=spont_vals,
        marker_color=PALETTE["spont"],
        text=[f"<b>{v}</b>%" for v in spont_vals],
        textposition="outside",
        textfont=dict(size=11),
        hovertemplate="<b>%{x}</b>  Spont: %{y:.1f}%<extra></extra>",
    ))
    fig.add_trace(go.Bar(
        name="Aided", x=zones, y=aided_vals,
        marker_color=PALETTE["aided"],
        text=[f"<b>{v}</b>%" for v in aided_vals],
        textposition="outside",
        textfont=dict(size=11),
        hovertemplate="<b>%{x}</b>  Aided: %{y:.1f}%<extra></extra>",
    ))

    # Zone avg reference lines
    for avg_val, color, lbl in [
        (avg_tom,   "#1a5d4d", f"TOM avg {avg_tom:.0f}%"),
        (avg_aided, "#86efac", f"Aided avg {avg_aided:.0f}%"),
    ]:
        fig.add_hline(y=avg_val, line_dash="dot", line_color=color, line_width=1.4,
                      annotation_text=lbl, annotation_position="right",
                      annotation_font_size=9, annotation_font_color=color)

    # Chart title annotations (no black background)
    fig.add_annotation(xref="paper", yref="paper", x=0.0, y=1.16,
                       text=f"<b>{brand_name}  ·  Awareness by Zone</b>",
                       showarrow=False, font=dict(size=11, color="#1a5d4d"), xanchor="left")
    fig.add_annotation(xref="paper", yref="paper", x=1.0, y=1.085,
                       text="<i>Data in %  ·  dashed = zone avg</i>",
                       showarrow=False, font=dict(size=9, color="#6366f1"), xanchor="right")

    # Zone-coloured x-tick callout annotations below each group
    # Use mixed xref="x"/yref="paper" so labels always show below bars
    for z in zones:
        zc = ZONE_COLORS.get(z, "#9ca3af")
        fig.add_annotation(
            x=z, xref="x", y=-0.08, yref="paper",
            text=f"<b>{z}</b>",
            showarrow=False,
            font=dict(size=11, color=zc), xanchor="center",
            bgcolor="rgba(255,255,255,0.85)", borderpad=2,
        )

    fig.update_layout(
        barmode="group", height=400,
        margin=dict(t=95, b=60, l=15, r=90),
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        xaxis=dict(gridcolor="rgba(0,0,0,0)", tickfont=dict(size=11, color="rgba(0,0,0,0)")),
        yaxis=dict(title="% of Respondents", gridcolor="#f3f4f6",
                   ticksuffix="%", range=[0, 105]),
        legend=dict(orientation="h", yanchor="bottom", y=1.23,
                    xanchor="center", x=0.5, font=dict(size=11)),
        font=dict(size=11),
        bargap=0.20,
        hoverlabel=dict(bgcolor="white", font_size=12),
    )
    return _theme_fig(fig)


def _zone_nps_chart(zone_data: dict, brand_name: str):
    """
    Horizontal NPS bar per zone. Zone-coloured bars. NPS zone backgrounds.
    """
    zones_with_nps = [(z, d) for z, d in zone_data.items() if d.get("nps") is not None]
    if not zones_with_nps:
        return None
    zones_with_nps.sort(key=lambda x: x[1]["nps"], reverse=True)
    names  = [z             for z, _ in zones_with_nps]
    scores = [d["nps"]      for _, d in zones_with_nps]
    bases  = [d.get("nps_base", 0) for _, d in zones_with_nps]
    colors = [ZONE_COLORS.get(z, "#9ca3af") for z in names]

    x_min = min(min(scores) - 12, -5)
    x_max = max(scores) + 14

    fig = go.Figure()

    # Zone-colored bars carry semantic meaning; omit NPS zone backgrounds to avoid conflict.
    fig.add_vline(x=0, line_color="#d1d5db", line_width=1.2)

    fig.add_trace(go.Bar(
        y=names, x=scores, orientation="h",
        marker=dict(color=colors, opacity=0.92, line=dict(color="white", width=1)),
        text=[f"<b>{s:+.0f}</b>" for s in scores], textposition="outside",
        textfont=dict(size=12, color="#374151"),
        customdata=bases,
        hovertemplate="<b>%{y} Zone</b><br>NPS: %{x:+.0f}  (n=%{customdata:,})<extra></extra>",
    ))
    fig.add_vline(x=NPS_INDUSTRY_AVG, line_dash="dot", line_color="#6b7280", line_width=1.8)
    fig.add_annotation(
        xref="x", yref="paper", x=NPS_INDUSTRY_AVG, y=-0.06,
        text=f"<b>Ind. avg +{NPS_INDUSTRY_AVG}</b>", showarrow=False,
        font=dict(size=9, color="#6b7280"), xanchor="center",
        bgcolor="rgba(255,255,255,0.85)", borderpad=2,
    )

    # Chart title — no black background
    fig.add_annotation(xref="paper", yref="paper", x=0.0, y=1.16,
                       text=f"<b>{brand_name}  ·  NPS by Zone</b>",
                       showarrow=False, font=dict(size=11, color="#1a5d4d"), xanchor="left")

    fig.update_layout(
        height=280,
        margin=dict(t=65, b=45, l=15, r=60),
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        xaxis=dict(title="Net Promoter Score", range=[x_min, x_max],
                   gridcolor="#f3f4f6", zeroline=False),
        yaxis=dict(gridcolor="rgba(0,0,0,0)", tickfont=dict(size=11)),
        font=dict(size=11),
        bargap=0.30,
        hoverlabel=dict(bgcolor="white", font_size=12),
    )
    return _theme_fig(fig)


def _india_zone_map(zone_data: dict, brand_name: str):
    """
    India zone awareness map — bubble size = respondent base,
    colour intensity = TOM%. Shows WHERE awareness is concentrated.
    Geographic footprint is an awareness story, not NPS.
    """
    ZONE_CENTERS = {
        "North": (28.6, 77.2),
        "South": (12.9, 77.6),
        "East":  (22.6, 88.4),
        "West":  (19.1, 72.9),
    }
    lats, lons, toms, aides, sponts, bases, hovers, bubble_labels = [], [], [], [], [], [], [], []

    for zone, d in zone_data.items():
        if zone not in ZONE_CENTERS:
            continue
        lat, lon = ZONE_CENTERS[zone]
        tom   = d.get("tom_pct", 0)
        aided = d.get("aided_pct", 0)
        spont = d.get("spont_pct", 0)
        base  = d.get("zone_base", 0)
        lats.append(lat); lons.append(lon)
        toms.append(tom); aides.append(aided); sponts.append(spont); bases.append(base)
        # Label: zone name only — bubble color/size already encodes the data
        bubble_labels.append(zone)
        hovers.append(
            f"<b>{zone} India</b><br>"
            f"Top of Mind: <b>{tom}%</b><br>"
            f"Spontaneous: {spont}%<br>"
            f"Aided: {aided}%<br>"
            f"Respondents: {base:,}"
        )

    if not lats:
        return None

    # Bubble size proportional to respondent base (market size)
    max_base = max(bases) if bases and max(bases) > 0 else 1
    bubble_sizes = [max(b / max_base * 55 + 20, 22) for b in bases]

    fig = go.Figure()

    # Main awareness bubbles
    fig.add_trace(go.Scattergeo(
        lat=lats, lon=lons,
        mode="markers+text",
        text=bubble_labels,
        textposition="top center",
        textfont=dict(size=10, color="#111827"),
        hovertext=hovers,
        hoverinfo="text",
        marker=dict(
            size=bubble_sizes,
            color=toms,
            colorscale=[[0, "#d1fae5"], [0.45, "#34d399"], [1, "#065f46"]],
            cmin=0,
            cmax=max(toms) * 1.05 if toms else 30,
            colorbar=dict(
                title=dict(text="TOM %", font=dict(size=10)),
                thickness=10, len=0.55, x=1.01,
                tickfont=dict(size=9), ticksuffix="%",
            ),
            line=dict(width=2.5, color="white"),
            sizemode="diameter",
            opacity=0.88,
        ),
    ))

    # No second text trace — zone name label above bubble is sufficient
    # TOM% is encoded in bubble colour (see colourbar) and full detail in hover

    fig.update_layout(
        height=380,
        margin=dict(t=8, b=8, l=0, r=70),
        paper_bgcolor="rgba(0,0,0,0)",
        title=dict(
            text=f"{brand_name} — Awareness by Zone  (bubble size = respondent base)",
            font=dict(size=11, color="#6b7280"), x=0.01, xanchor="left",
        ),
        geo=dict(
            showland=True,       landcolor="#f0fdf4",
            showocean=True,      oceancolor="#e0f2fe",
            showcountries=True,  countrycolor="#94a3b8",
            showsubunits=True,   subunitcolor="#cbd5e1",
            showcoastlines=True, coastlinecolor="#64748b",
            lonaxis=dict(range=[65, 100]),
            lataxis=dict(range=[5,  38]),
            bgcolor="rgba(0,0,0,0)",
            projection=dict(type="mercator"),
        ),
        font=dict(size=11),
        showlegend=False,
    )
    return _theme_fig(fig)


def _city_nps_chart(city_nps: list, brand_name: str, top_n: int = 10):
    """
    Horizontal NPS bar by city — zone-coloured bars, NPS zone backgrounds,
    black header strip. Sorted NPS best→worst.
    """
    if not city_nps:
        return None
    # Sort globally first, then slice — ensures correct top_n selection
    top = sorted(city_nps, key=lambda c: c["nps"], reverse=True)[:top_n]
    top_rev = list(reversed(top))  # chart renders bottom→top

    labels = [f"{c['city_name']} ({c.get('zone_name', '?')})" for c in top_rev]
    scores = [c["nps"]     for c in top_rev]
    bases  = [c["raters"]  for c in top_rev]
    zones  = [c.get("zone_name", "") for c in top_rev]

    # Zone-based bar colors; fallback to NPS-tier color
    colors = []
    for c, s in zip(top_rev, scores):
        z = c.get("zone_name", "")
        if z in ZONE_COLORS:
            colors.append(ZONE_COLORS[z])
        elif s >= NPS_INDUSTRY_AVG:
            colors.append(PALETTE["nps_promoter"])
        elif s < 0:
            colors.append(PALETTE["nps_detractor"])
        else:
            colors.append(PALETTE["nps_passive"])

    x_min = min(scores) - 18
    x_max = max(scores) + 28

    fig = go.Figure()

    # NPS zone backgrounds removed — zone-colored bars already carry semantic info.
    # Using only reference lines (zero and industry avg) to avoid color conflicts.
    fig.add_vline(x=0, line_color="#d1d5db", line_width=1.2)

    # ── Bars ──────────────────────────────────────────────────────────────────
    fig.add_trace(go.Bar(
        y=labels, x=scores, orientation="h",
        marker=dict(color=colors, opacity=0.88,
                    line=dict(color="white", width=1)),
        text=[f"{s:+.0f}" for s in scores], textposition="outside",
        textfont=dict(size=11, color="#374151"),
        customdata=bases,
        hovertemplate=(
            "<b>%{y}</b><br>"
            "NPS: %{x:+.0f}<br>"
            "Raters: %{customdata:,}"
            "<extra></extra>"
        ),
    ))

    # ── Reference lines ───────────────────────────────────────────────────────
    fig.add_vline(x=NPS_INDUSTRY_AVG, line_dash="dot", line_color="#6b7280", line_width=1.8,
                  annotation_text=f"Ind. avg +{NPS_INDUSTRY_AVG}",
                  annotation_position="top right", annotation_font_size=10)

    # ── Chart title annotations (no black background) ─────────────────────────
    fig.add_annotation(xref="paper", yref="paper", x=0.0, y=1.12,
                       text=f"<b>{brand_name}  ·  City NPS</b>  <i>(min 15 raters per city)</i>",
                       showarrow=False, font=dict(size=11, color="#1a5d4d"), xanchor="left")
    fig.add_annotation(xref="paper", yref="paper", x=1.0, y=1.12,
                       text="<i>Bar colour = Zone</i>",
                       showarrow=False, font=dict(size=9, color="#6b7280"), xanchor="right")

    # Zone legend chips (right side, outside plot)
    for i, (zone, zcolor) in enumerate(ZONE_COLORS.items()):
        fig.add_annotation(
            xref="paper", yref="paper",
            x=1.01, y=0.98 - i * 0.07,
            text=f"<b>{zone}</b>",
            showarrow=False, font=dict(size=9, color=zcolor),
            xanchor="left", bgcolor="rgba(255,255,255,0.8)", borderpad=2,
        )

    fig.update_layout(
        height=max(420, top_n * 36),
        margin=dict(t=75, b=45, l=10, r=80),
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        xaxis=dict(title="Net Promoter Score", range=[x_min, x_max],
                   gridcolor="#f3f4f6", zeroline=False),
        yaxis=dict(gridcolor="rgba(0,0,0,0)", tickfont=dict(size=10.5)),
        font=dict(size=11),
        bargap=0.24,
        hoverlabel=dict(bgcolor="white", font_size=12),
    )
    return _theme_fig(fig)


def _side_by_side_funnels(comparison_data: dict, primary_brand: str, funnel_type: str = "awareness"):
    """
    Side-by-side Plotly Funnel subplots — one funnel per (brand × segment) combination.
    funnel_type: 'awareness' or 'conversion'

    Layout:
      No segment  → 1 row, N cols (one funnel per brand)
      With segs   → rows = segments, cols = brands (brand × segment grid)
    """
    from plotly.subplots import make_subplots

    BRAND_COLORS = _chart_colors()
    if funnel_type == "conversion":
        FUNNEL_STAGES = ["Total Aware", "Ever Tried", "Current Use", "Preferred"]
    else:
        FUNNEL_STAGES = ["Total Aware", "Spont", "Consideration", "TOM"]

    brands   = list(comparison_data.keys())
    all_segs = list(next(iter(comparison_data.values())).keys()) if comparison_data else ["Overall"]
    n_brands = len(brands)
    n_segs   = len(all_segs)

    if n_brands == 0:
        return go.Figure()

    # Build grid: rows = segments, cols = brands
    rows, cols = n_segs, n_brands
    subplot_specs = [[{"type": "funnel"}] * cols for _ in range(rows)]

    # Titles: "BrandName (Segment)" for each cell
    subtitles = []
    for seg in all_segs:
        for brand in brands:
            label = f"<b>{brand}</b>" if n_segs == 1 else f"<b>{brand}</b><br><span style='font-size:9px'>{seg}</span>"
            subtitles.append(label)

    fig = make_subplots(
        rows=rows, cols=cols,
        specs=subplot_specs,
        subplot_titles=subtitles,
        horizontal_spacing=0.04,
        vertical_spacing=0.12,
    )

    def _hex_to_rgba(hex_c, alpha):
        h = hex_c.lstrip("#")
        r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
        return f"rgba({r},{g},{b},{alpha})"

    for r_idx, seg in enumerate(all_segs):
        for c_idx, brand in enumerate(brands):
            d = comparison_data.get(brand, {}).get(seg, {})
            aided = d.get("total_awareness_pct", d.get("aided_pct", 0)) or 0
            base  = d.get("base_n",    0) or 0
            color = BRAND_COLORS[c_idx % len(BRAND_COLORS)]

            if funnel_type == "conversion":
                ever_used = d.get("ever_used_pct", 0) or 0
                current   = d.get("current_pct", 0) or 0
                preferred = d.get("preferred_pct", 0) or 0

                conv_ever = round(ever_used / aided * 100, 1) if aided > 0 else 0
                conv_curr = round(current / ever_used * 100, 1) if ever_used > 0 else 0
                conv_pref = round(preferred / current * 100, 1) if current > 0 else 0

                stage_vals   = [aided, ever_used, current, preferred]
                stage_labels = [
                    f"{aided:.1f}%",
                    f"{ever_used:.1f}%  ↓{conv_ever:.0f}%",
                    f"{current:.1f}%  ↓{conv_curr:.0f}%",
                    f"{preferred:.1f}%  ↓{conv_pref:.0f}%",
                ]
            else:
                spont  = d.get("spont_pct", 0) or 0
                consid = d.get("consideration_pct", 0) or 0
                tom    = d.get("tom_pct",   0) or 0

                conv_spont = round(spont  / aided * 100, 1) if aided  > 0 else 0
                _conv_cons_raw = round(consid / spont * 100, 1) if spont  > 0 else 0
                conv_tom   = round(tom    / consid * 100, 1) if consid > 0 else 0

                _consid_lbl = f"{consid:.1f}%  (indep.)" if _conv_cons_raw > 100 else f"{consid:.1f}%  ↓{_conv_cons_raw:.0f}%"
                stage_vals   = [aided, spont, consid, tom]
                stage_labels = [
                    f"{aided:.1f}%",
                    f"{spont:.1f}%  ↓{conv_spont:.0f}%",
                    _consid_lbl,
                    f"{tom:.1f}%  ↓{conv_tom:.0f}%",
                ]

            marker_colors = [
                _hex_to_rgba(color, 0.25),
                _hex_to_rgba(color, 0.5),
                _hex_to_rgba(color, 0.75),
                _hex_to_rgba(color, 1.0),
            ]

            fig.add_trace(
                go.Funnel(
                    y=FUNNEL_STAGES,
                    x=stage_vals,
                    text=stage_labels,
                    textinfo="text",
                    textposition="inside",
                    textfont=dict(size=10, color="white"),
                    marker=dict(color=marker_colors),
                    connector=dict(line=dict(color="rgba(0,0,0,0)")),
                    hovertemplate=(
                        f"<b>{brand}</b> — {seg}<br>"
                        "%{y}: %{x:.1f}%<br>"
                        f"Base: {base:,}<extra></extra>"
                    ),
                    showlegend=False,
                ),
                row=r_idx + 1,
                col=c_idx + 1,
            )

    cell_h = max(240, 300 - n_brands * 15)
    total_h = cell_h * rows + 80

    fig.update_layout(
        height=total_h,
        margin=dict(t=60, b=40, l=80, r=20),
        paper_bgcolor="#ffffff",
        plot_bgcolor="#ffffff",
        font=dict(family=CHART_THEME["font_family"], size=11, color="#0f172a"),
        template="plotly_white",
    )

    # Ensure subplot titles are crisp dark slate
    fig.for_each_annotation(lambda a: a.update(font=dict(color="#0f172a", size=11, family=CHART_THEME["font_family"])))

    # Single shared Y-axis label via annotation — hide per-subplot axis labels
    fig.add_annotation(
        text="Funnel Stage",
        x=-0.05, xref="paper",
        y=0.5,   yref="paper",
        textangle=-90,
        showarrow=False,
        font=dict(size=12),
        xanchor="right",
    )

    # Hide all per-subplot yaxis titles (they clutter)
    for key in fig.layout:
        if key.startswith("yaxis"):
            fig.layout[key].title = ""

    # Subtitle styling
    for ann in fig.layout.annotations:
        if not ann.textangle:  # subplot title annotations have no rotation
            ann.font = dict(size=11)

    return _theme_fig(fig)


def _pooled_prop_sig_pct(pct_a, pct_b, base_n, alpha=0.05):
    """Pooled two-proportion z-test from two percentages sharing a common base N.
    Returns dict with flag ('higher'/'lower'/''), diff (+/-X.X%), and pval if significant at alpha, else dict with empty flag.
    """
    if not base_n or base_n < 1:
        return {"flag": "", "diff": 0, "pval": 1.0}
    import numpy as _np
    from scipy.stats import norm as _norm
    pa, pb = (pct_a or 0) / 100.0, (pct_b or 0) / 100.0
    diff = round(pct_b - pct_a, 1)
    pooled = (pa + pb) / 2.0
    se = _np.sqrt(pooled * (1 - pooled) * (2 / base_n))
    if se == 0:
        return {"flag": "", "diff": diff, "pval": 1.0}
    z = (pb - pa) / se
    pval = 2 * (1 - _norm.cdf(abs(z)))
    if pval >= alpha:
        return {"flag": "", "diff": diff, "pval": pval}
    flag = "higher" if pb > pa else "lower"
    return {"flag": flag, "diff": diff, "pval": pval, "z": round(z, 2)}


def _render_comparison_funnels_html(comparison_data: dict, primary_brand: str, alpha: float = 0.05, funnel_type: str = "awareness"):
    """
    Multi-brand comparison funnels using Streamlit columns (no squeeze).
    funnel_type: 'awareness' or 'conversion'
    """
    palette   = _chart_colors()   # respects active CHART_THEME palette
    font_fam  = CHART_THEME["font_family"]

    brands   = list(comparison_data.keys())
    all_segs = list(next(iter(comparison_data.values())).keys()) if comparison_data else ["Overall"]

    def _arrow_color(pct: float) -> str:
        return "#22c55e" if pct >= 50 else ("#f59e0b" if pct >= 25 else "#ef4444")

    def _hex_lighten(hex_c: str, strength: float) -> str:
        h = hex_c.lstrip("#")
        r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
        r2 = int(r + (255 - r) * strength)
        g2 = int(g + (255 - g) * strength)
        b2 = int(b + (255 - b) * strength)
        return f"#{r2:02x}{g2:02x}{b2:02x}"

    def _sig_badge(res: dict) -> str:
        if not isinstance(res, dict) or not res.get("flag"):
            return ""
        flag = res["flag"]
        diff = res.get("diff", 0)
        sign = "+" if diff > 0 else ""
        if flag == "higher":
            return f" <span style='color:#15803d;font-weight:800;font-size:0.65rem;background:#dcfce7;border:1px solid #bbf7d0;border-radius:6px;padding:1px 5px;'>▲ {sign}{diff:.1f}%</span>"
        if flag == "lower":
            return f" <span style='color:#b91c1c;font-weight:800;font-size:0.65rem;background:#fee2e2;border:1px solid #fca5a5;border-radius:6px;padding:1px 5px;'>▼ {sign}{diff:.1f}%</span>"
        return ""

    for seg in all_segs:
        if len(all_segs) > 1:
            st.markdown(f"**Segment: {seg}**")

        cols = st.columns(len(brands))
        for i, (brand, col) in enumerate(zip(brands, cols)):
            d = comparison_data.get(brand, {}).get(seg, {})
            aided  = float(d.get("total_awareness_pct", d.get("aided_pct", 0)) or 0)
            base_n = int(d.get("base_n", 0) or 0)

            _is_primary = (brand == primary_brand)
            _p = comparison_data.get(primary_brand, {}).get(seg, {})

            if funnel_type == "conversion":
                stg2 = float(d.get("ever_used_pct", 0) or 0)
                stg3 = float(d.get("current_pct", 0) or 0)
                stg4 = float(d.get("preferred_pct", 0) or 0)

                stg1_lbl, stg2_lbl, stg3_lbl, stg4_lbl = "Total Aware", "Ever Tried", "Current Use", "Preferred"
                conv_stg2 = round(stg2 / aided * 100, 1) if aided > 0 else 0
                conv_stg3 = round(stg3 / stg2  * 100, 1) if stg2  > 0 else 0
                conv_stg4 = round(stg4 / stg3  * 100, 1) if stg3  > 0 else 0
                arr2_lbl, arr3_lbl, arr4_lbl = "→ tried", "→ current", "→ preferred"

                if _is_primary:
                    _sig1 = _sig2 = _sig3 = _sig4 = ""
                else:
                    _sig1 = _pooled_prop_sig_pct(float(_p.get("total_awareness_pct", _p.get("aided_pct", 0)) or 0), aided, base_n, alpha)
                    _sig2 = _pooled_prop_sig_pct(float(_p.get("ever_used_pct", 0) or 0), stg2, base_n, alpha)
                    _sig3 = _pooled_prop_sig_pct(float(_p.get("current_pct", 0) or 0), stg3, base_n, alpha)
                    _sig4 = _pooled_prop_sig_pct(float(_p.get("preferred_pct", 0) or 0), stg4, base_n, alpha)
            else:
                stg2 = float(d.get("spont_pct", 0) or 0)
                stg3 = float(d.get("consideration_pct", 0) or 0)
                stg4 = float(d.get("tom_pct", 0) or 0)

                stg1_lbl, stg2_lbl, stg3_lbl, stg4_lbl = "Total Aware", "Spont", "Consid.", "Top of Mind"
                conv_stg2 = round(stg2 / aided * 100, 1) if aided > 0 else 0
                _conv_stg3_raw = round(stg3 / stg2 * 100, 1) if stg2 > 0 else 0
                conv_stg3 = min(_conv_stg3_raw, 100.0)
                conv_stg4 = round(stg4 / stg3  * 100, 1) if stg3  > 0 else 0
                # Consid. is measured independently; mark arrow when it exceeds spont base
                arr2_lbl = "→ spont"
                arr3_lbl = "indep." if _conv_stg3_raw > 100 else "→ consid."
                arr4_lbl = "→ TOM"

                if _is_primary:
                    _sig1 = _sig2 = _sig3 = _sig4 = ""
                else:
                    _sig1 = _pooled_prop_sig_pct(float(_p.get("total_awareness_pct", _p.get("aided_pct", 0)) or 0), aided, base_n, alpha)
                    _sig2 = _pooled_prop_sig_pct(float(_p.get("spont_pct", 0) or 0), stg2, base_n, alpha)
                    _sig3 = _pooled_prop_sig_pct(float(_p.get("consideration_pct", 0) or 0), stg3, base_n, alpha)
                    _sig4 = _pooled_prop_sig_pct(float(_p.get("tom_pct", 0) or 0), stg4, base_n, alpha)

            color   = palette[i % len(palette)]
            c_light = _hex_lighten(color, 0.88)
            c_mid1  = _hex_lighten(color, 0.70)
            c_mid2  = _hex_lighten(color, 0.50)

            txt_dark  = "#0f172a"

            stg2_w = max(22, min(100, round(stg2 / max(aided, 1) * 100))) if aided > 0 else 55
            stg3_w = max(22, min(100, round(stg3 / max(aided, 1) * 100))) if aided > 0 else 45
            stg4_w = max(42, min(100, round(stg4 / max(aided, 1) * 100))) if aided > 0 else 45

            ac = _arrow_color(conv_stg2)
            cc = _arrow_color(conv_stg3)
            tc = _arrow_color(conv_stg4)
            star        = " ★" if brand == primary_brand else ""
            hdr_outline = f"outline:2.5px solid white; outline-offset:-2px;" if brand == primary_brand else ""

            f_sp = max(0.42, min(0.60, round(0.55 * (stg2_w / 100) + 0.12, 2)))
            n_sp = max(1.0,  min(1.45, round(1.35 * (stg2_w / 100) + 0.30, 2)))

            f_co = max(0.42, min(0.60, round(0.55 * (stg3_w / 100) + 0.12, 2)))
            n_co = max(1.0,  min(1.45, round(1.35 * (stg3_w / 100) + 0.30, 2)))

            f_tm = max(0.52, min(0.65, round(0.55 * (stg4_w / 100) + 0.18, 2)))
            n_tm = max(1.15, min(1.50, round(1.35 * (stg4_w / 100) + 0.40, 2)))

            html = f"""
<div style="font-family:{font_fam}; width:100%;">
  <!-- header -->
  <div style="background:{color}; {hdr_outline} border-radius:10px 10px 0 0;
              padding:8px 6px; text-align:center; margin-bottom:0;">
    <div style="font-size:0.68rem; font-weight:900; color:#ffffff; text-transform:uppercase;
                letter-spacing:0.05em; overflow:hidden; text-overflow:ellipsis; white-space:nowrap; text-shadow:0 1px 2px rgba(0,0,0,0.4);">
      {brand}{star}
    </div>
    <div style="font-size:0.55rem; color:#f1f5f9; font-weight:600;">n={base_n:,}</div>
  </div>

  <!-- Stage 1 -->
  <div style="background:{c_light}; padding:10px 4px; text-align:center; border:1px solid #e2e8f0; border-top:none;">
    <div style="font-size:0.55rem; font-weight:800; color:{txt_dark}; text-transform:uppercase;
                letter-spacing:0.08em; margin-bottom:2px;">{stg1_lbl}</div>
    <div style="font-size:1.55rem; font-weight:900; color:{txt_dark}; line-height:1;">
      {aided:.1f}<span style="font-size:0.85rem;">%</span>{_sig_badge(_sig1)}
    </div>
  </div>

  <!-- arrow ↓ Stage 1→Stage 2 -->
  <div style="text-align:center; padding:4px 0; background:white;">
    <div style="display:inline-flex; align-items:center; gap:4px;
                border:1.5px solid {ac}; border-radius:20px; padding:2px 7px;
                background:white; box-shadow:0 1px 6px rgba(0,0,0,0.10);">
      <span style="font-size:0.85rem; color:{ac}; font-weight:700;">↓</span>
      <div>
        <div style="font-size:0.78rem; font-weight:800; color:{ac}; line-height:1;">{conv_stg2:.1f}%</div>
        <div style="font-size:0.4rem; color:#475569; text-transform:uppercase; font-weight:700;">{arr2_lbl}</div>
      </div>
    </div>
  </div>

  <!-- Stage 2 -->
  <div style="width:{stg2_w}%; margin:0 auto; background:{c_mid1}; border:1px solid #cbd5e1;
              padding:8px 4px; text-align:center; box-sizing:border-box; overflow:hidden;">
    <div style="font-size:{f_sp}rem; font-weight:800; color:{txt_dark}; text-transform:uppercase;
                letter-spacing:0.06em; margin-bottom:2px; white-space:nowrap; text-overflow:ellipsis; overflow:hidden;">{stg2_lbl}</div>
    <div style="font-size:{n_sp}rem; font-weight:900; color:{txt_dark}; line-height:1;">
      {stg2:.1f}<span style="font-size:0.75rem;">%</span>{_sig_badge(_sig2)}
    </div>
  </div>

  <!-- arrow ↓ Stage 2→Stage 3 -->
  <div style="text-align:center; padding:4px 0; background:white;">
    <div style="display:inline-flex; align-items:center; gap:4px;
                border:1.5px solid {cc}; border-radius:20px; padding:2px 7px;
                background:white; box-shadow:0 1px 6px rgba(0,0,0,0.10);">
      <span style="font-size:0.85rem; color:{cc}; font-weight:700;">↓</span>
      <div>
        <div style="font-size:0.78rem; font-weight:800; color:{cc}; line-height:1;">{conv_stg3:.1f}%</div>
        <div style="font-size:0.4rem; color:#475569; text-transform:uppercase; font-weight:700;">{arr3_lbl}</div>
      </div>
    </div>
  </div>

  <!-- Stage 3 -->
  <div style="width:{stg3_w}%; margin:0 auto; background:{c_mid2}; border:1px solid #94a3b8;
              padding:8px 4px; text-align:center; box-sizing:border-box; overflow:hidden;">
    <div style="font-size:{f_co}rem; font-weight:800; color:{txt_dark}; text-transform:uppercase;
                letter-spacing:0.06em; margin-bottom:2px; white-space:nowrap; text-overflow:ellipsis; overflow:hidden;">{stg3_lbl}</div>
    <div style="font-size:{n_co}rem; font-weight:900; color:{txt_dark}; line-height:1;">
      {stg3:.1f}<span style="font-size:0.75rem;">%</span>{_sig_badge(_sig3)}
    </div>
  </div>

  <!-- arrow ↓ Stage 3→Stage 4 -->
  <div style="text-align:center; padding:4px 0; background:white;">
    <div style="display:inline-flex; align-items:center; gap:4px;
                border:1.5px solid {tc}; border-radius:20px; padding:2px 7px;
                background:white; box-shadow:0 1px 6px rgba(0,0,0,0.10);">
      <span style="font-size:0.85rem; color:{tc}; font-weight:700;">↓</span>
      <div>
        <div style="font-size:0.78rem; font-weight:800; color:{tc}; line-height:1;">{conv_stg4:.1f}%</div>
        <div style="font-size:0.4rem; color:#9ca3af; text-transform:uppercase; font-weight:600;">{arr4_lbl}</div>
      </div>
    </div>
  </div>

  <!-- Stage 4 -->
  <div style="width:{stg4_w}%; min-width:130px; margin:0 auto; background:{color};
              border-radius:0 0 10px 10px; padding:10px 6px; text-align:center; box-sizing:border-box; overflow:hidden;">
    <div style="font-size:{f_tm}rem; font-weight:800; color:#ffffff;
                text-transform:uppercase; letter-spacing:0.06em; margin-bottom:3px; white-space:nowrap;">{stg4_lbl}</div>
    <div style="font-size:{n_tm}rem; font-weight:900; color:#ffffff; line-height:1;">
      {stg4:.1f}<span style="font-size:0.8rem;">%</span>{_sig_badge(_sig4).replace("color:#15803d", "color:#4ade80").replace("color:#b91c1c", "color:#fca5a5")}
    </div>
  </div>
</div>"""

            with col:
                st.markdown(html, unsafe_allow_html=True)


def _nps_stacked_bar(brand_data):
    """Compact full-width stacked bar: Promoters | Passives | Detractors."""
    p       = brand_data.get("nps_promoters_pct",  0) or 0
    pa      = brand_data.get("nps_passives_pct",   0) or 0
    d       = brand_data.get("nps_detractors_pct", 0) or 0
    nps_v   = brand_data.get("nps") or 0
    n_total = brand_data.get("nps_base", 0) or 0

    nps_color = "#22c55e" if nps_v >= NPS_INDUSTRY_AVG else "#f59e0b" if nps_v >= 0 else "#ef4444"
    delta     = round(nps_v - NPS_INDUSTRY_AVG, 1)
    sign      = "+" if delta >= 0 else ""

    fig = go.Figure()

    for seg, val, color in [
        ("Promoters",  p,  PALETTE["nps_promoter"]),
        ("Passives",   pa, PALETTE["nps_passive"]),
        ("Detractors", d,  PALETTE["nps_detractor"]),
    ]:
        n_approx = int(n_total * val / 100) if n_total else 0
        lbl = (f"<b>{seg}</b><br>{val:.0f}%" if val >= 14
               else f"<b>{val:.0f}%</b>" if val >= 7
               else "")
        fig.add_trace(go.Bar(
            y=[""], x=[val], name=seg, orientation="h",
            marker_color=color,
            text=[lbl],
            textposition="inside", insidetextanchor="middle",
            textfont=dict(color="white", size=13, family="Inter, Arial, sans-serif"),
            hovertemplate=f"{seg}: %{{x:.0f}}%  (~{n_approx:,} respondents)<extra></extra>",
        ))

    # NPS formula below the bar
    fig.add_annotation(xref="paper", yref="paper", x=0.0, y=-0.45,
                       text=f"<b>NPS = {p:.0f}% − {d:.0f}% = {nps_v:+.0f}</b>",
                       showarrow=False, font=dict(size=11, color=nps_color), xanchor="left")
    fig.add_annotation(xref="paper", yref="paper", x=1.0, y=-0.45,
                       text=f"<i>vs Industry avg +{NPS_INDUSTRY_AVG}: {sign}{delta}</i>",
                       showarrow=False, font=dict(size=10, color="#6b7280"), xanchor="right")

    fig.update_layout(
        barmode="stack",
        height=110,
        showlegend=False,
        margin=dict(t=6, b=42, l=4, r=4),
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        xaxis=dict(visible=False, range=[0, 100]),
        yaxis=dict(visible=False),
        bargap=0,
        hoverlabel=dict(bgcolor="white", font_size=12, bordercolor="#e5e7eb"),
    )
    return _theme_fig(fig)



# ── Brand Imagery chart builders ─────────────────────────────────────────────

def _multi_brand_radar(brands_data: list) -> go.Figure:
    """
    Overlay radar for multiple brands.
    brands_data: [{brand_name, tom_pct, spont_pct, aided_pct, nps, nps_base}, ...]
    """
    COLORS = _chart_colors()
    axes = ["Salience\n(TOM)", "Recall\n(SPONT)", "Total Reach\n(AIDED)",
            "Loyalty\n(NPS norm)", "Rater Depth"]

    def norm_nps(v):
        return round((v + 100) / 2, 1) if v is not None else 50.0

    # Rater Depth self-normalised against the largest base in this set (dynamic)
    _max_base = max((bd.get("nps_base", 0) or 0) for bd in brands_data) or 1

    fig = go.Figure()
    for i, bd in enumerate(brands_data):
        color = COLORS[i % len(COLORS)]
        vals = [
            bd.get("tom_pct", 0),
            bd.get("spont_pct", 0),
            bd.get("aided_pct", 0),
            norm_nps(bd.get("nps")),
            min(round((bd.get("nps_base", 0) or 0) / _max_base * 100, 1), 100),
        ]
        fill_color = color.replace("#", "")
        r, g, b = int(fill_color[0:2], 16), int(fill_color[2:4], 16), int(fill_color[4:6], 16)
        fill_rgba = f"rgba({r},{g},{b},0.12)"
        fig.add_trace(go.Scatterpolar(
            r=vals + [vals[0]],
            theta=axes + [axes[0]],
            fill="toself",
            name=bd["brand_name"],
            fillcolor=fill_rgba,
            line=dict(color=color, width=2),
            hovertemplate="%{theta}: %{r:.1f}<extra></extra>",
        ))

    fig.update_layout(
        polar=dict(
            bgcolor="rgba(0,0,0,0)",
            radialaxis=dict(visible=True, range=[0, 100], gridcolor="#e5e7eb",
                            tickfont=dict(size=8)),
            angularaxis=dict(gridcolor="#e5e7eb", tickfont=dict(size=9)),
        ),
        showlegend=True,
        legend=dict(orientation="h", yanchor="bottom", y=1.12, xanchor="center", x=0.5,
                    font=dict(size=10)),
        paper_bgcolor="rgba(0,0,0,0)",
        margin=dict(t=70, b=20, l=50, r=50),
        height=460,
        font=dict(size=11),
    )
    return _theme_fig(fig)


def _health_waterfall_chart(brand_data, brands_list):
    """Decompose the Health Score into its components (Aided Awareness, Recall, NPS)"""
    import plotly.graph_objects as go
    
    aided = brand_data.get("aided_pct", 0)
    spont = brand_data.get("spont_pct", 0)
    tom = brand_data.get("tom_pct", 0)
    nps = brand_data.get("nps", 0)
    
    # Normalize NPS to a comparable awareness-like scale
    nps_impact = (nps + 100) / 10  # -100 to 100 -> 0 to 20
    
    fig = go.Figure(go.Waterfall(
        name="Health Score Decomposition", orientation="v",
        measure=["relative", "relative", "relative", "total"],
        x=["Reach (Aided)", "Recall (Spont)", "Advocacy (NPS)", "Total Health Index"],
        textposition="outside",
        text=[f"+{aided}%", f"+{spont}%", f"+{nps_impact:.1f}", "Total"],
        y=[aided, spont, nps_impact, 0],
        connector={"line": {"color": "rgb(63, 63, 63)"}},
        increasing={"marker": {"color": "#30a76a"}},
        totals={"marker": {"color": "#1a5d4d"}}
    ))

    fig.update_layout(
        height=350, margin=dict(t=40, b=20, l=40, r=20),
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        font=dict(size=11),
        showlegend=False
    )
    return _theme_fig(fig)


def _driver_action_matrix(drivers_df, brand_name):
    """A 2x2 scatter plot (Performance vs. Derived Importance) for drivers of a specific brand."""
    import plotly.graph_objects as go
    
    brand_perf = drivers_df.get("norm", pd.Series())
    importance = drivers_df.get("importance", pd.Series())
    
    if brand_perf.empty or importance.empty:
        return None
        
    # Align indices
    common_idx = brand_perf.index.intersection(importance.index)
    if len(common_idx) < 2:
        return None
        
    brand_perf = brand_perf.loc[common_idx]
    importance = importance.loc[common_idx]
        
    fig = go.Figure()
    
    x_abs_q = max(abs(brand_perf.max()), abs(brand_perf.min()), 5) * 1.3
    y_min, y_max = importance.min(), importance.max()
    y_range = y_max - y_min if y_max != y_min else 1.0
    y_mid = y_min + y_range / 2.0
    
    # Add quadrants
    fig.add_shape(type="rect", x0=-x_abs_q, y0=y_mid, x1=0, y1=y_max + 0.1*y_range, fillcolor="rgba(239, 68, 68, 0.1)", line_width=0) # Ignore
    fig.add_shape(type="rect", x0=0, y0=y_mid, x1=x_abs_q, y1=y_max + 0.1*y_range, fillcolor="rgba(34, 197, 94, 0.1)", line_width=0) # Maintain
    fig.add_shape(type="rect", x0=-x_abs_q, y0=y_min - 0.1*y_range, x1=0, y1=y_mid, fillcolor="rgba(245, 158, 11, 0.1)", line_width=0) # Improve
    fig.add_shape(type="rect", x0=0, y0=y_min - 0.1*y_range, x1=x_abs_q, y1=y_mid, fillcolor="rgba(59, 130, 246, 0.1)", line_width=0) # Secondary

    fig.add_trace(go.Scatter(
        x=brand_perf.values,
        y=importance.values,
        mode="markers+text",
        text=brand_perf.index,
        textposition="top center",
        marker=dict(size=12, color="#1a5d4d", line=dict(width=1, color="white")),
        hovertemplate="<b>%{text}</b><br>Advantage (Norm Dev): %{x:+.1f}%<br>NPS Impact (Derived): %{y:.4f}<extra></extra>"
    ))
    
    fig.update_layout(
        xaxis=dict(title="Competitive Advantage (Norm Dev %)", range=[-x_abs_q, x_abs_q], zeroline=True, zerolinewidth=2, zerolinecolor="black"),
        yaxis=dict(title="Driver Importance (NPS Impact Score)", range=[y_min - 0.1*y_range, y_max + 0.1*y_range], zeroline=True, zerolinewidth=2, zerolinecolor="black"),
        height=500, margin=dict(t=40, b=50, l=50, r=30),
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        font=dict(size=11),
    )
    
    # Add quadrant labels
    _qx = x_abs_q * 0.65
    _qy_top = y_mid + 0.4 * (y_max - y_mid)
    _qy_bot = y_min + 0.4 * (y_mid - y_min)
    
    fig.add_annotation(x=_qx,  y=_qy_top, text="MAINTAIN (Strengths)", showarrow=False, font=dict(color="#166534", size=10, weight="bold"))
    fig.add_annotation(x=-_qx, y=_qy_top, text="IGNORE (Weaknesses)",  showarrow=False, font=dict(color="#991b1b", size=10, weight="bold"))
    fig.add_annotation(x=-_qx, y=_qy_bot, text="IMPROVE (Gaps)",       showarrow=False, font=dict(color="#92400e", size=10, weight="bold"))
    fig.add_annotation(x=_qx,  y=_qy_bot, text="SECONDARY (Niche)",    showarrow=False, font=dict(color="#1e40af", size=10, weight="bold"))

    return _theme_fig(fig)


def _correlation_heatmap(matrix_data: dict) -> go.Figure:
    """Annotated heatmap: brand × brand co-awareness %."""
    brands = matrix_data["brands"]
    matrix = matrix_data["matrix"]

    # Mask upper triangle for cleaner display (show lower + diagonal)
    import numpy as _np
    mat = _np.array(matrix, dtype=float)

    fig = go.Figure(go.Heatmap(
        z=mat,
        x=brands,
        y=brands,
        text=[[f"{v:.0f}%" for v in row] for row in mat],
        texttemplate="%{text}",
        textfont=dict(size=8),
        colorscale=[
            [0.0, "#f0fdf4"],
            [0.3, "#86efac"],
            [0.6, "#22c55e"],
            [1.0, "#14532d"],
        ],
        showscale=True,
        colorbar=dict(
            title=dict(text="Co-aware %", font=dict(size=10)),
            thickness=12, len=0.8,
            tickfont=dict(size=9), ticksuffix="%",
        ),
        hovertemplate="%{x} × %{y}: %{z:.1f}%<extra></extra>",
    ))
    n = len(brands)
    fig.update_layout(
        height=max(420, n * 28),
        margin=dict(t=30, b=80, l=10, r=20),
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        xaxis=dict(tickangle=-45, tickfont=dict(size=9)),
        yaxis=dict(tickfont=dict(size=9), autorange="reversed"),
        font=dict(size=11),
    )
    return _theme_fig(fig)


def _correspondence_map(zone_matrix: dict, highlight_brand: str = None) -> go.Figure:
    """
    PCA-based positioning map: brands clustered by zone awareness + NPS profile.
    Each brand = 8-dim vector [TOM_N, TOM_S, TOM_E, TOM_W, NPS_N, NPS_S, NPS_E, NPS_W].
    Projected to 2D via PCA.
    """
    import numpy as _np
    from sklearn.decomposition import PCA
    from sklearn.preprocessing import StandardScaler

    brands    = zone_matrix["brands"]
    tom_mat   = _np.array(zone_matrix["tom_matrix"], dtype=float)
    nps_mat   = _np.array(zone_matrix["nps_matrix"], dtype=float)
    X = _np.hstack([tom_mat, nps_mat])  # shape: (n_brands, 8)

    if len(brands) < 3 or X.shape[0] < 3:
        return None

    X_scaled = StandardScaler().fit_transform(X)
    pca = PCA(n_components=2)
    coords = pca.fit_transform(X_scaled)  # (n_brands, 2)

    ev = pca.explained_variance_ratio_
    x_label = f"Dim 1 ({ev[0]*100:.0f}% variance) — Geographic Reach Pattern"
    y_label = f"Dim 2 ({ev[1]*100:.0f}% variance)"

    colors = []
    sizes  = []
    for brand in brands:
        is_hl = highlight_brand and brand == highlight_brand
        colors.append("#1a5d4d" if is_hl else "#86efac")
        sizes.append(20 if is_hl else 12)

    fig = go.Figure()
    # Other brands
    other_idx = [i for i, b in enumerate(brands) if b != highlight_brand]
    sel_idx   = [i for i, b in enumerate(brands) if b == highlight_brand]

    fig.add_trace(go.Scatter(
        x=coords[other_idx, 0],
        y=coords[other_idx, 1],
        mode="markers+text",
        name="Brands",
        text=[brands[i] for i in other_idx],
        textposition="top center",
        textfont=dict(size=9, color="#6b7280"),
        marker=dict(size=11, color="#86efac", opacity=0.8,
                    line=dict(width=1, color="#30a76a")),
        hovertemplate="%{text}<extra></extra>",
    ))
    if sel_idx:
        i = sel_idx[0]
        fig.add_trace(go.Scatter(
            x=[coords[i, 0]], y=[coords[i, 1]],
            mode="markers+text",
            name=highlight_brand,
            text=[highlight_brand],
            textposition="top center",
            textfont=dict(size=11, color="#1a5d4d", family="Arial Black"),
            marker=dict(size=18, color="#1a5d4d", symbol="star",
                        line=dict(width=2, color="white")),
            hovertemplate=f"{highlight_brand}<extra></extra>",
        ))

    # Zone loading vectors (biplot)
    zones = zone_matrix["zones"]
    feat_names = [f"TOM_{z}" for z in zones] + [f"NPS_{z}" for z in zones]
    loadings = pca.components_.T  # (8, 2)
    scale = max(abs(coords).max(), 1) * 0.4
    for j, fname in enumerate(feat_names[:4]):  # show only TOM vectors for clarity
        lx, ly = loadings[j, 0] * scale, loadings[j, 1] * scale
        fig.add_annotation(
            x=lx, y=ly, ax=0, ay=0,
            xref="x", yref="y", axref="x", ayref="y",
            arrowhead=3, arrowwidth=1.5, arrowcolor="#d1d5db",
            text=zones[j], font=dict(size=8, color="#9ca3af"),
            showarrow=True,
        )

    fig.add_hline(y=0, line_dash="dot", line_color="#e5e7eb", line_width=1)
    fig.add_vline(x=0, line_dash="dot", line_color="#e5e7eb", line_width=1)

    fig.update_layout(
        height=500,
        margin=dict(t=40, b=50, l=50, r=30),
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        xaxis=dict(title=x_label, gridcolor="#f3f4f6", zeroline=False),
        yaxis=dict(title=y_label, gridcolor="#f3f4f6", zeroline=False),
        showlegend=False,
        font=dict(size=11),
    )
    return _theme_fig(fig)


# ── Cached imagery data ───────────────────────────────────────────────────────

@st.cache_data(ttl=3600)
def _get_cached_correlation(project_id: str = "project_1"):
    engine = BrandImageryEngine(project_id=project_id)
    return engine.get_brand_correlation_matrix(top_n=15)


@st.cache_data(ttl=3600)
def _get_cached_zone_matrix(project_id: str = "project_1"):
    engine = BrandImageryEngine(project_id=project_id)
    return engine.get_brand_zone_matrix(top_n=20)


# ── Main renderer ─────────────────────────────────────────────────────────────

@st.cache_data(ttl=3600)
def _get_cached_brand_health_data(cat_arg, zone_arg, city_arg, sel_months,
                                   gender_arg="all", age_band_arg="all", project_id=None):
    engine = BrandImageryEngine(project_id=project_id)
    return engine.get_brand_health(
        category=cat_arg, zone=zone_arg, city=city_arg, months=sel_months,
        gender=gender_arg, age_band=age_band_arg)



@st.cache_data(ttl=3600)
def _get_cached_narrative(brand_name, b_data, b_n, z_data, c_nps, r_list, b_all_str="[]", theme="All", focus=10):
    """Module-level cached narrative to ensure dict return type.
    b_all_str: JSON-serialized brands_list for cache key hashing.
    """
    import json as _json
    try:
        b_all = _json.loads(b_all_str)
    except Exception:
        b_all = []
    res = generate_brand_narrative(brand_name, b_data, b_n, z_data, c_nps, r_list, b_all, theme, focus)
    if isinstance(res, str):
        return {"overview": res, "geographic": "Update pending", "competitive": "", "nps": "",
                "funnel": "", "radar": "", "nps_league": "", "city_story": "", "positioning": "",
                "salience_finding": "", "loyalty_finding": "", "dynamics_finding": "", "imagery_finding": ""}
    return res


@st.cache_data(ttl=3600, show_spinner="Preparing executive briefing...")
def _get_cached_briefing(brand_name, b_data, b_n, z_data, c_nps, r_list,
                         b_all_str="[]", filter_key="all", schema_ver="v2-neutral"):
    """Cached Executive Command Briefing (neutral market/demand/experience).
    schema_ver busts the cache when the briefing schema changes."""
    import json as _json
    from lens.analytics.brand_narrative import generate_executive_briefing
    try:
        b_all = _json.loads(b_all_str)
    except Exception:
        b_all = []
    try:
        return generate_executive_briefing(brand_name, b_data, b_n, z_data,
                                           c_nps, r_list, b_all)
    except Exception as _e:
        print(f"[BH] briefing fetch failed: {_e}")
        return {}


# ── BQ3 Imagery Analysis (Sections 13 & 14) ──────────────────────────────────

_SIDEBAR_TO_PRODUCT = {
    "All": "All", "Ceiling Fans": "Ceiling Fans", "Air Cooler": "Air Cooler",
    "Mixer Grinder": "Mixer Grinder", "LED Batten": "LED Batten",
    "Water Heater": "Water Heater", "Water Pumps": "Water Pumps",
}

def _safe_chart(fig_or_none, fallback_msg="Chart unavailable."):
    """Render a Plotly figure (dict or go.Figure), stamping CHART_THEME font on it."""
    if fig_or_none and isinstance(fig_or_none, dict) and fig_or_none:
        try:
            f = go.Figure(fig_or_none)
            # Stamp active theme font so sidebar controls affect CAN MAP / BIP charts too
            f.update_layout(
                font=dict(
                    family=CHART_THEME["font_family"],
                    size=CHART_THEME["font_size"],
                ),
            )
            st.plotly_chart(f, use_container_width=True, config={"editable": True})
        except Exception:
            st.info(fallback_msg)
    elif fig_or_none is not None:
        try:
            _theme_fig(fig_or_none)
            st.plotly_chart(fig_or_none, use_container_width=True, config={"editable": True})
        except Exception:
            st.info(fallback_msg)
    else:
        st.info(fallback_msg)

@st.cache_data(ttl=3600, show_spinner="Generating AI interpretations...")
def _generate_interps_from_db(
    matrix_fingerprint: str,
    ca_product: str, ca_section: str, bip_pctile: int, n_brands: int,
    zone: str = "all", gender: str = "all", age_band: str = "all", city: str = "all",
    project_id: str = "project_1",
) -> dict:
    """Fetch live DB data and generate AI interpretations. Cached by fingerprint."""
    try:
        from lens.analytics.imagery_interpreter import generate_all
        sec = None if ca_section == "All" else ca_section
        ca_res  = _run_can_map(ca_product, sec, n_brands, zone, gender, age_band, city, project_id)
        bip_res = _run_bip(ca_product, sec, bip_pctile, zone, gender, age_band, city, top_brands=n_brands, project_id=project_id)
        if ca_res.get("status") != "ok":
            return {}
        return generate_all(ca_res, bip_res if bip_res else {})
    except Exception:
        return {}


def _get_cached_imagery_interpretations(
    ca_res: dict, bip_res, ca_product, ca_section, bip_pctile, n_brands,
    zone="all", gender="all", age_band="all", city="all",
) -> dict:
    """
    Return AI interpretations that match the data currently displayed.
    When ca_res already contains ok results (e.g. test data), call generate_all
    directly with those results so AI analysis reflects the shown data.
    Fall back to DB re-fetch path (cached) when ca_res is from the live DB.
    """
    from lens.analytics.imagery_interpreter import generate_all
    try:
        mat = ca_res.get("matrix")
        if mat is not None and hasattr(mat, "values") and not mat.empty:
            fp = f"{mat.shape}_{round(float(mat.values.sum()), 3)}"
        else:
            fp = f"{ca_product}_{ca_section}_{zone}_{gender}_{age_band}_{city}"
    except Exception:
        fp = f"{ca_product}_{ca_section}_{zone}"

    # Check if the displayed data matches live DB results by comparing fingerprint
    # to what _run_can_map would produce for the same params.
    db_fp_key = f"{ca_product}_{ca_section}_{n_brands}_{zone}_{gender}_{age_band}_{city}"
    if fp != db_fp_key and ca_res.get("status") == "ok":
        # Test/custom data is being shown — generate AI directly from it.
        # Use session state as lightweight cache keyed on fingerprint.
        ss_key = f"_interp_cache_{fp}"
        if ss_key not in st.session_state:
            try:
                st.session_state[ss_key] = generate_all(ca_res, bip_res if bip_res else {})
            except Exception:
                st.session_state[ss_key] = {}
        return st.session_state.get(ss_key, {})

    return _generate_interps_from_db(
        fp, ca_product, ca_section, bip_pctile, n_brands,
        zone, gender, age_band, city,
    )


@st.cache_data(ttl=3600, show_spinner="Running Correspondence Analysis...")
def _run_can_map(product, section, n_brands, zone="all", gender="all", age_band="all", city="all", project_id="project_1"):
    from lens.analytics.can_map_engine import run_ca_pipeline, PRODUCT_CODES
    cats = PRODUCT_CODES.get(product, [1, 2, 3, 4, 5, 6])
    return run_ca_pipeline(
        category_codes=cats,
        attr_section=section,
        top_brands=n_brands,
        top_attrs=50,
        zone=zone, gender=gender, age_band=age_band, city=city,
        project_id=project_id,
    )


@st.cache_data(ttl=3600, show_spinner="Running BIP Normalization...")
def _run_bip(product, section, pctile, zone="all", gender="all", age_band="all", city="all", top_brands=19, project_id="project_1"):
    from lens.analytics.bip_engine import BIPNormalizationEngine
    from lens.analytics.can_map_engine import PRODUCT_CODES
    cats = PRODUCT_CODES.get(product, [1, 2, 3, 4, 5, 6])
    engine = BIPNormalizationEngine(
        category_codes=cats, percentile_threshold=pctile,
        top_brands=top_brands,
        zone=zone, gender=gender, age_band=age_band, city=city,
        project_id=project_id,
    )
    if hasattr(engine, "run"):
        return engine.run(attr_section=section)
    else:
        matrix = engine.get_brand_attr_matrix(attr_section=section)
        tables = engine.compute_normalization(matrix)
        return {
            "matrix": matrix,
            "tables": tables,
            "significance": engine.get_significance_summary(),
            "chart_specs": engine.get_chart_specs()
        }


@st.cache_data(ttl=3600, show_spinner="Running driver analysis...")
def _run_driver_analysis(product, driver_ids_tuple, compare_by, top_brands, pctile, zone="all", gender="all", age_band="all", city="all", project_id="project_1", pooled=False):
    from lens.analytics.driver_analysis_engine import DriverAnalysisEngine
    from lens.analytics.can_map_engine import PRODUCT_CODES
    cats = PRODUCT_CODES.get(product, list(range(1, 13)))
    engine = DriverAnalysisEngine(
        category_codes=cats,
        zone=None if zone == "all" else zone,
        gender=None if gender == "all" else gender,
        age_band=None if age_band == "all" else age_band,
        city=None if city == "all" else city,
        project_id=project_id,
    )
    return engine.run(driver_ids=list(driver_ids_tuple), compare_by=compare_by, top_brands=top_brands, percentile_threshold=pctile, product_label=product, generate_ai_insight=True, pooled=pooled)


@st.cache_data(ttl=3600, show_spinner="Mapping category reach...")
def _run_cross_category_reach(driver_ids_tuple, pctile):
    import importlib, lens.analytics.driver_analysis_engine as _dae_mod
    importlib.reload(_dae_mod)
    DriverAnalysisEngine = _dae_mod.DriverAnalysisEngine
    eng = DriverAnalysisEngine()
    return eng.cross_category_driver_reach(
        attr_ids=list(driver_ids_tuple),
        percentile_threshold=pctile,
        top_brands=10,
    )


@st.cache_data(ttl=86400)
def _get_attr_groups():
    from lens.analytics.driver_analysis_engine import DriverAnalysisEngine
    eng = DriverAnalysisEngine()
    df = eng.get_all_attributes()
    groups = {}
    for feat, grp in df.groupby("broad_feature"):
        groups[feat] = {row["attr_label"]: int(row["attr_id"]) for _, row in grp.iterrows()}
    return groups


def _ca_ai_narrative(ca_res: dict, f1_pct: float, f2_pct: float, p_val: float, chi2: float):
    """
    Data-driven CA interpretation card — no LLM call needed.
    Reads directly from fit() DataFrames in ca_results.
    """
    try:
        ca          = ca_res["ca_results"]
        attr_names  = ca.get("_attr_names", [])

        # Contributions (DataFrames: index=brand/attr, cols=F1,F2,...)
        row_ctrs = ca["row_results"]["contributions"]
        col_ctrs = ca["col_results"]["contributions"]
        row_cos2 = ca["row_results"]["cos2"]

        f1_col = "F1" if "F1" in row_ctrs.columns else row_ctrs.columns[0]
        f2_col = "F2" if "F2" in row_ctrs.columns else (row_ctrs.columns[1] if len(row_ctrs.columns) > 1 else f1_col)

        top_brands_f1 = row_ctrs[f1_col].nlargest(3).index.tolist()
        top_brands_f2 = row_ctrs[f2_col].nlargest(3).index.tolist()
        top_attrs_f1  = col_ctrs[f1_col].nlargest(3).index.tolist() if f1_col in col_ctrs.columns else attr_names[:3]

        f1f2_cos2 = row_cos2[f1_col].add(row_cos2[f2_col]) if f2_col in row_cos2.columns else row_cos2[f1_col]
        well_rep  = f1f2_cos2.nlargest(3).index.tolist()
        poor_rep  = f1f2_cos2.nsmallest(2).index.tolist()

        f_coverage = f1_pct + f2_pct
        coverage_quality = (
            "high-quality 2D representation — map captures most structure."
            if f_coverage >= 70 else
            "moderate coverage — some nuance in higher dimensions."
            if f_coverage >= 50 else
            "low 2D coverage — interpret with caution, check higher factors."
        )

        findings = [
            (f"Association patterns are **{'significant' if p_val < 0.05 else 'NOT significant'}** "
             f"(χ²={chi2:.1f}, df={ca['chi2_test']['df']}, p={p_val:.4f}). "
             + ("Brands are perceptually distinct on these attributes." if p_val < 0.05
                else "Brands appear similar on tested attributes — look for sub-group differences.")),
            (f"F1+F2 explain **{f_coverage:.1f}%** of variation (F1={f1_pct:.1f}%, F2={f2_pct:.1f}%) — {coverage_quality}"),
            (f"**F1 (horizontal axis)** defined by brands: {', '.join(top_brands_f1)}; "
             f"top attributes: {', '.join(top_attrs_f1[:2])}."),
            (f"**F2 (vertical axis)** primarily shaped by: {', '.join(top_brands_f2)}."),
            (f"**Best-represented in 2D map** (high cos²): {', '.join(well_rep)}. "
             + (f"Treat {', '.join(poor_rep)} positions with caution (low cos²)." if poor_rep else "")),
        ]
        _insight_callout(findings, "CA Interpretation", "🧭")
    except Exception:
        pass  # narrative is optional — silent fail


def _coords_to_brand_df(pc):
    """Normalise CA principal_coords to a brand-indexed DataFrame with F1,F2."""
    if isinstance(pc, pd.DataFrame):
        df = pc.copy()
        if "F1" not in df.columns:
            df.columns = ["F1", "F2"] + list(df.columns[2:]) if df.shape[1] >= 2 else df.columns
        return df[["F1", "F2"]] if "F2" in df.columns else df.iloc[:, :2].rename(
            columns={df.columns[0]: "F1", df.columns[1]: "F2"})
    try:
        df = pd.DataFrame(pc)
        idx = "brand" if "brand" in df.columns else df.columns[0]
        return df.set_index(idx)[["F1", "F2"]]
    except Exception:
        return pd.DataFrame()


def _prefmap_figure(coords_df, pref_series, metric_label, sel_brand):
    """PREFMAP (vector model): regress external preference onto CA map dims, draw the
    preference vector + project brands onto it. Returns (fig, stats) or (None, msg)."""
    import numpy as _np
    common = [b for b in coords_df.index if b in pref_series.index and pd.notna(pref_series[b])]
    if len(common) < 3:
        return None, {"error": "Need ≥3 brands with both map coordinates and a preference value."}
    C = coords_df.loc[common]
    y = pref_series.loc[common].astype(float).values
    F1 = C["F1"].astype(float).values
    F2 = C["F2"].astype(float).values
    # OLS: pref ~ a + b1*F1 + b2*F2
    X = _np.column_stack([_np.ones(len(common)), F1, F2])
    coef, *_ = _np.linalg.lstsq(X, y, rcond=None)
    a, b1, b2 = coef
    yhat = X @ coef
    ss_res = float(_np.sum((y - yhat) ** 2))
    ss_tot = float(_np.sum((y - y.mean()) ** 2))
    r2 = 1 - ss_res / ss_tot if ss_tot > 0 else 0.0
    # Preference direction (unit vector of gradient) — points toward increasing preference
    grad = _np.array([b1, b2], dtype=float)
    gnorm = _np.linalg.norm(grad)
    udir = grad / gnorm if gnorm > 0 else _np.array([0.0, 0.0])
    # Scale arrow to map extent
    span = max(_np.ptp(F1), _np.ptp(F2), 0.1)
    arrow = udir * span * 0.55
    # Project brands onto preference axis (scalar score) for ranking
    proj = {b: float(_np.dot([C.loc[b, "F1"], C.loc[b, "F2"]], udir)) for b in common}

    fig = go.Figure()
    # brand points
    colors = ["#1a5d4d" if b == sel_brand else "#30a76a" for b in common]
    sizes = [16 if b == sel_brand else 11 for b in common]
    fig.add_trace(go.Scatter(
        x=F1, y=F2, mode="markers+text", text=common, textposition="top center",
        textfont=dict(size=10, color="#374151"),
        marker=dict(size=sizes, color=colors, line=dict(width=1, color="white")),
        hovertemplate="<b>%{text}</b><br>F1=%{x:.3f} F2=%{y:.3f}<extra></extra>",
        name="Brands", showlegend=False,
    ))
    # preference vector arrow from origin
    fig.add_annotation(x=arrow[0], y=arrow[1], ax=0, ay=0, xref="x", yref="y",
                       axref="x", ayref="y", showarrow=True, arrowhead=3,
                       arrowsize=1.4, arrowwidth=3, arrowcolor="#dc2626")
    fig.add_annotation(x=arrow[0], y=arrow[1], text=f"↑ {metric_label}",
                       showarrow=False, font=dict(size=11, color="#dc2626", family="Inter"),
                       xshift=int(_np.sign(arrow[0]) * 18), yshift=int(_np.sign(arrow[1]) * 10),
                       bgcolor="rgba(220,38,38,0.08)")
    fig.add_hline(y=0, line_dash="dot", line_color="#d1d5db")
    fig.add_vline(x=0, line_dash="dot", line_color="#d1d5db")
    base = {k: v for k, v in _chart_layout_base(500).items() if k not in ("xaxis", "yaxis", "legend")}
    fig.update_layout(
        **base,
        xaxis=dict(title="F1", zeroline=False),
        yaxis=dict(title="F2", zeroline=False, scaleanchor="x", scaleratio=1),
        title=dict(text=f"Preference Map (PREFMAP) — {metric_label} vector", font=dict(size=13)),
    )
    ranking = sorted(proj.items(), key=lambda kv: -kv[1])
    return fig, {"r2": r2, "b1": b1, "b2": b2, "ranking": ranking, "n": len(common),
                 "metric": metric_label}


def _render_section_13_can_map(sel_cat, zone_arg="all", gender_arg="all", age_band_arg="all", city_arg="all", sel_brand=None, project_id="project_1", attr_ids=None, awareness_stages=None):
    _section_header(
        "🗺️ Brand Perceptual Map (CAN MAP)",
        "Correspondence Analysis: how brands cluster by attribute associations. "
        "Close points = strong shared association pattern.",
    )

    ca_cat = _SIDEBAR_TO_PRODUCT.get(sel_cat, "All")

    using_test_data = False
    test_matrix = None

    # Active filter context pill
    active_filters = [f for f in [
        f"Product: {ca_cat}" if ca_cat != "All" else None,
        f"Zone: {zone_arg}" if zone_arg != "all" else None,
        f"Gender: {gender_arg}" if gender_arg != "all" else None,
        f"Age: {age_band_arg}" if age_band_arg != "all" else None,
        f"City: {city_arg}" if city_arg != "all" else None,
    ] if f]
    _pill = "  ·  ".join(active_filters) if active_filters else "All India — no filter applied"
    st.caption(f"Active filters: {_pill}")

    col1, col2, col3 = st.columns([2, 1, 1])
    with col1:
        # Load attribute themes from project_meta.json (dynamic — not project_1-only)
        try:
            from oxdata.db_loader import get_project_meta as _ca_get_meta
            _ca_themes = _ca_get_meta(project_id).get("attribute_themes") or ["All"]
            if "All" not in _ca_themes:
                _ca_themes = ["All"] + _ca_themes
        except Exception:
            _ca_themes = ["All", "Brand and Price", "Product Performance",
                          "Design & Body Material", "After Sales Support", "Advanced Features"]
        ca_section = st.selectbox(
            "Attribute Category",
            _ca_themes,
            key="ca_section_sel",
        )
    with col2:
        ca_max_brands = st.slider("Max Brands (by Aided Awareness)", 5, 19, 10, key="ca_brands_slider",
                                  help="Top N brands by aided awareness % included in the map")
    with col3:
        pass

    if attr_ids:
        # Driver-filtered mode: bypass cache, call engine directly with attr_ids
        from lens.analytics.can_map_engine import run_ca_pipeline as _ca_filtered, PRODUCT_CODES as _PC_CA
        _cats_ca = _PC_CA.get(ca_cat, [1, 2, 3, 4, 5, 6])
        ca_res = _ca_filtered(
            category_codes=_cats_ca, top_brands=ca_max_brands,
            top_attrs=len(attr_ids), attr_ids=attr_ids,
            zone=None if zone_arg == "all" else zone_arg,
            gender=None if gender_arg == "all" else gender_arg,
            age_band=None if age_band_arg == "all" else age_band_arg,
            city=None if city_arg == "all" else city_arg,
            project_id=project_id,
            awareness_stages=awareness_stages or None,
        )
    else:
        ca_res  = _run_can_map(ca_cat, None if ca_section == "All" else ca_section, ca_max_brands,
                               zone_arg, gender_arg, age_band_arg, city_arg, project_id=project_id)
    bip_res_for_interp = _run_bip(ca_cat, None if ca_section == "All" else ca_section,
                                   65, zone_arg, gender_arg, age_band_arg, city_arg, top_brands=ca_max_brands,
                                   project_id=project_id)
    interps = _get_cached_imagery_interpretations(
        ca_res, bip_res_for_interp, ca_cat, ca_section, 65, ca_max_brands,
        zone_arg, gender_arg, age_band_arg, city_arg,
    )

    if ca_res["status"] == "ok":
        eig_df = ca_res["ca_results"]["eigenvalues"]
        f1_pct = eig_df["Inertia_%"].iloc[0]
        f2_pct = eig_df["Inertia_%"].iloc[1] if len(eig_df) > 1 else 0
        p_val  = ca_res["ca_results"]["chi2_test"]["p_value"]
        chi2   = ca_res["ca_results"]["chi2_test"]["chi2"]
        chi2_t = ca_res["ca_results"]["chi2_test"]
        charts = {c["type"]: c["figure"] for c in ca_res["chart_specs"]}
        cm_t   = ca_res["tables"]

        ca_r   = cm_t["row_results"]
        ca_c   = cm_t["col_results"]

        # ── Significance banner ────────────────────────────────────────────────
        _sig_color = "#15803d" if chi2_t["significant"] else "#dc2626"
        _sig_icon  = "✓" if chi2_t["significant"] else "✗"
        _sig_label = "Significant" if chi2_t["significant"] else "Not significant"
        st.markdown(
            f"<div style='display:flex;gap:16px;align-items:center;padding:8px 14px;"
            f"background:rgba(0,0,0,0.03);border-radius:8px;border-left:4px solid {_sig_color};"
            f"font-size:0.82rem;margin-bottom:8px;'>"
            f"<span style='color:{_sig_color};font-weight:700;font-size:1rem;'>{_sig_icon} {_sig_label}</span>"
            f"<span>χ²&nbsp;=&nbsp;<b>{chi2:.2f}</b></span>"
            f"<span>df&nbsp;=&nbsp;<b>{chi2_t['df']}</b></span>"
            f"<span>p&nbsp;=&nbsp;<b>{p_val:.4f}</b></span>"
            f"<span style='color:#6b7280'>·</span>"
            f"<span>F1+F2&nbsp;=&nbsp;<b>{f1_pct+f2_pct:.1f}%</b> of inertia</span>"
            f"<span style='color:#6b7280'>·</span>"
            f"<span>Total inertia&nbsp;=&nbsp;<b>{ca_res['ca_results']['total_inertia']:.4f}</b></span>"
            f"</div>",
            unsafe_allow_html=True,
        )

        _INDEX_COLS = ("brand_name", "attr_label", "brand", "attribute", "index", "Factor")

        def _safe_df(obj):
            if isinstance(obj, pd.DataFrame): return obj
            try:
                df = pd.DataFrame(obj)
                if not df.empty and df.columns[0] in _INDEX_COLS:
                    df = df.set_index(df.columns[0])
                return df
            except Exception: return pd.DataFrame()

        tabs = st.tabs([
            "1. Input Table",
            "2. Chi-Square & Eigenvalues",
            "3. Symmetric Map",
            "4. Asymmetric Map",
            "5. Brand Results",
            "6. Attribute Results",
            "7. Diagnostics",
            "8. Preference Map (PREFMAP)",
        ])

        # Extract brand + attr names for specific highlighting
        _known = (
            list(ca_res["ca_results"].get("_brand_names", [])) +
            list(ca_res["ca_results"].get("_attr_names", []))
        )

        def _ca_ai(key, accent="#6366f1"):
            txt = interps.get(key, "")
            if txt:
                _structured_ai_card(txt, "AI Analysis", accent, known_names=_known)

        with tabs[0]:
            st.markdown("**Contingency Table — Brand × Attribute % Association (CA input)**")
            st.caption(
                "ℹ️ **Why zeros appear:** The database stores only positive brand-attribute associations "
                "(respondents who explicitly linked a brand to an attribute). A 0.00 means zero respondents "
                "in the current filter linked that brand to that attribute — this is a true data gap, not a "
                "missing value. Brands with sparse coverage (few attributes) are normal for niche or newer brands. "
                "Zero-variance attributes (all brands identical) are automatically excluded from the CA."
            )
            mat = _safe_df(cm_t["contingency_table"])
            if not mat.empty:
                # Transpose: attributes as rows, brands as columns
                mat = mat.T
                mat_with_totals = mat.copy()
                mat_with_totals["Row Total"] = mat.sum(axis=1)
                totals_row = mat.sum(axis=0)
                totals_row["Row Total"] = float(mat.values.sum())
                mat_with_totals.loc["Col Total"] = totals_row
                zero_count = int((mat == 0).sum().sum())
                if zero_count > 0:
                    st.caption(f"⚠️ {zero_count} zero cells ({zero_count / mat.size * 100:.1f}% of matrix). "
                               f"Brands with many zeros have sparse association data.")
                st.dataframe(mat_with_totals.round(2), use_container_width=True)
            else:
                st.info("Contingency table unavailable.")
            _ca_ai("tab_0", "#6366f1")

        with tabs[1]:
            k1, k2, k3, k4, k5 = st.columns(5)
            with k1: _metric_card("Total Inertia", f"{ca_res['ca_results']['total_inertia']:.4f}", "📐", "χ²/N — overall variation")
            with k2: _metric_card("F1 Inertia", f"{f1_pct:.1f}%", "🟦", f"Eigenvalue: {eig_df['Eigenvalue'].iloc[0]:.5f}")
            with k3: _metric_card("F2 Inertia", f"{f2_pct:.1f}%", "🟨", f"Eigenvalue: {eig_df['Eigenvalue'].iloc[1]:.5f}" if len(eig_df) > 1 else "")
            with k4: _metric_card("F1+F2", f"{f1_pct+f2_pct:.1f}%", "📊", "2D map coverage")
            with k5: _metric_card("Chi-Square", f"{chi2:.2f}", "⚖️", f"df={chi2_t['df']}  p={p_val:.4f}  {'✓ Sig.' if p_val < 0.05 else '✗ Not sig.'}")
            st.markdown("<div style='margin:10px 0;'></div>", unsafe_allow_html=True)
            c_col1, c_col2 = st.columns(2)
            with c_col1:
                st.markdown("**Test of Independence**")
                chi2_rows = [
                    {"Statistic": "Chi-square (obs)", "Value": f"{chi2_t['chi2']:.4f}"},
                    {"Statistic": "Chi-square (crit)", "Value": f"{chi2_t['critical']:.4f}"},
                    {"Statistic": "Degrees of freedom", "Value": str(chi2_t['df'])},
                    {"Statistic": "p-value", "Value": f"{chi2_t['p_value']:.6f}"},
                    {"Statistic": "alpha", "Value": str(chi2_t['alpha'])},
                    {"Statistic": "Significant?", "Value": "YES ✓" if chi2_t['significant'] else "NO"},
                ]
                st.dataframe(pd.DataFrame(chi2_rows).set_index("Statistic"), use_container_width=True)
            with c_col2:
                st.markdown("**Eigenvalues & Inertia**")
                eig_display = eig_df.copy().rename(columns={
                    "Eigenvalue": "λ", "Inertia_%": "% Inertia", "Cum_Inertia_%": "Cumul %"
                })
                st.dataframe(
                    eig_display.style.background_gradient(subset=["% Inertia"], cmap="Blues"),
                    use_container_width=True
                )
            _ca_ai("tab_1", "#0ea5e9")

        with tabs[2]:
            from lens.analytics.can_map_engine import build_perceptual_map as _bpm
            _sym_view = st.radio(
                "View mode", ["Aggregated (default)", "Custom (attribute filter)"],
                horizontal=True, key="can_sym_view_mode",
                help="Aggregated = pre-computed brand-aggregated map. Custom = pick which attribute labels to show."
            )
            if _sym_view == "Aggregated (default)":
                _safe_chart(charts.get("symmetric_map"))
            else:
                _ca_res_inner = ca_res["ca_results"]
                _all_attrs_sym = list(_ca_res_inner.get("_attr_names", []))
                _col_ri_sym = _ca_res_inner.get("_col_rel_inertia", None)
                if _col_ri_sym is not None and len(_col_ri_sym) == len(_all_attrs_sym):
                    import numpy as _np_sym
                    _sort_idx = list(_np_sym.argsort(_col_ri_sym)[::-1])
                    _sorted_attrs_sym = [_all_attrs_sym[i] for i in _sort_idx]
                else:
                    _sorted_attrs_sym = _all_attrs_sym
                _default_sym = _sorted_attrs_sym[:8] if len(_sorted_attrs_sym) >= 8 else _sorted_attrs_sym
                _sym_ctrl_l, _sym_ctrl_r = st.columns([3, 1])
                with _sym_ctrl_l:
                    _sel_attrs_sym = st.multiselect(
                        "Attribute labels to show",
                        options=_sorted_attrs_sym, default=_default_sym,
                        key="can_sym_attr_sel",
                        help="Attributes ranked by inertia (top = highest contribution). Select any subset."
                    )
                with _sym_ctrl_r:
                    _show_dots_sym = st.checkbox(
                        "Show all dots", value=True, key="can_sym_show_dots",
                        help="Uncheck to hide unlabelled attribute dots."
                    )
                if _sel_attrs_sym or _show_dots_sym:
                    _sym_fig = _bpm(_ca_res_inner, map_type="symmetric",
                                    label_attrs=_sel_attrs_sym if _sel_attrs_sym else [],
                                    show_all_dots=_show_dots_sym)
                    _theme_fig(_sym_fig)
                    st.plotly_chart(_sym_fig, use_container_width=True, config={"editable": True})
                    st.caption("💡 Labels are draggable — click and drag any label to reposition it.")
                else:
                    st.info("Select at least one attribute or enable 'Show all dots'.")
            st.caption("Brands (squares) and attributes (circles) on principal F1×F2 axes. Larger markers = higher contribution.")
            st.markdown("**Brand Coordinates (F1, F2)**")
            pc = ca_res["ca_results"]["row_results"]["principal_coords"]
            if isinstance(pc, pd.DataFrame):
                pc_show = pc[["F1", "F2"]].round(4) if "F1" in pc.columns else pc.iloc[:, :2].round(4)
            else:
                pc_show = pd.DataFrame(pc).set_index("brand")[["F1", "F2"]].round(4) if pc else pd.DataFrame()
            st.dataframe(pc_show.style.background_gradient(cmap="RdYlGn", axis=None), use_container_width=True)
            _ca_ai("tab_2", "#8b5cf6")

        with tabs[3]:
            asym_tabs = st.tabs(["Asymmetric Map", "Confidence Biplot"])
            with asym_tabs[0]:
                from lens.analytics.can_map_engine import build_perceptual_map as _bpm_a
                _asym_view = st.radio(
                    "View mode", ["Aggregated (default)", "Custom (attribute filter)"],
                    horizontal=True, key="can_asym_view_mode",
                    help="Aggregated = pre-computed brand-aggregated map. Custom = pick attribute labels."
                )
                if _asym_view == "Aggregated (default)":
                    _safe_chart(charts.get("asymmetric_map"))
                else:
                    _ca_res_asym = ca_res["ca_results"]
                    _all_attrs_asym = list(_ca_res_asym.get("_attr_names", []))
                    _col_ri_asym = _ca_res_asym.get("_col_rel_inertia", None)
                    if _col_ri_asym is not None and len(_col_ri_asym) == len(_all_attrs_asym):
                        import numpy as _np_asym
                        _sort_idx_a = list(_np_asym.argsort(_col_ri_asym)[::-1])
                        _sorted_attrs_asym = [_all_attrs_asym[i] for i in _sort_idx_a]
                    else:
                        _sorted_attrs_asym = _all_attrs_asym
                    _default_asym = _sorted_attrs_asym[:8] if len(_sorted_attrs_asym) >= 8 else _sorted_attrs_asym
                    _asym_ctrl_l, _asym_ctrl_r = st.columns([3, 1])
                    with _asym_ctrl_l:
                        _sel_attrs_asym = st.multiselect(
                            "Attribute labels to show",
                            options=_sorted_attrs_asym, default=_default_asym,
                            key="can_asym_attr_sel",
                            help="Attributes ranked by inertia contribution."
                        )
                    with _asym_ctrl_r:
                        _show_dots_asym = st.checkbox("Show all dots", value=True, key="can_asym_show_dots")
                    if _sel_attrs_asym or _show_dots_asym:
                        _asym_fig = _bpm_a(_ca_res_asym, map_type="asymmetric",
                                          label_attrs=_sel_attrs_asym if _sel_attrs_asym else [],
                                          show_all_dots=_show_dots_asym)
                        _theme_fig(_asym_fig)
                        st.plotly_chart(_asym_fig, use_container_width=True, config={"editable": True})
                        st.caption("💡 Labels are draggable — click and drag any label to reposition it.")
                    else:
                        st.info("Select at least one attribute or enable 'Show all dots'.")
                st.caption("Rows on principal coordinates, columns on standard coordinates (XLSTAT default asymmetric biplot).")
            with asym_tabs[1]:
                from lens.analytics.can_map_engine import build_biplot_map as _bbm
                _bip_view = st.radio(
                    "View mode", ["Aggregated (default)", "Custom (attribute filter)"],
                    horizontal=True, key="can_bip_view_mode",
                    help="Aggregated = pre-computed biplot. Custom = pick attribute labels with anti-overlap placement."
                )
                if _bip_view == "Aggregated (default)":
                    _safe_chart(charts.get("biplot_ellipses"))
                else:
                    _ca_res_bip = ca_res["ca_results"]
                    _all_attrs_bip = list(_ca_res_bip.get("_attr_names", []))
                    _col_ri_bip = _ca_res_bip.get("_col_rel_inertia", None)
                    if _col_ri_bip is not None and len(_col_ri_bip) == len(_all_attrs_bip):
                        import numpy as _np_bip
                        _sort_idx_b = list(_np_bip.argsort(_col_ri_bip)[::-1])
                        _sorted_attrs_bip = [_all_attrs_bip[i] for i in _sort_idx_b]
                    else:
                        _sorted_attrs_bip = _all_attrs_bip
                    _default_bip = _sorted_attrs_bip[:8] if len(_sorted_attrs_bip) >= 8 else _sorted_attrs_bip
                    _bip_ctrl_l, _bip_ctrl_r = st.columns([3, 1])
                    with _bip_ctrl_l:
                        _sel_attrs_bip = st.multiselect(
                            "Attribute labels to show",
                            options=_sorted_attrs_bip, default=_default_bip,
                            key="can_bip_attr_sel",
                            help="Attributes ranked by inertia contribution."
                        )
                    with _bip_ctrl_r:
                        _show_dots_bip = st.checkbox("Show all dots", value=True, key="can_bip_show_dots")
                    if _sel_attrs_bip or _show_dots_bip:
                        _bip_fig = _bbm(_ca_res_bip,
                                        label_attrs=_sel_attrs_bip if _sel_attrs_bip else [],
                                        show_all_dots=_show_dots_bip)
                        _theme_fig(_bip_fig)
                        st.plotly_chart(_bip_fig, use_container_width=True, config={"editable": True})
                    else:
                        st.info("Select at least one attribute or enable 'Show all dots'.")
                st.caption("Approx 95% confidence ellipses. Larger ellipse = less stable brand position.")
            _ca_ai("tab_3", "#f59e0b")

        with tabs[4]:
            itabs = st.tabs(["Principal Coords", "Standard Coords", "Contributions", "Cos²", "Row Profiles", "χ² Distances"])
            with itabs[0]:
                st.caption("Principal coordinates = brand positions on the map (XLSTAT: Scores F/Rows)")
                st.dataframe(_safe_df(ca_r["principal_coords"]).round(4), use_container_width=True)
            with itabs[1]:
                st.caption("Standard coordinates = normalised positions (XLSTAT: Standard scores/Rows)")
                st.dataframe(_safe_df(ca_r["standard_coords"]).round(4), use_container_width=True)
            with itabs[2]:
                st.caption("Contribution of each brand to each factor (0–1; sums to 1 per factor)")
                st.dataframe(_safe_df(ca_r["contributions"]).round(4), use_container_width=True)
            with itabs[3]:
                st.caption("Squared cosines: quality of 2D representation per brand (closer to 1 = better)")
                st.dataframe(_safe_df(ca_r["cos2"]).round(4), use_container_width=True)
            with itabs[4]:
                st.caption("Row profiles: brand's attribute distribution (each row sums to 1)")
                st.dataframe(_safe_df(ca_r["profiles"]).round(4), use_container_width=True)
            with itabs[5]:
                st.caption("Chi-square distance between brands. 0 = identical profiles")
                st.dataframe(_safe_df(ca_r["chisq_distances"]).round(4), use_container_width=True)
            _ca_ai("tab_4", "#10b981")

        with tabs[5]:
            itabs = st.tabs(["Principal Coords", "Standard Coords", "Contributions", "Cos²", "Col Profiles"])
            with itabs[0]:
                st.caption("Attribute positions on the map (XLSTAT: Scores F/Columns)")
                st.dataframe(_safe_df(ca_c["principal_coords"]).round(4), use_container_width=True)
            with itabs[1]:
                st.dataframe(_safe_df(ca_c["standard_coords"]).round(4), use_container_width=True)
            with itabs[2]:
                st.caption("Attribute contribution to each factor")
                st.dataframe(_safe_df(ca_c["contributions"]).round(4), use_container_width=True)
            with itabs[3]:
                st.dataframe(_safe_df(ca_c["cos2"]).round(4), use_container_width=True)
            with itabs[4]:
                st.dataframe(_safe_df(ca_c["profiles"]).round(4), use_container_width=True)
            _ca_ai("tab_5", "#f97316")

        with tabs[6]:
            diag_tabs = st.tabs(["Scree", "Brand→F1", "Brand→F2", "Attr→F1", "Cos² Heatmap", "χ² Heatmap", "Row Profiles", "Trajectory"])
            with diag_tabs[0]:
                _safe_chart(charts.get("scree"))
            with diag_tabs[1]:
                _safe_chart(charts.get("contribution_f1"))
            with diag_tabs[2]:
                _safe_chart(charts.get("contribution_f2"))
            with diag_tabs[3]:
                if "contribution_cols_f1" in charts:
                    _safe_chart(charts["contribution_cols_f1"])
                else:
                    st.info("Attribute contribution unavailable.")
            with diag_tabs[4]:
                _safe_chart(charts.get("cos2_heatmap"))
            with diag_tabs[5]:
                _safe_chart(charts.get("chisq_distance_heatmap"))
            with diag_tabs[6]:
                _safe_chart(charts.get("row_profiles_radar"))
            with diag_tabs[7]:
                _safe_chart(charts.get("brand_trajectory"))
            _ca_ai("tab_6", "#ef4444")

        with tabs[7]:
            # ── PREFMAP: overlay external preference vector on the perceptual map ──
            st.markdown("**Where does preference point on the perceptual map?**")
            st.caption(
                "PREFMAP (vector model): external preference data is regressed onto the CA map axes "
                "(pref ~ F1 + F2). The red arrow points toward increasing preference; brands projected "
                "onto it give the preference ranking the map implies. R² = how well the 2-D map explains "
                "that preference metric. This is XLSTAT's preference-mapping deliverable."
            )
            try:
                _pc_raw = ca_res["ca_results"]["row_results"]["principal_coords"]
                _coords = _coords_to_brand_df(_pc_raw)
                _bench = _get_competitive_benchmarking()
                if _coords.empty or _bench is None or _bench.empty:
                    st.info("Preference mapping needs both map coordinates and brand preference metrics.")
                else:
                    _pref_opts = {
                        "CONSIDERATION": "Consideration %", "PREFERRED": "Preference %",
                        "TOTAL_AIDED": "Total Awareness %", "TOM": "Top of Mind %",
                        "nps": "NPS Score", "csat": f"CSAT (0–{_get_csat_scale()})",
                    }
                    _pref_avail = {k: v for k, v in _pref_opts.items() if k in _bench.columns}
                    _pm_metric = st.selectbox(
                        "Preference metric to map", list(_pref_avail.keys()),
                        format_func=lambda k: _pref_avail[k], key="prefmap_metric")
                    _pref_series = _bench.set_index("brand_name")[_pm_metric]
                    _fig_pm, _stats_pm = _prefmap_figure(
                        _coords, _pref_series, _pref_avail[_pm_metric], sel_brand)
                    if _fig_pm is None:
                        st.info(_stats_pm.get("error", "PREFMAP unavailable."))
                    else:
                        st.plotly_chart(_theme_fig(_fig_pm), use_container_width=True)
                        _r2 = _stats_pm["r2"]
                        _fit_q = ("strong" if _r2 >= 0.7 else "moderate" if _r2 >= 0.4 else "weak")
                        _rank = _stats_pm["ranking"]
                        _findings_pm = [
                            f"The 2-D map explains **{_r2:.0%}** of variation in **{_pref_avail[_pm_metric]}** "
                            f"({_fit_q} fit). Higher R² = preference aligns cleanly with the perceptual axes.",
                            f"Map-implied preference leader: **{_rank[0][0]}**, then "
                            f"{', '.join(b for b, _ in _rank[1:4])}.",
                            "Brands further along the red arrow are perceived more favourably on this metric. "
                            "Brands off-axis are liked/disliked for reasons the 2-D map doesn't capture.",
                        ]
                        if sel_brand in dict(_rank):
                            _pos = [b for b, _ in _rank].index(sel_brand) + 1
                            _findings_pm.append(
                                f"**{sel_brand}** ranks #{_pos} of {_stats_pm['n']} on the preference axis.")
                        _insight_callout(_findings_pm, "Preference Mapping Read-out", "🧭")

                        _rank_df = pd.DataFrame(
                            [{"Rank": i + 1, "Brand": b, "Pref. axis score": round(s, 3)}
                             for i, (b, s) in enumerate(_rank)])
                        with st.expander("Preference-axis ranking (brand projections)", expanded=False):
                            st.dataframe(_rank_df, hide_index=True, use_container_width=True)
            except Exception as _e_pm:
                st.warning(f"PREFMAP unavailable: {_e_pm}")

        # ── Rule-based CA narrative (always shown, below tabs) ──────────────────
        _ca_ai_narrative(ca_res, f1_pct, f2_pct, p_val, chi2)

    else:
        st.info(f"CAN MAP: {ca_res['message']}")



def _emphasize_radar_brand(fig, sel_brand):
    """Post-process a multi-brand radar: spotlight the focus brand (bold, opaque,
    filled) and recede every other brand (thin, faint) so the chart reads as
    'focus brand vs the field' instead of unreadable spaghetti."""
    if fig is None or not sel_brand:
        return fig
    try:
        sb = str(sel_brand).strip().lower()
        for tr in fig.data:
            nm = str(getattr(tr, "name", "") or "").strip().lower()
            is_focus = nm == sb
            if hasattr(tr, "line"):
                tr.line.width = 3.4 if is_focus else 1.0
            tr.opacity = 1.0 if is_focus else 0.28
            if hasattr(tr, "fill"):
                tr.fill = "toself" if is_focus else "none"
                if is_focus:
                    tr.fillcolor = "rgba(26,93,77,0.16)"
            if is_focus:
                tr.legendrank = 1
    except Exception:
        return fig
    return fig


def _render_section_14_bip(sel_cat, zone_arg="all", gender_arg="all", age_band_arg="all", city_arg="all", sel_brand=None, project_id="project_1", attr_ids=None):
    _section_header(
        "📊 Brand Image Profiling (BIP Normalization)",
        "Which attributes each brand owns vs. the market average. "
        "YES = brand significantly over-associated with that attribute.",
    )

    bip_cat = _SIDEBAR_TO_PRODUCT.get(sel_cat, "All")

    bip_using_test = False
    bip_test_matrix = None

    active_filters = [f for f in [
        f"Product: {bip_cat}" if bip_cat != "All" else None,
        f"Zone: {zone_arg}" if zone_arg != "all" else None,
        f"Gender: {gender_arg}" if gender_arg != "all" else None,
        f"Age: {age_band_arg}" if age_band_arg != "all" else None,
        f"City: {city_arg}" if city_arg != "all" else None,
    ] if f]
    _bip_pill = "  ·  ".join(active_filters) if active_filters else "All India — no filter applied"
    st.caption(f"Active filters: {_bip_pill}")

    col1, col2, col3 = st.columns([2, 1, 1])
    with col1:
        bip_section = st.selectbox(
            "Attribute Category",
            ["All", "Brand and Price", "Product Performance",
             "Design & Body Material", "After Sales Support", "Advanced Features"],
            key="bip_section_sel",
        )
    with col2:
        bip_pctile = st.selectbox("Significance", [50, 60, 65, 70, 75, 80, 85, 90], index=2, key="bip_pctile_sel")
    with col3:
        bip_min_n = st.number_input("Min base N per brand", min_value=0, max_value=5000,
                                     value=0, step=50, key="bip_min_n",
                                     help="Exclude brands with fewer respondents than this threshold")

    # Resolve dynamic brand count for BIP (project_1 has ~56, AK has ~15)
    _actual_brand_count = 19
    try:
        from oxdata.db_loader import get_db_path as _bip_get_dbp
        _dbp = _bip_get_dbp(required_table="dim_brand", project_id=project_id)
        if _dbp and _dbp.exists():
            import sqlite3
            _c = sqlite3.connect(str(_dbp))
            _cnt = _c.execute("SELECT COUNT(*) FROM dim_brand").fetchone()[0]
            _c.close()
            if _cnt > 0:
                _actual_brand_count = _cnt
        else:
            from lens.data_layer import get_project_layer as _bip_get_lyr
            _lyr = _bip_get_lyr(project_id)
            if _lyr is not None and len(_lyr.all_brands) > 0:
                _actual_brand_count = len(_lyr.all_brands)
    except Exception:
        pass
    _bip_top_brands = min(_actual_brand_count, 19)

    if attr_ids:
        # Driver-filtered mode: bypass cache, call engine directly with attr_ids
        from lens.analytics.bip_engine import BIPNormalizationEngine as _BIPFiltered
        from lens.analytics.can_map_engine import PRODUCT_CODES as _PC_BIP
        _cats_bip = _PC_BIP.get(bip_cat, [1, 2, 3, 4, 5, 6])
        _bip_f_eng = _BIPFiltered(
            category_codes=_cats_bip, percentile_threshold=float(bip_pctile), top_brands=_bip_top_brands,
            zone=zone_arg, gender=gender_arg, age_band=age_band_arg, city=city_arg,
            project_id=project_id,
        )
        _bip_filt_mat = _bip_f_eng.get_brand_attr_matrix(attr_ids=attr_ids)
        bip_res = _bip_f_eng.run(matrix=_bip_filt_mat)
    else:
        bip_res       = _run_bip(bip_cat, None if bip_section == "All" else bip_section, bip_pctile,
                                  zone_arg, gender_arg, age_band_arg, city_arg, top_brands=_bip_top_brands,
                                  project_id=project_id)
    if attr_ids:
        from lens.analytics.can_map_engine import run_ca_pipeline as _ca_bip_filt, PRODUCT_CODES as _PC_CA_BIP
        _cats_ca_bip = _PC_CA_BIP.get(bip_cat, [1, 2, 3, 4, 5, 6])
        ca_res_for_bip = _ca_bip_filt(
            category_codes=_cats_ca_bip, top_brands=_bip_top_brands,
            top_attrs=len(attr_ids), attr_ids=attr_ids,
            zone=None if zone_arg == "all" else zone_arg,
            gender=None if gender_arg == "all" else gender_arg,
            age_band=None if age_band_arg == "all" else age_band_arg,
            city=None if city_arg == "all" else city_arg,
            project_id=project_id,
        )
    else:
        ca_res_for_bip = _run_can_map(bip_cat, None if bip_section == "All" else bip_section, _bip_top_brands,
                                       zone_arg, gender_arg, age_band_arg, city_arg, project_id=project_id)
    bip_interps   = _get_cached_imagery_interpretations(
        ca_res_for_bip, bip_res, bip_cat, bip_section, bip_pctile, _bip_top_brands,
        zone_arg, gender_arg, age_band_arg, city_arg,
    )

    if bip_res is None or bip_res.get("status") in ("error", "empty"):
        st.warning("No data for this combination. Try 'All' category or 'Brand and Price' section.")
        return

    bip_matrix = bip_res.get("matrix")
    bip_tables = bip_res.get("tables", {})
    bip_charts = bip_res.get("chart_specs", [])

    if bip_matrix is None or (hasattr(bip_matrix, "empty") and bip_matrix.empty):
        st.info("BIP Analysis: No data available for selected filters.")
        return

    # ── Apply min base-N filter ────────────────────────────────────────────────
    if bip_min_n > 0 and hasattr(bip_matrix, "attrs"):
        n_resp = bip_matrix.attrs.get("n_resp", None)
        if n_resp is not None and isinstance(n_resp, pd.Series):
            qualify = n_resp[n_resp >= bip_min_n].index
            if len(qualify) > 0 and len(qualify) < len(bip_matrix):
                excluded = set(bip_matrix.index) - set(qualify)
                st.warning(f"Min base N = {bip_min_n}: excluded {len(excluded)} brand(s) — {', '.join(str(e) for e in excluded)}")
                bip_matrix = bip_matrix.loc[bip_matrix.index.isin(qualify)]
                # Re-run normalization on filtered matrix (not cached — apply in-memory)
                from lens.analytics.bip_engine import BIPNormalizationEngine as _BIPE
                _eng_filt = _BIPE(percentile_threshold=float(bip_pctile))
                bip_filt = _eng_filt.run(matrix=bip_matrix)
                bip_tables = bip_filt.get("tables", bip_tables)
                bip_charts = bip_filt.get("chart_specs", bip_charts)

    # KPI row
    pcts = bip_tables.get("percentiles", {})
    k1, k2, k3, k4 = st.columns(4)
    pct_val = pcts.get(f"p{bip_pctile}", pcts.get("p65", 0))
    with k1: _metric_card("Brands", str(len(bip_matrix.index)), "🏢")
    with k2: _metric_card("Attributes", str(len(bip_matrix.columns)), "🎯")
    with k3: _metric_card(f"p{bip_pctile} Gate", f"{pct_val:.1f}", "📊", "Significance threshold")
    with k4: _metric_card("Avg Assoc", f"{bip_matrix.values.mean():.1f}%", "📈")

    st.markdown("<br>", unsafe_allow_html=True)

    def _bip_fig(idx):
        if idx < len(bip_charts):
            f = bip_charts[idx]
            return go.Figure(f["figure"] if "figure" in f else f)
        return None

    def _safe_round(df, n=2):
        return df.round(n) if df is not None and hasattr(df, "round") else (df or pd.DataFrame())

    tabs = st.tabs(["1. Raw Matrix", "2. Normalized Matrix", "3. Filtered Scores", "4. Significance (YES/NO)", "5. Visual Profiles", "6. Statistics"])

    # Extract brand + attr names from BIP matrix for specific highlighting
    _bip_known = (
        list(bip_matrix.index) if hasattr(bip_matrix, "index") else []
    ) + (
        list(bip_matrix.columns) if hasattr(bip_matrix, "columns") else []
    )

    def _bip_ai(key, accent="#6366f1"):
        txt = bip_interps.get(key, "")
        if txt:
            _structured_ai_card(txt, "AI Analysis", accent, known_names=_bip_known)

    # Shared transpose toggle for BIP tables — drivers as rows, brands as columns
    bip_transpose = st.checkbox("Drivers as Rows (brands as columns)", value=True,
                                 key="bip_transpose_toggle",
                                 help="Transpose tables so drivers are row labels and brands are column headers")

    def _bip_table(df, style_fn=None):
        """Show df or its transpose based on toggle; apply optional Styler."""
        if df is None or (hasattr(df, "empty") and df.empty):
            st.info("No data.")
            return
        display = df.T if bip_transpose else df
        if style_fn:
            st.dataframe(display.style.map(style_fn), use_container_width=True)
        else:
            st.dataframe(_safe_round(display), use_container_width=True)

    with tabs[0]:
        st.caption(
            "ℹ️ **Why zeros appear in the raw matrix:** Each cell = % of respondents who "
            "linked brand B to attribute A (value=1 in fact_brand_imagery). A **0.00** means "
            "zero respondents in the current filter made that association — this is **valid "
            "data, not missing data**. It reflects genuine lack of attribution (e.g. a niche "
            "brand not associated with a premium-price attribute). "
            "Zeros appear more when: (a) narrow category/demographic filter is applied, "
            "(b) brand has low overall awareness, or (c) attribute is outside the brand's "
            "perceived territory. The normalization (tabs 2–3) accounts for this by "
            "comparing each brand's score to the column mean."
        )
        _bip_table(bip_tables.get("table1_raw"))
        _bip_ai("bip_0", "#6366f1")

    with tabs[1]:
        _bip_table(bip_tables.get("table3_norm_pct"))
        _bip_ai("bip_1", "#8b5cf6")

    with tabs[2]:
        # ── TABLE 4: Threshold-filtered scores (Excel structure: appears before YES/NO) ──
        # Values below the significance gate are zeroed. Non-zero = significant association.
        # Mirrors the Excel table that precedes the YES/NO significance column.
        _pctile_val = bip_tables.get('percentiles', {}).get(f'p{bip_pctile}', bip_tables.get('percentiles', {}).get('p65', '—'))
        _pctile_str = f"{_pctile_val:.2f}" if isinstance(_pctile_val, (int, float)) else str(_pctile_val)
        st.caption(
            f"Threshold-filtered normalized deviation. Cells below p{bip_pctile} gate ({_pctile_str}) are zeroed. "
            "Non-zero = brand significantly over-indexed on that attribute. Matches Excel pre-YES/NO table."
        )
        _bip_table(bip_tables.get("table4_filtered"))
        _bip_ai("bip_2a", "#0891b2")

    with tabs[3]:
        # ── TABLE 14: YES / NO significance flags ─────────────────────────────
        st.caption(
            f"YES = brand's normalized deviation exceeds the p{bip_pctile} significance gate. "
            "Green = significant positive association. Red = below gate."
        )
        _bip_table(
            bip_tables.get("table14_significance"),
            lambda x: "background-color: #dcfce7; color: #166534" if x == "YES"
                      else "background-color: #fee2e2; color: #991b1b",
        )
        _bip_ai("bip_2", "#10b981")

    with tabs[4]:
        vtabs = st.tabs(["Normalized Heatmap", "Section Radar", "Brand Strength"])
        with vtabs[0]:
            _safe_chart(_bip_fig(0))
            st.caption("Heatmap: green = brand over-indexed on attribute, red = under-indexed. YES/NO = significance flag.")
        with vtabs[1]:
            _safe_chart(_emphasize_radar_brand(_bip_fig(4), sel_brand))
            _foc = f"**{sel_brand}** spotlighted; other brands receded for comparison. " if sel_brand else ""
            st.caption(f"{_foc}Section radar: normalised deviation by attribute group (Brand, Product, Service, Design, Features).")
        with vtabs[2]:
            _safe_chart(_bip_fig(1))
            st.caption("Brand strength: count of YES attributes per section. More green bars = broader positive ownership.")
        _bip_ai("bip_3", "#f59e0b")

    with tabs[5]:
        st_tabs = st.tabs(["Combined Score", "All Ranks", "Column Averages", "Significance Summary", "Attribute Insights", "Distribution"])
        with st_tabs[0]:
            df = bip_tables.get("table13_combined")
            if df is not None: st.dataframe(_safe_round(df), use_container_width=True)
        with st_tabs[1]:
            df = bip_tables.get("table7_ranks")
            if df is not None: st.dataframe(df, use_container_width=True)
        with st_tabs[2]:
            df = bip_tables.get("column_averages")
            if df is not None: st.dataframe(_safe_round(df), use_container_width=True)
        with st_tabs[3]:
            _sig_inner = st.tabs(["YES/NO Flags", "Cell p-values", "Summary"])
            with _sig_inner[0]:
                t14 = bip_tables.get("table14_significance")
                if t14 is not None and not t14.empty:
                    st.caption("YES = brand significantly over-indexed on that attribute (binomial z-test, threshold from percentile slider).")
                    st.dataframe(t14, use_container_width=True)
                else:
                    st.info("Significance flag matrix unavailable.")
            with _sig_inner[1]:
                t15 = bip_tables.get("table15_p_values")
                if t15 is not None and not t15.empty:
                    st.caption("Cell-level two-sided p-values (z-test). Values below 0.05 correspond to YES flags above.")
                    styled = _safe_round(t15)
                    st.dataframe(styled, use_container_width=True)
                else:
                    st.info("Cell-level p-values unavailable.")
            with _sig_inner[2]:
                sig_df = bip_res.get("significance")
                if sig_df is None or (hasattr(sig_df, "empty") and sig_df.empty):
                    sig_df = bip_res.get("significance_summary")
                if sig_df is not None and not (hasattr(sig_df, "empty") and sig_df.empty):
                    st.caption("Number of significantly over-indexed attributes per brand.")
                    st.dataframe(sig_df, use_container_width=True)
                else:
                    st.info("Significance summary unavailable.")
        with st_tabs[4]:
            st.markdown("**Attribute Competitiveness & Top/Bottom Performers**")
            c1, c2 = st.columns(2)
            with c1:
                _safe_chart(_bip_fig(2))
                st.caption("Attribute range across brands. Wide range = contested attribute.")
            with c2:
                _safe_chart(_bip_fig(3))
                st.caption("Top/bottom performers per attribute.")
        with st_tabs[5]:
            _safe_chart(_bip_fig(5))
            st.caption("Distribution of normalised deviations. Cells beyond gate are flagged YES.")
        _bip_ai("bip_4", "#ef4444")

    # ── BIP AI Narrative ──────────────────────────────────────────────────────
    try:
        t14 = bip_tables.get("table14_significance")
        if t14 is not None and not t14.empty:
            yes_counts = (t14 == "YES").sum(axis=1)
            top_brand  = yes_counts.idxmax() if len(yes_counts) > 0 else None
            top_n      = int(yes_counts.max()) if len(yes_counts) > 0 else 0
            total_attrs = len(t14.columns)
            avg_yes = yes_counts.mean()
            t3 = bip_tables.get("table3_norm_pct")
            if t3 is not None and not t3.empty:
                strongest_attr = t3.abs().max(axis=0).idxmax()
                strongest_brand_for_attr = t3[strongest_attr].abs().idxmax()
                strongest_val = round(float(t3.loc[strongest_brand_for_attr, strongest_attr]), 1)
            else:
                strongest_attr = strongest_brand_for_attr = strongest_val = None

            bip_findings = [
                (f"**{top_brand}** owns the most attributes significantly ({top_n}/{total_attrs}) — "
                 f"broadest positive/negative deviation from market average." if top_brand else ""),
                (f"Across all brands, average of **{avg_yes:.1f}** attributes flagged significant per brand "
                 f"(gate: p{bip_pctile} threshold = {pct_val:.1f})."),
            ]
            if strongest_attr:
                bip_findings.append(
                    f"Most contested attribute: **{strongest_attr}** — "
                    f"{strongest_brand_for_attr} deviates {strongest_val:+.1f}% from market average (normalized deviation)."
                )
            bip_findings.append(
                "YES = brand significantly differentiates on this attribute relative to market norm. "
                "NO = within normal range. Use this table to identify brand equities and vulnerabilities."
            )
            _insight_callout([f for f in bip_findings if f], "BIP Interpretation", "📊")
    except Exception:
        pass




def _resp_filter_cte(zone="all", gender="all", age_band="all", city="all"):
    """Build a respondent-filter CTE + params from segment filters.

    Returns (cte_sql, params, is_filtered) where cte_sql defines a CTE named `_base`
    holding the filtered respondent_id set. Mirrors brand_imagery_engine resp_where pattern
    so every section filters identically. Use:
        WITH {cte} SELECT ... WHERE f.respondent_id IN (SELECT respondent_id FROM _base)
    """
    parts, params = [], []
    if zone and str(zone).lower() not in ("all", ""):
        parts.append("zone_name = ?"); params.append(zone)
    if city and str(city).lower() not in ("all", ""):
        parts.append("LOWER(city_name) = LOWER(?)"); params.append(city)
    if gender and str(gender).lower() not in ("all", ""):
        parts.append("gender = ?"); params.append(gender)
    if age_band and str(age_band).lower() not in ("all", ""):
        parts.append("age_band = ?"); params.append(age_band)
    where = ("WHERE " + " AND ".join(parts)) if parts else ""
    cte = f"_base AS (SELECT respondent_id FROM v_respondents {where})"
    return cte, params, bool(parts)


_PROJECT_1_CATEGORY_LABELS = {1: "Ceiling Fans", 2: "Air Cooler", 3: "Mixer Grinder",
                               4: "LED Batten", 5: "Water Heater", 6: "Water Pumps"}


def _is_project_1() -> bool:
    """True only for project_1's own electrical-appliance category codes — other
    projects' fact_portfolio_awareness/fact_price_paid category_id values don't map to
    these labels and must fall back to the raw numeric code instead of a wrong name."""
    return st.session_state.get("active_project_id", "project_1") == "project_1"


@st.cache_data(ttl=3600)
def _get_portfolio_data(zone="all", gender="all", age_band="all", city="all"):
    """Load portfolio awareness: % of (filtered) respondents who associate each brand with each category."""
    import sqlite3
    from oxdata.db_loader import get_db_path
    db = sqlite3.connect(str(get_db_path()))
    cte, fparams, _ = _resp_filter_cte(zone, gender, age_band, city)
    try:
        # Denominator = filtered respondent base (BQ6 is asked of all respondents)
        total = pd.read_sql_query(
            f"WITH {cte} SELECT COUNT(DISTINCT respondent_id) AS n FROM _base", db, params=fparams
        ).iloc[0, 0]
        df = pd.read_sql_query(f"""
            WITH {cte}
            SELECT fp.brand_id, db.brand_name, fp.category_id AS category_code,
                   COUNT(DISTINCT fp.respondent_id) AS n_aware
            FROM fact_portfolio_awareness fp
            JOIN dim_brand db ON db.brand_id = fp.brand_id
            WHERE fp.respondent_id IN (SELECT respondent_id FROM _base)
            GROUP BY fp.brand_id, db.brand_name, fp.category_id
        """, db, params=fparams)
    finally:
        db.close()
    if total == 0 or df.empty:
        return pd.DataFrame(columns=["brand_id","brand_name","category_code","n_aware","category","pct","n_total"])
    cat_labels = _PROJECT_1_CATEGORY_LABELS if _is_project_1() else {}
    df["category"] = df["category_code"].map(cat_labels).fillna(df["category_code"].astype(str))
    # pct = % of ALL respondents who associate the brand with the category
    df["pct"] = (df["n_aware"] / total * 100).round(1)
    df["n_total"] = int(total)
    return df


_PRICE_TIER_LABELS = {
    1: {1: "< ₹1,500", 2: "₹1,501–2,000", 3: "₹2,001–3,000", 4: "₹3,001–4,000", 5: "₹4,000+"},
    2: {1: "< ₹3,000", 2: "₹3,001–5,000", 3: "₹5,001–6,000", 4: "₹6,001–8,000",
        5: "₹8,001–10,000", 6: "₹10,001–13,000", 7: "₹13,001–15,000", 8: "₹15,000+"},
    3: {1: "< ₹2,000", 2: "₹2,001–3,000", 3: "₹3,001–5,000", 4: "₹5,001–9,000", 5: "₹9,000+"},
    4: {1: "< ₹300", 2: "₹301–500", 3: "₹501–700", 4: "₹701–900", 5: "₹901+"},
    5: {1: "< ₹3,000", 2: "₹3,001–4,000", 3: "₹4,001–6,500", 4: "₹6,501–8,000", 5: "₹8,000+"},
    6: {1: "< ₹2,000", 2: "₹2,001–4,000", 3: "₹4,001–6,000", 4: "₹6,001–8,000",
        5: "₹8,001–10,000", 6: "₹10,001–15,000", 7: "₹15,000+"},
}

@st.cache_data(ttl=3600)
def _get_price_tier_data(zone="all", gender="all", age_band="all", city="all"):
    """Load price tier distribution per category from fact_price_paid (segment-filtered)."""
    import sqlite3
    from oxdata.db_loader import get_db_path
    db = sqlite3.connect(str(get_db_path()))
    cte, fparams, _ = _resp_filter_cte(zone, gender, age_band, city)
    try:
        df = pd.read_sql_query(f"""
            WITH {cte}
            SELECT fp.category_id AS category_code, fp.price_tier,
                   COUNT(*) AS n
            FROM fact_price_paid fp
            WHERE fp.price_tier IS NOT NULL
              AND fp.respondent_id IN (SELECT respondent_id FROM _base)
            GROUP BY fp.category_id, fp.price_tier
            ORDER BY fp.category_id, fp.price_tier
        """, db, params=fparams)
    finally:
        db.close()
    cat_labels = _PROJECT_1_CATEGORY_LABELS if _is_project_1() else {}
    df["category_code"] = df["category_code"].astype(int)
    df["category"] = df["category_code"].map(cat_labels).fillna(df["category_code"].astype(str))
    # price_tier is stored as TEXT in the generic multi-project schema (vs INTEGER in project_1's
    # fixed schema) — coerce here so downstream numeric ops (max/round/comparisons) work regardless
    # of which schema this project's DB came from.
    df["price_tier"] = pd.to_numeric(df["price_tier"], errors="coerce")
    df = df.dropna(subset=["price_tier"])
    df["price_tier"] = df["price_tier"].astype(int)
    tier_labels = _PRICE_TIER_LABELS if _is_project_1() else {}
    df["price_tier_label"] = df.apply(
        lambda r: tier_labels.get(r["category_code"], {}).get(int(r["price_tier"]), f"Tier {r['price_tier']}"),
        axis=1,
    )
    return df


def _render_portfolio_awareness(sel_brand: str, zone="all", gender="all", age_band="all", city="all"):
    """Portfolio awareness — detailed analysis of category-brand association (BQ6)."""
    df = _get_portfolio_data(zone, gender, age_band, city)
    if df.empty:
        st.info(
            "**Portfolio awareness data unavailable.** "
            "This section requires a brand-category portfolio question (BQ6). "
            "Either no data was collected for this question or the current filter "
            "returns no respondents."
        )
        return

    cat_order = ["Ceiling Fans", "LED Batten", "Water Heater", "Mixer Grinder", "Water Pumps", "Air Cooler"]
    total_resp = int(df["n_total"].iloc[0]) if "n_total" in df.columns else 6631

    # ── Auto-generated insight callout ───────────────────────────────────────
    brand_df = df[df["brand_name"] == sel_brand].sort_values("pct", ascending=False)
    if not brand_df.empty:
        strongest_cat  = brand_df.iloc[0]["category"]
        strongest_pct  = brand_df.iloc[0]["pct"]
        weakest_cat    = brand_df.iloc[-1]["category"]
        weakest_pct    = brand_df.iloc[-1]["pct"]
        # Category-leader for each of brand's categories
        cat_leaders = {}
        for cat in brand_df["category"].tolist():
            cat_top = df[df["category"] == cat].sort_values("pct", ascending=False)
            if not cat_top.empty:
                cat_leaders[cat] = (cat_top.iloc[0]["brand_name"], cat_top.iloc[0]["pct"])
        brand_is_leader = [c for c, (b, _) in cat_leaders.items() if b == sel_brand]
        insights_pa = [
            f"**{sel_brand}** is most strongly associated with **{strongest_cat}** — "
            f"{strongest_pct:.0f}% of all respondents connect this brand with that category.",
            f"Weakest category link is **{weakest_cat}** ({weakest_pct:.0f}%). "
            "Low score here signals a portfolio perception gap — opportunity to communicate category presence."
        ]
        if brand_is_leader:
            insights_pa.append(f"**Category leadership:** {sel_brand} ranks #1 in {', '.join(brand_is_leader)}.")
        else:
            top_cat = brand_df.iloc[0]["category"]
            leader, leader_pct = cat_leaders.get(top_cat, ("?", 0))
            gap = leader_pct - strongest_pct
            insights_pa.append(
                f"**No category leadership:** In its strongest category ({top_cat}), "
                f"{sel_brand} trails {leader} by {gap:.0f}pp. "
                "Closing this gap requires clearer communication of category credentials."
            )
        _insight_callout(insights_pa, "Portfolio Perception Insights", "🗂️")

    pa_t1, pa_t2, pa_t3 = st.tabs([
        "🔥 Category-Leader Matrix",
        f"📊 {sel_brand} Deep Dive",
        "🏅 Category Champions"
    ])

    with pa_t1:
        st.markdown("**What % of all respondents associate each brand with each product category?**")
        st.markdown(
            "<span style='font-size:0.8rem;color:#6b7280'>"
            "BQ6 (Portfolio Awareness): Respondents were shown each brand and asked which product categories it makes. "
            f"Values = % of the {total_resp:,} respondents in the current filter who said that brand makes that category. "
            "Higher = stronger category-brand link in consumer minds."
            "</span>", unsafe_allow_html=True
        )

        pivot = df.pivot_table(index="brand_name", columns="category", values="pct", aggfunc="first").fillna(0)
        pivot = pivot.reindex(columns=[c for c in cat_order if c in pivot.columns])
        # Sort: brand with highest total portfolio breadth at top
        pivot = pivot.loc[pivot.sum(axis=1).sort_values(ascending=False).index]

        # Compute category max for relative coloring
        actual_max = pivot.values.max() if pivot.values.max() > 0 else 100

        fig = go.Figure(go.Heatmap(
            z=pivot.values,
            x=pivot.columns.tolist(),
            y=pivot.index.tolist(),
            colorscale=[[0, "#f8fafc"], [0.2, "#d1fae5"], [0.5, "#34d399"], [0.8, "#059669"], [1.0, "#065f46"]],
            text=[[f"{v:.0f}%" for v in row] for row in pivot.values],
            texttemplate="%{text}",
            textfont=dict(size=11),
            hovertemplate=(
                "<b>%{y}</b> ↔ <b>%{x}</b><br>"
                "Association: <b>%{z:.1f}%</b><br>"
                f"of {total_resp:,} respondents<extra></extra>"
            ),
            zmin=0, zmax=actual_max,
        ))
        base = {k: v for k, v in _chart_layout_base(max(350, len(pivot)*28)).items()
                if k not in ("xaxis","yaxis","legend","margin")}
        fig.update_layout(**base,
                          xaxis=dict(side="top", tickangle=-20),
                          yaxis=dict(autorange="reversed", automargin=True),
                          margin=dict(l=130, r=40, t=80, b=40),
                          title=dict(text="Portfolio Awareness Matrix — % of All Respondents (BQ6)", font=dict(size=13)))
        st.plotly_chart(_theme_fig(fig), use_container_width=True)

        # Category leadership table
        st.markdown("**Category Leaders (brand with highest % per category):**")
        leader_rows = []
        for cat in [c for c in cat_order if c in pivot.columns]:
            cat_col = pivot[cat].sort_values(ascending=False)
            if len(cat_col) > 0:
                leader_rows.append({
                    "Category": cat,
                    "Leader": cat_col.index[0],
                    "Leader %": f"{cat_col.iloc[0]:.0f}%",
                    "2nd": cat_col.index[1] if len(cat_col) > 1 else "—",
                    "2nd %": f"{cat_col.iloc[1]:.0f}%" if len(cat_col) > 1 else "—",
                    f"{sel_brand} %": f"{pivot.loc[sel_brand, cat]:.0f}%" if sel_brand in pivot.index else "—",
                })
        if leader_rows:
            leader_df = pd.DataFrame(leader_rows)
            def style_leader(row):
                styles = ["background-color: #dcfce7; font-weight:bold"
                          if row["Leader"] == sel_brand else "" for _ in row]
                return pd.Series(styles, index=row.index)
            st.dataframe(
                leader_df.style.apply(style_leader, axis=1),
                hide_index=True, use_container_width=True
            )

    with pa_t2:
        if brand_df.empty:
            st.info(f"No portfolio data for {sel_brand}.")
        else:
            brand_df_sorted = brand_df.sort_values("pct", ascending=True)

            # Get category leader for comparison
            comparison_rows = []
            for _, row in brand_df_sorted.iterrows():
                cat = row["category"]
                cat_data = df[df["category"] == cat].sort_values("pct", ascending=False)
                leader = cat_data.iloc[0]["brand_name"] if not cat_data.empty else "—"
                leader_pct = cat_data.iloc[0]["pct"] if not cat_data.empty else 0
                rank_num = cat_data.reset_index(drop=True).index[cat_data.reset_index(drop=True)["brand_name"] == sel_brand].tolist()
                comparison_rows.append({
                    "category": cat,
                    "brand_pct": row["pct"],
                    "leader_pct": leader_pct,
                    "leader": leader,
                    "gap": round(leader_pct - row["pct"], 1),
                    "rank": rank_num[0] + 1 if rank_num else "-",
                    "n_aware": int(row["n_aware"]),
                })
            cmp_df = pd.DataFrame(comparison_rows)

            # Chart: brand % vs leader %
            fig = go.Figure()
            fig.add_trace(go.Bar(
                name=sel_brand,
                x=cmp_df["category"],
                y=cmp_df["brand_pct"],
                marker_color="#1a5d4d",
                text=[f"{v:.0f}%" for v in cmp_df["brand_pct"]],
                textposition="outside",
            ))
            fig.add_trace(go.Bar(
                name="Category Leader",
                x=cmp_df["category"],
                y=cmp_df["leader_pct"],
                marker_color="#d1fae5",
                marker_line=dict(color="#059669", width=1.5),
                text=[f"{r['leader']}<br>{r['leader_pct']:.0f}%" for _, r in cmp_df.iterrows()],
                textposition="outside",
                textfont=dict(size=9),
            ))
            base = {k: v for k, v in _chart_layout_base(380).items() if k not in ("xaxis","yaxis","legend")}
            fig.update_layout(**base,
                              barmode="group",
                              yaxis_title="% of All Respondents",
                              legend=dict(orientation="h", y=-0.15, xanchor="center", x=0.5),
                              title=dict(text=f"{sel_brand} — Portfolio Association vs. Category Leader", font=dict(size=13)))
            st.plotly_chart(_theme_fig(fig), use_container_width=True)

            # Detailed table
            show_cmp = cmp_df[["category","brand_pct","leader","leader_pct","gap","rank","n_aware"]].copy()
            show_cmp.columns = ["Category", f"{sel_brand} %", "Category Leader", "Leader %", "Gap (pp)", "Rank", "n (assoc.)"]

            def color_gap(val):
                if not isinstance(val, (int, float)): return ""
                if val <= 5: return "color:#15803d; font-weight:bold"
                if val <= 20: return "color:#d97706"
                return "color:#dc2626; font-weight:bold"

            styled = show_cmp.style.format({
                f"{sel_brand} %": "{:.1f}", "Leader %": "{:.1f}",
                "Gap (pp)": "{:.1f}", "n (assoc.)": "{:,}"
            }).map(color_gap, subset=["Gap (pp)"])
            st.dataframe(styled, hide_index=True, use_container_width=True)
            st.caption(
                f"Base = {total_resp:,} total respondents. "
                f"% = share of all respondents who associate {sel_brand} with that category. "
                "Gap = leader's % minus brand's %. Green = within 5pp of leader (competitive), red = major gap."
            )

    with pa_t3:
        st.markdown("**For each product category — which brands own it in the consumer mind?**")
        sel_cat_pa = st.selectbox("Select category:", [c for c in cat_order if c in df["category"].unique()],
                                  key="pa_cat_sel")
        cat_df = df[df["category"] == sel_cat_pa].sort_values("pct", ascending=False).head(15)
        colors_pa = ["#1a5d4d" if b == sel_brand else "#30a76a" for b in cat_df["brand_name"]]

        fig3 = go.Figure(go.Bar(
            x=cat_df["brand_name"],
            y=cat_df["pct"],
            marker_color=colors_pa,
            text=[f"{v:.0f}%" for v in cat_df["pct"]],
            textposition="outside",
            hovertemplate="<b>%{x}</b><br>%{y:.1f}% of respondents associate with " + sel_cat_pa + "<extra></extra>",
        ))
        base = {k: v for k, v in _chart_layout_base(360).items() if k not in ("xaxis","yaxis","legend")}
        fig3.update_layout(**base,
                           yaxis_title="% of All Respondents",
                           xaxis=dict(tickangle=-30),
                           title=dict(text=f"Category Champion — {sel_cat_pa}", font=dict(size=13)))
        st.plotly_chart(_theme_fig(fig3), use_container_width=True)
        st.caption(f"% of {total_resp:,} respondents who associate each brand with {sel_cat_pa}. Dark green = {sel_brand}.")


def _render_price_tier_distribution(sel_cat: str, zone="all", gender="all", age_band="all", city="all"):
    """Price tier distribution across categories."""
    df = _get_price_tier_data(zone, gender, age_band, city)
    if df.empty:
        st.info("Price tier data not available for the current filter selection.")
        return

    _CAT_ORDER = ["Ceiling Fans", "LED Batten", "Mixer Grinder", "Water Heater", "Water Pumps", "Air Cooler"]
    all_cats = [c for c in _CAT_ORDER if c in df["category"].values]
    if not all_cats:
        all_cats = sorted(df["category"].dropna().unique().tolist())

    pt_t1, pt_t2 = st.tabs(["Price Segment Overview", "Category Deep Dive"])

    with pt_t1:
        # Budget / Mid / Premium split per category (comparable across categories)
        summary_rows = []
        colors_bmp = {"Budget": "#15803d", "Mid-range": "#ca8a04", "Premium": "#1d4ed8"}
        for cat in all_cats:
            sub = df[df["category"] == cat].sort_values("price_tier")
            if sub.empty:
                continue
            total = sub["n"].sum()
            tiers = sub["price_tier"].tolist()
            max_t = max(tiers)
            # Budget = bottom third, premium = top third, mid = rest
            budget_cut = max(1, round(max_t / 3))
            premium_cut = max_t - round(max_t / 3) + 1
            budget_n  = sub[sub["price_tier"] <= budget_cut]["n"].sum()
            premium_n = sub[sub["price_tier"] >= premium_cut]["n"].sum()
            mid_n     = total - budget_n - premium_n
            summary_rows.append({
                "category": cat,
                "Budget":   round(budget_n / total * 100, 1),
                "Mid-range":round(mid_n    / total * 100, 1),
                "Premium":  round(premium_n / total * 100, 1),
                "n_total":  total,
            })
        smry = pd.DataFrame(summary_rows)

        fig_smry = go.Figure()
        for seg, clr in colors_bmp.items():
            fig_smry.add_trace(go.Bar(
                name=seg, x=smry["category"], y=smry[seg],
                marker_color=clr,
                text=[f"{v:.0f}%" for v in smry[seg]],
                textposition="inside",
                hovertemplate=f"<b>{seg}</b><br>%{{x}}<br>%{{y:.1f}}% (n=%{{customdata:,}})<extra></extra>",
                customdata=(smry[seg] / 100 * smry["n_total"]).round(0).astype(int),
            ))
        layout_smry = {k: v for k, v in _chart_layout_base(420).items() if k not in ("xaxis","yaxis","legend")}
        fig_smry.update_layout(
            **layout_smry,
            barmode="stack",
            yaxis=dict(title="% Buyers", range=[0, 105]),
            legend=dict(orientation="h", yanchor="bottom", y=-0.2, x=0.5, xanchor="center"),
            title=dict(text="Price Segment Mix by Category (BQ0b — Actual Price Paid)", font=dict(size=13)),
        )
        st.plotly_chart(_theme_fig(fig_smry), use_container_width=True)
        st.caption(
            "Budget = lower tier(s) · Mid-range = middle tier(s) · Premium = upper tier(s). "
            "Tier boundaries differ per category. Base = recent buyers who reported a price."
        )
        st.dataframe(
            smry.rename(columns={"category": "Category", "n_total": "Base (n)"})
                .style.format({"Budget": "{:.1f}%", "Mid-range": "{:.1f}%", "Premium": "{:.1f}%", "Base (n)": "{:,}"})
                .background_gradient(subset=["Premium"], cmap="Blues"),
            hide_index=True, use_container_width=True,
        )

    with pt_t2:
        focus_cat = sel_cat if sel_cat != "All" else all_cats[0]
        # Allow switching category within tab 2
        cat_sel2 = st.selectbox("Category", all_cats,
                                index=all_cats.index(focus_cat) if focus_cat in all_cats else 0,
                                key="pt_cat_detail_sel")
        sub = df[df["category"] == cat_sel2].sort_values("price_tier")
        if sub.empty:
            st.info(f"No price data for {cat_sel2}.")
            return
        total = sub["n"].sum()
        sub["pct"] = (sub["n"] / total * 100).round(1)

        # Horizontal bar chart with actual price labels
        top_tier = sub.loc[sub["n"].idxmax(), "price_tier_label"]
        bar_colors_pt = [
            "#1a5d4d" if lbl == top_tier else "#30a76a"
            for lbl in sub["price_tier_label"]
        ]
        fig_dt = go.Figure(go.Bar(
            x=sub["pct"],
            y=sub["price_tier_label"],
            orientation="h",
            marker_color=bar_colors_pt,
            text=[f"{v:.1f}%" for v in sub["pct"]],
            textposition="outside",
            hovertemplate="%{y}<br>%{x:.1f}% of buyers (n=%{customdata:,})<extra></extra>",
            customdata=sub["n"],
        ))
        # Annotate top tier
        top_pct = sub.loc[sub["n"].idxmax(), "pct"]
        fig_dt.add_annotation(
            x=top_pct, y=top_tier, text="★ Most common",
            xanchor="left", xshift=65, showarrow=False,
            font=dict(size=9, color="#1a5d4d"),
            bgcolor="rgba(26,93,77,0.08)",
        )
        layout_dt = {k: v for k, v in _chart_layout_base(max(300, len(sub) * 42)).items()
                     if k not in ("xaxis","yaxis","legend","margin")}
        fig_dt.update_layout(
            **layout_dt,
            title=dict(text=f"{cat_sel2} — Price Paid Distribution (n={total:,})", font=dict(size=13)),
            xaxis=dict(title="% Buyers", range=[0, sub["pct"].max() * 1.3]),
            yaxis=dict(autorange="reversed", automargin=True),
            margin=dict(l=160, r=100, t=60, b=40),
        )
        st.plotly_chart(_theme_fig(fig_dt), use_container_width=True)
        st.caption(f"Base = {total:,} {cat_sel2} recent buyers. Actual price paid (BQ0b). Dark green = most common tier.")


@st.cache_data(ttl=3600)
def _get_purchase_journey_data(zone="all", gender="all", age_band="all", city="all"):
    """Load aggregated purchase journey data from fact_purchase_journey (segment-filtered)."""
    import sqlite3
    from oxdata.db_loader import get_db_path

    # Codebook labels for each pq_var
    PQ1_LABELS = {
        1: "Needed replacement",
        2: "New purchase / first time",
        3: "Gift",
        4: "Home renovation",
        5: "Better features / upgrade",
        6: "Price deal / offer",
        7: "Other",
    }
    PQ2_LABELS = {
        1: "Online (e-commerce)",
        2: "Word of mouth / family & friends",
        3: "TV / newspaper ads",
        4: "In-store display",
        5: "Social media",
        6: "Brand website",
        7: "Other",
    }
    PQ3_LABELS = {
        1: "Offline retail (general trade)",
        2: "Offline retail (modern trade)",
        3: "Online marketplace",
        4: "Brand website / direct",
        5: "Company outlet / exclusive store",
        6: "Exchange/buyback scheme",
        7: "Other",
    }
    PQ4_LABELS = {
        1: "Self",
        2: "Spouse / partner",
        3: "Family member",
        4: "Other",
    }
    PQ5_LABELS = {
        1: "Myself",
        2: "Spouse",
        3: "Friends/Other Family",
        4: "Kids",
        5: "Electrician/Plumber",
        6: "Other",
    }

    try:
        db_path = get_db_path()
        conn = sqlite3.connect(str(db_path))
        cte, fparams, _ = _resp_filter_cte(zone, gender, age_band, city)
        try:
            # Detect schema: project_1 uses pq_var/value_code; generic loader uses question_code/answer
            col_names = {r[1] for r in conn.execute("PRAGMA table_info(fact_purchase_journey)").fetchall()}
            if "pq_var" in col_names and "value_code" in col_names:
                # project_1 / structured schema
                df = pd.read_sql_query(
                    f"WITH {cte} "
                    "SELECT pq_var, value_code, COUNT(*) as n "
                    "FROM fact_purchase_journey "
                    "WHERE pq_var IN ('pq1','pq2','pq3','pq4','pq5') AND value_code IS NOT NULL "
                    "  AND respondent_id IN (SELECT respondent_id FROM _base) "
                    "GROUP BY pq_var, value_code "
                    "ORDER BY pq_var, n DESC",
                    conn, params=fparams,
                )
            elif "question_code" in col_names and "answer" in col_names:
                # Generic-loader schema — remap to pq_var/value_code columns for uniform downstream
                # Only include rows whose question_text hints at a purchase journey topic
                df_raw = pd.read_sql_query(
                    f"WITH {cte} "
                    "SELECT question_code, question_text, answer, COUNT(*) as n "
                    "FROM fact_purchase_journey "
                    "WHERE answer IS NOT NULL "
                    "  AND respondent_id IN (SELECT respondent_id FROM _base) "
                    "GROUP BY question_code, answer "
                    "ORDER BY question_code, n DESC",
                    conn, params=fparams,
                )
                # Filter to journey-like questions by keyword in question_text
                _journey_kw = ["why", "reason", "bought", "purchase", "channel", "research",
                                "influenc", "decided", "who", "where", "shop", "online", "offline"]
                if not df_raw.empty:
                    mask = df_raw["question_text"].str.lower().apply(
                        lambda t: any(kw in str(t) for kw in _journey_kw)
                    )
                    df_raw = df_raw[mask]
                # Rename to match structured schema
                df = df_raw.rename(columns={"question_code": "pq_var", "answer": "value_code"}) if not df_raw.empty else df_raw
            else:
                df = pd.DataFrame()
        finally:
            conn.close()
        return df, PQ1_LABELS, PQ2_LABELS, PQ3_LABELS, PQ4_LABELS, PQ5_LABELS
    except Exception:
        return None, {}, {}, {}, {}, {}


def _render_purchase_journey(zone="all", gender="all", age_band="all", city="all"):
    """Render purchase journey charts (pq1 = why bought, pq2 = where researched,
    pq3 = channel, pq4 = who decided)."""
    result = _get_purchase_journey_data(zone, gender, age_band, city)
    df_all, PQ1, PQ2, PQ3, PQ4, PQ5 = result

    if df_all is None or df_all.empty:
        st.info(
            "**Purchase Journey data unavailable for this project.** "
            "This section requires pq1–pq5 structured purchase-journey variables "
            "in the survey. The current dataset uses a different schema or did not "
            "collect these variables in this format."
        )
        return

    LABEL_MAPS = {"pq1": PQ1, "pq2": PQ2, "pq3": PQ3, "pq4": PQ4, "pq5": PQ5}
    PQ_TITLES = {
        "pq1": "Why did you buy? (Purchase motivation)",
        "pq2": "Where did you research? (Research channels)",
        "pq3": "Where did you buy? (Purchase channel)",
        "pq4": "Who decided? (Decision maker)",
        "pq5": "Who went shopping? (Purchase party)",
    }
    PQ_TABS = ["Why Bought", "Research Channels", "Purchase Channel", "Decision Maker", "Purchase Party"]

    pj_tabs = st.tabs(PQ_TABS)

    colors = _chart_colors()
    font_fam = CHART_THEME["font_family"]

    for idx, pq_key in enumerate(["pq1", "pq2", "pq3", "pq4", "pq5"]):
        with pj_tabs[idx]:
            sub = df_all[df_all["pq_var"] == pq_key].copy()
            if sub.empty:
                st.info(f"No data for {pq_key}.")
                continue

            label_map = LABEL_MAPS[pq_key]
            sub["label"] = sub["value_code"].map(lambda c: label_map.get(int(c), f"Code {c}"))
            sub["pct"] = (sub["n"] / sub["n"].sum() * 100).round(1)
            sub = sub.sort_values("pct", ascending=True)  # horizontal bar — ascending for bottom-up

            top_item = sub.iloc[-1]  # highest pct after sort ascending
            base_color = colors[idx % len(colors)]
            bar_colors_pj = [
                "#1a5d4d" if i == len(sub) - 1 else base_color
                for i in range(len(sub))
            ]

            fig_pj = go.Figure(go.Bar(
                x=sub["pct"],
                y=sub["label"],
                orientation="h",
                marker_color=bar_colors_pj,
                text=[f"{v:.1f}%" for v in sub["pct"]],
                textposition="outside",
                hovertemplate="%{y}<br>%{x:.1f}% (%{customdata:,} respondents)<extra></extra>",
                customdata=sub["n"],
            ))

            # Top item annotation
            fig_pj.add_annotation(
                x=top_item["pct"], y=top_item["label"],
                text="★ Top",
                xanchor="left", xshift=60,
                showarrow=False,
                font=dict(size=9, color="#1a5d4d", family="Inter"),
                bgcolor="rgba(26,93,77,0.08)",
            )

            layout_base = {k: v for k, v in _chart_layout_base(max(320, len(sub) * 38)).items()
                           if k not in ("xaxis", "yaxis", "legend", "margin")}
            fig_pj.update_layout(
                **layout_base,
                title=dict(text=PQ_TITLES[pq_key], font=dict(size=13)),
                xaxis=dict(title="% respondents", range=[0, sub["pct"].max() * 1.25],
                           gridcolor="#f1f5f9", zeroline=False),
                yaxis=dict(title="", automargin=True),
                margin=dict(l=160, r=80, t=50, b=40),
            )
            st.plotly_chart(_theme_fig(fig_pj), use_container_width=True)
            st.caption(
                f"★ Top pick: **{top_item['label']}** ({top_item['pct']:.1f}%).  "
                f"Base n = {sub['n'].sum():,} respondents (recent buyers)."
            )


@st.cache_data(ttl=3600)
def _get_brand_csat(brand_name: str):
    """Return brand-specific CSAT avg + n."""
    _layer = _get_layer()
    if _layer is not None:
        csat_df = _layer.csat
        if not csat_df.empty:
            sub = csat_df[csat_df["brand_name"] == brand_name]["csat_score"].dropna()
            if len(sub) >= 1:
                raw = sub.mean()
                displayed = round(raw / 2 if raw > 5.5 else raw, 2)
                return {"avg_csat": displayed, "n": len(sub)}
        return None

    import sqlite3
    from oxdata.db_loader import get_db_path
    try:
        conn = sqlite3.connect(str(get_db_path()))
        try:
            row = conn.execute(
                "SELECT ROUND(AVG(s.score),2), COUNT(*) "
                "FROM fact_satisfaction s "
                "JOIN fact_brand_awareness ba ON s.respondent_id=ba.respondent_id AND ba.stage='LAST_PURCHASED' "
                "JOIN dim_brand b ON ba.brand_id=b.brand_id "
                "WHERE b.brand_name=?",
                (brand_name,)
            ).fetchone()
        finally:
            conn.close()
        if row and row[0] is not None:
            return {"avg_csat": round(float(row[0]), 2), "n": int(row[1])}
    except Exception:
        pass
    return None


@st.cache_data(ttl=3600)
def _get_brand_csat_distribution(brand_name: str):
    """CSAT score distribution for a specific brand's last purchasers."""
    import sqlite3
    from oxdata.db_loader import get_db_path
    try:
        conn = sqlite3.connect(str(get_db_path()))
        df = pd.read_sql_query(
            "SELECT s.score, COUNT(*) as n "
            "FROM fact_satisfaction s "
            "JOIN fact_brand_awareness ba ON s.respondent_id=ba.respondent_id AND ba.stage='LAST_PURCHASED' "
            "JOIN dim_brand b ON ba.brand_id=b.brand_id "
            "WHERE b.brand_name=? "
            "GROUP BY s.score ORDER BY s.score",
            conn, params=[brand_name],
        )
        conn.close()
        return df
    except Exception:
        return pd.DataFrame()


@st.cache_data(ttl=3600)
def _get_csat_overall():
    """Return overall CSAT avg + n from fact_satisfaction (market benchmark)."""
    import sqlite3
    from oxdata.db_loader import get_db_path
    try:
        db_path = get_db_path()
        conn = sqlite3.connect(str(db_path))
        try:
            row = conn.execute(
                "SELECT AVG(score) as avg_csat, COUNT(*) as n FROM fact_satisfaction"
            ).fetchone()
        finally:
            conn.close()
        if row and row[0] is not None:
            return {"avg_csat": round(float(row[0]), 2), "n": int(row[1])}
    except Exception:
        pass
    return None


@st.cache_data(ttl=3600)
def _get_csat_distribution():
    """CSAT score distribution (score × n) from fact_satisfaction."""
    import sqlite3
    from oxdata.db_loader import get_db_path
    try:
        conn = sqlite3.connect(str(get_db_path()))
        df = pd.read_sql_query(
            "SELECT score, COUNT(*) as n FROM fact_satisfaction GROUP BY score ORDER BY score",
            conn,
        )
        conn.close()
        return df
    except Exception:
        return pd.DataFrame()


@st.cache_data(ttl=3600, show_spinner="Loading IPA data...")
def _get_ipa_data(brand_name: str):
    """Fetch key-driver (importance × association) rows for a given brand.

    Was `SELECT ... FROM v_key_drivers WHERE brand_name = ?` — that view only exists in
    project_1's own database (a manually-created SQL view, never part of generic_loader.py's
    schema), so this crashed with "no such table: v_key_drivers" for EVERY generically-ingested
    project (found live on Akshayakalpa). The view's own definition only ever joins
    fact_need_importance/fact_brand_imagery/dim_bq3_attribute/dim_brand — all four of which DO
    exist in the generic schema — so inlining the same query works for any project's DB
    regardless of whether project_1's convenience view happens to exist in it.
    """
    import sqlite3
    from oxdata.db_loader import get_db_path
    db_path = get_db_path()
    conn = sqlite3.connect(str(db_path))
    try:
        # project_1's fixed schema names this column `score`; the generic multi-project schema
        # (generic_loader.py) names it `importance_score` — detect which one this DB actually has
        # rather than hardcoding either, so this works regardless of which pipeline wrote the DB.
        _cols = {r[1] for r in conn.execute("PRAGMA table_info(fact_need_importance)")}
        _score_col = "importance_score" if "importance_score" in _cols else "score"
        df = pd.read_sql_query(
            f"""
            WITH imp AS (
                SELECT attr_id, AVG({_score_col}) AS mean_importance,
                       COUNT(DISTINCT respondent_id) AS n_imp
                FROM fact_need_importance
                GROUP BY attr_id
            ),
            assoc AS (
                SELECT attr_id, brand_id,
                       COUNT(DISTINCT respondent_id) AS n_assoc,
                       (SELECT COUNT(DISTINCT respondent_id) FROM fact_brand_imagery b2
                        WHERE b2.brand_id = bi.brand_id) AS n_brand_total
                FROM fact_brand_imagery bi
                GROUP BY attr_id, brand_id
            )
            SELECT
                i.attr_id, a.attr_label, a.broad_feature, assoc.brand_id, db.brand_name,
                i.mean_importance, assoc.n_assoc * 100.0 / assoc.n_brand_total AS pct_association,
                i.n_imp, assoc.n_assoc
            FROM imp i
            JOIN dim_bq3_attribute a ON a.attr_id = i.attr_id
            JOIN assoc              ON assoc.attr_id = i.attr_id
            JOIN dim_brand db       ON db.brand_id  = assoc.brand_id
            WHERE db.brand_name = ?
            """,
            conn,
            params=(brand_name,),
        )
    finally:
        conn.close()
    return df


def _render_ipa_grid(sel_brand: str):
    """Importance-Performance Analysis (IPA) — readable scatter + quadrant analysis."""
    if not sel_brand:
        st.info("Select a brand to view the IPA grid.")
        return

    df = _get_ipa_data(sel_brand)
    if df.empty:
        st.info(
            f"**Importance-Performance Grid unavailable for this project.** "
            f"This chart requires attribute importance ratings (BQ3a) — "
            f"the current dataset does not include that battery. "
            f"Only brand imagery associations (BQ3b) were collected."
        )
        return

    # ── Filters ───────────────────────────────────────────────────────────────
    all_features = sorted(df["broad_feature"].dropna().unique().tolist())
    ipa_col1, ipa_col2 = st.columns([2, 3])
    with ipa_col1:
        ipa_features = st.multiselect(
            "Filter by attribute group",
            options=all_features,
            default=all_features,
            key="ipa_feature_filter",
            help="Show only selected attribute groups on the grid",
        )
    with ipa_col2:
        show_labels = st.toggle("Show attribute labels on chart", value=False, key="ipa_show_labels",
                                help="Turn on to see text labels — may overlap with many attributes")

    df_plot = df[df["broad_feature"].isin(ipa_features)].copy() if ipa_features else df.copy()
    if df_plot.empty:
        st.info("No attributes match current filter.")
        return

    # ── Quadrant assignment ───────────────────────────────────────────────────
    x_mid = df_plot["mean_importance"].median()
    y_mid = df_plot["pct_association"].median()

    def assign_quadrant(row):
        hi_imp = row["mean_importance"] >= x_mid
        hi_ass = row["pct_association"] >= y_mid
        if hi_imp and hi_ass:     return "🟢 Strengths (Maintain)"
        if not hi_imp and hi_ass: return "🔴 Opportunities (Improve)"
        if hi_imp and not hi_ass: return "🟡 Monitor"
        return "⬜ Low Priority"

    df_plot["quadrant"] = df_plot.apply(assign_quadrant, axis=1)

    # Quadrant insight callout
    strengths = df_plot[df_plot["quadrant"] == "🟢 Strengths (Maintain)"]["attr_label"].tolist()
    opportunities = df_plot[df_plot["quadrant"] == "🔴 Opportunities (Improve)"]["attr_label"].tolist()
    ipa_insights = []
    if strengths:
        ipa_insights.append(
            f"**{len(strengths)} strength attributes** consumers rate as both important AND strongly associate with {sel_brand}: "
            + ", ".join(f"*{a}*" for a in strengths[:4]) + ("..." if len(strengths) > 4 else "") + ". Defend these."
        )
    if opportunities:
        ipa_insights.append(
            f"**{len(opportunities)} opportunity attributes** are rated highly important by consumers but {sel_brand} has low association: "
            + ", ".join(f"*{a}*" for a in opportunities[:4]) + ("..." if len(opportunities) > 4 else "") + ". Priority investment areas."
        )
    _insight_callout(ipa_insights, f"IPA Insights — {sel_brand}", "📊")

    ipa_t1, ipa_t2, ipa_t3 = st.tabs(["📍 IPA Scatter", "📋 Quadrant Summary", "📊 Ranked Tables"])

    with ipa_t1:
        colors = _chart_colors()
        feature_list = sorted(df_plot["broad_feature"].dropna().unique().tolist())
        color_map = {feat: colors[i % len(colors)] for i, feat in enumerate(feature_list)}

        fig_ipa = go.Figure()

        x_min = max(0, df_plot["mean_importance"].min() - 0.2)
        x_max = min(7.2, df_plot["mean_importance"].max() + 0.2)
        y_max_val = min(100, df_plot["pct_association"].max() * 1.15 + 5)
        y_min_val = max(0, df_plot["pct_association"].min() - 3)

        # Quadrant shading
        quad_cfg = [
            (x_mid, x_max, y_mid, y_max_val, "rgba(34,197,94,0.08)",  "✦ STRENGTHS<br>Maintain",       x_max-0.02, y_max_val-1),
            (x_min, x_mid, y_mid, y_max_val, "rgba(239,68,68,0.08)",   "✦ OPPORTUNITIES<br>Improve",    x_min+0.02, y_max_val-1),
            (x_mid, x_max, y_min_val, y_mid, "rgba(251,191,36,0.08)",  "✦ MONITOR",                     x_max-0.02, y_min_val+1),
            (x_min, x_mid, y_min_val, y_mid, "rgba(156,163,175,0.05)", "✦ LOW PRIORITY",                x_min+0.02, y_min_val+1),
        ]
        for x0, x1, y0, y1, fill, label, lx, ly in quad_cfg:
            fig_ipa.add_shape(type="rect", x0=x0, x1=x1, y0=y0, y1=y1,
                              fillcolor=fill, line_width=0, layer="below")
            fig_ipa.add_annotation(
                x=lx, y=ly, text=f"<b>{label}</b>",
                xanchor="right" if "right" in label or lx > x_mid else "left",
                yanchor="top" if ly > y_mid else "bottom",
                showarrow=False, font=dict(size=9, color="#94a3b8"),
                bgcolor="rgba(255,255,255,0.7)"
            )

        fig_ipa.add_shape(type="line", x0=x_mid, x1=x_mid, y0=y_min_val, y1=y_max_val,
                          line=dict(color="#94a3b8", width=1.5, dash="dot"))
        fig_ipa.add_shape(type="line", x0=x_min, x1=x_max, y0=y_mid, y1=y_mid,
                          line=dict(color="#94a3b8", width=1.5, dash="dot"))

        for feat in feature_list:
            sub = df_plot[df_plot["broad_feature"] == feat]
            # Marker size = scaled importance (bigger = more important)
            sizes = (sub["mean_importance"] - sub["mean_importance"].min() + 1) * 3 + 9
            mode = "markers+text" if show_labels else "markers"
            fig_ipa.add_trace(go.Scatter(
                x=sub["mean_importance"],
                y=sub["pct_association"],
                mode=mode,
                name=feat,
                marker=dict(
                    size=sizes.clip(10, 22),
                    color=color_map[feat],
                    line=dict(width=1.5, color="white"),
                    opacity=0.85,
                ),
                text=sub["attr_label"] if show_labels else None,
                textposition="top center",
                textfont=dict(size=8),
                customdata=sub[["attr_label", "mean_importance", "pct_association", "broad_feature"]].values,
                hovertemplate=(
                    "<b>%{customdata[0]}</b><br>"
                    "Group: %{customdata[3]}<br>"
                    "Importance score: <b>%{customdata[1]:.2f}</b>/7<br>"
                    "Brand association: <b>%{customdata[2]:.1f}%</b> of respondents<br>"
                    "<i>Hover to explore — toggle labels above for names</i>"
                    "<extra></extra>"
                ),
            ))

        layout_base = {k: v for k, v in _chart_layout_base(580).items()
                       if k not in ("xaxis", "yaxis", "legend")}
        fig_ipa.update_layout(
            **layout_base,
            xaxis=dict(title="← Less Important | Mean Importance Score (1–7) | More Important →",
                       range=[x_min, x_max], gridcolor="#f1f5f9", zeroline=False, tickformat=".1f"),
            yaxis=dict(title="← Lower Association | Brand Association % | Higher Association →",
                       range=[y_min_val, y_max_val], gridcolor="#f1f5f9", zeroline=False),
            legend=dict(orientation="h", yanchor="bottom", y=-0.22, xanchor="center", x=0.5,
                        font=dict(size=10), title=dict(text="Attribute Group:")),
            showlegend=True,
            title=dict(text=f"IPA Grid — {sel_brand}  (hover dots for attribute names)", font=dict(size=13)),
        )
        st.plotly_chart(_theme_fig(fig_ipa), use_container_width=True)
        st.caption(
            f"Dividing lines at median importance ({x_mid:.2f}/7) and median association ({y_mid:.1f}%). "
            f"Dot size = relative importance score. Toggle labels above to see attribute names."
        )

    with ipa_t2:
        quadrant_order = ["🟢 Strengths (Maintain)", "🔴 Opportunities (Improve)", "🟡 Monitor", "⬜ Low Priority"]
        quad_colors   = {"🟢 Strengths (Maintain)": "#dcfce7", "🔴 Opportunities (Improve)": "#fee2e2",
                         "🟡 Monitor": "#fef9c3", "⬜ Low Priority": "#f1f5f9"}
        quad_desc = {
            "🟢 Strengths (Maintain)":     "High importance + high association. Core competitive moats. Protect & invest to maintain.",
            "🔴 Opportunities (Improve)":  "High importance + low association. Consumers care but brand is weak here. Fix these first.",
            "🟡 Monitor":                  "Low importance + high association. Brand over-delivers but consumers don't value it much.",
            "⬜ Low Priority":              "Low importance + low association. Neither a gap nor a strength. Deprioritise.",
        }
        for q in quadrant_order:
            q_df = df_plot[df_plot["quadrant"] == q].sort_values("mean_importance", ascending=False)
            if q_df.empty:
                continue
            with st.expander(f"{q} — {len(q_df)} attribute(s)", expanded=(q in ["🟢 Strengths (Maintain)", "🔴 Opportunities (Improve)"])):
                st.markdown(f"<div style='font-size:0.82rem;color:#6b7280;margin-bottom:8px'>{quad_desc[q]}</div>", unsafe_allow_html=True)
                show_q = q_df[["attr_label", "broad_feature", "mean_importance", "pct_association"]].copy()
                show_q.columns = ["Attribute", "Group", "Importance (avg)", "Association %"]
                show_q = show_q.reset_index(drop=True)
                styled_q = (
                    show_q.style
                    .set_properties(**{"background-color": quad_colors[q]})
                    .format({"Importance (avg)": "{:.2f}", "Association %": "{:.1f}"})
                    .background_gradient(subset=["Importance (avg)"], cmap="Blues", axis=0)
                    .background_gradient(subset=["Association %"], cmap="Greens", axis=0)
                )
                st.dataframe(styled_q, hide_index=True, use_container_width=True)

    with ipa_t3:
        rt1, rt2 = st.columns(2)
        with rt1:
            st.markdown(f"**Top 10 by Importance — {sel_brand}**")
            top_imp = df_plot.sort_values("mean_importance", ascending=False).head(10)[
                ["attr_label", "broad_feature", "mean_importance", "pct_association", "quadrant"]
            ]
            top_imp.columns = ["Attribute", "Group", "Importance", "Assoc %", "Quadrant"]
            st.dataframe(top_imp.style.format({"Importance": "{:.2f}", "Assoc %": "{:.1f}"}),
                         hide_index=True, use_container_width=True)
        with rt2:
            st.markdown(f"**Top 10 by Brand Association — {sel_brand}**")
            top_ass = df_plot.sort_values("pct_association", ascending=False).head(10)[
                ["attr_label", "broad_feature", "mean_importance", "pct_association", "quadrant"]
            ]
            top_ass.columns = ["Attribute", "Group", "Importance", "Assoc %", "Quadrant"]
            st.dataframe(top_ass.style.format({"Importance": "{:.2f}", "Assoc %": "{:.1f}"}),
                         hide_index=True, use_container_width=True)


def _render_what_if_simulator(brand_name, current_assoc, nps_impacts):
    """Interactive simulator to project NPS impact based on driver association changes."""
    st.markdown(f"### 🎯 What-If Impact Simulator: {brand_name}")
    st.markdown("Adjust the association levels (Market Presence %) for each driver to see the projected impact on NPS.")

    if current_assoc.empty or nps_impacts.empty:
        st.warning("Insufficient data for simulation.")
        return

    # Align drivers
    drivers = nps_impacts.index.intersection(current_assoc.index)
    if len(drivers) == 0:
        st.warning("No overlapping drivers found for simulation.")
        return

    col1, col2 = st.columns([1, 1])
    
    sim_changes = {}
    with col1:
        st.markdown("**Driver Association Adjustments**")
        for d in drivers:
            curr_val = float(current_assoc.loc[d])
            impact = float(nps_impacts.loc[d])
            
            # Label with impact direction
            label = f"{d} (Impact: {impact:+.3f})"
            new_val = st.slider(label, 0.0, 100.0, curr_val, step=1.0, key=f"sim_{brand_name}_{d}")
            sim_changes[d] = (new_val - curr_val) / 100.0 # Convert % to raw association delta

    with col2:
        st.markdown("**Projected Impact**")
        total_delta = sum(sim_changes[d] * nps_impacts.loc[d] for d in drivers)
        
        # Determine color
        color = "#16a34a" if total_delta >= 0 else "#dc2626"
        
        st.markdown(f"""
            <div style="background-color: {color}22; padding: 20px; border-radius: 10px; border: 2px solid {color}; text-align: center;">
                <h2 style="margin:0; color: {color};">Projected NPS Delta</h2>
                <div style="font-size: 3rem; font-weight: bold; color: {color};">
                    {total_delta:+.2f}
                </div>
                <p style="margin-top: 10px; font-size: 0.9rem; color: #4b5563;">
                    Based on Bayesian Ridge regression of attribute associations against respondent NPS scores.
                </p>
            </div>
        """, unsafe_allow_html=True)
        
        # Show top contributors to change
        st.markdown("<br>**Top Contributors to Change**", unsafe_allow_html=True)
        contributions = pd.Series({d: sim_changes[d] * nps_impacts.loc[d] for d in drivers})
        contributions = contributions[contributions != 0].sort_values(ascending=False)
        if not contributions.empty:
            for d, val in contributions.items():
                st.write(f"{'🟢' if val > 0 else '🔴'} **{d}**: {val:+.2f}")
        else:
            st.info("Adjust the sliders to see how NPS might move.")

def _render_section_16_driver_analysis(sel_cat, sel_brand=None, zone_arg="all", gender_arg="all", age_band_arg="all", city_arg="all", project_id="project_1"):
    _section_header(
        "🎯 Brand Driver Analysis",
        "Deep-dive into which attributes drive brand perception and how brands are positioned relative to those drivers.",
    )

    # Action Bar
    col1, col2, col3, col4 = st.columns([2, 1, 1, 1])
    with col1:
        attr_groups = _get_attr_groups()
        all_attrs = []
        for g, attrs in attr_groups.items():
            for label, aid in attrs.items():
                all_attrs.append(f"{g}: {label} ({aid})")
        
        # Default drivers if none selected
        default_selection = [f"Advanced Features: Good build quality & material (78)", 
                             f"Advanced Features: Easy to use (79)",
                             f"Advanced Features: Innovative features (80)",
                             f"Brand and Price: Value for money (85)",
                             f"Brand and Price: Affordable (87)"]
        # Filter default selection to what actually exists in all_attrs
        default_selection = [d for d in default_selection if d in all_attrs]

        selected_labels = st.multiselect(
            "Select Drivers (Attributes)",
            options=all_attrs,
            default=default_selection,
            key="da_drivers_sel",
            help="Choose attributes to analyze as brand drivers."
        )
        
        selected_drivers = []
        import re
        for label in selected_labels:
            match = re.search(r"\((\d+)\)$", label)
            if match:
                selected_drivers.append(int(match.group(1)))

    with col2:
        cats_list = ["All", "Ceiling Fans", "Air Cooler", "Mixer Grinder", "LED Batten", "Water Heater", "Water Pumps"]
        da_product = st.selectbox("Product Group", cats_list, key="da_product_sel", index=cats_list.index(sel_cat) if sel_cat in cats_list else 0)
    with col3:
        compare_by = st.selectbox("Compare by", ["Overall", "By Category", "By Zone"], key="da_compare_sel")
    with col4:
        da_pctile = st.selectbox("Significance", [50, 60, 65, 70, 75, 80, 85, 90], index=2, key="da_pctile_sel")

    _da_pooled = True  # always pooled — all brands stacked (XLSTAT category model)

    # ── Sample size filter ────────────────────────────────────────────────────
    with st.expander("Sample Filters", expanded=False):
        da_top_brands = st.slider("Max brands in analysis", 3, 19, 10, key="da_top_brands",
                                   help="Cap on number of brands included (ranked by aided awareness)")

    if len(selected_drivers) < 2:
        st.info("Select at least 2 drivers to run the analysis.")
        return

    # Run Analysis
    compare_map = {"Overall": "overall", "By Category": "category", "By Zone": "zone"}
    da_result = _run_driver_analysis(
        da_product, tuple(sorted(selected_drivers)),
        compare_map.get(compare_by, "overall"), da_top_brands, da_pctile,
        zone=zone_arg, gender=gender_arg, age_band=age_band_arg, city=city_arg,
        project_id=project_id,
        pooled=_da_pooled,
    )

    if da_result.get("status") != "ok":
        st.warning(f"Driver Analysis: {da_result.get('message', 'No data available for selected filters.')}")
        return

    summary = da_result.get("summary_table")
    bip_o   = da_result["bip_overall"]

    if bip_o.get("status") != "ok":
        st.warning("No BIP data available for the selected drivers/filters.")
        return

    bip_tables = bip_o.get("tables", {})
    nps_impacts = da_result.get("nps_impacts", pd.Series())
    nps_impact_stats = da_result.get("nps_impact_stats", {})
    nps_impact_drivers = da_result.get("nps_impact_drivers", {})

    # KPI Row
    k1, k2, k3, k4 = st.columns(4)
    with k1: _metric_card("Drivers Analysed", str(len(selected_drivers)), "🎯")
    with k2: _metric_card("Brands Compared", str(len(summary)) if summary is not None else "0", "🏢")
    with k3: _metric_card(f"p{da_pctile} Threshold", f"{bip_tables.get('percentiles', {}).get(f'p{da_pctile}', 0):.1f}%", "📊")
    with k4: _metric_card("Analysis Mode", compare_by, "🔍")

    st.markdown("<br>", unsafe_allow_html=True)

    # Independent `if`s, not st.tabs() — some of these sub-tabs (Category Reach)
    # trigger their own fresh compute; lazy rendering avoids paying for all 10
    # on every rerun.
    _dd_labels = [
        "💡 Summary & Insights",
        "🎯 Driver Matrix",
        "🎯 What-If Simulator",
        "🕸️ Perceptual Map",
        "📊 Detailed Tables",
        "🤖 AI Narrative",
        "🌐 Category Reach",
        "🏆 Brand Battle",
        "🎯 White Space",
        "🕸️ Rival Radar",
    ]
    _dd_active = st.radio("Driver Deep-Dive view", _dd_labels, horizontal=True,
                          label_visibility="collapsed", key="da_subtab_selector")

    if _dd_active == _dd_labels[0]:
        col_a, col_b = st.columns([2, 1])
        with col_a:
            st.markdown("**Driver Ownership Summary**")
            if summary is not None and not summary.empty:
                st.dataframe(
                    summary.style.background_gradient(subset=["YES count", "% of drivers owned"], cmap="Greens"),
                    hide_index=True, use_container_width=True
                )
            else:
                st.info("No summary data available.")
        with col_b:
            st.markdown("**Top Findings**")
            ai_txt = da_result.get("ai_insight", "")
            if ai_txt and not ai_txt.startswith("_AI insight"):
                # Extract first few lines of AI insight for summary
                summary_text = ai_txt.split("\n\n")[0] if "\n\n" in ai_txt else ai_txt[:300]
                st.info(summary_text)
            else:
                st.info("Select more drivers to see detailed brand positioning insights.")

        # ── NPS Impact regression — goodness-of-fit + per-driver significance ──
        # nps_impacts (used by the Driver Matrix / What-If tabs below) is a raw
        # regression coefficient with no context on whether it's statistically
        # reliable. Surface the same fit stats the Key Driver Regression tab
        # shows, so a coefficient here is never displayed unqualified.
        if nps_impact_stats:
            st.divider()
            st.markdown("**NPS Impact Regression — Goodness of Fit** "
                        "*(the model behind the importance scores used in Driver Matrix / What-If Simulator)*")
            _pct_dropped_dd = nps_impact_stats.get("pct_dropped_na", 0) or 0
            if _pct_dropped_dd > 5:
                st.warning(f"⚠️ {_pct_dropped_dd:.1f}% of respondents were dropped for missing data "
                          f"(n={nps_impact_stats.get('n_before_na_omit', '?')} → n={nps_impact_stats.get('n', 0)}). "
                          f"Importance scores below may be unstable.")
            _dd_sw = nps_impact_stats.get("shapiro_wilk") or {}
            _dd_sw_p = _dd_sw.get("p_value")
            _dd_fp = nps_impact_stats.get("f_p_value")
            def _nis(k, d=0):
                try: return float(nps_impact_stats.get(k, d) or d)
                except: return float(d)
            _dd_gof_rows = [
                {"Statistic": "R²",               "Value": f"{_nis('r_squared'):.4f}"},
                {"Statistic": "Adjusted R²",      "Value": f"{_nis('adj_r_squared'):.4f}"},
                {"Statistic": "F (model)",        "Value":
                 f"{_nis('f_statistic'):.2f}"
                 + (f"  p={_dd_fp:.4f}" if _dd_fp is not None else "")},
                {"Statistic": "AIC",              "Value": f"{_nis('aic'):.2f}"},
                {"Statistic": "BIC",              "Value": f"{_nis('bic'):.2f}"},
                {"Statistic": "Shapiro-Wilk (residuals)", "Value":
                 f"W={_dd_sw.get('w_stat', 0):.4f}  p={_dd_sw_p:.4f}" if _dd_sw_p is not None else "—"},
                {"Statistic": "Observations (n)", "Value": f"{nps_impact_stats.get('n', 0):,}"},
                {"Statistic": "Predictors",       "Value": f"{nps_impact_stats.get('n_attrs', 0)}"},
            ]
            st.dataframe(pd.DataFrame(_dd_gof_rows), hide_index=True, use_container_width=True)

            if nps_impact_drivers:
                _dd_rows = []
                for _lbl, _coef in nps_impacts.items():
                    _d = nps_impact_drivers.get(_lbl, {}) or {}
                    _p = _d.get("p_value")
                    _dd_rows.append({
                        "Driver": _lbl,
                        "Std. coef (importance)": round(float(_coef), 4),
                        "p-value": f"{_p:.4f}" if _p is not None else "—",
                        "Significant (p<0.05)": "✓" if _d.get("significant") else ("—" if _p is not None else "?"),
                        "95% CI": f"[{_d.get('ci_low', 0):.3f}, {_d.get('ci_high', 0):.3f}]" if _d.get("ci_low") is not None else "—",
                    })
                if _dd_rows:
                    st.markdown("**Per-driver significance** *(is each importance score statistically reliable?)*")
                    st.dataframe(pd.DataFrame(_dd_rows).sort_values("Std. coef (importance)", key=abs, ascending=False),
                                 hide_index=True, use_container_width=True)

        if compare_by != "Overall" and da_result.get("bip_by_split"):
            st.divider()
            st.markdown(f"**Detailed Breakdown: {compare_by}**")
            split_tabs = st.tabs(list(da_result["bip_by_split"].keys()))
            for i, (split_name, split_res) in enumerate(da_result["bip_by_split"].items()):
                with split_tabs[i]:
                    sig_table = split_res["tables"].get("table14_significance", pd.DataFrame())
                    if not sig_table.empty:
                        st.dataframe(sig_table.style.map(lambda x: "background-color: #dcfce7; color: #166534" if x == "YES" else ""), use_container_width=True)

    if _dd_active == _dd_labels[1]:
        if sel_brand:
            norm_all = bip_tables.get("table3_norm_pct", pd.DataFrame())
            
            if not norm_all.empty and sel_brand in norm_all.index:
                drivers_df = {
                    "norm":       norm_all.loc[sel_brand],
                    "importance": nps_impacts
                }
                fig_matrix = _driver_action_matrix(drivers_df, sel_brand)
                if fig_matrix:
                    st.plotly_chart(_theme_fig(fig_matrix), use_container_width=True)
                    st.caption(f"Action priority for **{sel_brand}** based on normalized competitive advantage vs. derived NPS impact (importance).")
                else:
                    st.info(f"Action matrix unavailable for {sel_brand}.")
            else:
                st.info(f"Insufficient driver data for {sel_brand}.")
        else:
            st.info("Select a brand in the 'Deep-Dive Brand' sidebar to see the action matrix.")

    if _dd_active == _dd_labels[2]:
        if sel_brand:
            raw_all = bip_tables.get("table1_raw", pd.DataFrame())
            if not raw_all.empty and sel_brand in raw_all.index:
                _render_what_if_simulator(sel_brand, raw_all.loc[sel_brand], nps_impacts)
            else:
                st.info(f"Insufficient data for simulator for {sel_brand}.")
        else:
            st.info("Select a brand in the sidebar to use the simulator.")

    if _dd_active == _dd_labels[3]:
        c1, c2 = st.columns([2, 1])
        with c1:
            ca_o = da_result["ca_overall"]
            if ca_o.get("status") == "ok" and ca_o.get("chart_specs"):
                st.plotly_chart(ca_o["chart_specs"][0]["figure"], use_container_width=True)
        with c2:
            if bip_o.get("chart_specs"):
                st.plotly_chart(bip_o["chart_specs"][0]["figure"], use_container_width=True)
            if ca_o.get("status") == "ok":
                eig_df = ca_o["ca_results"]["eigenvalues"]
                f1_pct = eig_df["Inertia_%"].iloc[0]
                f2_pct = eig_df["Inertia_%"].iloc[1] if len(eig_df) > 1 else 0
                st.caption(f"F1+F2 explain {f1_pct+f2_pct:.1f}% of total variation.")

    if _dd_active == _dd_labels[4]:
        col_ts1, col_ts2 = st.columns([3, 1])
        with col_ts1:
            table_select = st.selectbox("Select View", ["YES/NO Significance", "Raw % Association", "Normalized Deviation", "All Statistical Tables"], key="da_table_view")
        with col_ts2:
            transpose_view = st.checkbox("Drivers as Rows (transposed)", value=True,
                                          key="da_transpose_view",
                                          help="Brands as columns, drivers as rows — easier to compare")

        def _show_transposed(df, style_fn=None):
            if df is None or df.empty:
                st.info("No data.")
                return
            display_df = df.T if transpose_view else df
            if style_fn:
                st.dataframe(display_df.style.map(style_fn), use_container_width=True)
            else:
                st.dataframe(display_df.round(2), use_container_width=True)

        sig_table = bip_tables.get("table14_significance", pd.DataFrame())
        if table_select == "YES/NO Significance":
            _show_transposed(sig_table, lambda x: "background-color: #dcfce7; color: #166534" if x == "YES" else "background-color: #fee2e2; color: #991b1b")
        elif table_select == "Raw % Association":
            _show_transposed(bip_tables.get("table1_raw", pd.DataFrame()))
        elif table_select == "Normalized Deviation":
            _show_transposed(bip_tables.get("table3_norm_pct", pd.DataFrame()))
        else:
            inner_tabs = st.tabs(["Raw %", "Abs Diff", "Norm %", "Filtered", "Positive", "Negative", "Ranks", "Combined"])
            table_keys = ["table1_raw", "table2_abs_diff", "table3_norm_pct", "table4_filtered", "table5_positive", "table6_negative", "table7_ranks", "table13_combined"]
            for i, (tab, k) in enumerate(zip(inner_tabs, table_keys)):
                with tab:
                    _show_transposed(bip_tables.get(k))

    if _dd_active == _dd_labels[5]:
        ai_txt = da_result.get("ai_insight", "")
        if ai_txt:
            st.markdown(ai_txt)
        else:
            st.info("AI analysis requires an active OpenRouter API key.")

    if _dd_active == _dd_labels[6]:
        _section_header(
            "🌐 Cross-Category Driver Reach",
            "Is this driver universally important, or specific to one category? "
            "Market-norm association % shows how strongly consumers link this attribute "
            "to brands in each category — regardless of which brand they pick.",
        )
        reach_data = _run_cross_category_reach(tuple(sorted(selected_drivers)), da_pctile)
        reach_df     = reach_data.get("reach_df", pd.DataFrame())
        ownership_df = reach_data.get("ownership_df", pd.DataFrame())

        if reach_df.empty:
            st.info("Cross-category data unavailable for selected drivers.")
        else:
            fig_heat = go.Figure(data=go.Heatmap(
                z=reach_df.values.tolist(),
                x=list(reach_df.columns),
                y=list(reach_df.index),
                colorscale="RdYlGn",
                text=[[f"{v:.1f}%" for v in row] for row in reach_df.values],
                texttemplate="%{text}",
                showscale=True,
                colorbar=dict(title="Market Norm %"),
            ))
            fig_heat.update_layout(
                title="Market Association % by Category",
                xaxis_title="Driver",
                yaxis_title="Category",
                height=320,
                margin=dict(l=10, r=10, t=40, b=10),
                plot_bgcolor="white",
                paper_bgcolor="white",
            )
            st.plotly_chart(fig_heat, use_container_width=True)
            st.caption(
                "Color = market-norm association %. Green = high consumer linkage in that category. "
                "Drivers that are green across all rows are **universal**; single-row green = **niche**."
            )

            universality = reach_df.mean(axis=0).sort_values(ascending=False)
            fig_bar = go.Figure(go.Bar(
                x=universality.index.tolist(),
                y=universality.values.tolist(),
                marker_color="#6366f1",
                text=[f"{v:.1f}%" for v in universality.values],
                textposition="outside",
            ))
            fig_bar.update_layout(
                title="Universality Score — Average Market Norm % Across All Categories",
                yaxis_title="Avg Market Norm %",
                height=280,
                margin=dict(l=10, r=10, t=40, b=10),
                plot_bgcolor="white",
                paper_bgcolor="white",
                yaxis=dict(gridcolor="#f0f0f0"),
            )
            st.plotly_chart(fig_bar, use_container_width=True)

            top_universal = universality.idxmax() if not universality.empty else None
            top_niche_cat = reach_df.idxmax(axis=0)
            findings = []
            if top_universal:
                findings.append(
                    f"**{top_universal}** is the most universal driver (avg {universality.iloc[0]:.1f}% "
                    f"market norm across all categories) — safe equity to build brand-wide."
                )
            for drv in list(universality.index[:3]):
                best_cat = top_niche_cat.get(drv, "—")
                best_val = reach_df.loc[best_cat, drv] if best_cat in reach_df.index else 0
                avg_val  = universality.get(drv, 0)
                if best_val > avg_val * 1.5:
                    findings.append(
                        f"**{drv}** skews heavily toward **{best_cat}** ({best_val:.1f}%) "
                        f"vs overall avg ({avg_val:.1f}%) — category-specific opportunity."
                    )
            _insight_callout(findings, "Category Reach Interpretation", "🌐")

    if _dd_active == _dd_labels[7]:
        _section_header(
            "🏆 Brand Battle — Per-Driver Competition",
            "Who wins each driver head-to-head? Green bars = brand significantly "
            "over-associated (BIP YES). Dashed line = significance gate. "
            "Tight clustering = commodity driver; spread = owned territory.",
        )
        raw_df   = bip_tables.get("table1_raw", pd.DataFrame())
        sig_df   = bip_tables.get("table14_significance", pd.DataFrame())
        col_avgs_battle = bip_tables.get("column_averages", pd.Series(dtype=float))

        if raw_df.empty:
            st.info("Raw association data unavailable.")
        else:
            from plotly.subplots import make_subplots

            driver_labels_available = [
                lbl for lbl in da_result.get("driver_labels", [])
                if lbl in raw_df.columns
            ]
            if not driver_labels_available:
                driver_labels_available = list(raw_df.columns)

            use_subplot = len(driver_labels_available) <= 4

            if use_subplot:
                n_cols = min(2, len(driver_labels_available))
                n_rows = (len(driver_labels_available) + n_cols - 1) // n_cols
                fig_battle = make_subplots(
                    rows=n_rows, cols=n_cols,
                    subplot_titles=driver_labels_available,
                    horizontal_spacing=0.12,
                    vertical_spacing=0.18,
                )
                for idx, drv in enumerate(driver_labels_available):
                    row = idx // n_cols + 1
                    col = idx % n_cols + 1
                    brand_scores = raw_df[drv].sort_values(ascending=True)
                    colors = []
                    for brand in brand_scores.index:
                        is_yes = (
                            not sig_df.empty
                            and drv in sig_df.columns
                            and brand in sig_df.index
                            and sig_df.loc[brand, drv] == "YES"
                        )
                        colors.append("#22c55e" if is_yes else "#d1d5db")
                    fig_battle.add_trace(
                        go.Bar(
                            x=brand_scores.values.tolist(),
                            y=brand_scores.index.tolist(),
                            orientation="h",
                            marker_color=colors,
                            text=[f"{v:.1f}%" for v in brand_scores.values],
                            textposition="outside",
                            showlegend=False,
                        ),
                        row=row, col=col,
                    )
                    mkt_avg_val = float(col_avgs_battle.get(drv, 0.0)) if not col_avgs_battle.empty else 0.0
                    fig_battle.add_vline(
                        x=mkt_avg_val, line_dash="dash", line_color="#6366f1",
                        annotation_text=f"Mkt avg {mkt_avg_val:.1f}%",
                        annotation_position="top right",
                        row=row, col=col,
                    )
                fig_battle.update_layout(
                    height=280 * n_rows,
                    margin=dict(l=10, r=60, t=40, b=10),
                    plot_bgcolor="white",
                    paper_bgcolor="white",
                )
                st.plotly_chart(fig_battle, use_container_width=True)
            else:
                sel_drv = st.selectbox(
                    "Select Driver", driver_labels_available, key="da_battle_drv_sel"
                )
                brand_scores = raw_df[sel_drv].sort_values(ascending=True)
                colors = []
                for brand in brand_scores.index:
                    is_yes = (
                        not sig_df.empty
                        and sel_drv in sig_df.columns
                        and brand in sig_df.index
                        and sig_df.loc[brand, sel_drv] == "YES"
                    )
                    colors.append("#22c55e" if is_yes else "#d1d5db")
                fig_single = go.Figure(go.Bar(
                    x=brand_scores.values.tolist(),
                    y=brand_scores.index.tolist(),
                    orientation="h",
                    marker_color=colors,
                    text=[f"{v:.1f}%" for v in brand_scores.values],
                    textposition="outside",
                ))
                mkt_avg_single = float(col_avgs_battle.get(sel_drv, 0.0)) if not col_avgs_battle.empty else 0.0
                fig_single.add_vline(
                    x=mkt_avg_single, line_dash="dash", line_color="#6366f1",
                    annotation_text=f"Mkt avg: {mkt_avg_single:.1f}%",
                    annotation_position="top right",
                )
                fig_single.update_layout(
                    xaxis_title="Raw Association %",
                    height=max(300, 40 * len(brand_scores)),
                    margin=dict(l=10, r=60, t=30, b=10),
                    plot_bgcolor="white",
                    paper_bgcolor="white",
                    xaxis=dict(gridcolor="#f0f0f0"),
                )
                st.plotly_chart(fig_single, use_container_width=True)

            st.caption(
                "🟢 Green bar = BIP significant (brand over-indexes vs market norm). "
                "🔵 Dashed line = market average for that driver. "
                "Bars clustered near average = commodity space; clear winner above average = owned territory."
            )

            battle_findings = []
            for drv in driver_labels_available[:4]:
                if drv not in raw_df.columns:
                    continue
                scores = raw_df[drv].sort_values(ascending=False)
                if len(scores) >= 2:
                    gap = float(scores.iloc[0]) - float(scores.iloc[1])
                    leader = scores.index[0]
                    if gap > 8:
                        battle_findings.append(
                            f"**{drv}**: {leader} leads by {gap:.1f}pp over #2 — clear ownership."
                        )
                    else:
                        battle_findings.append(
                            f"**{drv}**: tight contest — top 2 within {gap:.1f}pp — contested commodity space."
                        )
            _insight_callout(battle_findings, "Battle Analysis", "🏆")

    if _dd_active == _dd_labels[8]:
        _section_header(
            "🎯 Driver White Space Map",
            "Importance (x) vs. Ownership (y). Drivers in the bottom-right quadrant "
            "are high-market-norm but claimed by few brands — maximum untapped opportunity.",
        )
        col_avgs = bip_tables.get("column_averages", pd.Series(dtype=float))
        sig_df_ws = bip_tables.get("table14_significance", pd.DataFrame())

        if col_avgs.empty:
            st.info("Column averages unavailable for white space analysis.")
        else:
            driver_labels_ws = da_result.get("driver_labels", list(col_avgs.index))
            importance = {lbl: float(col_avgs.get(lbl, 0.0)) for lbl in driver_labels_ws}
            ownership  = {}
            for lbl in driver_labels_ws:
                if not sig_df_ws.empty and lbl in sig_df_ws.columns:
                    ownership[lbl] = int((sig_df_ws[lbl] == "YES").sum())
                else:
                    ownership[lbl] = 0

            x_vals = [importance.get(lbl, 0.0) for lbl in driver_labels_ws]
            y_vals = [ownership.get(lbl, 0)    for lbl in driver_labels_ws]

            x_mid = sum(x_vals) / max(len(x_vals), 1)
            y_mid = sum(y_vals) / max(len(y_vals), 1)

            quadrant_colors = []
            quadrant_labels = []
            for x, y in zip(x_vals, y_vals):
                if x >= x_mid and y < y_mid:
                    quadrant_colors.append("#22c55e")
                    quadrant_labels.append("White Space")
                elif x >= x_mid and y >= y_mid:
                    quadrant_colors.append("#ef4444")
                    quadrant_labels.append("Battleground")
                elif x < x_mid and y >= y_mid:
                    quadrant_colors.append("#f59e0b")
                    quadrant_labels.append("Owned Niche")
                else:
                    quadrant_colors.append("#d1d5db")
                    quadrant_labels.append("Deprioritize")

            fig_ws = go.Figure()
            for zone, color, symbol in [
                ("White Space",  "#22c55e", "circle"),
                ("Battleground", "#ef4444", "diamond"),
                ("Owned Niche",  "#f59e0b", "square"),
                ("Deprioritize", "#d1d5db", "x"),
            ]:
                idxs = [i for i, ql in enumerate(quadrant_labels) if ql == zone]
                if idxs:
                    fig_ws.add_trace(go.Scatter(
                        x=[x_vals[i] for i in idxs],
                        y=[y_vals[i] for i in idxs],
                        mode="markers+text",
                        name=zone,
                        text=[driver_labels_ws[i] for i in idxs],
                        textposition="top center",
                        marker=dict(size=14, color=color, symbol=symbol, line=dict(width=1, color="white")),
                    ))

            fig_ws.add_vline(x=x_mid, line_dash="dot", line_color="#9ca3af")
            fig_ws.add_hline(y=y_mid, line_dash="dot", line_color="#9ca3af")

            for quad, (ax, ay, text) in {
                "White Space":  (max(x_vals + [x_mid + 1]) * 0.95, y_mid * 0.15, "🟢 White Space<br>(build equity)"),
                "Battleground": (max(x_vals + [x_mid + 1]) * 0.95, max(y_vals + [y_mid + 1]) * 0.9, "🔴 Battleground<br>(defend/attack)"),
                "Owned Niche":  (min(x_vals + [x_mid - 1]) * 1.05 if min(x_vals) > 0 else x_mid * 0.1, max(y_vals + [y_mid + 1]) * 0.9, "🟡 Owned Niche<br>(maintain)"),
                "Deprioritize": (min(x_vals + [x_mid - 1]) * 1.05 if min(x_vals) > 0 else x_mid * 0.1, y_mid * 0.15, "⚪ Deprioritize"),
            }.items():
                fig_ws.add_annotation(
                    x=ax, y=ay, text=text, showarrow=False,
                    font=dict(size=10, color="#6b7280"),
                    align="center",
                )

            fig_ws.update_layout(
                xaxis_title="Market Importance (Avg Association % across all brands)",
                yaxis_title="Ownership Breadth (# Brands with BIP = YES)",
                height=480,
                margin=dict(l=10, r=10, t=20, b=10),
                plot_bgcolor="white",
                paper_bgcolor="white",
                legend=dict(orientation="h", yanchor="bottom", y=1.01, xanchor="right", x=1),
                xaxis=dict(gridcolor="#f0f0f0"),
                yaxis=dict(gridcolor="#f0f0f0", dtick=1),
            )
            st.plotly_chart(fig_ws, use_container_width=True)
            st.caption(
                "🟢 **White Space** = high consumer importance, few brands own it — build equity here. "
                "🔴 **Battleground** = everyone fights here — only enter if you can win. "
                "🟡 **Owned Niche** = brands claim it but consumers don't care much — maintain efficiently. "
                "⚪ **Deprioritize** = low importance, low ownership — skip."
            )

            ws_drivers = [
                driver_labels_ws[i] for i, ql in enumerate(quadrant_labels) if ql == "White Space"
            ]
            if ws_drivers:
                _insight_callout(
                    [f"**{d}** is unclaimed high-value territory — no brand owns it above the gate yet." for d in ws_drivers[:3]],
                    "White Space Opportunities", "🎯"
                )

    if _dd_active == _dd_labels[9]:
        _section_header(
            "🕸️ Rival Radar — Competitive Positioning",
            "Normalized deviation per driver. Above 0 = brand over-associates vs. market average. "
            "Shows YOUR brand vs. top 3 rivals on every selected driver simultaneously.",
        )
        norm_df = bip_tables.get("table3_norm_pct", pd.DataFrame())
        sig_df_rr = bip_tables.get("table14_significance", pd.DataFrame())

        if norm_df.empty:
            st.info("Normalized deviation data unavailable for radar chart.")
        else:
            if not sig_df_rr.empty:
                yes_counts = (sig_df_rr == "YES").sum(axis=1).sort_values(ascending=False)
                top_brands_rr = list(yes_counts.index[:4])
            else:
                top_brands_rr = list(norm_df.index[:4])

            if sel_brand and sel_brand in norm_df.index and sel_brand not in top_brands_rr:
                top_brands_rr = [sel_brand] + top_brands_rr[:3]
            elif sel_brand and sel_brand in top_brands_rr:
                top_brands_rr = [sel_brand] + [b for b in top_brands_rr if b != sel_brand][:3]
            else:
                top_brands_rr = top_brands_rr[:4]

            driver_cols_rr = [c for c in norm_df.columns if c in (da_result.get("driver_labels") or norm_df.columns)]
            if not driver_cols_rr:
                driver_cols_rr = list(norm_df.columns)

            RADAR_COLORS = ["#6366f1", "#ef4444", "#f59e0b", "#22c55e", "#8b5cf6", "#06b6d4"]

            fig_radar = go.Figure()
            for i, brand in enumerate(top_brands_rr):
                if brand not in norm_df.index:
                    continue
                vals = [float(norm_df.loc[brand, c]) if c in norm_df.columns else 0.0 for c in driver_cols_rr]
                vals_closed = vals + [vals[0]]
                cats_closed = driver_cols_rr + [driver_cols_rr[0]]
                fig_radar.add_trace(go.Scatterpolar(
                    r=vals_closed,
                    theta=cats_closed,
                    fill="toself" if brand == sel_brand else "none",
                    name=brand,
                    line=dict(color=RADAR_COLORS[i % len(RADAR_COLORS)], width=2 if brand == sel_brand else 1.5),
                    opacity=0.85 if brand == sel_brand else 0.65,
                ))

            r_min = min(-5, norm_df[driver_cols_rr].min().min() - 2) if driver_cols_rr else -5
            r_max = max(5,  norm_df[driver_cols_rr].max().max() + 2) if driver_cols_rr else 5

            fig_radar.update_layout(
                polar=dict(
                    radialaxis=dict(
                        visible=True,
                        range=[r_min, r_max],
                        tickfont=dict(size=9),
                        gridcolor="#e5e7eb",
                    ),
                    angularaxis=dict(tickfont=dict(size=10)),
                    bgcolor="white",
                ),
                showlegend=True,
                legend=dict(orientation="h", yanchor="bottom", y=-0.15, xanchor="center", x=0.5),
                height=480,
                margin=dict(l=40, r=40, t=40, b=60),
                paper_bgcolor="white",
            )
            st.plotly_chart(fig_radar, use_container_width=True)
            st.caption(
                "Normalized deviation: **above 0** = brand over-associates vs. market average; "
                "**below 0** = under-associates. Filled shape = selected deep-dive brand. "
                "Larger filled area = stronger ownership of the selected driver set."
            )

            radar_findings = []
            if sel_brand and sel_brand in norm_df.index:
                brand_row = norm_df.loc[sel_brand, driver_cols_rr] if driver_cols_rr else pd.Series()
                strongest  = brand_row.idxmax() if not brand_row.empty else None
                weakest    = brand_row.idxmin()  if not brand_row.empty else None
                if strongest:
                    radar_findings.append(
                        f"**{sel_brand}** strongest on **{strongest}** "
                        f"(+{float(brand_row[strongest]):.1f} normalized deviation) — protect this."
                    )
                if weakest and float(brand_row[weakest]) < 0:
                    rival_best = norm_df[weakest].drop(sel_brand, errors="ignore").idxmax() if weakest in norm_df.columns else None
                    radar_findings.append(
                        f"**{weakest}** is a gap for {sel_brand} ({float(brand_row[weakest]):.1f})"
                        + (f" — {rival_best} owns it." if rival_best else ".")
                    )
            _insight_callout(radar_findings, "Positioning Insights", "🕸️")


# ══════════════════════════════════════════════════════════════════════════════
# INDUSTRY-STANDARD BRAND HEALTH SECTIONS
# 1. Competitive Benchmarking Panel
# 2. Brand Equity Index (BEI)
# 3. Funnel Conversion & Leakage
# 4. Consumer Demographics Profile
# 5. Attribute Ownership Map
# ══════════════════════════════════════════════════════════════════════════════

@st.cache_data(ttl=3600)
def _competitive_benchmarking_from_layer(_layer: "ProjectDataLayer", zone="all", gender="all", age_band="all", city="all") -> pd.DataFrame:
    """
    data_layer version of _get_competitive_benchmarking — no SQL.
    Computes all-brand funnel % + NPS + CSAT from raw Excel via ProjectDataLayer.
    Output columns match SQL version exactly so downstream fns need no changes.
    """
    layer = _layer
    aw = layer.awareness
    nps_df = layer.nps
    csat_df = layer.csat
    demo_df = layer.demographics

    resp_ids = set(demo_df["respondent_id"].tolist())
    if zone.lower() not in ("all", "") and "zone_name" in demo_df.columns:
        resp_ids &= set(demo_df[demo_df["zone_name"].str.lower() == zone.lower()]["respondent_id"])
    if city.lower() not in ("all", "") and "city_name" in demo_df.columns:
        resp_ids &= set(demo_df[demo_df["city_name"].str.lower() == city.lower()]["respondent_id"])
    if gender.lower() not in ("all", "") and "gender" in demo_df.columns:
        resp_ids &= set(demo_df[demo_df["gender"].str.lower() == gender.lower()]["respondent_id"])
    if age_band.lower() not in ("all", "") and "age_band" in demo_df.columns:
        resp_ids &= set(demo_df[demo_df["age_band"].str.lower() == age_band.lower()]["respondent_id"])

    base_n = len(resp_ids)
    if base_n < 10 or aw.empty:
        return pd.DataFrame()

    if resp_ids != set(demo_df["respondent_id"]):
        aw = aw[aw["respondent_id"].isin(resp_ids)]
        if not nps_df.empty:
            nps_df = nps_df[nps_df["respondent_id"].isin(resp_ids)]
        if not csat_df.empty:
            csat_df = csat_df[csat_df["respondent_id"].isin(resp_ids)]

    rows = []
    for brand_name, grp in aw.groupby("brand_name"):
        if brand_name in ("Don't Know / None", "DK/None"):
            continue
        _awa = grp[grp["stage"].isin(["TOM", "SPONT", "AIDED"])]
        total_awa_n = _awa["respondent_id"].nunique()
        spont_n = _awa[_awa["stage"].isin(["SPONT", "TOM"])]["respondent_id"].nunique()
        tom_n = _awa[_awa["stage"] == "TOM"]["respondent_id"].nunique()
        aided_only_n = grp[grp["stage"] == "AIDED"]["respondent_id"].nunique()

        def _stage_n(s): return grp[grp["stage"] == s]["respondent_id"].nunique()
        ever_used_n = _stage_n("EVER_USED")
        current_n = _stage_n("CURRENT_USER") or _stage_n("CURRENT_USE")
        consid_n = _stage_n("CONSIDERATION")
        pref_n = _stage_n("PREFERRED")
        last_n = _stage_n("LAST_PURCHASED")

        if max(total_awa_n, ever_used_n, consid_n) < 5:
            continue

        rows.append({
            "brand_name": str(brand_name),
            "TOM": round(tom_n / base_n * 100, 1),
            "SPONT": round(spont_n / base_n * 100, 1),
            "AIDED": round(aided_only_n / base_n * 100, 1),
            "TOTAL_AIDED": round(total_awa_n / base_n * 100, 1),
            "EVER_USED": round(ever_used_n / base_n * 100, 1),
            "CURRENT_USER": round(current_n / base_n * 100, 1),
            "CONSIDERATION": round(consid_n / base_n * 100, 1),
            "PREFERRED": round(pref_n / base_n * 100, 1),
            "LAST_PURCHASED": round(last_n / base_n * 100, 1),
            "aided_n": total_awa_n,
        })

    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)

    n_threshold = 5 if len(getattr(layer, "all_brands", [])) < 50 else 20

    if not nps_df.empty:
        def _nps_score(g):
            n = len(g)
            if n < n_threshold: return None
            return round((g >= 9).sum() / n * 100 - (g <= 6).sum() / n * 100, 1)
        nps_agg = nps_df.groupby("brand_name")["nps_score"].apply(_nps_score).reset_index(name="nps")
        df = df.merge(nps_agg, on="brand_name", how="left")
    else:
        df["nps"] = None

    if not csat_df.empty:
        def _csat_score(g):
            if len(g) < n_threshold: return None
            raw = g.mean()
            return round(raw / 2 if raw > 5.5 else raw, 2)
        csat_agg = csat_df.groupby("brand_name")["csat_score"].apply(_csat_score).reset_index(name="csat")
        df = df.merge(csat_agg, on="brand_name", how="left")
    else:
        df["csat"] = None

    df["nps"] = pd.to_numeric(df.get("nps"), errors="coerce")
    df["csat"] = pd.to_numeric(df.get("csat"), errors="coerce")
    df["aided_n"] = pd.to_numeric(df.get("aided_n"), errors="coerce")
    return df.sort_values("TOTAL_AIDED", ascending=False, na_position="last").reset_index(drop=True)


@st.cache_data(ttl=3600)
def _get_competitive_benchmarking(zone="all", gender="all", age_band="all", city="all") -> pd.DataFrame:
    """All brands × funnel stage % + NPS + brand-level CSAT."""
    # Try data layer first (raw Excel path)
    _layer = _get_layer()
    if _layer is not None:
        return _competitive_benchmarking_from_layer(_layer, zone=zone, gender=gender, age_band=age_band, city=city)

    import sqlite3
    from oxdata.db_loader import get_db_path
    conn = sqlite3.connect(get_db_path())

    cte, fparams, is_filtered = _resp_filter_cte(zone, gender, age_band, city)
    base_restrict = " AND a.respondent_id IN (SELECT respondent_id FROM _base)" if is_filtered else ""
    cte_prefix = f"WITH {cte} " if is_filtered else ""

    if is_filtered:
        total = int(pd.read_sql(f"WITH {cte} SELECT COUNT(DISTINCT respondent_id) n FROM _base",
                                conn, params=fparams).iloc[0, 0])
    else:
        total = int(pd.read_sql("SELECT COUNT(DISTINCT respondent_id) n FROM fact_respondents",
                                conn).iloc[0, 0])

    if total < 10:
        conn.close()
        return pd.DataFrame()

    stage_df = pd.read_sql(f"""
        {cte_prefix}
        SELECT 
            b.brand_name,
            COUNT(DISTINCT CASE WHEN a.stage = 'TOM' THEN a.respondent_id END) AS tom_n,
            COUNT(DISTINCT CASE WHEN a.stage IN ('TOM', 'SPONT') THEN a.respondent_id END) AS spont_n,
            COUNT(DISTINCT CASE WHEN a.stage = 'AIDED' THEN a.respondent_id END) AS aided_n,
            COUNT(DISTINCT CASE WHEN a.stage IN ('TOM', 'SPONT', 'AIDED') THEN a.respondent_id END) AS total_aware_n,
            COUNT(DISTINCT CASE WHEN a.stage = 'EVER_USED' THEN a.respondent_id END) AS ever_used_n,
            COUNT(DISTINCT CASE WHEN a.stage IN ('CURRENT_USER', 'CURRENT_USE') THEN a.respondent_id END) AS current_n,
            COUNT(DISTINCT CASE WHEN a.stage = 'CONSIDERATION' THEN a.respondent_id END) AS consid_n,
            COUNT(DISTINCT CASE WHEN a.stage = 'PREFERRED' THEN a.respondent_id END) AS pref_n,
            COUNT(DISTINCT CASE WHEN a.stage = 'LAST_PURCHASED' THEN a.respondent_id END) AS last_n
        FROM dim_brand b
        JOIN fact_brand_awareness a ON b.brand_id = a.brand_id
        WHERE b.brand_name NOT IN ('Don''t Know / None')
          {base_restrict}
        GROUP BY b.brand_name
        HAVING total_aware_n >= 5 OR ever_used_n >= 5 OR consid_n >= 5
    """, conn, params=fparams)

    if stage_df.empty:
        conn.close()
        return pd.DataFrame()

    stage_df["TOM"] = (stage_df["tom_n"] * 100.0 / total).round(1)
    stage_df["SPONT"] = (stage_df["spont_n"] * 100.0 / total).round(1)
    stage_df["AIDED"] = (stage_df["aided_n"] * 100.0 / total).round(1)
    stage_df["TOTAL_AIDED"] = (stage_df["total_aware_n"] * 100.0 / total).round(1)
    stage_df["EVER_USED"] = (stage_df["ever_used_n"] * 100.0 / total).round(1)
    stage_df["CURRENT_USER"] = (stage_df["current_n"] * 100.0 / total).round(1)
    stage_df["CONSIDERATION"] = (stage_df["consid_n"] * 100.0 / total).round(1)
    stage_df["PREFERRED"] = (stage_df["pref_n"] * 100.0 / total).round(1)
    stage_df["LAST_PURCHASED"] = (stage_df["last_n"] * 100.0 / total).round(1)
    stage_df["aided_n"] = stage_df["total_aware_n"]

    # Dynamic threshold: 5 if project has <50 brands, 20 otherwise
    try:
        b_count = int(pd.read_sql("SELECT COUNT(*) FROM dim_brand", conn).iloc[0, 0])
    except Exception:
        b_count = 56
    nps_min_threshold = 5 if b_count < 50 else 20

    nps_where_clause = " WHERE n.nps_score IS NOT NULL"
    if is_filtered:
        nps_where_clause += " AND n.respondent_id IN (SELECT respondent_id FROM _base)"
    nps_df = pd.read_sql(f"""
        {cte_prefix}
        SELECT b.brand_name,
               ROUND(SUM(CASE WHEN n.nps_score>=9 THEN 1.0 ELSE 0 END)*100.0/COUNT(n.nps_score) -
                     SUM(CASE WHEN n.nps_score<=6 THEN 1.0 ELSE 0 END)*100.0/COUNT(n.nps_score), 1) AS nps,
               COUNT(n.nps_score) AS nps_n
        FROM fact_brand_nps n JOIN dim_brand b ON n.brand_id = b.brand_id
        {nps_where_clause}
        GROUP BY b.brand_name HAVING nps_n >= {nps_min_threshold}
    """, conn, params=fparams)

    csat_restrict = " AND s.respondent_id IN (SELECT respondent_id FROM _base)" if is_filtered else ""
    csat_df = pd.read_sql(f"""
        {cte_prefix}
        SELECT b.brand_name, ROUND(AVG(s.score), 2) AS csat, COUNT(*) AS csat_n
        FROM fact_satisfaction s
        JOIN fact_brand_awareness ba ON s.respondent_id = ba.respondent_id AND ba.stage = 'LAST_PURCHASED'
        JOIN dim_brand b ON ba.brand_id = b.brand_id
        WHERE b.brand_name NOT IN ('Don''t Know / None')
          {csat_restrict}
        GROUP BY b.brand_name HAVING csat_n >= {nps_min_threshold}
    """, conn, params=fparams)
    conn.close()

    df = stage_df.merge(nps_df[["brand_name", "nps"]], on="brand_name", how="left")
    df = df.merge(csat_df[["brand_name", "csat"]], on="brand_name", how="left")
    df["nps"]     = pd.to_numeric(df["nps"],     errors="coerce")
    df["csat"]    = pd.to_numeric(df["csat"],    errors="coerce")
    df["aided_n"] = pd.to_numeric(df["aided_n"], errors="coerce")
    return df.sort_values("TOTAL_AIDED", ascending=False, na_position="last").reset_index(drop=True)


@st.cache_data(ttl=3600)
def _get_brand_stage_counts(zone="all", gender="all", age_band="all", city="all"):
    """Per-brand respondent COUNT + common base per awareness stage, for significance testing."""
    layer = _get_layer()
    if layer is not None:
        aw = layer.awareness.copy()
        demo = layer.demographics
        filt = demo.copy()
        if zone != "all":     filt = filt[filt["zone"]     == zone]
        if gender != "all":   filt = filt[filt["gender"]   == gender]
        if age_band != "all": filt = filt[filt["age_band"] == age_band]
        if city != "all":     filt = filt[filt["city"]     == city]
        base = len(filt)
        aw = aw[aw["respondent_id"].isin(filt["respondent_id"])]
        aw = aw[~aw["brand_name"].isin(["Don't Know / None", "DK/None"])]

        rows = []
        for brand_name, grp in aw.groupby("brand_name"):
            tom_n = grp[grp["stage"] == "TOM"]["respondent_id"].nunique()
            spont_n = grp[grp["stage"].isin(["TOM", "SPONT"])]["respondent_id"].nunique()
            total_awa_n = grp[grp["stage"].isin(["TOM", "SPONT", "AIDED"])]["respondent_id"].nunique()
            aided_n = grp[grp["stage"] == "AIDED"]["respondent_id"].nunique()
            ever_used_n = grp[grp["stage"] == "EVER_USED"]["respondent_id"].nunique()
            current_n = grp[grp["stage"].isin(["CURRENT_USER", "CURRENT_USE"])]["respondent_id"].nunique()
            consid_n = grp[grp["stage"] == "CONSIDERATION"]["respondent_id"].nunique()
            pref_n = grp[grp["stage"] == "PREFERRED"]["respondent_id"].nunique()
            last_n = grp[grp["stage"] == "LAST_PURCHASED"]["respondent_id"].nunique()

            rows.extend([
                {"brand_name": brand_name, "stage": "TOTAL_AIDED", "n": total_awa_n},
                {"brand_name": brand_name, "stage": "SPONT", "n": spont_n},
                {"brand_name": brand_name, "stage": "TOM", "n": tom_n},
                {"brand_name": brand_name, "stage": "AIDED", "n": aided_n},
                {"brand_name": brand_name, "stage": "EVER_USED", "n": ever_used_n},
                {"brand_name": brand_name, "stage": "CURRENT_USER", "n": current_n},
                {"brand_name": brand_name, "stage": "CONSIDERATION", "n": consid_n},
                {"brand_name": brand_name, "stage": "PREFERRED", "n": pref_n},
                {"brand_name": brand_name, "stage": "LAST_PURCHASED", "n": last_n},
            ])
        return pd.DataFrame(rows), base

    import sqlite3
    from oxdata.db_loader import get_db_path
    conn = sqlite3.connect(get_db_path())
    cte, fparams, is_filtered = _resp_filter_cte(zone, gender, age_band, city)
    base_restrict = " AND a.respondent_id IN (SELECT respondent_id FROM _base)" if is_filtered else ""
    cte_prefix = f"WITH {cte} " if is_filtered else ""
    if is_filtered:
        base = int(pd.read_sql(f"WITH {cte} SELECT COUNT(DISTINCT respondent_id) n FROM _base",
                                conn, params=fparams).iloc[0, 0])
    else:
        base = int(pd.read_sql("SELECT COUNT(DISTINCT respondent_id) n FROM fact_respondents",
                                conn).iloc[0, 0])
    counts_df = pd.read_sql(f"""
        {cte_prefix}
        SELECT 
            b.brand_name,
            COUNT(DISTINCT CASE WHEN a.stage = 'TOM' THEN a.respondent_id END) AS tom_n,
            COUNT(DISTINCT CASE WHEN a.stage IN ('TOM', 'SPONT') THEN a.respondent_id END) AS spont_n,
            COUNT(DISTINCT CASE WHEN a.stage = 'AIDED' THEN a.respondent_id END) AS aided_n,
            COUNT(DISTINCT CASE WHEN a.stage IN ('TOM', 'SPONT', 'AIDED') THEN a.respondent_id END) AS total_aware_n,
            COUNT(DISTINCT CASE WHEN a.stage = 'EVER_USED' THEN a.respondent_id END) AS ever_used_n,
            COUNT(DISTINCT CASE WHEN a.stage IN ('CURRENT_USER', 'CURRENT_USE') THEN a.respondent_id END) AS current_n,
            COUNT(DISTINCT CASE WHEN a.stage = 'CONSIDERATION' THEN a.respondent_id END) AS consid_n,
            COUNT(DISTINCT CASE WHEN a.stage = 'PREFERRED' THEN a.respondent_id END) AS pref_n,
            COUNT(DISTINCT CASE WHEN a.stage = 'LAST_PURCHASED' THEN a.respondent_id END) AS last_n
        FROM dim_brand b
        JOIN fact_brand_awareness a ON b.brand_id = a.brand_id
        WHERE b.brand_name NOT IN ('Don''t Know / None')
          {base_restrict}
        GROUP BY b.brand_name
    """, conn, params=fparams)
    conn.close()

    rows = []
    for _, r in counts_df.iterrows():
        b = r["brand_name"]
        rows.extend([
            {"brand_name": b, "stage": "TOTAL_AIDED", "n": int(r["total_aware_n"] or 0)},
            {"brand_name": b, "stage": "SPONT", "n": int(r["spont_n"] or 0)},
            {"brand_name": b, "stage": "TOM", "n": int(r["tom_n"] or 0)},
            {"brand_name": b, "stage": "AIDED", "n": int(r["aided_n"] or 0)},
            {"brand_name": b, "stage": "EVER_USED", "n": int(r["ever_used_n"] or 0)},
            {"brand_name": b, "stage": "CURRENT_USER", "n": int(r["current_n"] or 0)},
            {"brand_name": b, "stage": "CONSIDERATION", "n": int(r["consid_n"] or 0)},
            {"brand_name": b, "stage": "PREFERRED", "n": int(r["pref_n"] or 0)},
            {"brand_name": b, "stage": "LAST_PURCHASED", "n": int(r["last_n"] or 0)},
        ])
    return pd.DataFrame(rows), base


@st.cache_data(ttl=3600)
def _get_brand_equity_scores():
    """Composite BEI: TOM×0.25 + CONSIDERATION×0.30 + NPS_norm×0.25 + CSAT_norm×0.20, scaled 0–100."""
    df = _get_competitive_benchmarking()
    if df.empty:
        return df

    def norm(s):
        valid = s.dropna()
        if valid.empty:
            return pd.Series(50.0, index=s.index)
        mn, mx = valid.min(), valid.max()
        if mx == mn:
            return pd.Series(50.0, index=s.index)
        return ((s.fillna(mn) - mn) / (mx - mn) * 100).round(1)

    out = df[["brand_name"]].copy()
    # Total awareness = TOM + SPONT + incremental AIDED (AIDED alone is only incremental)
    _tom   = pd.to_numeric(df["TOM"],   errors="coerce").fillna(0)
    _spont = pd.to_numeric(df["SPONT"], errors="coerce").fillna(0) if "SPONT" in df.columns else 0
    _aided = pd.to_numeric(df["AIDED"], errors="coerce").fillna(0)
    out["aided_pct"] = (_tom + _spont + _aided).round(1)
    # 2026-07-28: found live testing a sparse new project (a brand with zero TOM/CONSIDERATION/
    # NPS/CSAT rows) — `df["TOM"].round(1)` crashed with "type NoneType doesn't define __round__"
    # because a column with no data for some brands comes back as object dtype holding raw `None`
    # (not a proper NaN float), and pandas' `.round()` on object dtype calls Python's built-in
    # `round()` on each element rather than a vectorized numeric round. `_tom` two lines above
    # already goes through `pd.to_numeric(..., errors="coerce")` for exactly this reason (turns
    # None into a real NaN float, which `.round()` handles fine) — these four lines just weren't
    # updated to match when that fix was made, so the same source columns crashed here instead.
    # NOT using `.fillna(0)` here (unlike `_tom`/`_spont`/`_aided` above): those three are being
    # SUMMED into aided_pct, where a missing tier should read as "0 contribution"; these four are
    # standalone DISPLAY/normalization columns, where a genuinely missing NPS/CSAT/TOM value
    # should stay blank (NaN), not be shown as a fake 0.
    _con   = pd.to_numeric(df["CONSIDERATION"], errors="coerce")
    _nps   = pd.to_numeric(df["nps"], errors="coerce")
    _csat  = pd.to_numeric(df["csat"], errors="coerce")
    out["TOM_pct"]           = pd.to_numeric(df["TOM"], errors="coerce").round(1)
    out["CONSIDERATION_pct"] = _con.round(1)
    out["NPS"]               = _nps.round(1)
    out["CSAT"]              = _csat.round(2)
    out["tom_n"]             = norm(pd.to_numeric(df["TOM"], errors="coerce"))
    out["con_n"]             = norm(_con)
    out["nps_n"]             = norm(_nps)
    out["csat_n"]            = norm(_csat)
    out["brand_equity_index"] = (
        out["tom_n"] * 0.25 +
        out["con_n"] * 0.30 +
        out["nps_n"] * 0.25 +
        out["csat_n"] * 0.20
    ).round(1)
    return out.sort_values("brand_equity_index", ascending=False).reset_index(drop=True)


@st.cache_data(ttl=3600)
def _get_funnel_conversion_data():
    """Stage-to-stage conversion rates per brand — two chained funnels sharing TOTAL_AIDED as
    their anchor: the awareness-depth funnel (TOTAL_AIDED → Spontaneous → Consideration → TOM)
    and the behavioral funnel (TOTAL_AIDED → EVER_USED → CONSIDERATION → LAST_PURCHASED → PREFERRED)."""
    df = _get_competitive_benchmarking()
    if df.empty:
        return pd.DataFrame()

    # TOM/SPONT/AIDED are mutually EXCLUSIVE tiers in fact_brand_awareness (each respondent's
    # awareness lands in exactly one), not cumulative counts — TOTAL_AIDED = TOM + SPONT + AIDED
    # (see _get_competitive_benchmarking). The classic brand-tracking awareness funnel is nested
    # and CUMULATIVE (Total Aware ⊇ Spontaneous ⊇ Top-of-Mind), so "Spontaneous" here must be
    # built as TOM + SPONT summed, not read off the raw SPONT column alone — using the raw
    # column would understate spontaneous recall by excluding respondents whose recall was
    # spontaneous AND top-of-mind.
    df = df.copy()
    df["SPONT_CUML"] = (df["TOM"].fillna(0) + df["SPONT"].fillna(0)).round(1)

    rows = []
    for _, row in df.iterrows():
        brand = row["brand_name"]

        # Awareness-depth funnel: how much of total awareness survives at each stricter recall bar.
        # NOTE: CONSIDERATION is an independent survey question (not a nested subset of
        # Spontaneous recall — see comment on the behavioral funnel below), so unlike the
        # TOTAL_AIDED→SPONT→TOM chain, the Spont→Consideration and Consideration→TOM legs are
        # NOT guaranteed monotonic and can show >100% "conversion". Kept per explicit request.
        _aware_stages = [("TOTAL_AIDED", "TOTAL_AIDED"), ("SPONT_CUML", "SPONT"),
                          ("CONSIDERATION", "CONSIDERATION"), ("TOM", "TOM")]
        _aware_labels = ["Total Aided", "Spontaneous", "Consideration", "Top of Mind"]
        for i in range(len(_aware_stages) - 1):
            (c1, s1), (c2, s2) = _aware_stages[i], _aware_stages[i + 1]
            v1, v2 = row.get(c1), row.get(c2)
            if pd.notna(v1) and pd.notna(v2) and v1 > 0:
                rows.append({
                    "brand_name": brand,
                    "transition": f"{s1} → {s2}",
                    "transition_label": f"{_aware_labels[i]} → {_aware_labels[i+1]}",
                    "from_stage": s1, "to_stage": s2,
                    "from_pct": round(float(v1), 1), "to_pct": round(float(v2), 1),
                    "conversion_rate": round(float(v2) / float(v1) * 100, 1),
                })

        # Behavioral funnel — LAST_PURCHASED and PREFERRED swapped per explicit request;
        # both are independent questions (CONSIDERATION can exceed EVER_USED, etc.), not a
        # strictly nested funnel — conversion_rate here is a ratio, not literal drop-off.
        _beh_stages = ["TOTAL_AIDED", "EVER_USED", "CONSIDERATION", "LAST_PURCHASED", "PREFERRED"]
        for i in range(len(_beh_stages) - 1):
            s1, s2 = _beh_stages[i], _beh_stages[i + 1]
            v1 = row[s1] if s1 in row.index else None
            v2 = row[s2] if s2 in row.index else None
            if pd.notna(v1) and pd.notna(v2) and v1 > 0:
                rows.append({
                    "brand_name": brand,
                    "transition": f"{s1} → {s2}",
                    "transition_label": f"{s1} → {s2}",
                    "from_stage": s1, "to_stage": s2,
                    "from_pct": round(float(v1), 1), "to_pct": round(float(v2), 1),
                    "conversion_rate": round(float(v2) / float(v1) * 100, 1),
                })
    return pd.DataFrame(rows)


@st.cache_data(ttl=3600)
def _get_demographic_profile(brand_name: str, zone="all", gender="all", age_band="all", city="all"):
    """Brand consideration % broken down by gender, age_band, and zone vs. overall brand avg.

    Segment filters restrict the respondent base (e.g. gender=Female → age/zone breakdown among women).
    Filtering by the same dimension being charted is degenerate (shows one segment) — expected.
    """
    import sqlite3
    from oxdata.db_loader import get_db_path
    conn = sqlite3.connect(get_db_path())

    # Respondent base restriction from segment filters
    cte, fparams, is_filtered = _resp_filter_cte(zone, gender, age_band, city)
    base_restrict = " AND r.respondent_id IN (SELECT respondent_id FROM _base)" if is_filtered else ""
    cte_prefix = f"WITH {cte} " if is_filtered else ""

    results = {}
    for dim in ["gender", "age_band", "zone_id"]:
        if dim == "zone_id":
            sql = f"""
                {cte_prefix}
                SELECT z.zone_name AS segment,
                       COUNT(DISTINCT r.respondent_id) AS seg_n,
                       COUNT(DISTINCT ba.respondent_id) AS brand_n,
                       ROUND(COUNT(DISTINCT ba.respondent_id)*100.0/COUNT(DISTINCT r.respondent_id), 1) AS brand_pct
                FROM fact_respondents r
                JOIN dim_zone z ON r.zone_id = z.zone_id
                LEFT JOIN (
                    SELECT fa.respondent_id FROM fact_brand_awareness fa
                    JOIN dim_brand b ON fa.brand_id = b.brand_id
                    WHERE fa.stage = 'CONSIDERATION' AND b.brand_name = ?
                ) ba ON r.respondent_id = ba.respondent_id
                WHERE r.zone_id IS NOT NULL{base_restrict}
                GROUP BY z.zone_name
                ORDER BY brand_pct DESC
            """
        else:
            sql = f"""
                {cte_prefix}
                SELECT r.{dim} AS segment,
                       COUNT(DISTINCT r.respondent_id) AS seg_n,
                       COUNT(DISTINCT ba.respondent_id) AS brand_n,
                       ROUND(COUNT(DISTINCT ba.respondent_id)*100.0/COUNT(DISTINCT r.respondent_id), 1) AS brand_pct
                FROM fact_respondents r
                LEFT JOIN (
                    SELECT fa.respondent_id FROM fact_brand_awareness fa
                    JOIN dim_brand b ON fa.brand_id = b.brand_id
                    WHERE fa.stage = 'CONSIDERATION' AND b.brand_name = ?
                ) ba ON r.respondent_id = ba.respondent_id
                WHERE r.{dim} IS NOT NULL AND r.{dim} != ''{base_restrict}
                GROUP BY r.{dim}
                ORDER BY brand_pct DESC
            """
        results[dim] = pd.read_sql(sql, conn, params=fparams + [brand_name])

    # Overall brand consideration % for index calculation (within filtered base)
    if is_filtered:
        overall = pd.read_sql(f"""
            WITH {cte}
            SELECT ROUND(COUNT(DISTINCT ba.respondent_id)*100.0/
                   (SELECT COUNT(DISTINCT respondent_id) FROM _base), 1) AS overall_pct
            FROM fact_brand_awareness ba
            JOIN dim_brand b ON ba.brand_id = b.brand_id
            WHERE ba.stage = 'CONSIDERATION' AND b.brand_name = ?
              AND ba.respondent_id IN (SELECT respondent_id FROM _base)
        """, conn, params=fparams + [brand_name])
    else:
        overall = pd.read_sql("""
            SELECT ROUND(COUNT(DISTINCT ba.respondent_id)*100.0/
                   (SELECT COUNT(DISTINCT respondent_id) FROM fact_respondents), 1) AS overall_pct
            FROM fact_brand_awareness ba
            JOIN dim_brand b ON ba.brand_id = b.brand_id
            WHERE ba.stage = 'CONSIDERATION' AND b.brand_name = ?
        """, conn, params=[brand_name])
    conn.close()

    overall_pct = float(overall.iloc[0, 0]) if not overall.empty and pd.notna(overall.iloc[0, 0]) else 1.0
    for k in results:
        df = results[k]
        if not df.empty:
            df["index_score"] = (df["brand_pct"] / max(overall_pct, 0.1) * 100).round(0).astype(int)
    return results, overall_pct


@st.cache_data(ttl=3600)
def _get_attribute_ownership():
    """For each attribute: brand with highest association %, plus margin over 2nd place."""
    import sqlite3
    from oxdata.db_loader import get_db_path
    conn = sqlite3.connect(get_db_path())
    total = int(pd.read_sql("SELECT COUNT(DISTINCT respondent_id) as n FROM fact_respondents", conn).iloc[0, 0])

    df = pd.read_sql("""
        SELECT da.attr_label, da.broad_feature,
               b.brand_name,
               ROUND(COUNT(DISTINCT bi.respondent_id)*100.0 / ?, 1) AS assoc_pct,
               COUNT(DISTINCT bi.respondent_id) AS n
        FROM fact_brand_imagery bi
        JOIN dim_brand b ON bi.brand_id = b.brand_id
        JOIN dim_bq3_attribute da ON bi.attr_id = da.attr_id
        WHERE b.brand_name NOT IN ('Don''t Know / None')
        GROUP BY da.attr_label, da.broad_feature, b.brand_name
        HAVING n >= 20
    """, conn, params=[total])
    conn.close()

    if df.empty:
        return pd.DataFrame()

    df_sorted = df.sort_values(["attr_label", "assoc_pct"], ascending=[True, False])
    # pandas 3.x: groupby().nth() keeps original integer index; attr_label stays as column
    first  = df_sorted.groupby("attr_label", sort=False).nth(0).reset_index(drop=True)
    second = df_sorted.groupby("attr_label", sort=False).nth(1).reset_index(drop=True)
    second = second[["attr_label", "assoc_pct"]].rename(columns={"assoc_pct": "second_pct"})
    first["attr_label"]  = first["attr_label"].astype(str)
    second["attr_label"] = second["attr_label"].astype(str)
    ownership = first.merge(second, on="attr_label", how="left")
    ownership["margin"] = (ownership["assoc_pct"] - ownership["second_pct"].fillna(0)).round(1)
    return ownership.sort_values(["broad_feature", "assoc_pct"], ascending=[True, False]).reset_index(drop=True)


@st.cache_data(ttl=3600)
def _get_ownership_matrix(top_attrs: int = 20, top_brands: int = 8, attr_order: tuple = None):
    """Attribute × brand ownership matrix — which brand, if any, is SIGNIFICANTLY ahead of
    every other brand on each attribute statement (pooled two-proportion z-test per attribute,
    not just "whoever has the highest raw %").

    Row order: if `attr_order` is given (a tuple of attr_labels, already ranked), use it as-is —
    this is the driver-regression's own importance ranking for whatever Y the caller picked (NPS/
    CSAT/an imagery attribute), since which attributes matter most genuinely changes depending on
    what outcome you're explaining, not a fixed property of the attribute itself. Falls back to
    static respondent-stated importance (bq3a, fact_need_importance) only when no regression has
    been run yet — a reasonable default, not the "real" ranking.
    """
    import sqlite3
    from oxdata.db_loader import get_db_path
    conn = sqlite3.connect(get_db_path())
    total = int(pd.read_sql("SELECT COUNT(DISTINCT respondent_id) as n FROM fact_respondents", conn).iloc[0, 0])

    if attr_order:
        imp_df = pd.DataFrame({"attr_label": list(attr_order)})
    else:
        imp_df = pd.read_sql("""
            SELECT da.attr_label, ROUND(AVG(ni.score), 2) AS mean_importance
            FROM fact_need_importance ni
            JOIN dim_bq3_attribute da ON ni.attr_id = da.attr_id
            GROUP BY da.attr_label
            ORDER BY mean_importance DESC
        """, conn)

    brand_totals = pd.read_sql("""
        SELECT b.brand_name, COUNT(DISTINCT bi.respondent_id) AS n
        FROM fact_brand_imagery bi JOIN dim_brand b ON bi.brand_id = b.brand_id
        WHERE b.brand_name NOT IN ('Don''t Know / None')
        GROUP BY b.brand_name ORDER BY n DESC
    """, conn)
    _top_brand_names = brand_totals.head(top_brands)["brand_name"].tolist()

    assoc = pd.read_sql(f"""
        SELECT da.attr_label, b.brand_name,
               COUNT(DISTINCT bi.respondent_id) AS n,
               ROUND(COUNT(DISTINCT bi.respondent_id)*100.0/?, 1) AS assoc_pct
        FROM fact_brand_imagery bi
        JOIN dim_brand b ON bi.brand_id = b.brand_id
        JOIN dim_bq3_attribute da ON bi.attr_id = da.attr_id
        WHERE b.brand_name IN ({",".join(["?"]*len(_top_brand_names))})
        GROUP BY da.attr_label, b.brand_name
    """, conn, params=[total] + _top_brand_names)
    conn.close()

    if imp_df.empty or assoc.empty:
        return pd.DataFrame(), [], []

    _top_attrs = imp_df.head(top_attrs)["attr_label"].tolist()
    assoc = assoc[assoc["attr_label"].isin(_top_attrs)]

    rows = []
    for attr in _top_attrs:
        sub = assoc[assoc["attr_label"] == attr]
        items = [(r["brand_name"], int(r["n"]), total) for _, r in sub.iterrows()]
        if len(items) < 2:
            continue
        letters, beats, pct = _sig_letters_proportions(items, alpha=0.05)
        if not pct:
            continue
        _top_brand = max(pct, key=pct.get)
        _n_beaten = len(beats.get(_top_brand, []))
        _n_others = len(items) - 1
        # "++" = significantly ahead of EVERY other brand shown (uniquely, clearly owned).
        # "+"  = leads but doesn't clear every rival at 95% (partial/contested ownership).
        mark = "++" if (_n_others > 0 and _n_beaten == _n_others) else ("+" if _n_beaten > 0 else "")
        row = {"Attribute": attr, "_owner": _top_brand if mark else None, "_mark": mark}
        for b in _top_brand_names:
            row[b] = pct.get(b, None)
        rows.append(row)

    matrix = pd.DataFrame(rows)
    return matrix, _top_attrs, _top_brand_names


# ── Render: Competitive Benchmarking Panel ───────────────────────────────────

def _render_competitive_benchmarking(sel_brand: str, min_base_n: int = 30,
                                     zone="all", gender="all", age_band="all", city="all"):
    df = _get_competitive_benchmarking(zone=zone, gender=gender, age_band=age_band, city=city)
    if df is None or df.empty:
        st.info("Competitive data unavailable.")
        return

    DISPLAY_COLS = {
        "brand_name":     "Brand",
        "TOTAL_AIDED":    "Total Aware %",
        "SPONT":          "Spontaneous %",
        "TOM":            "Top of Mind %",
        "EVER_USED":      "Ever Tried %",
        "CURRENT_USER":   "Current Usage %",
        "CONSIDERATION":  "Consideration %",
        "PREFERRED":      "Preferred %",
        "LAST_PURCHASED": "Last Purchased %",
        "nps":            "NPS Score",
        "csat":           "CSAT (0–10)",
    }
    cb_t1, cb_t2, cb_t3, cb_t4 = st.tabs(
        ["📊 Heatmap Table", "🏆 Metric Rankings", "📊 Stage Coverage", "🔬 Significance Test"])

    with cb_t1:
        show_df = df[[c for c in DISPLAY_COLS if c in df.columns]].copy()
        show_df = show_df.rename(columns=DISPLAY_COLS)

        metric_cols = [v for k, v in DISPLAY_COLS.items() if k != "brand_name"]
        avail_metric_cols = [c for c in metric_cols if c in show_df.columns]

        # Ensure numeric types for gradient
        for _mc in avail_metric_cols:
            show_df[_mc] = pd.to_numeric(show_df[_mc], errors="coerce")

        def _cell_color(val, col_min, col_max, is_nps=False):
            if pd.isna(val):
                return "background-color: #f3f4f6; color: #9ca3af;"
            if col_max == col_min:
                ratio = 0.5
            else:
                ratio = (val - col_min) / (col_max - col_min)
            if is_nps:
                # Red–amber–green: negative=red, zero=amber, positive=green
                if val < 0:
                    r = int(220 + (1 + val / 20) * 35) if val >= -20 else 220
                    return f"background-color: rgba(239,68,68,{0.3+ratio*0.5:.2f}); color: #7f1d1d; font-weight:600;"
                elif val < 45:
                    return f"background-color: rgba(251,191,36,{0.25+ratio*0.4:.2f}); color: #78350f; font-weight:600;"
                else:
                    return f"background-color: rgba(34,197,94,{0.3+ratio*0.4:.2f}); color: #14532d; font-weight:600;"
            else:
                intensity = 0.15 + ratio * 0.65
                return f"background-color: rgba(26,93,77,{intensity:.2f}); color: {'white' if intensity > 0.45 else '#1a5d4d'}; font-weight:600;"

        def _style_row(row):
            styles = []
            for col in row.index:
                if col == "Brand":
                    if row["Brand"] == sel_brand:
                        styles.append("background-color: #1a5d4d; color: #e5e7eb; font-weight:800;")
                    else:
                        styles.append("font-weight:500;")
                elif col in avail_metric_cols:
                    is_nps = col == "NPS Score"
                    col_min = show_df[col].min()
                    col_max = show_df[col].max()
                    styles.append(_cell_color(row[col], col_min, col_max, is_nps))
                else:
                    styles.append("")
            return pd.Series(styles, index=row.index)

        # Suppress cells where AIDED respondent count is below min_base_n
        if "aided_n" in df.columns:
            _aided_n_map = df.set_index("brand_name")["aided_n"].to_dict()
            for col in [c for c in avail_metric_cols if c not in ("NPS Score", "CSAT (0–10)")]:
                show_df[col] = show_df.apply(
                    lambda row: row[col] if _aided_n_map.get(row["Brand"], 999) >= min_base_n else float("nan"),
                    axis=1,
                )

        styled = (
            show_df.style
            .apply(_style_row, axis=1)
            .format({c: "{:.1f}" for c in avail_metric_cols}, na_rep="—")
        )
        st.dataframe(styled, use_container_width=True, hide_index=True)
        st.caption(f"Darker = higher value. Dark green row = selected brand. Min {min_base_n} respondents (cells below threshold shown as —). NPS: red (<0) · amber (0–45) · green (>45).")

    with cb_t2:
        metric_opts = [c for c in ["TOTAL_AIDED", "SPONT", "TOM", "EVER_USED", "CURRENT_USER", "CONSIDERATION", "PREFERRED", "LAST_PURCHASED", "nps", "csat"]
                       if c in df.columns]
        metric_labels = {"TOTAL_AIDED": "Total Awareness %", "SPONT": "Spontaneous Recall %", "TOM": "Top of Mind %",
                         "EVER_USED": "Ever Tried %", "CURRENT_USER": "Current Usage %",
                         "CONSIDERATION": "Consideration %", "PREFERRED": "Preference %",
                         "LAST_PURCHASED": "Last Purchased %",
                         "nps": "NPS Score", "csat": "CSAT (0–10)"}
        sel_metric = st.selectbox("Sort and rank by:", metric_opts,
                                  format_func=lambda x: metric_labels.get(x, x),
                                  key="cb_rank_metric")
        rank_df = df[["brand_name", sel_metric]].dropna().sort_values(sel_metric, ascending=False).head(20)
        colors = ["#1a5d4d" if b == sel_brand else "#30a76a" for b in rank_df["brand_name"]]
        fig = go.Figure(go.Bar(
            x=rank_df[sel_metric], y=rank_df["brand_name"],
            orientation="h",
            marker_color=list(reversed(colors)),
            text=rank_df[sel_metric].round(1),
            textposition="outside",
        ))
        base = {k: v for k, v in _chart_layout_base(420).items() if k not in ("xaxis","yaxis","legend","margin")}
        fig.update_layout(**base,
                          xaxis_title=metric_labels.get(sel_metric, sel_metric),
                          yaxis=dict(autorange="reversed", automargin=True),
                          margin=dict(l=140, r=80, t=40, b=40),
                          title=dict(text=f"Brand Rankings — {metric_labels.get(sel_metric, sel_metric)}", font=dict(size=13)))
        st.plotly_chart(_theme_fig(fig), use_container_width=True)

    with cb_t3:
        # Stage coverage: independent survey questions, NOT a sequential funnel.
        # Total Awareness = TOM + Spontaneous + incremental Aided.
        _cb_df = df.copy()
        top_brands_df = _cb_df.dropna(subset=["TOTAL_AIDED"]).nlargest(12, "TOTAL_AIDED")
        show_stages = [s for s in ["TOTAL_AIDED", "SPONT", "TOM", "EVER_USED", "CURRENT_USER", "CONSIDERATION", "PREFERRED"] if s in _cb_df.columns]
        show_labels = {
            "TOTAL_AIDED": "Total Awareness",
            "SPONT": "Spontaneous",
            "TOM": "Top of Mind",
            "EVER_USED": "Ever Tried",
            "CURRENT_USER": "Current Usage",
            "CONSIDERATION": "Consideration",
            "PREFERRED": "Preferred",
        }
        st.caption("⚠️ These are **independent survey questions**, not a sequential funnel. "
                   "Total Awareness = TOM + Spontaneous + Aided. "
                   "Consideration/Usage stages are asked to all respondents regardless of awareness.")
        fig = go.Figure()
        for _, row in top_brands_df.iterrows():
            brand = row["brand_name"]
            _y = [row[s] if s in row.index and pd.notna(row[s]) else None for s in show_stages]
            _x = [show_labels.get(s, s) for s in show_stages]
            color = "#1a5d4d" if brand == sel_brand else None
            width = 3 if brand == sel_brand else 1.5
            fig.add_trace(go.Bar(
                name=brand, x=_x, y=_y,
                marker_color=color if color else "#93c5fd",
                opacity=1.0 if brand == sel_brand else 0.6,
                text=[f"{v:.1f}%" if v is not None else "" for v in _y],
                textposition="outside",
                hovertemplate=f"<b>{brand}</b><br>%{{x}}: %{{y:.1f}}%<extra></extra>",
            ))
        base = {k: v for k, v in _chart_layout_base(450).items() if k not in ("xaxis","yaxis","legend")}
        fig.update_layout(**base, barmode="group",
                          yaxis_title="% of All Respondents",
                          legend=dict(orientation="h", yanchor="bottom", y=-0.38, xanchor="center", x=0.5),
                          title=dict(text="Stage Coverage — Top 12 Brands by Total Awareness", font=dict(size=13)))
        st.plotly_chart(_theme_fig(fig), use_container_width=True)

    with cb_t4:
        # ── Column-proportion significance test (XLSTAT / pValue-style sig letters) ──
        st.markdown("**Pairwise significance — which brands are *significantly* ahead?**")
        st.caption(
            "Two-proportion z-test between every pair of brands on the chosen awareness metric "
            "(pooled variance, two-sided). Each brand gets a letter (A = highest). A brand's "
            "**Sig. higher than** column lists the letters of the brands it significantly beats at the "
            "chosen confidence — the standard XLSTAT / pValue column-comparison table."
        )
        sig_c1, sig_c2 = st.columns([2, 1])
        _sig_stage_opts = {
            "TOTAL_AIDED": "Total Awareness",
            "SPONT": "Spontaneous Recall",
            "TOM": "Top of Mind",
            "EVER_USED": "Ever Tried",
            "CURRENT_USER": "Current Usage",
            "CONSIDERATION": "Consideration",
            "PREFERRED": "Preferred",
            "LAST_PURCHASED": "Last Purchased",
        }
        with sig_c1:
            _sig_stage = st.selectbox("Metric", list(_sig_stage_opts.keys()),
                                      format_func=lambda s: _sig_stage_opts[s],
                                      index=0, key="cb_sig_stage")
        with sig_c2:
            _sig_conf = st.selectbox("Confidence", [0.90, 0.95, 0.99], index=1,
                                     format_func=lambda a: f"{a:.0%}", key="cb_sig_conf")
        _sig_alpha = round(1 - _sig_conf, 2)

        # Counts are computed from the ACTIVE segment filters — fetch first so the
        # segment + base are shown explicitly (makes filter application visible).
        _counts, _sig_base = _get_brand_stage_counts(zone, gender, age_band, city)
        _is_filt = any(x != "all" for x in (zone, gender, age_band, city))
        _sig_seg = " · ".join(
            v for v in [
                None if zone == "all" else zone,
                None if gender == "all" else gender,
                None if age_band == "all" else age_band,
                None if city == "all" else city,
            ] if v
        ) or "All-India"
        st.markdown(
            f"<div style='display:flex;gap:10px;align-items:center;margin:2px 0 8px;'>"
            f"<span style='background:{'#1a5d4d' if _is_filt else '#6b7280'};color: #e5e7eb;"
            f"font-size:0.66rem;font-weight:800;padding:3px 12px;border-radius:20px;'>"
            f"SEGMENT · {_h_escape(_sig_seg)}</span>"
            f"<span style='font-size:0.74rem;color:#6b7280;'>Filtered base N = "
            f"<b>{_sig_base:,}</b> respondents{' (filters applied)' if _is_filt else ''}</span></div>",
            unsafe_allow_html=True,
        )
        # Adaptive min-N: cap to the available base so a small segment can't gate everything out.
        _minn_cap = max(20, min(200, _sig_base // 10)) if _sig_base else 200
        _minn_def = min(30, _minn_cap)
        _sig_minn = st.slider("Min base N per brand", 10, _minn_cap, _minn_def, 5, key="cb_sig_minn",
                              help="Exclude brands with too few respondents to test reliably. "
                                   "Capped to the filtered segment's base.")

        _stg = _counts[_counts["stage"] == _sig_stage]
        if _stg.empty or _sig_base < 1:
            st.info(f"No respondents for **{_sig_stage_opts[_sig_stage]}** in segment **{_sig_seg}** "
                    f"(base N={_sig_base:,}). Widen the filters or pick another metric.")
        else:
            # base is common (awareness asked of all); filter brands by their own count >= min N
            _n_with_data = len(_stg)
            _items = [(r["brand_name"], int(r["n"]), _sig_base)
                      for _, r in _stg.iterrows() if int(r["n"]) >= _sig_minn]
            if len(_items) < 2:
                _maxn = int(_stg["n"].max()) if not _stg.empty else 0
                st.info(f"Only {len(_items)} brand(s) have ≥{_sig_minn} respondents on "
                        f"**{_sig_stage_opts[_sig_stage]}** in segment **{_sig_seg}** "
                        f"(largest brand has {_maxn}). Lower the Min base N slider or widen the filters.")
            else:
                _letters, _beats, _pct = _sig_letters_proportions(_items, alpha=_sig_alpha)
                _rows = []
                for lbl in sorted(_pct, key=lambda x: -_pct[x]):
                    # preserve rank order from the test (do NOT alpha-sort — breaks
                    # multi-letter labels AA/AB…); space-separate for readability
                    beats = " ".join(_beats[lbl])
                    _rows.append({
                        "": _letters[lbl],
                        "Brand": lbl,
                        f"{_sig_stage_opts[_sig_stage]} %": round(_pct[lbl], 1),
                        "Sig. higher than": beats if beats else "—",
                        "# beaten": len(_beats[lbl]),
                    })
                _sig_df = pd.DataFrame(_rows)

                def _hl_sig(row):
                    base_style = ("background-color:#1a5d4d;color: #e5e7eb;font-weight:800;"
                                  if row["Brand"] == sel_brand else "")
                    return pd.Series([base_style] * len(row), index=row.index)

                _styled_sig = (
                    _sig_df.style
                    .apply(_hl_sig, axis=1)
                    .background_gradient(subset=[f"{_sig_stage_opts[_sig_stage]} %"], cmap="Greens")
                    .format({f"{_sig_stage_opts[_sig_stage]} %": "{:.1f}", "# beaten": "{:d}"})
                )
                st.dataframe(_styled_sig, use_container_width=True, hide_index=True)

                # Headline read-out
                _leader = _sig_df.iloc[0]
                _n_beaten = int(_leader["# beaten"])
                _findings = [
                    f"**{_leader['Brand']}** (letter {_leader['']}) leads on "
                    f"**{_sig_stage_opts[_sig_stage]}** at {_leader[f'{_sig_stage_opts[_sig_stage]} %']:.1f}% — "
                    f"significantly ahead of **{_n_beaten}** of {len(_items)-1} rival(s) at {_sig_conf:.0%} confidence.",
                ]
                if sel_brand in _pct:
                    _sb_beats = len(_beats[sel_brand])
                    _sb_letter = _letters[sel_brand]
                    _findings.append(
                        f"**{sel_brand}** (letter {_sb_letter}) at {_pct[sel_brand]:.1f}% is significantly "
                        f"higher than {_sb_beats} rival(s)."
                    )
                _findings.append(
                    f"Base = {_sig_base:,} respondents ({_sig_seg}). Letters share = no significant "
                    f"difference at {_sig_conf:.0%}. Test: pooled two-proportion z, α={_sig_alpha}."
                )
                _insight_callout(_findings, "Significance Read-out", "🔬")


# ── Render: Brand Equity Index ───────────────────────────────────────────────

def _render_brand_equity_index(sel_brand: str):
    df = _get_brand_equity_scores()
    if df is None or df.empty:
        st.info("Brand equity data unavailable.")
        return

    bei_t1, bei_t2 = st.tabs(["🏅 BEI Ranking", "📐 Component Breakdown"])

    with bei_t1:
        col_gauge, col_rank = st.columns([1, 2])
        with col_gauge:
            brand_row = df[df["brand_name"] == sel_brand]
            bei_score = float(brand_row["brand_equity_index"].iloc[0]) if not brand_row.empty else 0.0
            rank = int(df[df["brand_equity_index"] > bei_score].shape[0]) + 1
            n_total = len(df)
            market_avg_bei = float(df["brand_equity_index"].mean()) if not df.empty else 50.0

            fig_gauge = go.Figure(go.Indicator(
                mode="gauge+number+delta",
                value=bei_score,
                delta={
                    "reference": market_avg_bei,
                    "increasing": {"color": "#30a76a"},
                    "decreasing": {"color": "#e53e3e"},
                    "suffix": f" vs mkt avg {market_avg_bei:.0f}",
                },
                gauge={
                    "axis": {"range": [0, 100], "tickwidth": 1},
                    "bar": {"color": "#1a5d4d"},
                    "steps": [
                        {"range": [0, 33],  "color": "#fee2e2"},
                        {"range": [33, 66], "color": "#fef9c3"},
                        {"range": [66, 100],"color": "#dcfce7"},
                    ],
                    "threshold": {"line": {"color": "#374151", "width": 3}, "value": market_avg_bei},
                },
                title={"text": f"{sel_brand}<br><span style='font-size:11px'>Brand Equity Index</span>"},
                number={"suffix": "/100"},
            ))
            fig_gauge.update_layout(height=280, margin=dict(t=40, b=20, l=20, r=20))
            st.plotly_chart(_theme_fig(fig_gauge), use_container_width=True)
            st.markdown(
                f"<div style='text-align:center;font-size:0.85rem;color:#6b7280'>"
                f"Ranked <b>{rank}</b> of {n_total} brands &nbsp;·&nbsp; "
                f"Market avg: <b>{market_avg_bei:.1f}</b></div>",
                unsafe_allow_html=True,
            )

        with col_rank:
            # Top N by aided awareness, then ordered by BEI for display
            _aided_col = "aided_pct" if "aided_pct" in df.columns else "brand_equity_index"
            top_df = df.dropna(subset=[_aided_col]).nlargest(15, _aided_col).sort_values(
                "brand_equity_index", ascending=False
            )
            # Always include selected brand
            if sel_brand not in top_df["brand_name"].values:
                _sb_row = df[df["brand_name"] == sel_brand]
                if not _sb_row.empty:
                    top_df = pd.concat([top_df, _sb_row]).drop_duplicates("brand_name")
            bar_colors = ["#1a5d4d" if b == sel_brand else "#30a76a" for b in top_df["brand_name"]]
            fig_rank = go.Figure(go.Bar(
                x=top_df["brand_name"], y=top_df["brand_equity_index"],
                marker_color=bar_colors,
                text=top_df["brand_equity_index"].round(1),
                textposition="outside",
                hovertemplate="<b>%{x}</b><br>BEI: %{y:.1f}<extra></extra>",
            ))
            base = {k: v for k, v in _chart_layout_base(300).items() if k not in ("xaxis","yaxis","legend")}
            fig_rank.update_layout(**base,
                                   yaxis_title="Brand Equity Index (0–100)",
                                   xaxis=dict(tickangle=-35),
                                   title=dict(text="Brand Equity Index — Top 15 by Aided Awareness, Ranked by BEI",
                                              font=dict(size=13)))
            st.plotly_chart(_theme_fig(fig_rank), use_container_width=True)

        st.caption("BEI formula: Top-of-Mind% × 0.25 + Consideration% × 0.30 + NPS (norm) × 0.25 + CSAT (norm) × 0.20. Each component normalized 0–100 within dataset before weighting.")

    with bei_t2:
        brand_row = df[df["brand_name"] == sel_brand]
        if brand_row.empty:
            st.info(f"No BEI data for {sel_brand}.")
            return

        def _safe_float(val, default=0.0):
            try:
                return float(val) if pd.notna(val) else default
            except (TypeError, ValueError):
                return default

        # has_data must reflect MISSING (NaN), not zero — NPS 0 / CSAT 0 are valid values.
        components = {
            "Top of Mind":   ("TOM_pct",          "tom_n",  "25% weight", "#1a5d4d"),
            "Consideration": ("CONSIDERATION_pct", "con_n", "30% weight", "#30a76a"),
            "NPS":           ("NPS",               "nps_n", "25% weight", "#059669"),
            "CSAT":          ("CSAT",              "csat_n","20% weight", "#10b981"),
        }
        c1, c2, c3, c4 = st.columns(4)
        for col, (label, (raw_col, norm_col, weight, color)) in zip([c1, c2, c3, c4], components.items()):
            _raw_cell = brand_row[raw_col].iloc[0]
            has_data = pd.notna(_raw_cell)
            raw_val  = _safe_float(_raw_cell)
            norm_val = _safe_float(brand_row[norm_col].iloc[0])
            _cs = _get_csat_scale()
            unit = f"/{_cs}" if label == "CSAT" else ("pts" if label == "NPS" else "%")
            col.metric(
                label=f"{label} ({weight})",
                value=f"{raw_val:.1f}{unit}" if has_data else "N/A",
                delta=f"Norm: {norm_val:.0f}/100" if has_data else "No data",
                delta_color="normal" if has_data else "off",
            )

        # Spider/radar: selected brand vs top 5 by aided awareness
        _aided_col_r = "aided_pct" if "aided_pct" in df.columns else "brand_equity_index"
        top5 = df.dropna(subset=[_aided_col_r]).nlargest(5, _aided_col_r)
        if sel_brand not in top5["brand_name"].values:
            _sb_r = df[df["brand_name"] == sel_brand]
            if not _sb_r.empty:
                top5 = pd.concat([top5.head(4), _sb_r]).drop_duplicates("brand_name")
        comp_cols = ["tom_n", "con_n", "nps_n", "csat_n"]
        comp_labels = ["TOM (norm)", "Consideration (norm)", "NPS (norm)", "CSAT (norm)"]
        fig_radar = go.Figure()
        for _, row in top5.iterrows():
            brand = row["brand_name"]
            vals = [float(row[c]) if pd.notna(row[c]) else 0 for c in comp_cols]
            vals += [vals[0]]  # close polygon
            fig_radar.add_trace(go.Scatterpolar(
                r=vals, theta=comp_labels + [comp_labels[0]],
                name=brand,
                line=dict(width=3 if brand == sel_brand else 1.5,
                          color="#1a5d4d" if brand == sel_brand else None),
                fill="toself" if brand == sel_brand else None,
                fillcolor="rgba(26,93,77,0.1)" if brand == sel_brand else None,
            ))
        fig_radar.update_layout(
            polar=dict(radialaxis=dict(visible=True, range=[0, 100])),
            showlegend=True,
            height=380,
            legend=dict(orientation="h", yanchor="bottom", y=-0.25, xanchor="center", x=0.5),
            title=dict(text=f"BEI Components — {sel_brand} vs Top Brands", font=dict(size=13)),
            margin=dict(t=50, b=80, l=60, r=60),
        )
        st.plotly_chart(_theme_fig(fig_radar), use_container_width=True)


# ── Render: Funnel Conversion & Leakage ──────────────────────────────────────

def _render_funnel_leakage(sel_brand: str, zone="all", gender="all", age_band="all", city="all"):
    bench_df = _get_competitive_benchmarking()
    conv_df  = _get_funnel_conversion_data()
    if bench_df is None or bench_df.empty:
        st.info("Funnel data unavailable.")
        return

    fl_t1, fl_t2, fl_t3 = st.tabs([f"🔍 {sel_brand} Funnel Detail", "⚔️ Competitor Conversion Comparison",
                                   "📐 Base vs Up-to-6 Brands (significance)"])

    with fl_t1:
        brand_row = bench_df[bench_df["brand_name"] == sel_brand]
        if brand_row.empty:
            st.info(f"No funnel data for {sel_brand}.")
        else:
            row = brand_row.iloc[0]
            stages       = ["TOTAL_AIDED", "EVER_USED", "CONSIDERATION", "LAST_PURCHASED", "PREFERRED"]
            stage_labels = ["Total Aware", "Ever Used", "Consideration", "Last Purchased", "Preferred"]
            vals     = [row[s] if s in row.index else None for s in stages]
            cat_avgs = [bench_df[s].mean() for s in stages]

            # ── Awareness depth: Total Aided → Spontaneous → Consideration → Top of Mind ──
            # Was only visible buried in the Conversion Detail expander + the Competitor
            # Conversion Comparison dropdown further down — surfaced here explicitly since
            # it's one of the most standard brand-tracking metrics and was easy to miss.
            _aw_stage_keys = ["TOTAL_AIDED → SPONT", "SPONT → CONSIDERATION", "CONSIDERATION → TOM"]
            _aw_stage_lbls = ["Total Aided → Spontaneous", "Spontaneous → Consideration", "Consideration → Top of Mind"]
            _aw_cols = st.columns(3)
            for _awi, (_key, _lbl) in enumerate(zip(_aw_stage_keys, _aw_stage_lbls)):
                _rate, _cat_rate = None, None
                if not conv_df.empty:
                    _bc0 = conv_df[(conv_df["brand_name"] == sel_brand) & (conv_df["transition"] == _key)]
                    _rate = float(_bc0["conversion_rate"].iloc[0]) if not _bc0.empty else None
                    _cat_rate = conv_df[conv_df["transition"] == _key]["conversion_rate"].mean()
                if _rate is None:
                    continue
                _gap = (_rate - _cat_rate) if pd.notna(_cat_rate) else None
                _col_c = "#15803d" if (_gap or 0) > 5 else "#dc2626" if (_gap or 0) < -5 else "#6b7280"
                _gap_str = (f"<span style='color:{_col_c};font-size:0.65rem;'>"
                           f"{'+' if (_gap or 0) > 0 else ''}{_gap:.0f}pp vs cat</span>" if _gap is not None else "")
                _aw_cols[_awi].markdown(
                    f"<div style='text-align:center;border:1px solid #e5e7eb;border-radius:8px;"
                    f"padding:8px 4px;'>"
                    f"<div style='font-size:0.58rem;color:#9ca3af;text-transform:uppercase;"
                    f"font-weight:700;letter-spacing:0.06em;'>{_lbl}</div>"
                    f"<div style='font-size:1.2rem;font-weight:900;color:#1a5d4d;'>{_rate:.0f}%</div>"
                    f"{_gap_str}</div>",
                    unsafe_allow_html=True,
                )
            st.caption(
                "Awareness depth — of everyone aided-aware, what % recall spontaneously (unprompted); "
                "of those spontaneous recallers, what % have this brand top-of-mind (first-mentioned). "
                "Cumulative, not the raw exclusive-tier %s (Spontaneous here = TOM + SPONT combined)."
            )
            st.markdown("<div style='height:6px;'></div>", unsafe_allow_html=True)

            # ── KPI row: conversion rates between stages ──────────────────────
            _conv_pairs = list(zip(range(len(stages)-1), range(1, len(stages))))
            _conv_cols  = st.columns(len(_conv_pairs))
            _avg_conv   = conv_df.groupby("transition")["conversion_rate"].mean().to_dict() if not conv_df.empty else {}
            _brand_conv_map = {}
            if not conv_df.empty:
                _bc = conv_df[conv_df["brand_name"] == sel_brand]
                _brand_conv_map = _bc.set_index("transition")["conversion_rate"].to_dict()

            for _ci, (fi, ti) in enumerate(_conv_pairs):
                _from_v, _to_v = vals[fi], vals[ti]
                _trans_key = f"{stages[fi]} → {stages[ti]}"
                _conv_rate = _brand_conv_map.get(_trans_key, ((_to_v/_from_v*100) if _from_v and _from_v > 0 and _to_v is not None else None))
                _cat_c = _avg_conv.get(_trans_key)
                if _conv_rate is not None:
                    _gap = (_conv_rate - _cat_c) if _cat_c is not None else None
                    _col_c = "#15803d" if (_gap or 0) > 5 else "#dc2626" if (_gap or 0) < -5 else "#6b7280"
                    _gap_str = f"<span style='color:{_col_c};font-size:0.65rem;'>{'+' if (_gap or 0)>0 else ''}{_gap:.0f}pp vs cat</span>" if _gap is not None else ""
                    _conv_cols[_ci].markdown(
                        f"<div style='text-align:center;border:1px solid #e5e7eb;border-radius:8px;"
                        f"padding:8px 4px;'>"
                        f"<div style='font-size:0.58rem;color:#9ca3af;text-transform:uppercase;"
                        f"font-weight:700;letter-spacing:0.06em;'>{stage_labels[fi][:4]}→{stage_labels[ti][:4]}</div>"
                        f"<div style='font-size:1.2rem;font-weight:900;color:#1a5d4d;'>{_conv_rate:.0f}%</div>"
                        f"{_gap_str}</div>",
                        unsafe_allow_html=True,
                    )

            # ── Dual-bar chart: brand vs category avg per stage ───────────────
            valid = [(lbl, v, ca) for lbl, v, ca in zip(stage_labels, vals, cat_avgs)
                     if pd.notna(v) and pd.notna(ca)]
            if valid:
                _lbls, _bvals, _cavals = zip(*valid)
                fig = go.Figure()
                fig.add_trace(go.Bar(
                    name=sel_brand, x=list(_lbls), y=list(_bvals),
                    marker_color="#1a5d4d",
                    text=[f"{v:.1f}%" for v in _bvals], textposition="outside",
                    hovertemplate="<b>%{x}</b><br>" + sel_brand + ": %{y:.1f}%<extra></extra>",
                ))
                fig.add_trace(go.Bar(
                    name="Category Avg", x=list(_lbls), y=list(_cavals),
                    marker_color="#cbd5e1",
                    text=[f"{v:.1f}%" for v in _cavals], textposition="outside",
                    hovertemplate="<b>%{x}</b><br>Category avg: %{y:.1f}%<extra></extra>",
                ))
                base = {k: v for k, v in _chart_layout_base(340).items()
                        if k not in ("xaxis", "yaxis", "legend")}
                fig.update_layout(
                    **base,
                    barmode="group",
                    yaxis_title="% of All Respondents",
                    legend=dict(orientation="h", y=-0.18, x=0.5, xanchor="center"),
                    title=dict(text=f"{sel_brand} vs Category Average — Funnel Penetration",
                               font=dict(size=13)),
                )
                st.plotly_chart(_theme_fig(fig), use_container_width=True)

            # ── Conversion detail table ────────────────────────────────────────
            brand_conv = conv_df[conv_df["brand_name"] == sel_brand] if not conv_df.empty else pd.DataFrame()
            cat_conv_avg = conv_df.groupby("transition")["conversion_rate"].mean().reset_index() if not conv_df.empty else pd.DataFrame()

            if not brand_conv.empty and not cat_conv_avg.empty:
                conv_show = brand_conv[["transition", "from_pct", "to_pct", "conversion_rate"]].merge(
                    cat_conv_avg.rename(columns={"conversion_rate": "cat_avg_conv"}),
                    on="transition", how="left",
                )
                conv_show.columns = ["Transition", "From %", "To %", "Conversion %", "Cat Avg Conv %"]
                conv_show["Gap vs Cat"] = (conv_show["Conversion %"] - conv_show["Cat Avg Conv %"]).round(1)

                def color_gap(val):
                    if pd.isna(val): return ""
                    if val > 5: return "color: #15803d; font-weight:bold"
                    if val < -5: return "color: #dc2626; font-weight:bold"
                    return "color: #6b7280"

                styled = conv_show.style.format({
                    "From %": "{:.1f}", "To %": "{:.1f}",
                    "Conversion %": "{:.1f}", "Cat Avg Conv %": "{:.1f}", "Gap vs Cat": "{:+.1f}",
                }).map(color_gap, subset=["Gap vs Cat"])
                with st.expander("📋 Conversion Detail Table", expanded=False):
                    st.dataframe(styled, hide_index=True, use_container_width=True)
                    st.caption("Conversion % = (To Stage %) / (From Stage %) × 100. Green gap = above category, red = below.")

    with fl_t2:
        if conv_df.empty:
            st.info("Conversion comparison data unavailable.")
            return
        transitions = conv_df["transition"].unique().tolist()
        sel_transition = st.selectbox("Select funnel transition:", transitions, key="fl_transition")
        sub = conv_df[conv_df["transition"] == sel_transition].sort_values("conversion_rate", ascending=False).head(15)
        colors = ["#1a5d4d" if b == sel_brand else "#30a76a" for b in sub["brand_name"]]
        cat_avg = sub["conversion_rate"].mean()

        fig2 = go.Figure()
        fig2.add_trace(go.Bar(
            x=sub["brand_name"], y=sub["conversion_rate"],
            marker_color=colors,
            text=sub["conversion_rate"].apply(lambda v: f"{v:.1f}%"),
            textposition="outside",
            hovertemplate="<b>%{x}</b><br>Conversion: %{y:.1f}%<extra></extra>",
        ))
        fig2.add_hline(y=cat_avg, line_dash="dash", line_color="#6b7280",
                       annotation_text=f"Cat avg: {cat_avg:.1f}%",
                       annotation_position="top right")
        base = {k: v for k, v in _chart_layout_base(380).items() if k not in ("xaxis","yaxis","legend")}
        fig2.update_layout(**base,
                           yaxis_title="Conversion Rate %",
                           xaxis=dict(tickangle=-35),
                           title=dict(text=f"Conversion Rate — {sel_transition}", font=dict(size=13)))
        st.plotly_chart(_theme_fig(fig2), use_container_width=True)
        st.caption(f"Conversion = (Stage 2 %) / (Stage 1 %) × 100. Dashed line = category average {cat_avg:.1f}%.")

    with fl_t3:
        st.markdown(f"**{sel_brand} (base) vs up to 6 brands — full funnel, with significance per stage**")
        st.caption(
            "Left = base brand's funnel. Right = up to 6 comparison brands, 2 rows × 3 columns, "
            "same stages. Each comparison brand's bar is flagged if it's significantly higher or "
            "lower than the base brand AT THAT STAGE (pooled two-proportion z-test, common base — "
            "the same test used in the Significance Test tab above)."
        )
        # Controls get their own full-width row — NOT the same column split as the chart
        # layout below, otherwise the 1:1 split from a 2-control row squeezes the 2x3 grid
        # (which needs 3x the width of the single base-brand chart) into half the page.
        _ctrl_c1, _ctrl_c2 = st.columns([2, 1])
        with _ctrl_c1:
            _all_brands_grid = sorted(bench_df["brand_name"].unique().tolist())
            _other_brands = [b for b in _all_brands_grid if b != sel_brand]
            _default_cmp = (bench_df[bench_df["brand_name"] != sel_brand]
                            .sort_values("TOTAL_AIDED", ascending=False)["brand_name"].head(6).tolist())
            # Widget key is scoped to sel_brand — without this, switching the base brand (e.g.
            # Crompton -> Bajaj) can leave a previously-selected brand in session_state that is
            # no longer valid once it becomes the new base and drops out of _other_brands,
            # which breaks/resets the multiselect unpredictably. A fresh key per base brand
            # means each base gets its own clean default instead of colliding with stale state.
            _cmp_brands = st.multiselect(
                "Compare against (up to 6)", _other_brands, default=_default_cmp,
                max_selections=6, key=f"fl_grid_brands_{sel_brand}",
            )
        with _ctrl_c2:
            _grid_conf = st.selectbox("Flag differences at", [0.95, 0.90, 0.80], index=0,
                                      format_func=lambda a: f"{a:.0%}", key="fl_grid_conf")
        _grid_alpha = round(1 - _grid_conf, 2)

        _grid_stages = ["TOTAL_AIDED", "EVER_USED", "CONSIDERATION", "LAST_PURCHASED", "PREFERRED"]
        _grid_labels = ["Total Aided", "Ever Used", "Consideration", "Last Purchased", "Preferred"]

        _gcounts, _gbase = _get_brand_stage_counts(zone, gender, age_band, city)

        def _stage_count(brand: str, stage: str) -> int:
            if stage == "TOTAL_AIDED":
                return int(_gcounts[(_gcounts["brand_name"] == brand)
                                    & (_gcounts["stage"].isin(["TOM", "SPONT", "AIDED"]))]["n"].sum())
            _row = _gcounts[(_gcounts["brand_name"] == brand) & (_gcounts["stage"] == stage)]
            return int(_row["n"].iloc[0]) if not _row.empty else 0

        def _prop_sig(cnt_a, cnt_b, base):
            # Pooled two-proportion z — same formula as _sig_letters_proportions, direct
            # base-vs-one-brand comparison instead of all-pairs letters.
            if base < 1:
                return ""
            import numpy as _np
            from scipy.stats import norm as _norm
            pa, pb = cnt_a / base, cnt_b / base
            pooled = (cnt_a + cnt_b) / (2 * base)
            se = _np.sqrt(pooled * (1 - pooled) * (2 / base))
            if se == 0:
                return ""
            z = (pb - pa) / se
            pval = 2 * (1 - _norm.cdf(abs(z)))
            if pval >= _grid_alpha:
                return ""
            return "higher" if pb > pa else "lower"

        if not _cmp_brands:
            st.info("Pick at least one comparison brand above.")
        elif _gbase < 1:
            st.info("No respondent base available for the active filters.")
        else:
            def _funnel_bar_fig(brand: str, is_base: bool):
                _vals = [_stage_count(brand, s) / _gbase * 100 for s in _grid_stages]
                _flags = ["" if is_base else _prop_sig(_stage_count(sel_brand, s), _stage_count(brand, s), _gbase)
                          for s in _grid_stages]
                _colors = ["#1a5d4d" if is_base else
                          ("#16a34a" if f == "higher" else ("#dc2626" if f == "lower" else "#94a3b8"))
                          for f in _flags]
                _texts = [f"{v:.1f}%" + ({"higher": " ↑", "lower": " ↓"}.get(f, ""))
                         for v, f in zip(_vals, _flags)]
                _fig = go.Figure(go.Bar(
                    x=_vals, y=_grid_labels, orientation="h", marker_color=_colors,
                    text=_texts, textposition="outside",
                ))
                _bfig = {k: v for k, v in _chart_layout_base(230).items()
                        if k not in ("xaxis", "yaxis", "margin")}
                _fig.update_layout(**_bfig, xaxis=dict(title="", range=[0, 100]),
                                   yaxis=dict(automargin=True, categoryorder="array",
                                              categoryarray=_grid_labels[::-1]),
                                   title=dict(text=brand, font=dict(size=12)),
                                   margin=dict(t=32, b=10, l=10, r=40))
                return _fig

            # 20:80 split — base brand narrow on the left, comparison grid (up to 2 rows × 3
            # cols) wide on the right. A 1:1 or 1:3 split still starves the 3-wide grid.
            _base_col, _grid_col = st.columns([1, 4])
            with _base_col:
                st.plotly_chart(_theme_fig(_funnel_bar_fig(sel_brand, is_base=True)),
                                use_container_width=True)
            with _grid_col:
                # Ordered strongest → weakest (by TOTAL_AIDED, same ranking as the default
                # comparison set) so the grid reads top-left-to-bottom-right in descending
                # strength — the natural reading order — instead of arbitrary pick order.
                _cmp_rank = (bench_df[bench_df["brand_name"].isin(_cmp_brands)]
                            .set_index("brand_name")["TOTAL_AIDED"])
                _cmp_brands_ordered = sorted(_cmp_brands,
                                             key=lambda b: -(_cmp_rank.get(b) or 0))
                _rows_of_3 = [_cmp_brands_ordered[i:i + 3] for i in range(0, len(_cmp_brands_ordered), 3)]
                for _row_brands in _rows_of_3:
                    _grid_cols = st.columns(3)
                    for _gi, _b in enumerate(_row_brands):
                        with _grid_cols[_gi]:
                            st.plotly_chart(_theme_fig(_funnel_bar_fig(_b, is_base=False)),
                                            use_container_width=True)
            st.caption(
                f"Green = significantly higher than {sel_brand} at that stage, red = significantly "
                f"lower, grey = no significant difference, at {_grid_conf:.0%} confidence "
                f"(base N = {_gbase:,} respondents, common across all brands since awareness is "
                f"asked of everyone)."
            )


# ── Render: Consumer Demographics Profile ────────────────────────────────────

def _render_demographic_profile(sel_brand: str, zone="all", gender="all", age_band="all", city="all"):
    try:
        demo_data, overall_pct = _get_demographic_profile(sel_brand, zone, gender, age_band, city)
    except Exception as e:
        st.warning(f"Demographics unavailable: {e}")
        return

    if not demo_data:
        st.info("No demographic data.")
        return

    # Insight callout: best over-index and worst under-index across all dimensions
    _demo_all_segs = []
    for _dk, _dn in zip(["gender", "age_band", "zone_id"], ["Gender", "Age Band", "Zone"]):
        _ddf = demo_data.get(_dk, pd.DataFrame())
        if not _ddf.empty and "index_score" in _ddf.columns:
            for _, _dr in _ddf.iterrows():
                _demo_all_segs.append((float(_dr["index_score"]), str(_dr["segment"]), _dn))
    if _demo_all_segs:
        _demo_best  = max(_demo_all_segs, key=lambda x: x[0])
        _demo_worst = min(_demo_all_segs, key=lambda x: x[0])
        _dci1, _dci2 = st.columns(2)
        with _dci1:
            st.markdown(
                f"<div style='background:#f0fdf4;border:1px solid #bbf7d0;border-radius:10px;"
                f"padding:10px 14px;margin-bottom:10px;'>"
                f"<div style='font-size:0.58rem;font-weight:700;text-transform:uppercase;"
                f"letter-spacing:0.08em;color:#15803d;margin-bottom:3px;'>Strongest Segment</div>"
                f"<div style='font-size:1.1rem;font-weight:800;color:#14532d;'>{_demo_best[1]}</div>"
                f"<div style='font-size:0.7rem;color:#166534;'>Index {_demo_best[0]:.0f} · {_demo_best[2]} · "
                f"over-indexes by {_demo_best[0]-100:.0f} pts</div>"
                f"</div>",
                unsafe_allow_html=True,
            )
        with _dci2:
            st.markdown(
                f"<div style='background:#fef2f2;border:1px solid #fecaca;border-radius:10px;"
                f"padding:10px 14px;margin-bottom:10px;'>"
                f"<div style='font-size:0.58rem;font-weight:700;text-transform:uppercase;"
                f"letter-spacing:0.08em;color:#dc2626;margin-bottom:3px;'>Underperforming Segment</div>"
                f"<div style='font-size:1.1rem;font-weight:800;color:#7f1d1d;'>{_demo_worst[1]}</div>"
                f"<div style='font-size:0.7rem;color:#991b1b;'>Index {_demo_worst[0]:.0f} · {_demo_worst[2]} · "
                f"under-indexes by {100-_demo_worst[0]:.0f} pts</div>"
                f"</div>",
                unsafe_allow_html=True,
            )

    dim_tabs = st.tabs(["👥 Gender", "📅 Age Band", "🗺️ Zone"])
    dim_keys  = ["gender", "age_band", "zone_id"]
    dim_names = ["Gender", "Age Band", "Zone"]

    for tab, dim_key, dim_name in zip(dim_tabs, dim_keys, dim_names):
        with tab:
            df = demo_data.get(dim_key, pd.DataFrame())
            if df.empty:
                st.info(f"No {dim_name} data.")
                continue

            col_left, col_right = st.columns([3, 2])
            with col_left:
                bar_colors = []
                for idx_score in df["index_score"]:
                    if idx_score >= 110: bar_colors.append("#15803d")
                    elif idx_score <= 90: bar_colors.append("#dc2626")
                    else: bar_colors.append("#6b7280")

                fig = go.Figure()
                fig.add_trace(go.Bar(
                    x=df["brand_pct"], y=df["segment"],
                    orientation="h",
                    marker_color=bar_colors,
                    text=[f"{v:.1f}% (idx {i})" for v, i in zip(df["brand_pct"], df["index_score"])],
                    textposition="outside",
                    hovertemplate="<b>%{y}</b><br>Consideration: %{x:.1f}%<extra></extra>",
                ))
                fig.add_vline(x=overall_pct, line_dash="dash", line_color="#374151",
                              annotation_text=f"Overall: {overall_pct:.1f}%",
                              annotation_position="top right")
                base = {k: v for k, v in _chart_layout_base(max(200, len(df)*50)).items()
                        if k not in ("xaxis","yaxis","legend","margin")}
                fig.update_layout(**base,
                                  xaxis_title=f"{sel_brand} Consideration %",
                                  yaxis=dict(automargin=True),
                                  margin=dict(l=120, r=100, t=50, b=40),
                                  title=dict(text=f"{sel_brand} — Consideration by {dim_name}", font=dict(size=13)))
                st.plotly_chart(_theme_fig(fig), use_container_width=True)

            with col_right:
                show_df = df[["segment", "brand_pct", "index_score", "seg_n"]].copy()
                show_df.columns = [dim_name, "Consideration %", "Index (100=avg)", "N"]

                def color_index(val):
                    if val >= 110: return "color: #15803d; font-weight: bold"
                    if val <= 90:  return "color: #dc2626; font-weight: bold"
                    return ""

                styled = show_df.style.format({
                    "Consideration %": "{:.1f}", "N": "{:,}"
                }).map(color_index, subset=["Index (100=avg)"])
                st.dataframe(styled, hide_index=True, use_container_width=True)
                st.caption("Index = segment consideration% ÷ overall brand consideration% × 100. >110 = over-index (green), <90 = under-index (red).")


# ── Render: Attribute Ownership Map ─────────────────────────────────────────

def _render_attribute_ownership(sel_brand: str):
    df = _get_attribute_ownership()
    if df is None or df.empty:
        st.info("Attribute ownership data unavailable.")
        return
    # Ensure numeric columns are float (guard against object dtype from merge)
    for _c in ("assoc_pct", "margin", "second_pct"):
        if _c in df.columns:
            df[_c] = pd.to_numeric(df[_c], errors="coerce")

    ao_t1, ao_t2 = st.tabs(["🏆 Ownership Summary", f"🎯 {sel_brand} Position"])

    with ao_t1:
        # Group by brand ownership count
        brand_counts = df.groupby("brand_name").size().sort_values(ascending=False).reset_index()
        brand_counts.columns = ["Brand", "Attributes Owned"]

        c1, c2 = st.columns([1, 2])
        with c1:
            fig_count = go.Figure(go.Bar(
                x=brand_counts["Attributes Owned"],
                y=brand_counts["Brand"],
                orientation="h",
                marker_color=["#1a5d4d" if b == sel_brand else "#30a76a" for b in brand_counts["Brand"]],
                text=brand_counts["Attributes Owned"],
                textposition="outside",
            ))
            base = {k: v for k, v in _chart_layout_base(400).items()
                    if k not in ("xaxis","yaxis","legend","margin")}
            fig_count.update_layout(**base,
                                    xaxis_title="# Attributes Owned",
                                    yaxis=dict(autorange="reversed", automargin=True),
                                    margin=dict(l=130, r=60, t=40, b=40),
                                    title=dict(text="Attributes Owned per Brand", font=dict(size=12)))
            st.plotly_chart(_theme_fig(fig_count), use_container_width=True)

        with c2:
            # Full ownership table grouped by attribute category
            show_cols = ["broad_feature", "attr_label", "brand_name", "assoc_pct", "margin"]
            show_df = df[show_cols].copy()
            show_df.columns = ["Attribute Group", "Attribute", "Owned By", "Association %", "Lead Margin"]

            def highlight_brand(row):
                color = "background-color: #e8f5e9; font-weight:bold" if row["Owned By"] == sel_brand else ""
                return pd.Series([color] * len(row), index=row.index)

            styled = (
                show_df.style
                .apply(highlight_brand, axis=1)
                .background_gradient(subset=["Association %"], cmap="YlGn", axis=0)
                .format({"Association %": "{:.1f}", "Lead Margin": "{:.1f}"}, na_rep="—")
            )
            st.dataframe(styled, hide_index=True, use_container_width=True, height=380)
            st.caption(f"Highlighted rows = attributes owned by {sel_brand}. Lead margin = gap over 2nd-place brand.")

    with ao_t2:
        brand_owned   = df[df["brand_name"] == sel_brand].copy()
        brand_not_own = df[df["brand_name"] != sel_brand].copy()

        # Attributes brand owns
        if brand_owned.empty:
            st.info(f"{sel_brand} does not own any attributes in this dataset.")
        else:
            st.markdown(f"**{sel_brand} owns {len(brand_owned)} attribute(s):**")
            owned_fig = go.Figure(go.Bar(
                x=brand_owned["assoc_pct"],
                y=brand_owned["attr_label"],
                orientation="h",
                marker_color="#1a5d4d",
                text=brand_owned["assoc_pct"].apply(lambda v: f"{v:.1f}%" if pd.notna(v) else ""),
                textposition="outside",
                hovertemplate="<b>%{y}</b><br>%{x:.1f}% association<extra></extra>",
            ))
            base = {k: v for k, v in _chart_layout_base(max(200, len(brand_owned)*35 + 80)).items()
                    if k not in ("xaxis","yaxis","legend","margin")}
            owned_fig.update_layout(**base,
                                    xaxis_title="Association %",
                                    yaxis=dict(automargin=True),
                                    margin=dict(l=220, r=80, t=40, b=40))
            st.plotly_chart(_theme_fig(owned_fig), use_container_width=True)

        # Attributes brand could challenge (it's in top 3 but doesn't own it)
        st.markdown(f"**Attributes NOT owned by {sel_brand} (competitor owns them):**")
        # Show top 10 attributes owned by others with highest brand's own association (competitive gaps)
        # For this we need per-brand association for sel_brand on all attributes
        # Use the full df before idxmax aggregation — re-query
        try:
            import sqlite3
            from oxdata.db_loader import get_db_path
            conn = sqlite3.connect(get_db_path())
            total = int(pd.read_sql("SELECT COUNT(DISTINCT respondent_id) as n FROM fact_respondents", conn).iloc[0, 0])
            brand_attrs = pd.read_sql("""
                SELECT da.attr_label, da.broad_feature,
                       ROUND(COUNT(DISTINCT bi.respondent_id)*100.0 / ?, 1) AS brand_assoc
                FROM fact_brand_imagery bi
                JOIN dim_brand b ON bi.brand_id = b.brand_id
                JOIN dim_bq3_attribute da ON bi.attr_id = da.attr_id
                WHERE b.brand_name = ?
                GROUP BY da.attr_label, da.broad_feature
            """, conn, params=[total, sel_brand])
            conn.close()
            brand_attrs["brand_assoc"] = pd.to_numeric(brand_attrs["brand_assoc"], errors="coerce")

            gaps = brand_not_own[["attr_label", "brand_name", "assoc_pct"]].merge(
                brand_attrs, on="attr_label", how="inner"
            )
            gaps = gaps.rename(columns={"brand_name": "Owner", "assoc_pct": "Owner Assoc %", "brand_assoc": f"{sel_brand} Assoc %"})
            gaps["Gap to Leader"] = (gaps["Owner Assoc %"] - gaps[f"{sel_brand} Assoc %"]).round(1)
            gaps = gaps.sort_values(f"{sel_brand} Assoc %", ascending=False).head(15)

            st.dataframe(gaps[["broad_feature","attr_label","Owner","Owner Assoc %",f"{sel_brand} Assoc %","Gap to Leader"]]
                         .rename(columns={"broad_feature":"Group","attr_label":"Attribute"}),
                         hide_index=True, use_container_width=True)
            st.caption(f"Sorted by {sel_brand}'s own association %. Small 'Gap to Leader' = most winnable attributes.")
        except Exception as e:
            st.warning(f"Competitive gap data: {e}")


@st.cache_data(ttl=1800, show_spinner=False)
def _filtered_resp_ids(zone="all", gender="all", age_band="all", city="all", project_id="project_1"):
    """Return set of respondent_ids matching segment filters (None = all)."""
    if all(x == "all" for x in (zone, gender, age_band, city)):
        return None
    try:
        from lens.data_layer import get_project_layer as _gpl
        _lyr = _gpl(project_id)
    except Exception:
        _lyr = None
    if _lyr is not None:
        dem = _lyr.demographics  # respondent_id, gender, age_band, city, zone, ...
        mask = pd.Series(True, index=dem.index)
        if zone != "all" and "zone_name" in dem.columns:
            mask &= dem["zone_name"].str.lower() == zone.lower()
        if gender != "all" and "gender" in dem.columns:
            mask &= dem["gender"].str.lower() == gender.lower()
        if age_band != "all" and "age_band" in dem.columns:
            mask &= dem["age_band"].str.lower() == age_band.lower()
        if city != "all" and "city_name" in dem.columns:
            mask &= dem["city_name"].str.lower() == city.lower()
        return set(dem[mask]["respondent_id"].tolist())
    import sqlite3
    from oxdata.db_loader import get_db_path
    cte, params, is_filt = _resp_filter_cte(zone, gender, age_band, city)
    if not is_filt:
        return None
    conn = sqlite3.connect(f"file:{get_db_path(project_id=project_id)}?mode=ro", uri=True)
    try:
        df = pd.read_sql(f"WITH {cte} SELECT respondent_id FROM _base", conn, params=params)
    finally:
        conn.close()
    return set(df["respondent_id"].tolist())


@st.cache_data(ttl=1800, show_spinner=False)
def _brand_attr_options(brand, project_id="project_1"):
    """Imagery attributes with real variance for a brand — DV-attribute picker."""
    try:
        from lens.data_layer import get_project_layer as _gpl
        _lyr = _gpl(project_id)
    except Exception:
        _lyr = None
    if _lyr is not None:
        img = _lyr.imagery
        img = img[img["brand_name"] == brand] if not str(brand).startswith("All Brands") else img
        if img.empty:
            return []
        cnt = img.groupby("attr_label")["respondent_id"].nunique()
        nb = img["respondent_id"].nunique() or 1
        return sorted(a for a, n in cnt.items() if 0.02 < n / nb < 0.98 and n >= 30)
    import sqlite3
    from oxdata.db_loader import get_db_path
    conn = sqlite3.connect(f"file:{get_db_path(project_id=project_id)}?mode=ro", uri=True)
    try:
        cnt = pd.read_sql(
            "SELECT da.attr_label, COUNT(DISTINCT bi.respondent_id) AS n "
            "FROM fact_brand_imagery bi JOIN dim_brand db ON bi.brand_id = db.brand_id "
            "JOIN dim_bq3_attribute da ON bi.attr_id = da.attr_id "
            "WHERE db.brand_name = ? GROUP BY da.attr_label", conn, params=[brand])
        base = pd.read_sql(
            "SELECT COUNT(DISTINCT bi.respondent_id) n FROM fact_brand_imagery bi "
            "JOIN dim_brand db ON bi.brand_id = db.brand_id WHERE db.brand_name = ?",
            conn, params=[brand])
    except Exception:
        conn.close()
        return []
    conn.close()
    if cnt.empty or base.empty:
        return []
    nb = int(base.iloc[0, 0]) or 1
    opts = [r["attr_label"] for _, r in cnt.iterrows()
            if 0.02 < (r["n"] / nb) < 0.98 and r["n"] >= 30]
    return sorted(opts)


def _driver_regression_from_layer(layer, brand, dv_kind, topbox_min, zone="all", gender="all",
                                   age_band="all", city="all", model="linear", dv_attr=None,
                                   project_id="project_1"):
    """Pandas-based regression path for Excel projects (no SQLite)."""
    from oxdata.skills.r_bridge import run_r_stat
    is_pooled = str(brand).startswith("All Brands")

    # --- Build imagery pivot ---
    img = layer.imagery.copy()  # respondent_id, brand_name, attr_label (value=1 implicit)
    if not is_pooled:
        img = img[img["brand_name"] == brand]
    if img.empty:
        return None
    img["value"] = 1
    piv = img.pivot_table(index="respondent_id", columns="attr_label", values="value",
                          aggfunc="max", fill_value=0)

    # --- Build DV series ---
    _FUNNEL_STAGES = {"TOM", "SPONT", "AIDED", "CONSIDERATION", "EVER_USED", "CURRENT_USER", "PREFERRED"}
    is_binary_dv = False

    if dv_kind == "ATTR":
        if not dv_attr or dv_attr not in piv.columns:
            return {"error": f"attribute '{dv_attr}' not available for {brand}"}
        dv_series = piv[dv_attr].astype(int)
        piv2 = piv.drop(columns=[dv_attr])
        rdf = piv2.copy()
        rdf.insert(0, "dv", dv_series)
        is_binary_dv = True

    elif dv_kind in _FUNNEL_STAGES:
        aw = layer.awareness.copy()  # respondent_id, brand_name, stage
        if not is_pooled:
            aw = aw[aw["brand_name"] == brand]
        hits = aw[aw["stage"] == dv_kind].drop_duplicates("respondent_id")["respondent_id"]
        dv_series = piv.index.isin(hits).astype(int)
        rdf = piv.copy()
        rdf.insert(0, "dv", pd.Series(dv_series, index=piv.index))
        is_binary_dv = True

    elif dv_kind == "CSAT":
        csat = layer.csat.copy()  # respondent_id, brand_name, score
        if not is_pooled:
            csat = csat[csat["brand_name"] == brand]
        csat = csat.drop_duplicates("respondent_id").set_index("respondent_id")["score"]
        if topbox_min and topbox_min > 0:
            csat = (csat >= topbox_min).astype(int)
            is_binary_dv = True
        rdf = csat.rename("dv").to_frame().join(piv, how="inner")

    else:  # NPS
        nps = layer.nps.copy()  # respondent_id, brand_name, nps_score
        if not is_pooled:
            nps = nps[nps["brand_name"] == brand]
        nps = nps.drop_duplicates("respondent_id").set_index("respondent_id")["nps_score"]
        nps = pd.to_numeric(nps, errors="coerce")
        if topbox_min and topbox_min > 0:
            nps = (nps >= topbox_min).astype(int)
            is_binary_dv = True
        rdf = nps.rename("dv").to_frame().dropna().join(piv, how="inner")

    # --- Segment filter ---
    ids = _filtered_resp_ids(zone, gender, age_band, city, project_id=project_id)
    if ids is not None:
        rdf = rdf[rdf.index.isin(ids)]

    rdf = rdf.reset_index().rename(columns={"dv": "nps_score"})
    if "index" in rdf.columns:
        rdf = rdf.drop(columns=["index"])
    if len(rdf) < 30:
        return {"error": "insufficient", "n": len(rdf)}

    # Pre-filter top 50 predictors by |correlation|
    _pred_cols = [c for c in rdf.columns if c not in ("respondent_id", "nps_score")]
    if len(_pred_cols) > 50:
        _corr = rdf[_pred_cols].corrwith(rdf["nps_score"]).abs()
        rdf = rdf[["respondent_id", "nps_score"] + _corr.nlargest(50).index.tolist()]

    if model == "logistic" and is_binary_dv:
        return run_r_stat("logistic_regression", rdf, timeout=120)
    if model == "random_forest":
        return run_r_stat("random_forest", rdf, timeout=180)
    return run_r_stat("driver_regression", rdf, timeout=120)


@st.cache_data(ttl=1800, show_spinner=False)
def _driver_regression_for_brand(brand, dv_kind, topbox_min, zone="all", gender="all",
                                 age_band="all", city="all", model="linear", dv_attr=None,
                                 project_id="project_1"):
    """Build the per-respondent regression frame for one brand and run the R driver
    regression. Returns the R result dict (with importance/std_coef) or {'error':..} / None.

    dv_kind: 'NPS' | 'CSAT' | 'ATTR'.
      - NPS/CSAT: a 0-10 outcome; topbox_min recodes it to a binary top-box (0 = raw).
      - ATTR: the dependent variable is an imagery attribute (dv_attr, binary 0/1);
        every OTHER attribute becomes a predictor. topbox is ignored.
    model: 'linear' (OLS / linear-probability) or 'logistic' (binary outcome).
    """
    import sqlite3, pathlib
    from oxdata.db_loader import get_db_path
    from oxdata.skills.r_bridge import run_r_stat
    is_pooled = str(brand).startswith("All Brands")

    # --- Pandas layer path (Excel-based projects) ---
    try:
        from lens.data_layer import get_project_layer as _gpl
        _layer = _gpl(project_id)
    except Exception:
        _layer = None
    if _layer is not None:
        return _driver_regression_from_layer(
            _layer, brand, dv_kind, topbox_min, zone, gender, age_band, city, model, dv_attr, project_id
        )

    conn = sqlite3.connect(f"file:{get_db_path(project_id=project_id)}?mode=ro", uri=True)

    # --- Pre-ingested regression matrix shortcut ---
    # If <project_data_dir>/regression_matrix.csv exists, use it directly for pooled mode.
    # Format: respondent_id, brand_id, dv_ever_used, [attr cols...]
    # This guarantees XLSTAT-identical n and variable values without runtime CROSS JOIN.
    _matrix_path = pathlib.Path(get_db_path(project_id=project_id)).parent / "regression_matrix.csv"
    if is_pooled and _matrix_path.exists() and dv_kind not in ("CSAT", "ATTR"):
        try:
            df_m = pd.read_csv(_matrix_path)
            _attr_cols = [c for c in df_m.columns if c not in ("respondent_id", "brand_id", "dv_ever_used")]
            # Composite key = respondent_id + brand_id for uniqueness across brands
            df_m["_rk"] = df_m["respondent_id"].astype(str) + "_" + df_m["brand_id"].astype(str)
            piv_m = df_m.set_index("_rk")[_attr_cols].fillna(0)
            dv_m  = df_m.set_index("_rk")["dv_ever_used"].astype(int)
            if dv_kind in {"TOM", "SPONT", "AIDED", "CONSIDERATION", "EVER_USED", "CURRENT_USER", "PREFERRED"}:
                # Use pre-ingested dv_ever_used (always EVER_USED in matrix file)
                rdf_m = piv_m.copy()
                rdf_m.insert(0, "dv", dv_m)
            else:
                # NPS fallback — join from DB; keep only rows with actual NPS data
                dv_nps = pd.read_sql(
                    "SELECT n.respondent_id || '_' || n.brand_id AS respondent_id, n.nps_score AS dv "
                    "FROM fact_brand_nps n WHERE n.nps_score IS NOT NULL", conn)
                dv_nps = dv_nps.drop_duplicates("respondent_id").set_index("respondent_id")["dv"]
                rdf_m = piv_m.copy()
                # Use NaN (not 0) for missing NPS so dropna() removes them cleanly
                rdf_m.insert(0, "dv", dv_nps.reindex(piv_m.index).astype(float))
                rdf_m = rdf_m.dropna(subset=["dv"])  # keep only raters
            is_binary_dv = dv_kind in {"TOM","SPONT","AIDED","CONSIDERATION","EVER_USED","CURRENT_USER","PREFERRED"}
            rdf_m = rdf_m.reset_index().rename(columns={"dv": "nps_score", "_rk": "respondent_id"})
            _pred_cols = [c for c in rdf_m.columns if c not in ("respondent_id", "nps_score")]
            if len(_pred_cols) > 50:
                _corr = rdf_m[_pred_cols].corrwith(rdf_m["nps_score"]).abs()
                rdf_m = rdf_m[["respondent_id", "nps_score"] + _corr.nlargest(50).index.tolist()]
            if len(rdf_m) < 30:
                conn.close()
                return {"error": "insufficient", "n": len(rdf_m)}
            conn.close()
            if model == "logistic" and is_binary_dv:
                return run_r_stat("logistic_regression", rdf_m, timeout=120)
            if model == "random_forest":
                return run_r_stat("random_forest", rdf_m, timeout=180)
            return run_r_stat("driver_regression", rdf_m, timeout=120)
        except Exception as _me:
            pass  # fall through to DB path on any error

    try:
        if is_pooled:
            # Base = all AIDED-aware respondent×brand pairs (matches XLSTAT: every aware pair
            # gets a row even if respondent ticked zero imagery attributes for that brand).
            # Cross-join with all attrs, LEFT JOIN imagery → non-associated attrs = NULL = 0.
            bi = pd.read_sql(
                "SELECT aw.respondent_id || '_' || aw.brand_id AS respondent_id, "
                "       da.attr_label, COALESCE(bi.value, 0) AS value "
                "FROM fact_brand_awareness aw "
                "CROSS JOIN dim_bq3_attribute da "
                "LEFT JOIN fact_brand_imagery bi "
                "  ON bi.respondent_id = aw.respondent_id "
                "  AND bi.brand_id = aw.brand_id "
                "  AND bi.attr_id = da.attr_id "
                "WHERE aw.stage = 'AIDED'", conn)
        else:
            bi = pd.read_sql(
                "SELECT bi.respondent_id, da.attr_label, bi.value "
                "FROM fact_brand_imagery bi "
                "JOIN dim_brand db ON bi.brand_id = db.brand_id "
                "JOIN dim_bq3_attribute da ON bi.attr_id = da.attr_id "
                "WHERE db.brand_name = ?", conn, params=[brand])

        _FUNNEL_STAGES = {"TOM", "SPONT", "AIDED", "CONSIDERATION", "EVER_USED", "CURRENT_USER", "PREFERRED"}
        dv = None
        if dv_kind != "ATTR":
            if dv_kind == "CSAT":
                if is_pooled:
                    dv = pd.read_sql(
                        "SELECT s.respondent_id || '_' || s.brand_id AS respondent_id, s.score AS dv "
                        "FROM fact_satisfaction s", conn)
                else:
                    dv = pd.read_sql(
                        "SELECT ba.respondent_id, s.score AS dv "
                        "FROM fact_satisfaction s "
                        "JOIN fact_brand_awareness ba ON s.respondent_id = ba.respondent_id "
                        "  AND ba.stage = 'LAST_PURCHASED' "
                        "JOIN dim_brand db ON ba.brand_id = db.brand_id "
                        "WHERE db.brand_name = ?", conn, params=[brand])
            elif dv_kind in _FUNNEL_STAGES:
                if is_pooled:
                    dv = pd.read_sql(
                        "SELECT fa.respondent_id || '_' || fa.brand_id AS respondent_id, 1 AS dv "
                        "FROM fact_brand_awareness fa "
                        "WHERE fa.stage = ?", conn, params=[dv_kind])
                else:
                    dv = pd.read_sql(
                        "SELECT fa.respondent_id, 1 AS dv "
                        "FROM fact_brand_awareness fa "
                        "JOIN dim_brand db ON fa.brand_id = db.brand_id "
                        "WHERE db.brand_name = ? AND fa.stage = ?",
                        conn, params=[brand, dv_kind])
            else:
                if is_pooled:
                    dv = pd.read_sql(
                        "SELECT n.respondent_id || '_' || n.brand_id AS respondent_id, n.nps_score AS dv "
                        "FROM fact_brand_nps n", conn)
                else:
                    dv = pd.read_sql(
                        "SELECT n.respondent_id, n.nps_score AS dv "
                        "FROM fact_brand_nps n JOIN dim_brand db ON n.brand_id = db.brand_id "
                        "WHERE db.brand_name = ?", conn, params=[brand])
    except Exception as e:
        conn.close()
        return {"error": str(e)}
    conn.close()
    if bi.empty:
        return None
    piv = bi.pivot_table(index="respondent_id", columns="attr_label", values="value",
                         aggfunc="max", fill_value=0)

    if dv_kind == "ATTR":
        # Dependent variable IS an imagery attribute; predictors = all the others.
        if not dv_attr or dv_attr not in piv.columns:
            return {"error": f"attribute '{dv_attr}' not available for {brand}"}
        dv_series = piv[dv_attr].astype(int)
        X = piv.drop(columns=[dv_attr])
        rdf = X.copy()
        rdf.insert(0, "dv", dv_series)
        is_binary_dv = True
    else:
        _FUNNEL_STAGES_CHECK = {"TOM", "SPONT", "AIDED", "CONSIDERATION",
                                 "EVER_USED", "CURRENT_USER", "PREFERRED"}
        if dv_kind in _FUNNEL_STAGES_CHECK:
            # Base = all respondents in imagery matrix; 1 = funnel hit, 0 = not.
            if dv is None or dv.empty:
                dv_series = pd.Series(0, index=piv.index, name="dv")
            else:
                dv_hits = dv.drop_duplicates("respondent_id").set_index("respondent_id")["dv"]
                dv_series = dv_hits.reindex(piv.index, fill_value=0).astype(int)
            rdf = piv.copy()
            rdf.insert(0, "dv", dv_series)
            is_binary_dv = True
        else:
            if dv is None or dv.empty:
                return None
            dv["dv"] = pd.to_numeric(dv["dv"], errors="coerce")
            dv = dv.dropna(subset=["dv"]).drop_duplicates("respondent_id")
            if topbox_min and topbox_min > 0:
                dv["dv"] = (dv["dv"] >= topbox_min).astype(int)
            rdf = dv.set_index("respondent_id")[["dv"]].join(piv, how="inner")
            is_binary_dv = bool(topbox_min and topbox_min > 0)

    # Apply segment filters
    ids = _filtered_resp_ids(zone, gender, age_band, city, project_id=project_id)
    if ids is not None:
        rdf = rdf[rdf.index.isin(ids)]
    rdf = rdf.reset_index().rename(columns={"dv": "nps_score"})  # R script expects 'nps_score' as DV
    if "index" in rdf.columns:
        rdf = rdf.drop(columns=["index"])
    if len(rdf) < 30:
        return {"error": "insufficient", "n": len(rdf)}

    # Pre-filter predictors: keep top 50 by |correlation| with outcome.
    # Reduces R VIF computation from O(p³) on 87 predictors to ~40 — ~5× faster.
    _pred_cols = [c for c in rdf.columns if c not in ("respondent_id", "nps_score")]
    if len(_pred_cols) > 50:
        _corr = rdf[_pred_cols].corrwith(rdf["nps_score"]).abs()
        _top  = _corr.nlargest(50).index.tolist()
        rdf   = rdf[["respondent_id", "nps_score"] + _top]

    # Logistic needs a binary outcome.
    if model == "logistic" and is_binary_dv:
        return run_r_stat("logistic_regression", rdf, timeout=120)
    if model == "random_forest":
        return run_r_stat("random_forest", rdf, timeout=180)
    return run_r_stat("driver_regression", rdf, timeout=120)


def _norm_key(s: str) -> str:
    """Canonical key for matching attribute names across R's name-mangling
    (R read.csv turns 'A b (c)' into 'A.b..c.'). Lowercase alphanumerics only."""
    import re as _re3
    return _re3.sub(r"[^a-z0-9]", "", str(s or "").lower())


def _clean_attr(name: str) -> str:
    """Human-readable attribute label: dots/underscores → spaces, fix mojibake, tidy."""
    if not name:
        return ""
    s = str(name).replace(".", " ").replace("_", " ")
    # common latin-1/utf-8 mojibake seen in the imagery labels
    for bad, good in (("Ã©", "é"), ("Ã¨", "è"), ("Ã³", "ó"), ("Ã±", "ñ"),
                      ("Ã¤", "ä"), ("â€™", "'"), ("Â", "")):
        s = s.replace(bad, good)
    return " ".join(s.split()).strip()


@st.cache_data(ttl=3600, show_spinner=False)
def _attr_feature_map():
    """attr_label → broad_feature (theme) for grouping driver importance."""
    import sqlite3
    from oxdata.db_loader import get_db_path
    conn = sqlite3.connect(f"file:{get_db_path()}?mode=ro", uri=True)
    try:
        df = pd.read_sql("SELECT attr_label, broad_feature FROM dim_bq3_attribute", conn)
    except Exception:
        conn.close()
        return {}
    conn.close()
    return {r["attr_label"]: (r["broad_feature"] or "Other") for _, r in df.iterrows()}


def _brand_assoc_pct(brand: str) -> dict:
    """Each attribute's association % for a brand (the 'performance' axis for the
    importance × performance key-driver quadrant). Cached."""
    return _brand_assoc_pct_cached(brand)


@st.cache_data(ttl=1800, show_spinner=False)
def _brand_assoc_pct_cached(brand: str) -> dict:
    """Each attribute's association % for a brand = (# respondents associating this
    brand with attr A) / (# respondents who evaluated attr A on ANY brand).

    The denominator is the PER-ATTRIBUTE evaluation base (matches bip_engine's
    column base), NOT the brand's own imagery base — using the brand base would
    make %s non-comparable across attributes and inflate them vs. the BIP table.
    fact_brand_imagery.value is always 1, so AVG(value) would be a meaningless 100%."""
    import sqlite3
    from oxdata.db_loader import get_db_path
    conn = sqlite3.connect(f"file:{get_db_path()}?mode=ro", uri=True)
    try:
        df = pd.read_sql(
            "SELECT da.attr_label, COUNT(DISTINCT bi.respondent_id) AS n "
            "FROM fact_brand_imagery bi JOIN dim_brand db ON bi.brand_id = db.brand_id "
            "JOIN dim_bq3_attribute da ON bi.attr_id = da.attr_id "
            "WHERE db.brand_name = ? GROUP BY da.attr_label", conn, params=[brand])
        # Per-attribute denominator: distinct respondents who associated ANY brand
        # with that attribute (the attribute's evaluation base).
        base_df = pd.read_sql(
            "SELECT da.attr_label, COUNT(DISTINCT bi.respondent_id) AS nb "
            "FROM fact_brand_imagery bi "
            "JOIN dim_bq3_attribute da ON bi.attr_id = da.attr_id "
            "GROUP BY da.attr_label", conn)
    except Exception:
        conn.close()
        return {}
    conn.close()
    attr_base = dict(zip(base_df["attr_label"], base_df["nb"]))
    return {
        r["attr_label"]: float(r["n"]) / attr_base[r["attr_label"]] * 100
        for _, r in df.iterrows()
        if attr_base.get(r["attr_label"], 0) > 0
    }


def _render_random_forest_driver(sel_brand, dv_kind, topbox, zone, gender, age_band, city, dv_attr, _out_lbl_for_rf="", project_id="project_1"):
    """Random Forest key-driver view: OOB fit stats, variable-importance ranking,
    OOB-error convergence, predicted-vs-actual. No coefficients/equation — a
    nonlinear ensemble ranks drivers by how much OOB error rises when a
    predictor's values are permuted (XLSTAT: 'Mean increase error')."""
    with st.spinner(f"Growing random forest for {sel_brand}…"):
        res = _driver_regression_for_brand(sel_brand, dv_kind, topbox, zone, gender, age_band, city, "random_forest", dv_attr, project_id=project_id)

    if not res:
        st.info(f"No {dv_kind} + imagery data for {sel_brand}.")
        return
    if "error" in res:
        if res.get("error") == "insufficient":
            st.info(f"Need ≥30 respondents with {dv_kind} + imagery data for {sel_brand} "
                    f"(found {res.get('n', 0)} after filters).")
        else:
            st.warning(f"Random forest error: {res['error']}")
        return

    importance = res.get("importance", [])
    if not importance:
        st.info("Random forest returned no importance results.")
        return

    _workbench_capture(
        f"Driver Regression — {sel_brand} (Random Forest)",
        {"brand": sel_brand, "model": "random_forest", "outcome": dv_kind,
         "topbox_min": topbox, "attribute_dv": dv_attr, "result": res, "importance": importance},
    )

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("OOB R²", f"{res.get('oob_r2', 0):.3f}",
              help="Variance explained by out-of-bag predictions — the forest's own honest "
                   "held-out estimate, no separate test set needed.")
    m2.metric("OOB RMSE", f"{res.get('oob_rmse', 0):.3f}")
    m3.metric("Trees", res.get("ntree", 0))
    m4.metric("Respondents", res.get("n", 0))

    st.caption(f"Model: **Random Forest** (regression, {res.get('mtry','?')} predictors sampled per split) "
               f"· outcome = {_out_lbl_for_rf} · predictors = {res.get('n_attrs', len(importance))} "
               f"imagery attributes · n = {res.get('n',0):,} respondents.")

    if res.get("missing_pct", 0) and res["missing_pct"] > 0:
        st.caption(f"ℹ️ {res['missing_pct']:.1f}% of predictor cells had missing values, filled via "
                   f"{res.get('missing_data_method', 'imputation')}.")

    with st.expander("📐 How Random Forest ranks drivers — no equation, a permutation test", expanded=False):
        st.markdown(
            "Unlike linear/logistic regression, a random forest has no single coefficient per "
            "predictor — it's an ensemble of decision trees. **Variable importance** is computed by "
            "measuring how much the model's out-of-bag prediction error *increases* when a predictor's "
            "values are randomly shuffled (breaking its relationship with the outcome) while every "
            "other predictor stays intact. A large error increase means the model relied heavily on "
            "that predictor; near-zero means it barely mattered.")
        st.caption("This matches XLSTAT's 'Mean increase error' variable-importance metric. "
                   "Values are NOT comparable to standardized-beta importance % from the Linear/Logistic "
                   "models — different scale, same ranking intent.")

    _imp_df = pd.DataFrame(importance)
    _imp_df["attribute_clean"] = _imp_df["attribute"].apply(_clean_attr)
    _top_n = min(20, len(_imp_df))
    _plot_df = _imp_df.sort_values("pct_inc_mse", ascending=True).tail(_top_n)
    fig_imp = go.Figure(go.Bar(
        x=_plot_df["pct_inc_mse"], y=_plot_df["attribute_clean"], orientation="h",
        marker_color=_chart_colors()[0] if _chart_colors() else "#4C78A8",
    ))
    fig_imp.update_layout(**_chart_layout_base(height=max(400, 24 * _top_n)))
    fig_imp.update_xaxes(title_text="Mean increase in OOB error when predictor is permuted (higher = more important)")
    fig_imp.update_yaxes(title_text="")
    fig_imp.update_layout(title=f"Variable importance — top {_top_n} drivers of {_out_lbl_for_rf}")
    st.plotly_chart(_theme_fig(fig_imp), use_container_width=True)
    st.caption("Bar length = how much prediction error rises if this attribute's values were "
               "randomly scrambled. Longer bar = the model depends on it more to predict the outcome.")

    _oob_curve = res.get("oob_error_curve", [])
    _pva = res.get("pred_vs_actual")
    if _oob_curve or _pva:
        cc1, cc2 = st.columns(2)
        if _oob_curve:
            with cc1:
                _oc_df = pd.DataFrame(_oob_curve)
                fig_oob = go.Figure(go.Scatter(x=_oc_df["ntree"], y=_oc_df["mse"], mode="lines"))
                fig_oob.update_layout(**_chart_layout_base(height=350))
                fig_oob.update_xaxes(title_text="Number of trees grown")
                fig_oob.update_yaxes(title_text="OOB mean squared error")
                fig_oob.update_layout(title="OOB error convergence")
                st.plotly_chart(_theme_fig(fig_oob), use_container_width=True)
                st.caption("Error should flatten as more trees are added — a still-falling curve at "
                           "the right edge means more trees could still help; a flat curve means the "
                           f"forest has converged at {res.get('ntree',0)} trees.")
        if _pva:
            with cc2:
                fig_pva = go.Figure(go.Scatter(x=_pva["x"], y=_pva["y"], mode="markers",
                                               marker=dict(size=5, opacity=0.5)))
                _lo = min(min(_pva["x"]), min(_pva["y"]))
                _hi = max(max(_pva["x"]), max(_pva["y"]))
                fig_pva.add_shape(type="line", x0=_lo, y0=_lo, x1=_hi, y1=_hi,
                                  line=dict(dash="dash", color="gray"))
                fig_pva.update_layout(**_chart_layout_base(height=350))
                fig_pva.update_xaxes(title_text=f"Actual {_out_lbl_for_rf}")
                fig_pva.update_yaxes(title_text=f"Predicted {_out_lbl_for_rf} (OOB)")
                fig_pva.update_layout(title="Predicted vs. actual (out-of-bag)")
                st.plotly_chart(_theme_fig(fig_pva), use_container_width=True)
                st.caption("Each dot is one respondent's OOB-predicted vs. actual outcome. Dots hugging "
                           "the dashed diagonal = accurate predictions; scatter away from it = model error.")

    with st.expander("Full variable importance table"):
        _tbl = _imp_df[["attribute_clean", "pct_inc_mse", "inc_node_purity"]].rename(columns={
            "attribute_clean": "Attribute", "pct_inc_mse": "Mean increase error", "inc_node_purity": "Node purity increase"
        }).sort_values("Mean increase error", ascending=False).reset_index(drop=True)
        st.dataframe(_tbl, use_container_width=True, hide_index=True)

    _rf_corrm = res.get("correlation_matrix") or {}
    _rf_corr_attrs = _rf_corrm.get("attributes") or []
    _rf_corr_mat = _rf_corrm.get("matrix") or []
    if _rf_corr_attrs and _rf_corr_mat:
        with st.expander(f"Correlation matrix — predictor pairs ({len(_rf_corr_attrs)}×{len(_rf_corr_attrs)})"):
            _rf_corr_labels = [_clean_attr(a) for a in _rf_corr_attrs]
            _rf_fig_corr = go.Figure(go.Heatmap(
                z=_rf_corr_mat, x=_rf_corr_labels, y=_rf_corr_labels,
                colorscale="RdBu", zmid=0, zmin=-1, zmax=1,
                text=[[f"{v:.2f}" for v in row] for row in _rf_corr_mat],
                texttemplate="%{text}", textfont=dict(size=8),
                colorbar=dict(title="r")))
            _rf_bcorr = {k: v for k, v in _chart_layout_base(max(320, 24 * len(_rf_corr_attrs))).items()
                         if k not in ("xaxis", "yaxis")}
            _rf_fig_corr.update_layout(**_rf_bcorr,
                                       xaxis=dict(tickfont=dict(size=9)),
                                       yaxis=dict(tickfont=dict(size=9), autorange="reversed"),
                                       title=dict(text="Predictor correlation matrix (Pearson)", font=dict(size=12)))
            st.plotly_chart(_theme_fig(_rf_fig_corr), use_container_width=True)
            st.caption("Pairwise Pearson correlation between predictor attributes. Random Forest is robust to "
                      "correlated predictors (unlike linear/logistic), but this still helps identify redundant "
                      "attributes that could be consolidated.")


def _build_driver_regression_workbook(res, drivers, prow_fn, sel_brand, dv_kind, is_logit,
                                       segment_rows=None, segment_dim=None):
    """Rich multi-sheet .xlsx export of a Driver Regression run — every table
    the UI shows (parameters, ANOVA/goodness-of-fit, Type II, global tests,
    correlation matrix, VIF, influence diagnostics + DFBeta, confusion matrix,
    predictions) with formatted headers, frozen panes, number formats and a
    conditional-format heatmap on the correlation matrix. Returns raw .xlsx bytes."""
    import io as _io
    import datetime as _dt
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule

    _BRAND = "1A5D4D"
    _HEADER_FILL = PatternFill(start_color=_BRAND, end_color=_BRAND, fill_type="solid")
    _HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    _TITLE_FONT  = Font(bold=True, size=14, color=_BRAND)
    _SUB_FONT    = Font(italic=True, size=9, color="6B7280")
    _THIN = Side(style="thin", color="D1D5DB")
    _BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
    _SIG_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")

    wb = Workbook()

    def _write_table(ws, headers, rows, start_row=1, sig_col=None, sig_test=None,
                      col_formats=None):
        col_formats = col_formats or {}
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=c, value=h)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _BORDER
        for r_i, row in enumerate(rows, start=start_row + 1):
            for c, h in enumerate(headers, start=1):
                v = row.get(h)
                cell = ws.cell(row=r_i, column=c, value=v)
                cell.border = _BORDER
                if h in col_formats and isinstance(v, (int, float)):
                    cell.number_format = col_formats[h]
                if sig_col and h == sig_col and sig_test is not None and sig_test(row):
                    cell.fill = _SIG_FILL
        for c, h in enumerate(headers, start=1):
            width = max(len(str(h)), *(len(str(row.get(h, ""))) for row in rows)) if rows else len(str(h))
            ws.column_dimensions[get_column_letter(c)].width = min(max(width + 2, 10), 42)
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1).coordinate
        return start_row + len(rows) + 1

    # ── Summary / cover sheet ─────────────────────────────────────────────
    ws0 = wb.active
    ws0.title = "Summary"
    ws0["A1"] = "InfoLeap Pulse — Driver Regression Report"
    ws0["A1"].font = _TITLE_FONT
    ws0["A2"] = f"Generated {_dt.datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws0["A2"].font = _SUB_FONT
    _out_lbl = dv_kind if dv_kind != "ATTR" else "imagery attribute"
    _model_lbl = "Logistic regression" if is_logit else "Linear regression (OLS)"
    _summary_rows = [
        ("Brand", sel_brand),
        ("Model", _model_lbl),
        ("Dependent variable", _out_lbl),
        ("Respondents (n)", res.get("n")),
        ("Predictors", res.get("n_attrs")),
        ("McFadden R² / R²", res.get("mcfadden_r2") if is_logit else res.get("r_squared")),
        ("AIC", res.get("aic")),
        ("BIC", res.get("bic")),
        ("% rows dropped (missing data)", res.get("pct_dropped_na")),
    ]
    for i, (k, v) in enumerate(_summary_rows, start=4):
        ws0.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws0.cell(row=i, column=2, value=v)
    ws0.column_dimensions["A"].width = 32
    ws0.column_dimensions["B"].width = 30

    # ── Model Parameters ──────────────────────────────────────────────────
    ws1 = wb.create_sheet("Model Parameters")
    _prows = [prow_fn(d) for d in drivers]
    if _prows:
        _headers = list(_prows[0].keys())
        _fmt = {"Importance %": "0.0", "Std β": "+0.000", "Coef (b)": "+0.0000",
                 "p": "0.0000", "Odds ratio": "0.000", "t": "0.00"}
        _write_table(ws1, _headers, _prows, sig_col="Sig.",
                     sig_test=lambda r: r.get("Sig.") not in (None, "ns"),
                     col_formats=_fmt)

    # ── Goodness of fit / ANOVA ────────────────────────────────────────────
    ws2 = wb.create_sheet("Goodness of Fit")
    if is_logit:
        _gof_rows = [
            {"Metric": "McFadden R²", "Value": res.get("mcfadden_r2")},
            {"Metric": "Cox & Snell R²", "Value": res.get("cox_snell_r2")},
            {"Metric": "Nagelkerke R²", "Value": res.get("nagelkerke_r2")},
            {"Metric": "AIC", "Value": res.get("aic")},
            {"Metric": "BIC/SBC", "Value": res.get("bic")},
            {"Metric": "-2 Log-Likelihood (LRT) χ²", "Value": res.get("lrt_chi2")},
            {"Metric": "LRT df", "Value": res.get("lrt_df")},
            {"Metric": "LRT p-value", "Value": res.get("lrt_p")},
            {"Metric": "AUC", "Value": res.get("auc")},
            {"Metric": "Accuracy", "Value": res.get("accuracy")},
            {"Metric": "Sensitivity", "Value": res.get("sensitivity")},
            {"Metric": "Specificity", "Value": res.get("specificity")},
        ]
        _hl = res.get("hosmer_lemeshow") or {}
        if _hl:
            _gof_rows += [
                {"Metric": "Hosmer-Lemeshow χ²", "Value": _hl.get("chi2")},
                {"Metric": "Hosmer-Lemeshow df", "Value": _hl.get("df")},
                {"Metric": "Hosmer-Lemeshow p-value", "Value": _hl.get("p_value")},
            ]
    else:
        _gof_rows = [
            {"Metric": "R²", "Value": res.get("r_squared")},
            {"Metric": "Adjusted R²", "Value": res.get("adj_r_squared")},
            {"Metric": "F-statistic", "Value": res.get("f_statistic")},
            {"Metric": "F p-value", "Value": res.get("f_p_value")},
            {"Metric": "RMSE", "Value": res.get("rmse")},
            {"Metric": "AIC", "Value": res.get("aic")},
            {"Metric": "BIC", "Value": res.get("bic")},
        ]
        _sw = res.get("shapiro_wilk") or {}
        if _sw:
            _gof_rows += [
                {"Metric": "Shapiro-Wilk W (residual normality)", "Value": _sw.get("w_stat")},
                {"Metric": "Shapiro-Wilk p-value", "Value": _sw.get("p_value")},
            ]
        if res.get("ss_total"):
            _write_table(ws2, ["Metric", "Value"], _gof_rows)
            _r = ws2.max_row + 2
            ws2.cell(row=_r, column=1, value="ANOVA").font = Font(bold=True, size=12)
            _ms_m = (res.get("ss_model", 0) / res.get("df_model", 1)) if res.get("df_model") else 0
            _ms_r = (res.get("ss_resid", 0) / res.get("df_resid", 1)) if res.get("df_resid") else 0
            _anova_rows = [
                {"Source": "Model", "SS": res.get("ss_model"), "df": res.get("df_model"), "MS": round(_ms_m, 4)},
                {"Source": "Residual", "SS": res.get("ss_resid"), "df": res.get("df_resid"), "MS": round(_ms_r, 4)},
                {"Source": "Total", "SS": res.get("ss_total"),
                 "df": (res.get("df_model") or 0) + (res.get("df_resid") or 0), "MS": None},
            ]
            _write_table(ws2, ["Source", "SS", "df", "MS"], _anova_rows, start_row=_r + 1,
                         col_formats={"SS": "0.0", "MS": "0.0000"})
            _gof_rows = None  # already written
    if _gof_rows:
        _write_table(ws2, ["Metric", "Value"], _gof_rows)

    # ── Type II analysis ───────────────────────────────────────────────────
    _type2 = res.get("type2") or []
    if _type2:
        ws3 = wb.create_sheet("Type II Analysis")
        _t2_key = "lr_chi2" if is_logit else "f_stat"
        _t2_lbl = "Chi-square (LR)" if is_logit else "F"
        _t2_rows = [{"Attribute": t.get("attribute"), "df": t.get("df"),
                      _t2_lbl: t.get(_t2_key), "p-value": t.get("p_value")}
                     for t in _type2]
        _write_table(ws3, ["Attribute", "df", _t2_lbl, "p-value"], _t2_rows,
                     sig_col="p-value", sig_test=lambda r: (r.get("p-value") or 1) < 0.05,
                     col_formats={_t2_lbl: "0.0000", "p-value": "0.000000"})

    # ── Global tests (logistic only) ───────────────────────────────────────
    _gtests = res.get("global_tests") or {}
    if _gtests:
        ws4 = wb.create_sheet("Global Tests")
        _gt_rows = []
        for _key, _lbl in (("minus2ll", "-2 Log(Likelihood)"), ("score", "Score"), ("wald", "Wald")):
            _t = _gtests.get(_key)
            if _t:
                _gt_rows.append({"Statistic": _lbl, "df": _t.get("df"),
                                  "Chi-square": _t.get("chi2"), "p-value": _t.get("p_value")})
        _write_table(ws4, ["Statistic", "df", "Chi-square", "p-value"], _gt_rows,
                     col_formats={"Chi-square": "0.0000", "p-value": "0.000000"})

    # ── Correlation matrix (with color-scale conditional formatting) ───────
    _corrmat = res.get("correlation_matrix") or {}
    _corr_attrs = _corrmat.get("attributes") or []
    if _corr_attrs and _corrmat.get("matrix"):
        ws5 = wb.create_sheet("Correlation Matrix")
        ws5.cell(row=1, column=1, value="")
        for c, a in enumerate(_corr_attrs, start=2):
            cell = ws5.cell(row=1, column=c, value=a)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(wrap_text=True, horizontal="center")
        for r, (a, row) in enumerate(zip(_corr_attrs, _corrmat["matrix"]), start=2):
            ws5.cell(row=r, column=1, value=a).font = Font(bold=True)
            for c, v in enumerate(row, start=2):
                cell = ws5.cell(row=r, column=c, value=v)
                cell.number_format = "0.00"
        n_attrs = len(_corr_attrs)
        data_range = f"B2:{get_column_letter(n_attrs + 1)}{n_attrs + 1}"
        ws5.conditional_formatting.add(
            data_range,
            ColorScaleRule(start_type="num", start_value=-1, start_color="DC2626",
                           mid_type="num", mid_value=0, mid_color="F9FAFB",
                           end_type="num", end_value=1, end_color="1A5D4D"))
        ws5.column_dimensions["A"].width = 34
        for c in range(2, n_attrs + 2):
            ws5.column_dimensions[get_column_letter(c)].width = 12
        ws5.freeze_panes = "B2"

    # ── VIF ─────────────────────────────────────────────────────────────
    _vif = res.get("vif") or {}
    if _vif:
        ws6 = wb.create_sheet("VIF")
        _vif_rows = sorted(
            [{"Attribute": k, "VIF": v, "Flag": "high" if v > 10 else ("moderate" if v > 5 else "ok")}
             for k, v in _vif.items()], key=lambda r: -r["VIF"])
        _write_table(ws6, ["Attribute", "VIF", "Flag"], _vif_rows,
                     sig_col="Flag", sig_test=lambda r: r.get("Flag") in ("high", "moderate"),
                     col_formats={"VIF": "0.00"})

    # ── Influence diagnostics (+ DFBeta) ───────────────────────────────────
    _infl = res.get("influence") or {}
    _top_infl = _infl.get("top_influential") or []
    if _top_infl:
        ws7 = wb.create_sheet("Influence Diagnostics")
        _infl_rows = [{"Respondent": r["obs"], "Cook's D": r["cooks_d"], "Leverage": r["leverage"],
                        "High leverage": "yes" if r.get("high_leverage") else "",
                        "High Cook's D": "yes" if r.get("high_cooks_d") else ""}
                       for r in _top_infl]
        _end_row = _write_table(ws7, ["Respondent", "Cook's D", "Leverage", "High leverage", "High Cook's D"],
                     _infl_rows, sig_col="High Cook's D", sig_test=lambda r: r.get("High Cook's D") == "yes",
                     col_formats={"Cook's D": "0.00000", "Leverage": "0.00000"})
        if any(r.get("dfbeta") for r in _top_infl):
            ws7.cell(row=_end_row + 1, column=1, value="DFBeta per predictor").font = Font(bold=True, size=12)
            _dfb_rows = []
            for r in _top_infl:
                _row = {"Respondent": r["obs"]}
                _row.update(r.get("dfbeta") or {})
                _dfb_rows.append(_row)
            if _dfb_rows:
                _dfb_headers = list(_dfb_rows[0].keys())
                _write_table(ws7, _dfb_headers, _dfb_rows, start_row=_end_row + 2,
                             col_formats={h: "0.00000" for h in _dfb_headers if h != "Respondent"})

    # ── Confusion matrix (logistic only) ────────────────────────────────
    _cm = res.get("confusion") or {}
    if is_logit and _cm:
        ws8 = wb.create_sheet("Confusion Matrix")
        _tp, _tn, _fp, _fn = _cm.get("tp", 0), _cm.get("tn", 0), _cm.get("fp", 0), _cm.get("fn", 0)
        _r0, _r1 = _tn + _fp, _fn + _tp
        _cm_rows = [
            {"from \\ to": "Observed 0", "Predicted 0": _tn, "Predicted 1": _fp, "Total": _r0,
             "% correct": (_tn / _r0) if _r0 else None},
            {"from \\ to": "Observed 1", "Predicted 0": _fn, "Predicted 1": _tp, "Total": _r1,
             "% correct": (_tp / _r1) if _r1 else None},
            {"from \\ to": "Total", "Predicted 0": _tn + _fn, "Predicted 1": _fp + _tp,
             "Total": res.get("n", _r0 + _r1), "% correct": res.get("accuracy")},
        ]
        _write_table(ws8, ["from \\ to", "Predicted 0", "Predicted 1", "Total", "% correct"], _cm_rows,
                     col_formats={"% correct": "0.0%"})
        _r = ws8.max_row + 2
        ws8.cell(row=_r, column=1, value="Discrimination metrics").font = Font(bold=True, size=12)
        _disc_rows = [
            {"Metric": "Accuracy", "Value": res.get("accuracy")},
            {"Metric": "Sensitivity (recall)", "Value": res.get("sensitivity")},
            {"Metric": "Specificity", "Value": res.get("specificity")},
            {"Metric": "PPV (precision)", "Value": res.get("ppv")},
            {"Metric": "NPV", "Value": res.get("npv")},
        ]
        _write_table(ws8, ["Metric", "Value"], _disc_rows, start_row=_r + 1,
                     col_formats={"Value": "0.0%"})

    # ── Predictions (downsampled predicted-vs-actual, same points shown on screen) ──
    _pc = res.get("prob_chart") if is_logit else None
    _pva = res.get("pred_vs_actual") if not is_logit else None
    if _pc and _pc.get("pred"):
        ws9 = wb.create_sheet("Predictions")
        _pred_rows = [{"Obs (sorted by predicted prob)": x, "Predicted Pr(1)": p, "Actual": a}
                       for x, p, a in zip(_pc.get("x", []), _pc.get("pred", []), _pc.get("actual", []))]
        _write_table(ws9, ["Obs (sorted by predicted prob)", "Predicted Pr(1)", "Actual"], _pred_rows,
                     col_formats={"Predicted Pr(1)": "0.0000"})
    elif _pva and _pva.get("x"):
        ws9 = wb.create_sheet("Predictions")
        _pred_rows = [{"Actual": x, "Predicted": y} for x, y in zip(_pva.get("x", []), _pva.get("y", []))]
        _write_table(ws9, ["Actual", "Predicted"], _pred_rows, col_formats={"Predicted": "0.0000"})

    # ── Segment comparison (Importance by Zone/Gender, on-screen expander) ────
    # Was on-screen only until now — exported so the Excel report matches what the driver-
    # regression tab actually shows, instead of the reader having to re-derive it manually.
    if segment_rows:
        ws10 = wb.create_sheet(f"Importance by {segment_dim or 'Segment'}")
        _seg_cols = ["Driver", "All"] + [k for k in segment_rows[0].keys()
                                          if k not in ("Driver", "All") and not k.endswith("_flag")]
        _seg_headers = ["Driver", "All"] + [f"{c} (↑/↓ vs All @ flagged confidence)" for c in _seg_cols[2:]]
        _seg_out_rows = []
        for r in segment_rows:
            out = {"Driver": r["Driver"], "All": r["All"]}
            for c in _seg_cols[2:]:
                flag = r.get(f"{c}_flag", "")
                val = r.get(c, 0)
                out[f"{c} (↑/↓ vs All @ flagged confidence)"] = f"{val:.1f}%{' ' + flag if flag else ''}"
            _seg_out_rows.append(out)
        _write_table(ws10, _seg_headers, _seg_out_rows, col_formats={"All": "0.0"})

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _render_driver_regression(sel_brand, brands_list, zone="all", gender="all",
                              age_band="all", city="all", project_id="project_1"):
    """XLSTAT-grade Key Driver REGRESSION: user picks the dependent variable and a
    top-box scale recode; imagery attributes are the independent drivers. Outputs a
    standardized-importance ranking, a driver-impact bar, and a cross-brand
    importance heatmap (which drivers matter most, for which brand)."""
    _CSAT_SCALE = _get_csat_scale()  # local so function is self-contained

    def _rf(d, key, default=0):
        """Safely get a float from result dict — guards against str errors from R bridge."""
        try: return float(d.get(key, default) or default)
        except (TypeError, ValueError): return float(default)
    st.markdown(
        "<div style='font-size:0.8rem;color:#6b7280;margin:2px 0 8px;'>"
        "Regression of the chosen dependent variable on the brand-imagery attributes. "
        "The DV can be an outcome (NPS / CSAT) <b>or any imagery attribute</b> — the remaining "
        "attributes become the predictors. Standardized coefficients (β); |β| share = relative "
        "importance (XLSTAT Key Driver convention).</div>", unsafe_allow_html=True)

    _target_brand = "All Brands (Pooled Category - XLSTAT)"  # always pooled — category model (XLSTAT)
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        _dv_type = st.selectbox(
            "Dependent variable",
            ["NPS", "CSAT", "Imagery attribute",
             "TOM", "SPONT", "AIDED", "CONSIDERATION", "EVER_USED", "CURRENT_USER", "PREFERRED"],
            key="dr_dv",
            help="NPS/CSAT = satisfaction outcomes. Imagery attribute = another attribute predicts this one. "
                 "Funnel stages (TOM/SPONT/AIDED/CONSIDERATION/EVER_USED/CURRENT_USER/PREFERRED) = "
                 "binary 0/1 — which imagery drivers predict reaching that funnel stage for this brand.")
    _FUNNEL_STAGES = {"TOM", "SPONT", "AIDED", "CONSIDERATION", "EVER_USED", "CURRENT_USER", "PREFERRED"}
    _dv_attr = None
    if _dv_type == "Imagery attribute":
        _attr_opts = _brand_attr_options(sel_brand, project_id=project_id)
        with c2:
            if not _attr_opts:
                st.info("No attributes with enough variance for this brand.")
                return
            _dv_attr = st.selectbox("Which attribute (DV)", _attr_opts, key="dr_dvattr",
                                    help="This attribute becomes the outcome; all others predict it.")
        _dv_kind = "ATTR"
        _topbox = 0  # attribute DV is already binary
    elif _dv_type in _FUNNEL_STAGES:
        _dv_kind = _dv_type   # passed through to _driver_regression_for_brand as funnel stage name
        _topbox  = 1          # funnel is already binary (0/1 hit)
        with c2:
            st.caption(f"Binary DV: 1 = respondent reached **{_dv_type}** for {_target_brand}, 0 = did not.")
    else:
        _dv_kind = _dv_type
        _dr_scale = _CSAT_SCALE if _dv_type == "CSAT" else 10
        _dr_max   = int(_dr_scale)
        _dr_t2, _dr_t3 = _dr_max - 1, _dr_max - 2
        _TOPBOX = {
            f"Top-2 Box ({_dr_t2}–{_dr_max} → 1)": _dr_t2,
            f"Top-3 Box ({_dr_t3}–{_dr_max} → 1)": _dr_t3,
            f"Top-1 Box ({_dr_max} → 1)":           _dr_max,
            f"Raw 0–{_dr_max} score":                0,
        }
        with c2:
            _tb_lbl = st.selectbox("Scale recode", list(_TOPBOX), index=0, key="dr_topbox",
                                   help="Scale the rating down to a binary outcome. "
                                        "Standard for promoter/top-box key-driver models.")
        _topbox = _TOPBOX[_tb_lbl]
    with c3:
        _model_lbl = st.selectbox("Model", ["Linear (LPM)", "Logistic", "Random Forest"], key="dr_model",
                                  help="Linear = OLS (linear-probability for a binary outcome). "
                                       "Logistic = correct model for a binary outcome (odds ratios). "
                                       "Random Forest = nonlinear ensemble; no coefficients, ranks drivers "
                                       "by how much prediction error rises without each one.")
    _model = "logistic" if _model_lbl.startswith("Logistic") else ("random_forest" if _model_lbl.startswith("Random") else "linear")
    with c4:
        _xbrands = st.slider("Cross-brand count", 3, 8, 5, 1, key="dr_xbrands")

    # Outcome is binary for an imagery-attribute DV (0/1) or any top-box recode.
    _dv_is_binary = (_dv_kind == "ATTR") or (_topbox > 0)
    if _model == "logistic" and not _dv_is_binary:
        st.warning("Logistic regression needs a binary outcome — pick a top-box recode "
                   "(e.g. 9–10 → 1), or switch to Linear for the raw 0–10 score. Using Linear for now.")
        _model = "linear"
    _is_logit = (_model == "logistic")

    if _model == "random_forest":
        _render_random_forest_driver(_target_brand, _dv_kind, _topbox, zone, gender, age_band, city, _dv_attr,
                                     project_id=project_id,
                                     _out_lbl_for_rf=(
            f"'{_clean_attr(_dv_attr)}'" if _dv_kind == "ATTR" else f"{_dv_kind}" + ("" if _topbox == 0 else f" top-box (≥{_topbox})")
        ))
        return

    # ── 1. Focus-brand regression ────────────────────────────────────────────
    with st.spinner(f"Running {'logistic' if _is_logit else 'linear'} driver regression for {_target_brand}…"):
        res = _driver_regression_for_brand(_target_brand, _dv_kind, _topbox, zone, gender, age_band, city, _model, _dv_attr, project_id=project_id)

    if not res:
        st.info(f"No {_dv_kind} + imagery data for {sel_brand}.")
        return
    if "error" in res:
        if res.get("error") == "insufficient":
            st.info(f"Need ≥30 respondents with {_dv_kind} + imagery data for {sel_brand} "
                    f"(found {res.get('n', 0)} after filters).")
        else:
            st.warning(f"Regression error: {res['error']}")
        return

    drivers = res.get("significant_drivers", [])
    if not drivers:
        st.info("Regression returned no drivers.")
        return

    # Store significant driver attr_ids in session state so BIP + CAN MAP tabs can filter to them.
    # R mangles column names (spaces/parens → dots), so exact label match fails.
    # Use _norm_key (strips all non-alphanumeric) to match mangled R names to DB labels.
    try:
        import sqlite3 as _drsql
        from oxdata.db_loader import get_db_path as _drgdb
        _dr_active_pid = st.session_state.get("active_project_id", "project_1")
        _dr_conn = _drsql.connect(f"file:{_drgdb(project_id=_dr_active_pid)}?mode=ro", uri=True)
        _all_attr_rows = _dr_conn.execute(
            "SELECT attr_id, attr_label FROM dim_bq3_attribute"
        ).fetchall()
        _dr_conn.close()
        _norm_to_id = {_norm_key(row[1]): row[0] for row in _all_attr_rows}
        _dr_ids = []
        for _d in drivers:
            _aid = _norm_to_id.get(_norm_key(_d["attribute"]))
            if _aid is not None:
                _dr_ids.append(_aid)
        st.session_state["driver_flow_attr_ids"] = _dr_ids
        st.session_state["driver_flow_labels"]   = [_clean_attr(d["attribute"]) for d in drivers]
        st.session_state["driver_flow_brand"]    = sel_brand
        st.session_state["driver_flow_outcome"]  = _dv_kind
        # Awareness stages used in regression — passed to driven CAN MAP for consistent respondent universe
        _FUNNEL_STAGES_INNER = {"TOM", "SPONT", "AIDED", "CONSIDERATION", "EVER_USED", "CURRENT_USER", "PREFERRED"}
        if _dv_kind in _FUNNEL_STAGES_INNER:
            # Gate: must be aware at AIDED level to be eligible as 0 case
            _dr_aware_stages = ["TOM", "SPONT", "AIDED"]
        elif _dv_kind == "CSAT":
            _dr_aware_stages = ["LAST_PURCHASED"]
        else:
            _dr_aware_stages = ["TOM", "SPONT", "AIDED"]
        st.session_state["driver_flow_awareness_stages"] = _dr_aware_stages
    except Exception:
        pass

    _workbench_capture(
        f"Driver Regression — {sel_brand} ({'Logistic' if _is_logit else 'Linear/OLS'})",
        {"brand": sel_brand, "model": "logistic" if _is_logit else "linear",
         "outcome": _dv_kind, "topbox_min": _topbox, "attribute_dv": _dv_attr,
         "result": res, "drivers": drivers},
    )

    m1, m2, m3, m4 = st.columns(4)
    def _safe_f(v, default=0):
        try:
            return float(v)
        except (TypeError, ValueError):
            return float(default)
    if _is_logit:
        m1.metric("McFadden R²", f"{_safe_f(res.get('mcfadden_r2', res.get('r_squared', 0))):.3f}",
                  help="Pseudo-R² for the logistic model")
        m2.metric("Base rate", f"{_safe_f(res.get('base_rate', 0)):.1%}", help="Share of respondents in the top box")
        m3.metric("Model", "Logistic")
    else:
        m1.metric("R² (fit)", f"{_safe_f(res.get('r_squared', 0)):.3f}",
                  help="Variance in the outcome explained by all imagery drivers")
        m2.metric("Adj. R²", f"{_safe_f(res.get('adj_r_squared', 0)):.3f}")
        m3.metric("F-stat", f"{_safe_f(res.get('f_statistic', 0)):.1f}")
    m4.metric("Respondents", res.get("n", 0))

    if _dv_kind == "ATTR":
        _out_lbl = f"'{_clean_attr(_dv_attr)}'"
    else:
        _out_lbl = f"{_dv_kind}" + ("" if _topbox == 0 else f" top-box (≥{_topbox})")

    _model_note = "Logistic regression (binary outcome)" if _is_logit else "Linear regression (OLS)"
    st.caption(f"Model: **{_model_note}** · outcome = {_out_lbl} · predictors = "
               f"{res.get('n_attrs', len(drivers))} imagery attributes · n = {res.get('n',0):,} respondents.")

    _pct_dropped = res.get("pct_dropped_na", 0) or 0
    if _pct_dropped > 5:
        st.warning(f"⚠️ {_pct_dropped:.1f}% of respondents were dropped for missing data on one or more "
                   f"selected attributes (n={res.get('n_before_na_omit', '?')} → n={res.get('n', 0)}). "
                   f"Results below may be unstable — try fewer/more common attributes.")

    # ── Calculation Showcase — transparency: show the actual fitted equation and
    # the importance-% formula with real substituted numbers, not just the results.
    # Formulas verified against a manual XLSTAT-workbook calculation (Wald CI and
    # |std β| / Σ|std β| × 100) — see .regression_reference/AUDIT.md.
    with st.expander("📐 How this is calculated — Model Equation", expanded=False):
        _intercept = res.get("intercept")
        _eq_drivers = sorted(drivers, key=lambda d: abs(d.get("coef", 0)), reverse=True)
        _EQ_CAP = 8
        _eq_terms = [f"{d.get('coef', 0):+.4f}×{_clean_attr(d['attribute'])}"
                    for d in _eq_drivers[:_EQ_CAP]]
        _eq_body = " ".join(_eq_terms)
        _more_note = f"  + {len(_eq_drivers) - _EQ_CAP} more terms" if len(_eq_drivers) > _EQ_CAP else ""
        if _intercept is not None:
            if _is_logit:
                st.markdown(f"**Fitted model** (top {min(_EQ_CAP, len(_eq_drivers))} of "
                           f"{len(_eq_drivers)} predictors by |coefficient|):")
                st.code(f"Pr({_out_lbl}=1) = 1 / (1 + exp(-({_intercept:+.4f} {_eq_body}{_more_note})))",
                       language=None)
            else:
                st.markdown(f"**Fitted model** (top {min(_EQ_CAP, len(_eq_drivers))} of "
                           f"{len(_eq_drivers)} predictors by |coefficient|):")
                st.code(f"{_out_lbl} = {_intercept:+.4f} {_eq_body}{_more_note}", language=None)
            st.caption("Coefficients and predictor names are substituted live from this run's fitted model — "
                      "not a template.")
        else:
            st.info("Model equation unavailable (intercept not returned by this analysis).")

        st.divider()
        st.markdown("**Importance % formula** — how the ranking above is derived:")
        _top2 = [d for d in _eq_drivers if d.get("std_coef") is not None][:2]
        if _top2:
            def _fcoef(v):
                try: return float(v)
                except: return 0.0
            _tot_abs_std = sum(abs(_fcoef(d.get("std_coef", 0))) for d in drivers if d.get("std_coef") is not None)
            for d in _top2:
                _std = _fcoef(d.get("std_coef", 0))
                _imp = _fcoef(d.get("importance", 0))
                st.markdown(
                    f"Importance % = |std β| ÷ Σ|std β| × 100 — for **{_clean_attr(d['attribute'])}**: "
                    f"|{_std:.4f}| ÷ {_tot_abs_std:.4f} × 100 = **{_imp:.2f}%**"
                )
        st.caption("Same formula XLSTAT uses (verified against the audited sample workbook's manual "
                  "cross-check formula, cell-for-cell).")

    # ── Goodness of fit (XLSTAT full parity) ─────────────────────────────────
    def _p_flag(p):
        if p is None: return ""
        if p < 0.001: return " ***"
        if p < 0.01:  return " **"
        if p < 0.05:  return " *"
        return " ns"

    if _is_logit:
        _lrt_p  = res.get("lrt_p", None)
        _hl     = res.get("hosmer_lemeshow") or {}
        _hl_p   = _hl.get("p_value")
        _hl_ok  = _hl.get("good_fit")
        _gof_rows = [
            {"Statistic": "McFadden pseudo-R²",      "Value": f"{_rf(res,'mcfadden_r2'):.4f}",
             "Note": "0.2+ = adequate, 0.4+ = good"},
            {"Statistic": "Cox & Snell R²",           "Value": f"{_rf(res,'cox_snell_r2'):.4f}",
             "Note": "cannot reach 1.0 for binary DV"},
            {"Statistic": "Nagelkerke R²",            "Value": f"{_rf(res,'nagelkerke_r2'):.4f}",
             "Note": "scaled to [0,1] — most cited"},
            {"Statistic": "AUC (c-statistic)",        "Value": f"{_rf(res,'auc'):.4f}",
             "Note": "0.7+ acceptable, 0.8+ good"},
            {"Statistic": "AIC",                      "Value": f"{_rf(res,'aic'):.2f}",
             "Note": "lower = better fit (penalises complexity)"},
            {"Statistic": "BIC",                      "Value": f"{_rf(res,'bic'):.2f}",
             "Note": "stricter penalty than AIC"},
            {"Statistic": "LRT χ² (model sig.)",      "Value":
             f"{_rf(res,'lrt_chi2'):.3f}  df={int(_rf(res,'lrt_df'))}"
             f"  p={_lrt_p:.4f}{_p_flag(_lrt_p)}" if _lrt_p is not None else "—",
             "Note": "overall model vs null (intercept-only)"},
            {"Statistic": "Hosmer-Lemeshow χ²",       "Value":
             f"{_rf(_hl,'chi2'):.3f}  df={int(_rf(_hl,'df',8))}"
             f"  p={_hl_p:.4f}{_p_flag(_hl_p)}" if _hl_p is not None else "—",
             "Note": "p>0.05 = good calibration ✓" if _hl_ok else ("p<0.05 = calibration gap ⚠️" if _hl_ok is False else "")},
            {"Statistic": "Base rate (top-box share)", "Value": f"{_rf(res,'base_rate'):.1%}",
             "Note": ""},
            {"Statistic": "Observations (n)",          "Value": f"{res.get('n', 0):,}", "Note": ""},
            {"Statistic": "Predictors",                "Value": f"{res.get('n_attrs', 0)}", "Note": ""},
        ]
    else:
        _fp  = res.get("f_p_value")
        _sw  = res.get("shapiro_wilk") or {}
        _sw_p = _sw.get("p_value")
        _gof_rows = [
            {"Statistic": "R²",           "Value": f"{_rf(res,'r_squared'):.4f}",
             "Note": "variance explained"},
            {"Statistic": "Adjusted R²",  "Value": f"{_rf(res,'adj_r_squared'):.4f}",
             "Note": "penalises added predictors"},
            {"Statistic": "F (model)",    "Value":
             f"{_rf(res,'f_statistic'):.2f}"
             + (f"  p={_fp:.4f}{_p_flag(_fp)}" if _fp is not None else ""),
             "Note": "overall model significance"},
            {"Statistic": "RMSE",         "Value": f"{_rf(res,'rmse'):.4f}", "Note": ""},
            {"Statistic": "AIC",          "Value": f"{_rf(res,'aic'):.2f}", "Note": ""},
            {"Statistic": "BIC",          "Value": f"{_rf(res,'bic'):.2f}", "Note": ""},
            {"Statistic": "Shapiro-Wilk (residuals)", "Value":
             f"W={_sw.get('w_stat', 0):.4f}  p={_sw_p:.4f}{_p_flag(_sw_p)}" if _sw_p is not None else "—",
             "Note": "p>0.05 = residuals normal ✓" if _sw.get("normal") else (
                 "p<0.05 = non-normal residuals ⚠️" if _sw.get("normal") is False else "")},
            {"Statistic": "Observations (n)", "Value": f"{res.get('n', 0):,}", "Note": ""},
            {"Statistic": "Predictors",       "Value": f"{res.get('n_attrs', 0)}", "Note": ""},
        ]

    # Headline metrics (McFadden R²/R², n, etc.) already shown as cards above — this full table
    # is the same numbers plus the ones nobody asked for by name (AIC/BIC/Hosmer-Lemeshow/
    # Shapiro-Wilk), each with a plain-English "Note" already, but a wall of 8-12 statistics is
    # still a lot to land on by default. One click away, not hidden.
    with st.expander("📊 Full goodness-of-fit statistics", expanded=False):
        _gof_df = pd.DataFrame(_gof_rows)
        st.dataframe(_gof_df, hide_index=True, use_container_width=True)

    if _is_logit:
        # Classification table (confusion matrix at 0.5 threshold) — laid out
        # exactly like XLSTAT's "Classification table for the training sample":
        # rows = observed ("from"), columns = predicted ("to"), with a Total
        # column/row and a "% correct" column per observed class.
        # Own full-width row (not squeezed beside the GOF table) — heatmap and
        # tables each get real room instead of fighting for a 1.4/3 column.
        _cm = res.get("confusion") or {}
        if _cm:
            st.divider()
            st.markdown("**Classification table** *(threshold = 0.5)*")
            _tp, _tn, _fp2, _fn = _cm.get("tp",0), _cm.get("tn",0), _cm.get("fp",0), _cm.get("fn",0)
            # Observed 0 row: predicted-0 = tn, predicted-1 = fp2
            # Observed 1 row: predicted-0 = fn, predicted-1 = tp
            _row0_total = _tn + _fp2
            _row1_total = _fn + _tp
            _cls_c1, _cls_c2 = st.columns([1, 1.2])
            with _cls_c1:
                _z = go.Figure(go.Heatmap(
                    z=[[_tn, _fp2], [_fn, _tp]],
                    x=["Predicted 0", "Predicted 1"],
                    y=["Observed 0", "Observed 1"],
                    text=[[f"{_tn}<br>({_tn/_row0_total:.1%})" if _row0_total else "0",
                           f"{_fp2}<br>({_fp2/_row0_total:.1%})" if _row0_total else "0"],
                          [f"{_fn}<br>({_fn/_row1_total:.1%})" if _row1_total else "0",
                           f"{_tp}<br>({_tp/_row1_total:.1%})" if _row1_total else "0"]],
                    texttemplate="%{text}", textfont=dict(size=13),
                    colorscale=[[0, "#fee2e2"], [1, "#1a5d4d"]],
                    showscale=False,
                ))
                _bcm = {k: v for k, v in _chart_layout_base(300).items() if k not in ("xaxis", "yaxis")}
                _z.update_layout(**_bcm,
                    xaxis=dict(title="Predicted class", side="bottom"),
                    yaxis=dict(title="Observed class", autorange="reversed"))
                st.plotly_chart(_theme_fig(_z), use_container_width=True)
            with _cls_c2:
                _cm_df = pd.DataFrame({
                    "from \\ to":  ["Observed 0", "Observed 1", "Total"],
                    "Predicted 0": [_tn, _fn, _tn+_fn],
                    "Predicted 1": [_fp2, _tp, _fp2+_tp],
                    "Total":       [_row0_total, _row1_total, res.get("n", _row0_total+_row1_total)],
                    "% correct":   [f"{_tn/_row0_total:.1%}" if _row0_total else "—",
                                    f"{_tp/_row1_total:.1%}" if _row1_total else "—",
                                    f"{res.get('accuracy',0):.1%}"],
                })
                st.dataframe(_cm_df, hide_index=True, use_container_width=True)
                _disc_rows = [
                    {"Metric": "Accuracy",    "Value": f"{res.get('accuracy',0):.1%}"},
                    {"Metric": "Sensitivity (recall)", "Value": f"{res.get('sensitivity',0):.1%}"},
                    {"Metric": "Specificity", "Value": f"{res.get('specificity',0):.1%}"},
                    {"Metric": "PPV (precision)", "Value": f"{res.get('ppv',0):.1%}"},
                    {"Metric": "NPV",         "Value": f"{res.get('npv',0):.1%}"},
                ]
                st.dataframe(pd.DataFrame(_disc_rows), hide_index=True, use_container_width=True)
            st.caption("Rows = actual outcome, columns = the model's predicted outcome at the 0.5 cutoff "
                      "(same orientation XLSTAT uses). Diagonal cells (top-left, bottom-right) are correct "
                      "calls; off-diagonal are misses. % correct = how often that row's actual class was "
                      "predicted correctly.")
    else:
        # OLS ANOVA decomposition — own full-width row too
        if res.get("ss_total"):
            st.divider()
            _fp3 = res.get("f_p_value")
            st.markdown("**Analysis of variance (ANOVA)**")
            _ms_m = (_rf(res,"ss_model") / (_rf(res,"df_model") or 1))
            _ms_r = (_rf(res,"ss_resid") / (_rf(res,"df_resid") or 1))
            _anova_df = pd.DataFrame([
                {"Source":"Model",    "SS": _rf(res,"ss_model"), "df": int(_rf(res,"df_model")),
                 "MS": round(_ms_m, 2),
                 "F / p": f"F={_rf(res,'f_statistic'):.2f}  p={_fp3:.4f}{_p_flag(_fp3)}" if _fp3 is not None else f"F={_rf(res,'f_statistic'):.2f}"},
                {"Source":"Residual", "SS": _rf(res,"ss_resid"), "df": int(_rf(res,"df_resid")),
                 "MS": round(_ms_r, 2), "F / p": "—"},
                {"Source":"Total",    "SS": res.get("ss_total",0),
                 "df": int((res.get("df_model",0) or 0)+(res.get("df_resid",0) or 0)),
                 "MS": float("nan"), "F / p": "—"},
            ])
            st.dataframe(_anova_df.style.format({"SS": "{:,.1f}", "MS": "{:,.2f}"}, na_rep="—"),
                         hide_index=True, use_container_width=True)

    # ── VIF table (multicollinearity check) ──────────────────────────────────
    _vif = res.get("vif") or {}
    if _vif:
        _vif_rows = sorted(
            [{"Attribute": _clean_attr(k), "VIF": v,
              "Flag": "⚠️ high" if v > 10 else ("⚡ moderate" if v > 5 else "✓ ok")}
             for k, v in _vif.items()],
            key=lambda r: -r["VIF"])
        _high_vif = [r for r in _vif_rows if r["VIF"] > 5]
        with st.expander(f"Multicollinearity check — VIF"
                         + (f"  ⚠️ {len(_high_vif)} attributes with VIF > 5" if _high_vif else "  ✓ all clear")):
            _vif_s = pd.DataFrame(_vif_rows).style
            _vif_cell = getattr(_vif_s, "map", getattr(_vif_s, "applymap", None))
            _vif_s = _vif_cell(lambda v: "color:#dc2626;font-weight:700" if "high" in str(v) else
                                          ("color:#d97706" if "moderate" in str(v) else ""),
                               subset=["Flag"]).format({"VIF": "{:.2f}"})
            st.dataframe(_vif_s,
                hide_index=True, use_container_width=True)
            st.caption("VIF > 5 = moderate multicollinearity; > 10 = severe. "
                       "High VIF means two attributes co-occur so strongly that their individual "
                       "coefficients (and importance shares) become unreliable.")

    # ── Type II analysis (XLSTAT "Type II analysis") ─────────────────────────
    # Per-predictor significance from single-term deletion — independent of entry
    # order, unlike the Wald/t-test already shown per driver above.
    _type2 = res.get("type2") or []
    if _type2:
        _t2_rows = sorted(
            [{"Attribute": _clean_attr(t["attribute"]), "df": t.get("df", 1),
              ("Chi-square (LR)" if _is_logit else "F"):
                  t.get("lr_chi2") if _is_logit else t.get("f_stat"),
              "p-value": t.get("p_value"),
              "Significant": "✓" if _rf(t, "p_value", 1) < 0.05 else ""}
             for t in _type2],
            key=lambda r: (_rf(r, "p-value", 1) if r.get("p-value") is not None else 1))
        with st.expander("Type II analysis — per-predictor significance (order-independent)"):
            st.dataframe(pd.DataFrame(_t2_rows), hide_index=True, use_container_width=True)
            st.caption("Tests each predictor's contribution with every other predictor already in the model — "
                      "unlike entry-order-dependent Type I tests, this matches XLSTAT's default 'Type II "
                      "analysis' and confirms the Wald/t-test significance shown per driver above.")

    # ── Global significance tests (XLSTAT "Test of H0") ──────────────────────
    # XLSTAT reports 3 overall-model tests (-2LL/LRT, Score, Wald); the app
    # previously only surfaced LRT. Show all 3 side by side.
    _gtests = res.get("global_tests") or {}
    if _gtests:
        _gt_rows = []
        _gt_labels = [("minus2ll", "-2 Log-Likelihood (LRT)"), ("score", "Score"), ("wald", "Wald")]
        for _key, _label in _gt_labels:
            _g = _gtests.get(_key)
            if _g:
                _gt_rows.append({"Test": _label, "Chi-square": _g.get("chi2"),
                                  "df": _g.get("df"), "p-value": _g.get("p_value"),
                                  "Significant": "✓" if (_g.get("p_value") or 1) < 0.05 else ""})
        if _gt_rows:
            with st.expander("Global tests — overall model significance (3-test parity)"):
                st.dataframe(pd.DataFrame(_gt_rows), hide_index=True, use_container_width=True)
                st.caption("Three equivalent tests of H0: all predictor coefficients = 0. XLSTAT reports all "
                          "three; they should broadly agree — large divergence between them is itself a "
                          "diagnostic signal (e.g. quasi-complete separation).")

    # ── Correlation matrix of predictors (XLSTAT "Correlation matrix") ───────
    _corrm = res.get("correlation_matrix") or {}
    _corr_attrs = _corrm.get("attributes") or []
    _corr_mat = _corrm.get("matrix") or []
    if _corr_attrs and _corr_mat:
        with st.expander(f"Correlation matrix — predictor pairs ({len(_corr_attrs)}×{len(_corr_attrs)})"):
            _corr_labels = [_clean_attr(a) for a in _corr_attrs]
            _fig_corr = go.Figure(go.Heatmap(
                z=_corr_mat, x=_corr_labels, y=_corr_labels,
                colorscale="RdBu", zmid=0, zmin=-1, zmax=1,
                text=[[f"{v:.2f}" for v in row] for row in _corr_mat],
                texttemplate="%{text}", textfont=dict(size=8),
                colorbar=dict(title="r")))
            _bcorr = {k: v for k, v in _chart_layout_base(max(320, 24 * len(_corr_attrs))).items()
                      if k not in ("xaxis", "yaxis")}
            _fig_corr.update_layout(**_bcorr,
                                    xaxis=dict(tickfont=dict(size=9)),
                                    yaxis=dict(tickfont=dict(size=9), autorange="reversed"),
                                    title=dict(text="Predictor correlation matrix (Pearson)", font=dict(size=12)))
            st.plotly_chart(_theme_fig(_fig_corr), use_container_width=True)
            _corr_df = pd.DataFrame(_corr_mat, index=_corr_labels, columns=_corr_labels)
            st.dataframe(_corr_df.style.format("{:.3f}").background_gradient(
                cmap="RdBu_r", vmin=-1, vmax=1), use_container_width=True)
            st.caption("Pairwise Pearson correlation between predictor attributes (not with the outcome). "
                      "High |r| between two predictors (rule of thumb: |r| > 0.7) signals redundancy that "
                      "shows up as inflated VIF above — this table lets you see which specific pairs are driving it.")

    # ── Global significance tests (XLSTAT "Test of H0") — logistic only ──────
    # XLSTAT reports 3 overall-model tests side by side; the app previously only
    # surfaced the -2LL/LRT row. OLS's equivalent overall test is the F-statistic
    # already shown in the ANOVA table above, so this block is logistic-only.
    _gtests = res.get("global_tests") or {}
    if _gtests:
        with st.expander("Test of H0 — global significance (3 tests)"):
            _gt_rows = []
            for _key, _lbl in (("minus2ll", "−2 Log(Likelihood)"), ("score", "Score"), ("wald", "Wald")):
                _t = _gtests.get(_key)
                if _t:
                    _gt_rows.append({"Statistic": _lbl, "df": _t.get("df"),
                                      "Chi-square": _t.get("chi2"),
                                      "p-value": _t.get("p_value"),
                                      "Significant": "✓" if (_t.get("p_value") if _t.get("p_value") is not None else 1) < 0.05 else ""})
            st.dataframe(pd.DataFrame(_gt_rows), hide_index=True, use_container_width=True)
            st.caption("All 3 tests ask the same question — does the full model fit meaningfully better than "
                      "an intercept-only model? — via different statistical approximations. −2LL (likelihood "
                      "ratio) is the standard/most-cited; Score and Wald are shown for full XLSTAT parity. "
                      "They should (and normally do) agree qualitatively on significance.")

    # ── Correlation matrix of predictors (XLSTAT "Correlation matrix") ───────
    _corrmat = res.get("correlation_matrix") or {}
    _corr_attrs = _corrmat.get("attributes") or []
    if _corr_attrs and _corrmat.get("matrix"):
        _corr_labels = [_clean_attr(a) for a in _corr_attrs]
        _corr_z = _corrmat["matrix"]
        with st.expander(f"Correlation matrix — {len(_corr_attrs)} predictors"):
            _fig_corr = go.Figure(go.Heatmap(
                z=_corr_z, x=_corr_labels, y=_corr_labels,
                zmid=0, zmin=-1, zmax=1,
                colorscale=[[0, "#dc2626"], [0.5, "#f9fafb"], [1, "#1a5d4d"]],
                colorbar=dict(title="r"),
            ))
            _bcorr = {k: v for k, v in _chart_layout_base(max(360, 24 * len(_corr_attrs))).items()
                      if k not in ("xaxis", "yaxis")}
            _fig_corr.update_layout(**_bcorr,
                xaxis=dict(tickfont=dict(size=9), tickangle=-45),
                yaxis=dict(tickfont=dict(size=9), autorange="reversed"))
            st.plotly_chart(_theme_fig(_fig_corr), use_container_width=True)
            _corr_df = pd.DataFrame(_corr_z, columns=_corr_labels, index=_corr_labels)
            st.dataframe(_corr_df.style.format("{:.2f}").background_gradient(
                cmap="RdYlGn", vmin=-1, vmax=1) if importlib.util.find_spec("matplotlib") else _corr_df,
                use_container_width=True)
            st.caption("Pearson correlation between predictor attributes. High off-diagonal values (near ±1) "
                      "flag redundant predictors — cross-check against the VIF table above.")

    # ── Influence diagnostics (XLSTAT "Influence diagnostics") ───────────────
    # Cook's distance + leverage, summarised to the most influential respondents
    # rather than a full per-row dump.
    _infl = res.get("influence") or {}
    _top_infl = _infl.get("top_influential") or []
    if _top_infl:
        _n_hl = _infl.get("n_high_leverage", 0)
        _n_hc = _infl.get("n_high_cooks_d", 0)
        with st.expander(f"Influence diagnostics — outlier check"
                         + (f"  ⚠️ {_n_hc} high-influence respondents" if _n_hc else "  ✓ no high-influence outliers")):
            _lev_cut = _infl.get("leverage_cutoff")
            _cd_cut  = _infl.get("cooksd_cutoff")
            _fig_infl = go.Figure(go.Scatter(
                x=[r["leverage"] for r in _top_infl], y=[r["cooks_d"] for r in _top_infl],
                mode="markers",
                marker=dict(size=8,
                           color=["#dc2626" if r["high_cooks_d"] else "#1a5d4d" for r in _top_infl],
                           opacity=0.75),
                text=[f"Respondent #{r['obs']}" for r in _top_infl], hoverinfo="text+x+y"))
            # This panel only ever shows the top-20 by Cook's D, so on a large sample
            # every plotted point legitimately clears the population-level cutoff —
            # that's correct, not a bug. But it means the cutoff sits far outside the
            # data range, so Plotly's autorange silently clips the dashed reference
            # line off-screen. Force both axes to start at 0 and extend past the
            # cutoff so the line — the whole point of showing it — is actually visible.
            _lev_vals = [_rf(r, "leverage") for r in _top_infl]
            _cd_vals  = [_rf(r, "cooks_d") for r in _top_infl]
            _x_hi = max(_lev_vals + ([float(_lev_cut)] if _lev_cut else [0.001])) * 1.12
            _y_hi = max(_cd_vals  + ([float(_cd_cut)]  if _cd_cut  else [0.001])) * 1.12
            if _lev_cut:
                _fig_infl.add_vline(x=_lev_cut, line=dict(color="#9ca3af", dash="dot"),
                                     annotation_text=f"leverage cutoff {_lev_cut:.3f}",
                                     annotation_position="top", annotation_font=dict(size=9, color="#6b7280"))
            if _cd_cut:
                _fig_infl.add_hline(y=_cd_cut, line=dict(color="#9ca3af", dash="dot"),
                                     annotation_text=f"Cook's D cutoff {_cd_cut:.4f}",
                                     annotation_position="right", annotation_font=dict(size=9, color="#6b7280"))
            _binfl = {k: v for k, v in _chart_layout_base(320).items() if k not in ("xaxis", "yaxis")}
            _fig_infl.update_layout(**_binfl, xaxis=dict(title="Leverage (hat value)", range=[0, _x_hi]),
                                    yaxis=dict(title="Cook's distance", range=[0, _y_hi]),
                                    title=dict(text="Top 20 most-influential respondents", font=dict(size=12)))
            st.plotly_chart(_theme_fig(_fig_infl), use_container_width=True)
            st.dataframe(pd.DataFrame([
                {"Respondent": r["obs"], "Cook's D": r["cooks_d"], "Leverage": r["leverage"],
                 "Flag": "⚠️ high influence" if r["high_cooks_d"] else ("high leverage" if r["high_leverage"] else "")}
                for r in _top_infl]), hide_index=True, use_container_width=True)
            _has_dfbeta = any(r.get("dfbeta") for r in _top_infl)
            if _has_dfbeta:
                with st.expander("DFBeta per predictor — top 20 most-influential respondents"):
                    _dfb_rows = []
                    for r in _top_infl:
                        _row = {"Respondent": r["obs"]}
                        for _k, _v in (r.get("dfbeta") or {}).items():
                            _row[_clean_attr(_k) if _k != "(Intercept)" else _k] = _v
                        _dfb_rows.append(_row)
                    st.dataframe(pd.DataFrame(_dfb_rows), hide_index=True, use_container_width=True)
                    st.caption("DFBeta = how much each coefficient shifts if that respondent were dropped from "
                              "the model — one column per predictor, one row per top-20-by-Cook's-D respondent. "
                              "Large |DFBeta| on a specific predictor flags that respondent as disproportionately "
                              "driving that one coefficient, not just the model overall.")
            _lev_cut_txt = f"leverage > {_lev_cut:.3f}" if _lev_cut else "leverage > 2p/n"
            _cd_cut_txt  = f"Cook's D > {_cd_cut:.4f}" if _cd_cut else "Cook's D > 4/n"
            st.caption(f"Leverage = how unusual a respondent's predictor values are; Cook's distance = how much "
                      f"the fitted model would change if that respondent were removed. This panel lists only the "
                      f"**top 20 by Cook's D** — so on a large sample, most or all of them will legitimately clear "
                      f"XLSTAT's cutoffs (dashed lines: {_lev_cut_txt}, {_cd_cut_txt}) — that's expected, not an "
                      f"error. The number in the expander title ({_n_hc} of {res.get('n', 0):,} respondents) is "
                      f"the real population-wide count worth acting on; the chart below just shows which "
                      f"respondents those are.")

    # ── Diagnostic filters ────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("**Diagnostic filters**")
    _df1, _df2, _df3, _df4, _df5 = st.columns(5)
    with _df1:
        _sig_thresh = st.selectbox("Significance threshold", [0.05, 0.01, 0.001, 0.10],
                                   format_func=lambda x: f"p < {x}",
                                   key="dr_sig_thresh",
                                   help="Only drivers below this p-value are 'significant'")
    with _df2:
        _min_imp = st.slider("Min importance %", 0.0, 10.0, 0.0, 0.5,
                             key="dr_min_imp",
                             help="Hide drivers below this importance share")
    with _df3:
        _dir_filter = st.selectbox("Direction", ["Both", "Positive only (raises)", "Negative only (lowers)"],
                                   key="dr_dir",
                                   help="Filter by coefficient direction")
    with _df4:
        _fmap2 = {_norm_key(k): v for k, v in _attr_feature_map().items()}
        _all_themes = sorted({str(_fmap2.get(_norm_key(d["attribute"]), "Other") or "Other") for d in drivers})
        _theme_filter = st.selectbox("Attribute theme", ["All themes"] + _all_themes,
                                     key="dr_theme",
                                     help="Show only drivers from one attribute theme")
    with _df5:
        _top_n_chart = st.slider("Top N in charts", 5, 30, 15, 1,
                                 key="dr_topn",
                                 help="How many top drivers to show in quadrant and heatmap")

    # Apply filters to get working driver set
    def _apply_dr_filters(drv_list):
        out = []
        for d in drv_list:
            p = d.get("p_value")
            p = 1.0 if p is None else p
            imp = d.get("importance")
            imp = 0 if imp is None else imp
            sc = d.get("std_coef")
            sc = 0 if sc is None else sc
            theme = _fmap2.get(_norm_key(d["attribute"]), "Other")
            if p > _sig_thresh: continue
            if imp < _min_imp: continue
            if _dir_filter == "Positive only (raises)" and sc < 0: continue
            if _dir_filter == "Negative only (lowers)" and sc > 0: continue
            if _theme_filter != "All themes" and theme != _theme_filter: continue
            out.append(d)
        return out

    # Filtered significant set — drives BOTH table and all actionable views.
    _sig = _apply_dr_filters(drivers)
    _pos = sorted([d for d in _sig if (d.get("std_coef") or 0) > 0], key=lambda d: -d.get("importance", 0))
    _neg = sorted([d for d in _sig if (d.get("std_coef") or 0) < 0], key=lambda d: -d.get("importance", 0))

    # ── Model parameters table ────────────────────────────────────────────────
    _filter_active = (_sig_thresh != 0.05 or _min_imp > 0 or
                      _dir_filter != "Both" or _theme_filter != "All themes")
    _tbl_label = (f"**Drivers matching current filters** — {len(_sig)} of {len(drivers)} total"
                  if _filter_active else
                  f"**Significant drivers of {_out_lbl}** — {len(_sig)} of {len(drivers)} total")
    st.markdown(_tbl_label)
    st.caption(
        f"We tested all {len(drivers)} imagery attributes as possible drivers. Only **{len(_sig)}** cleared "
        f"the significance bar (p < {_sig_thresh}) — meaning we're confident their effect on {_out_lbl} is real, "
        f"not just noise from this sample. The other {len(drivers)-len(_sig)} may still show an 'Importance %' "
        f"number, but it's not reliable enough to act on. Lower the significance threshold above to see more "
        f"(less strict), or raise it to see fewer (more strict).")

    # Flag when a higher-importance driver got excluded purely on significance —
    # importance % is computed across ALL predictors regardless of p-value, so it
    # does NOT rank the same as the significant-only list. This is expected, but
    # confusing without calling it out.
    _excluded = [d for d in drivers if d not in _sig]
    if _sig and _excluded:
        _min_sig_imp = min((d.get("importance") or 0) for d in _sig)
        _bigger_excluded = sorted(
            [d for d in _excluded if (d.get("importance") or 0) > _min_sig_imp],
            key=lambda d: -(d.get("importance") or 0))
        if _bigger_excluded:
            _ex0 = _bigger_excluded[0]
            st.warning(
                f"⚠️ **'{_clean_attr(_ex0['attribute'])}'** has a higher Importance % "
                f"({_ex0.get('importance',0):.1f}%) than some drivers shown above, but it's **not** in this list "
                f"because its p-value ({_ex0.get('p_value'):.3f}) didn't clear the significance bar. "
                f"In plain terms: its estimated effect looks big, but we can't be statistically confident it's "
                f"real rather than random noise — often because it overlaps heavily with other attributes "
                f"(respondents who pick one tend to pick several similar ones, which makes each one's individual "
                f"effect hard to pin down precisely). Treat its number as unreliable, not as a hidden driver. "
                + (f"{len(_bigger_excluded)-1} other driver(s) have the same issue." if len(_bigger_excluded) > 1 else ""))

    def _prow(d):
        theme = _fmap2.get(_norm_key(d["attribute"]), "Other")
        sc = d.get("std_coef") or 0
        direction = "▲ raises" if sc > 0 else ("▼ lowers" if sc < 0 else "— neutral")
        r = {"Direction": direction,
             "Driver": _clean_attr(d["attribute"]),
             "Theme": theme,
             "Importance %": d.get("importance", 0),
             "Std β": d.get("std_coef"),
             "Coef (b)": d.get("coef")}
        if _is_logit:
            r["Odds ratio"] = d.get("odds_ratio")
            r["OR 95% CI"] = (f"{d.get('or_ci_low')}–{d.get('or_ci_high')}"
                              if d.get("or_ci_low") is not None else None)
        else:
            r["95% CI"] = (f"{d.get('ci_low')}–{d.get('ci_high')}"
                           if d.get("ci_low") is not None else None)
            r["t"] = d.get("t_stat")
        r["p"] = d.get("p_value")
        _pv = d.get("p_value")
        _pv = 1.0 if _pv is None else _pv
        r["Sig."] = ("***" if _pv < 0.001 else
                     "**"  if _pv < 0.01  else
                     "*"   if _pv < 0.05  else "ns")
        return r

    _TOP_TABLE = 20
    _pfmt = {"Importance %": "{:.1f}", "Std β": "{:+.3f}", "Coef (b)": "{:+.4f}", "p": "{:.4f}"}
    if _is_logit: _pfmt["Odds ratio"] = "{:.3f}"
    else:         _pfmt["t"] = "{:.2f}"

    def _style_ptbl(df):
        # pandas Styler defers background_gradient's matplotlib import until the
        # Styler is actually rendered (st.dataframe -> styler._compute()), well
        # outside a try/except wrapped around the .background_gradient() call —
        # so check availability upfront instead of relying on try/except here.
        if importlib.util.find_spec("matplotlib") is not None:
            _s = df.style.background_gradient(subset=["Importance %"], cmap="Greens")
        else:
            _s = df.style  # matplotlib not installed — skip gradient
        _cell = getattr(_s, "map", getattr(_s, "applymap", None))
        _s = _cell(lambda v: "color:#16a34a;font-weight:700" if v and "▲" in str(v) else
                             ("color:#dc2626;font-weight:700" if v and "▼" in str(v) else ""),
                   subset=["Direction"])
        _cell = getattr(_s, "map", getattr(_s, "applymap", None))
        _s = _cell(lambda v: "font-weight:700;color:#111827" if v in ("*","**","***") else
                             "color:#9ca3af", subset=["Sig."])
        return _s.format(_pfmt, na_rep="—")

    _ROW_H = 36   # px per row
    _HDR_H = 40   # px header

    if _sig:
        _ptbl_sig = [_prow(d) for d in _sig]
        _ptbl_show = _ptbl_sig[:_TOP_TABLE]
        _tbl_h = min(_HDR_H + len(_ptbl_show) * _ROW_H, 600)
        st.dataframe(_style_ptbl(pd.DataFrame(_ptbl_show)),
                     hide_index=True, use_container_width=True, height=_tbl_h)
        if len(_ptbl_sig) > _TOP_TABLE:
            with st.expander(f"Show all {len(_ptbl_sig)} matching drivers"):
                _exp_h = min(_HDR_H + len(_ptbl_sig) * _ROW_H, 600)
                st.dataframe(_style_ptbl(pd.DataFrame(_ptbl_sig)),
                             hide_index=True, use_container_width=True, height=_exp_h)
    else:
        st.info("No drivers match the current filter settings. Adjust the filters above.")

    with st.expander(f"All {len(drivers)} predictors (unfiltered, includes non-significant)"):
        _all_h = min(_HDR_H + min(len(drivers), 30) * _ROW_H, 600)
        st.dataframe(_style_ptbl(pd.DataFrame([_prow(d) for d in drivers])),
                     hide_index=True, use_container_width=True, height=_all_h)

    st.caption(
        "**Direction**: ▲ = raises outcome (positive β)  ▼ = lowers outcome (negative β)  |  "
        "**Importance %** = share of total |std β| — how much this driver explains  |  "
        "**Std β** = standardised effect size (comparable across drivers)  |  "
        + ("**Odds ratio** > 1 = increases top-box probability  |  " if _is_logit else "")
        + "**Sig.**: * p<0.05  ** p<0.01  *** p<0.001  ns = not significant")

    # ── Headline: the single biggest lever + biggest drag ────────────────────
    _h1, _h2 = st.columns(2)
    def _headline(col, title, d, color, arrow):
        with col:
            if not d:
                st.markdown(f"<div style='border:1px solid #e5e7eb;border-radius:10px;padding:14px 16px;'>"
                            f"<div style='font-size:0.62rem;font-weight:800;text-transform:uppercase;"
                            f"letter-spacing:0.08em;color:#9ca3af;'>{title}</div>"
                            f"<div style='color:#9ca3af;font-size:0.85rem;margin-top:6px;'>none significant</div></div>",
                            unsafe_allow_html=True)
                return
            st.markdown(
                f"<div style='border:1px solid {color}33;border-top:3px solid {color};border-radius:10px;"
                f"padding:14px 16px;background:{color}08;'>"
                f"<div style='font-size:0.62rem;font-weight:800;text-transform:uppercase;letter-spacing:0.08em;"
                f"color:{color};'>{arrow} {title}</div>"
                f"<div style='font-size:0.98rem;font-weight:700;color:#111827;margin:6px 0 2px;line-height:1.3;'>"
                f"{_clean_attr(d['attribute'])}</div>"
                f"<div style='font-size:0.74rem;color:#6b7280;'>importance {_rf(d,'importance'):.1f}% · "
                f"std β {_rf(d,'std_coef'):+.2f}" + (f" · OR {_rf(d,'odds_ratio'):.2f}" if _is_logit else "") +
                f" · p={_rf(d,'p_value'):.3f}</div></div>", unsafe_allow_html=True)
    _headline(_h1, f"Biggest lever (raises {_out_lbl})", _pos[0] if _pos else None, "#16a34a", "▲")
    _headline(_h2, f"Biggest drag (lowers {_out_lbl})", _neg[0] if _neg else None, "#dc2626", "▼")

    # ── Importance by theme (broad_feature) — what TYPE of driver matters ────
    # Match via normalized keys — R mangles attribute names (spaces/() → dots).
    _fmap = {_norm_key(k): v for k, v in _attr_feature_map().items()}
    _theme_imp = {}
    for d in _sig:
        th = _fmap.get(_norm_key(d["attribute"]), "Other")
        _theme_imp[th] = _theme_imp.get(th, 0) + d.get("importance", 0)
    if _theme_imp:
        _tt = sorted(_theme_imp.items(), key=lambda x: x[1])
        _fig_t = go.Figure(go.Bar(
            x=[v for _, v in _tt], y=[k for k, _ in _tt], orientation="h",
            marker_color="#1a5d4d", text=[f"{v:.2g}%" for _, v in _tt], textposition="outside"))
        _bt = {k: v for k, v in _chart_layout_base(max(240, len(_tt) * 42)).items() if k not in ("xaxis", "yaxis")}
        _fig_t.update_layout(**_bt, xaxis=dict(title="Share of total driver importance (%)"),
                             yaxis=dict(automargin=True),
                             title=dict(text=f"Which TYPE of driver moves {_out_lbl} most — {sel_brand}",
                                        font=dict(size=12)))
        st.plotly_chart(_theme_fig(_fig_t), use_container_width=True)
        st.caption(
            f"These bars sum to 100% across all {len(_sig)} significant drivers combined (grouped into "
            f"themes) — they are **not** a % of respondents and don't need to look like a big number to "
            f"matter. E.g. a theme at 5% still means it's the single largest theme if everything else "
            f"splits the remaining 95% across many attributes. Read this as *relative* weight — which "
            f"theme to prioritize first — not an absolute score.")

    # ── Standardized coefficients with 95% CI (XLSTAT "Standardized coefficients
    # (95% conf. interval)" chart) — per-driver, not aggregated by theme.
    if _is_logit and _sig:
        _sc_rows = [d for d in _sig[:_top_n_chart] if d.get("std_coef_ci_low") is not None]
        if _sc_rows:
            _sc_sorted = sorted(_sc_rows, key=lambda d: d.get("std_coef", 0))
            _sc_names = [_clean_attr(d["attribute"]) for d in _sc_sorted]
            _sc_vals  = [d.get("std_coef", 0) for d in _sc_sorted]
            _sc_lo    = [d.get("std_coef", 0) - d.get("std_coef_ci_low", 0) for d in _sc_sorted]
            _sc_hi    = [d.get("std_coef_ci_high", 0) - d.get("std_coef", 0) for d in _sc_sorted]
            _fig_sc = go.Figure(go.Bar(
                x=_sc_vals, y=_sc_names, orientation="h",
                marker_color=["#16a34a" if v >= 0 else "#dc2626" for v in _sc_vals],
                error_x=dict(type="data", symmetric=False, array=_sc_hi, arrayminus=_sc_lo, color="#374151"),
            ))
            _bsc = {k: v for k, v in _chart_layout_base(max(260, len(_sc_names) * 34)).items()
                    if k not in ("xaxis", "yaxis")}
            _fig_sc.update_layout(**_bsc, xaxis=dict(title="Standardized coefficient (β)"),
                                  yaxis=dict(automargin=True),
                                  title=dict(text=f"Standardized coefficients — 95% CI — {_out_lbl}",
                                             font=dict(size=12)))
            # Explicit zero-reference line — the caption used to just tell the reader to "check
            # whether the whisker spans zero" with no visual marker to check it against. A bar
            # whose error whisker crosses this line is directionally uncertain even if it cleared
            # the significance bar shown elsewhere.
            _fig_sc.add_vline(x=0, line=dict(color="#374151", width=1.5, dash="dash"))
            st.plotly_chart(_theme_fig(_fig_sc), use_container_width=True)
            _n_crosses_zero = sum(
                1 for d in _sc_sorted
                if (d.get("std_coef_ci_low", 0) or 0) < 0 < (d.get("std_coef_ci_high", 0) or 0)
            )
            st.caption(
                f"The dashed vertical line = zero (no effect). Whiskers = 95% confidence interval on "
                f"the coefficient (β). A bar whose whisker crosses the dashed line is directionally "
                f"uncertain — we can't rule out zero effect — even if it's flagged significant "
                f"elsewhere on this page (a different test, e.g. Wald p-value, can disagree with the "
                f"CI at the margin)."
                + (f" **{_n_crosses_zero} of {len(_sc_sorted)}** shown here cross zero."
                   if _n_crosses_zero else " None of the drivers shown here cross zero.")
            )

    # ── Importance by segment — same model re-fit per segment value, with a real
    # significance test flagging when a segment's coefficient differs from the "All"
    # baseline, plus per-segment AUC (XLSTAT-style "Brand Drivers: All / Metro / Non-Metro"
    # comparison — this dataset has no metro/non-metro flag, Zone/Gender are the closest
    # real segment cuts available). ──────────────────────────────────────────
    _seg_rows, _seg_dim = None, None  # populated below when the segment fit succeeds; stay None
                                       # otherwise so the Excel export can safely skip the sheet.
    with st.expander("📐 Importance by segment — compare across Zone or Gender", expanded=False):
        _seg_c1, _seg_c2 = st.columns([1, 1])
        with _seg_c1:
            _seg_dim = st.selectbox("Segment dimension", ["Zone", "Gender"], key="dr_seg_dim")
        with _seg_c2:
            _seg_conf = st.selectbox("Flag differences at", [0.95, 0.90, 0.80], index=0,
                                     format_func=lambda a: f"{a:.0%}", key="dr_seg_conf")
        _Z_CRIT = {0.95: 1.959964, 0.90: 1.644854, 0.80: 1.281552}[_seg_conf]

        _seg_values = (["North", "South", "East", "West"] if _seg_dim == "Zone"
                       else ["Male", "Female"])
        _seg_kwarg = "zone" if _seg_dim == "Zone" else "gender"

        with st.spinner(f"Fitting the model separately for each {_seg_dim.lower()}…"):
            _seg_results = {}
            for _sv in _seg_values:
                _kwargs = dict(zone=zone, gender=gender, age_band=age_band, city=city)
                _kwargs[_seg_kwarg] = _sv
                _r = _driver_regression_for_brand(sel_brand, _dv_kind, _topbox, model=_model,
                                                   dv_attr=_dv_attr, project_id=project_id, **_kwargs)
                if _r and "error" not in _r and _r.get("significant_drivers"):
                    _seg_results[_sv] = _r

        if not _seg_results:
            st.info(f"Not enough respondents per {_seg_dim.lower()} to fit a separate model "
                    f"(each cut needs ≥30 respondents with {_out_lbl} + imagery data).")
        else:
            def _se_from_ci(d):
                # std_coef_ci_low/high only exist for logistic models (that's what the "95% CI"
                # chart above reads). For Linear (the default model choice), the R output only
                # returns a CI on the RAW coefficient (ci_low/ci_high, used in the driver table's
                # "95% CI" column) — without this fallback, every SE lookup silently returns None
                # for Linear and the significance flag never fires, with no visible error.
                lo, hi = d.get("std_coef_ci_low"), d.get("std_coef_ci_high")
                if lo is not None and hi is not None:
                    return (hi - lo) / (2 * 1.959964)
                lo2, hi2 = d.get("ci_low"), d.get("ci_high")
                coef, std_coef = d.get("coef"), d.get("std_coef")
                if lo2 is not None and hi2 is not None and coef not in (None, 0) and std_coef is not None:
                    se_coef = (hi2 - lo2) / (2 * 1.959964)
                    return se_coef * abs(std_coef / coef)
                return None

            _all_by_attr = {d["attribute"]: d for d in _sig}
            # Sort order matches the rest of this page — highest-importance driver first
            # (same "order of importance" convention used throughout this section).
            _seg_drivers_ordered = sorted(_sig, key=lambda d: -d.get("importance", 0))[:_top_n_chart]

            _seg_rows = []
            for d in _seg_drivers_ordered:
                _attr = d["attribute"]
                _row = {"Driver": _clean_attr(_attr), "All": d.get("importance", 0)}
                _se_all = _se_from_ci(d)
                for _sv, _r in _seg_results.items():
                    _sd = next((x for x in _r["significant_drivers"] if x["attribute"] == _attr), None)
                    if _sd is None:
                        _row[_sv] = 0.0
                        _row[f"{_sv}_flag"] = ""
                        continue
                    _row[_sv] = _sd.get("importance", 0)
                    _se_sv = _se_from_ci(_sd)
                    if _se_all and _se_sv:
                        _z = (_sd.get("std_coef", 0) - d.get("std_coef", 0)) / ((_se_all**2 + _se_sv**2) ** 0.5)
                        _row[f"{_sv}_flag"] = ("↑" if _z > _Z_CRIT else ("↓" if _z < -_Z_CRIT else ""))
                    else:
                        _row[f"{_sv}_flag"] = ""
                _seg_rows.append(_row)

            _fig_seg = go.Figure()
            _seg_colors = ["#1a5d4d", "#30a76a", "#7c9885", "#a3c9a8"]
            _fig_seg.add_trace(go.Bar(
                y=[r["Driver"] for r in _seg_rows], x=[r["All"] for r in _seg_rows],
                orientation="h", name="All", marker_color="#111827",
                text=[f"{r['All']:.1f}%" for r in _seg_rows], textposition="outside",
            ))
            for _i, _sv in enumerate(_seg_results):
                _fig_seg.add_trace(go.Bar(
                    y=[r["Driver"] for r in _seg_rows], x=[r.get(_sv, 0) for r in _seg_rows],
                    orientation="h", name=_sv, marker_color=_seg_colors[_i % len(_seg_colors)],
                    text=[f"{r.get(_sv,0):.1f}%{r.get(f'{_sv}_flag','')}" for r in _seg_rows],
                    textposition="outside",
                ))
            _bseg = {k: v for k, v in _chart_layout_base(max(320, len(_seg_rows) * 46)).items()
                    if k not in ("xaxis", "yaxis", "legend")}
            _fig_seg.update_layout(**_bseg, barmode="group",
                                   xaxis=dict(title="Importance % (within that segment's own model)"),
                                   yaxis=dict(automargin=True, categoryorder="array",
                                              categoryarray=[r["Driver"] for r in _seg_rows][::-1]),
                                   title=dict(text=f"Importance by {_seg_dim} — {_out_lbl}", font=dict(size=12)),
                                   legend=dict(orientation="h", y=-0.12))
            st.plotly_chart(_theme_fig(_fig_seg), use_container_width=True)

            if _is_logit:
                _auc_bits = ["All AUC = " + (f"{res.get('auc',0):.2f}" if res.get('auc') is not None else "—")]
                for _sv, _r in _seg_results.items():
                    _auc_bits.append(f"{_sv} AUC = " + (f"{_r.get('auc',0):.2f}" if _r.get('auc') is not None else "—"))
                st.caption(" · ".join(_auc_bits))

            st.caption(
                f"Each driver's importance %, refit separately within each {_seg_dim.lower()} "
                f"(not the same coefficients re-sliced — a genuinely separate model per segment, "
                f"same formula as the main table above). ↑/↓ = that segment's coefficient is "
                f"significantly higher/lower than the **All** model's coefficient at {_seg_conf:.0%} "
                f"confidence (two-sample test on the coefficients' own standard errors) — not just "
                f"a bigger bar, a real statistical difference. Bars ordered by the All model's "
                f"importance, same as the driver table above."
            )

    # ── Ownership matrix — which brand uniquely owns each attribute, ranked by THIS
    # regression's own importance %, not a fixed/generic ranking. Lives here (not under
    # Attribute Ownership) because "most important" genuinely depends on which outcome (Y)
    # this regression is explaining — switching the DV/topbox/model above changes this order
    # too, using the exact same `_sig` this tab already computed, no separate DV picker needed.
    with st.expander("🧩 Ownership matrix — who uniquely owns each attribute (ranked by this regression)",
                     expanded=False):
        _mx_top_n = st.slider("Attributes shown", 8, 30, 20, 1, key="dr_matrix_n")
        # R mangles attribute names (spaces/() -> dots), so d["attribute"] doesn't match
        # dim_bq3_attribute.attr_label directly — map back via the same _norm_key() convention
        # used everywhere else on this page (_attr_feature_map, _brand_assoc_pct) before handing
        # names to _get_ownership_matrix, which queries the DB by real attr_label.
        _real_label_by_norm = {_norm_key(lbl): lbl for lbl in _attr_feature_map().keys()}
        _attr_order = tuple(
            _real_label_by_norm[_norm_key(d["attribute"])]
            for d in sorted(_sig, key=lambda d: -d.get("importance", 0))
            if _norm_key(d["attribute"]) in _real_label_by_norm
        ) if _sig else None
        _matrix, _mx_attrs, _mx_brands = _get_ownership_matrix(
            top_attrs=_mx_top_n, top_brands=8, attr_order=_attr_order)

        if _matrix.empty:
            st.info("Not enough imagery data to build the ownership matrix.")
        else:
            _disp = _matrix.set_index("Attribute")[_mx_brands].copy()
            _owner_map = _matrix.set_index("Attribute")["_owner"].to_dict()
            _mark_map = _matrix.set_index("Attribute")["_mark"].to_dict()

            def _fmt_cell(val, attr, brand):
                if pd.isna(val):
                    return ""
                if _owner_map.get(attr) == brand and _mark_map.get(attr):
                    return f"{val:.0f}% {_mark_map[attr]}"
                return ""  # non-owning cells shown blank, matching the reference layout

            _cell_df = pd.DataFrame(
                {b: [_fmt_cell(_disp.loc[a, b], a, b) for a in _disp.index] for b in _mx_brands},
                index=_disp.index,
            )

            def _style_owned(_):
                styles = pd.DataFrame("", index=_disp.index, columns=_mx_brands)
                for a in _disp.index:
                    owner = _owner_map.get(a)
                    if owner and owner in _mx_brands:
                        mark = _mark_map.get(a)
                        bg = "#1a5d4d" if mark == "++" else "#30a76a"
                        styles.loc[a, owner] = f"background-color:{bg};color:white;font-weight:800;text-align:center"
                return styles

            st.dataframe(
                _cell_df.style.apply(_style_owned, axis=None),
                use_container_width=True,
                height=min(48 * (len(_disp) + 1), 700),
            )
            st.caption(
                f"Rows ranked by {sel_brand}'s own {_out_lbl} driver-regression importance — "
                f"same numbers as the driver table above, not a fixed/generic ranking. ++ = "
                f"significantly ahead of every other brand shown on this statement (pooled "
                f"two-proportion z-test, 95%). + = leads but doesn't clear every rival. "
                f"Base: aided-aware respondents per brand, top {len(_mx_brands)} brands shown."
            )

    # ── Key Driver Quadrant: importance × brand performance ──────────────────
    _assoc = {_norm_key(k): v for k, v in _brand_assoc_pct(sel_brand).items()}
    _qd_all = _sig[:_top_n_chart]   # respect top-N filter
    if _qd_all and _assoc:
        import numpy as _np2
        _pos_qd = [d for d in _qd_all if (d.get("std_coef") or 0) >= 0]
        _neg_qd = [d for d in _qd_all if (d.get("std_coef") or 0) < 0]

        def _qd_xy(lst):
            xi = [d.get("importance", 0) for d in lst]
            yi = [_assoc.get(_norm_key(d["attribute"]), 0) for d in lst]
            nm = [_clean_attr(d["attribute"]) for d in lst]
            imp_str = [f"Importance: {_rf(d,'importance'):.1f}%<br>"
                       f"Performance: {_assoc.get(_norm_key(d['attribute']),0):.0f}%<br>"
                       f"Std β: {_rf(d,'std_coef'):+.3f}<br>"
                       f"p = {_rf(d,'p_value',1):.4f}" for d in lst]
            return xi, yi, nm, imp_str

        # label only top-5 by importance to avoid overlap; alternate label position
        # around each dot (top/bottom/left/right) so nearby points don't collide.
        _label_top = 5
        _label_positions = ["top center", "bottom center", "middle right", "middle left", "bottom center"]
        def _dot_labels_pos(lst, nm):
            top_idx = sorted(range(len(lst)), key=lambda i: -lst[i].get("importance", 0))[:_label_top]
            rank_of = {i: r for r, i in enumerate(top_idx)}
            labels = [nm[i] if i in top_idx else "" for i in range(len(nm))]
            positions = [_label_positions[rank_of[i] % len(_label_positions)] if i in top_idx else "top center"
                         for i in range(len(nm))]
            return labels, positions

        _xi_p, _yi_p, _nm_p, _ht_p = _qd_xy(_pos_qd)
        _xi_n, _yi_n, _nm_n, _ht_n = _qd_xy(_neg_qd)

        all_xi = _xi_p + _xi_n
        all_yi = _yi_p + _yi_n
        _xmed  = float(_np2.median(all_xi)) if all_xi else 0
        _ymed  = float(_np2.median(all_yi)) if all_yi else 0
        _xmax  = max(all_xi) * 1.15 if all_xi else 1
        # y-axis: zoom to the actual data band instead of always anchoring at 0 —
        # driver association % often clusters tightly (e.g. 45-52%), and forcing
        # the axis to start at 0 compresses those real differences into noise.
        _y_data_min = min(all_yi) if all_yi else 0
        _y_data_max = max(all_yi) if all_yi else 1
        _y_pad = max((_y_data_max - _y_data_min) * 0.25, 1.0)
        _ymin  = max(0.0, _y_data_min - _y_pad)
        _ymax  = _y_data_max + _y_pad

        _pos_labels, _pos_pos = _dot_labels_pos(_pos_qd, _nm_p)
        _neg_labels, _neg_pos = _dot_labels_pos(_neg_qd, _nm_n)

        _qfig = go.Figure()
        # quadrant backgrounds
        _qfig.add_shape(type="rect", x0=_xmed, x1=_xmax, y0=_ymed, y1=_ymax,
                        fillcolor="rgba(16,185,129,0.07)", line_width=0, layer="below")
        _qfig.add_shape(type="rect", x0=_xmed, x1=_xmax, y0=_ymin, y1=_ymed,
                        fillcolor="rgba(245,158,11,0.08)", line_width=0, layer="below")
        _qfig.add_shape(type="rect", x0=0, x1=_xmed, y0=_ymed, y1=_ymax,
                        fillcolor="rgba(148,163,184,0.04)", line_width=0, layer="below")
        _qfig.add_shape(type="rect", x0=0, x1=_xmed, y0=_ymin, y1=_ymed,
                        fillcolor="rgba(148,163,184,0.04)", line_width=0, layer="below")
        # positive drivers (green)
        if _pos_qd:
            _qfig.add_trace(go.Scatter(
                x=_xi_p, y=_yi_p, mode="markers+text", name="▲ Raises outcome",
                text=_pos_labels,
                textposition=_pos_pos, textfont=dict(size=9, color="#15803d"),
                marker=dict(size=14, color="#16a34a", opacity=0.85,
                            line=dict(width=2, color="white")),
                customdata=_ht_p,
                hovertemplate="<b>%{text}</b><br>%{customdata}<extra>▲ raises outcome</extra>"))
        # negative drivers (red)
        if _neg_qd:
            _qfig.add_trace(go.Scatter(
                x=_xi_n, y=_yi_n, mode="markers+text", name="▼ Lowers outcome",
                text=_neg_labels,
                textposition=_neg_pos, textfont=dict(size=9, color="#b91c1c"),
                marker=dict(size=14, color="#dc2626", opacity=0.85,
                            line=dict(width=2, color="white")),
                customdata=_ht_n,
                hovertemplate="<b>%{text}</b><br>%{customdata}<extra>▼ lowers outcome</extra>"))
        # median lines
        _qfig.add_vline(x=_xmed, line_dash="dot", line_color="#94a3b8", line_width=1)
        _qfig.add_hline(y=_ymed, line_dash="dot", line_color="#94a3b8", line_width=1)
        # quadrant labels
        for ann_x, ann_y, ann_txt, ann_col, ann_xa, ann_ya in [
            (_xmax, _ymax, "● STRENGTH (green) / ⚠ Risky association (red)",  "#16a34a", "right", "top"),
            (_xmax, _ymin, "PRIORITY GAP — important & under-owned",           "#d97706", "right", "bottom"),
            (0,     _ymax, "Low priority (monitor)",                            "#94a3b8", "left",  "top"),
            (0,     _ymin, "Low priority (low relevance)",                      "#94a3b8", "left",  "bottom"),
        ]:
            _qfig.add_annotation(x=ann_x, y=ann_y, text=ann_txt, showarrow=False,
                                 font=dict(size=9, color=ann_col),
                                 xanchor=ann_xa, yanchor=ann_ya,
                                 bgcolor="white", borderpad=3, opacity=0.85)

        _bq = {k: v for k, v in _chart_layout_base(520).items() if k not in ("xaxis", "yaxis", "legend")}
        _qfig.update_layout(
            **_bq,
            xaxis=dict(title="← Less important    Driver importance %    More important →",
                       range=[0, _xmax]),
            yaxis=dict(title=f"← Low    {sel_brand} association %    High →",
                       range=[_ymin, _ymax]),
            title=dict(text=(f"Key Driver Quadrant — {sel_brand}  ·  outcome: {_out_lbl}  ·  "
                             f"top {len(_qd_all)} significant drivers  ·  only top 5 labelled"),
                       font=dict(size=12)),
            legend=dict(orientation="h", y=-0.12, x=0.5, xanchor="center",
                        font=dict(size=11), bgcolor="rgba(0,0,0,0)"),
            showlegend=True)
        st.plotly_chart(_theme_fig(_qfig), use_container_width=True)
        st.caption(
            "**How to read this chart** — "
            "X = driver importance (how much it shifts the outcome). "
            "Y = brand's association % (how often customers link this to the brand).  "
            "**Green dot, top-right** = high-importance attribute that *raises* outcome + brand already owns it → **defend**.  "
            "**Red dot, top-right** = high-importance attribute that *lowers* outcome + brand is strongly associated with it → "
            "**investigate** (customers link the brand to this but it hurts NPS — a reputational risk).  "
            "**Green dot, bottom-right** = important lever + low ownership → **priority gap to close**.  "
            "**Hover** any dot for full stats. Only top 5 by importance are labelled.")

    # ── AI read-out ──────────────────────────────────────────────────────────
    _fit = res.get("mcfadden_r2", res.get("r_squared", 0)) if _is_logit else res.get("r_squared", 0)
    _bits = [f"For **{sel_brand}**, imagery attributes explain {_fit:.0%} of {_out_lbl} "
             f"({'pseudo-R²' if _is_logit else 'R²'}, n={res.get('n',0)})."]
    if _pos:
        _bits.append(f"The strongest positive lever is **{_clean_attr(_pos[0]['attribute'])}** "
                     f"({_pos[0].get('importance',0):.1f}% importance)" +
                     (f", and the top theme is **{_tt[-1][0]}** ({_tt[-1][1]:.2g}%)." if _theme_imp else "."))
    if _neg:
        _bits.append(f"The biggest drag is **{_clean_attr(_neg[0]['attribute'])}** — associations here "
                     f"coincide with lower {_out_lbl}.")
    # priority gap = high importance, low performance
    if _pos and _assoc:
        _prio = sorted(_pos, key=lambda d: (-d.get("importance", 0), _assoc.get(_norm_key(d["attribute"]), 0)))
        _gap = next((d for d in _prio if _assoc.get(_norm_key(d["attribute"]), 100) < 40), None)
        if _gap:
            _bits.append(f"Clearest opportunity: **{_clean_attr(_gap['attribute'])}** matters "
                         f"({_gap.get('importance',0):.1f}%) but {sel_brand} only owns it at "
                         f"{_assoc.get(_norm_key(_gap['attribute']),0):.0f}% — invest to close the gap.")
    _ai_card(" ".join(_bits), "AI Driver Read-out", "#1a5d4d")

    # ── Logistic diagnostic charts (XLSTAT: ROC Curve / Confusion plot / Probabilities) ──
    if _is_logit:
        _roc = res.get("roc_points") or {}
        _pc  = res.get("prob_chart") or {}
        if _roc.get("fpr") and _roc.get("tpr"):
            _rc1, _rc2 = st.columns(2)
            with _rc1:
                _fig_roc = go.Figure()
                _fig_roc.add_trace(go.Scatter(
                    x=_roc["fpr"], y=_roc["tpr"], mode="lines", name="ROC",
                    line=dict(color="#1a5d4d", width=2)))
                _fig_roc.add_trace(go.Scatter(
                    x=[0, 1], y=[0, 1], mode="lines", name="Chance",
                    line=dict(color="#9ca3af", width=1, dash="dot")))
                _broc = {k: v for k, v in _chart_layout_base(360).items() if k not in ("xaxis", "yaxis")}
                _fig_roc.update_layout(**_broc, xaxis=dict(title="1 - Specificity", range=[0, 1]),
                                       yaxis=dict(title="Sensitivity", range=[0, 1]),
                                       title=dict(text=f"ROC Curve (AUC={res.get('auc', 0):.3f})",
                                                  font=dict(size=12)), showlegend=False)
                st.plotly_chart(_theme_fig(_fig_roc), use_container_width=True)
                st.caption(f"Curve above the diagonal = model separates the two outcomes better than chance. "
                          f"AUC {res.get('auc', 0):.3f} = probability a random respondent who IS in the top-box "
                          f"is scored higher by the model than a random respondent who ISN'T. 0.5 = no better "
                          f"than a coin flip, 1.0 = perfect separation.")
            with _rc2:
                # Replaced the jittered "calibration scatter" (predicted probability vs. actual
                # class squeezed onto a 0/1 axis with artificial jitter to separate overlapping
                # dots) with a plain sortable table — same underlying per-respondent data
                # (predicted probability + actual outcome + correct/incorrect), but readable
                # without first understanding what jitter means. Click any column header to
                # sort; high and low predicted-probability respondents are both visible and
                # distinctly colored, not lost under an overlap of dots.
                if _pc.get("pred") and _pc.get("actual"):
                    _correct = [(1 if p >= 0.5 else 0) == a for p, a in zip(_pc["pred"], _pc["actual"])]
                    _prob_df = pd.DataFrame({
                        "Respondent": [f"#{i+1}" for i in range(len(_pc["pred"]))],
                        "Predicted probability": _pc["pred"],
                        "Actual outcome": ["Top-box" if a == 1 else "Not top-box" for a in _pc["actual"]],
                        "Model called it": ["✓ correct" if c else "✗ wrong" for c in _correct],
                    }).sort_values("Predicted probability", ascending=False)
                    st.markdown("**Predicted probability per respondent** *(sortable — click a column header)*")

                    def _style_prob_row(row):
                        color = "#16a34a" if row["Model called it"].startswith("✓") else "#dc2626"
                        return [f"color:{color};font-weight:600" if col == "Model called it" else ""
                                for col in row.index]

                    st.dataframe(
                        _prob_df.style.apply(_style_prob_row, axis=1)
                                       .format({"Predicted probability": "{:.1%}"})
                                       .background_gradient(subset=["Predicted probability"], cmap="Greens")
                        if importlib.util.find_spec("matplotlib") else _prob_df,
                        hide_index=True, use_container_width=True, height=340,
                    )
                    _n_correct = sum(_correct)
                    st.caption(
                        f"Each row is one respondent — predicted probability of being top-box, what they "
                        f"actually were, and whether the model's 0.5-cutoff call matched reality "
                        f"({_n_correct} of {len(_correct)} correct, matches the classification table above). "
                        f"Darker green = higher predicted probability, at either end of the sort — nothing "
                        f"is hidden or averaged away by re-sorting."
                    )
            st.caption("ROC Curve / predicted-probability table reproduce XLSTAT's standard logistic-regression "
                      "diagnostic set — AUC and classification counts here match the goodness-of-fit table above.")

    # Download a full multi-sheet Excel workbook (XLSTAT-style export, formatted)
    try:
        _wb_bytes = _build_driver_regression_workbook(
            res=res, drivers=drivers, prow_fn=_prow, sel_brand=sel_brand,
            dv_kind=_dv_kind, is_logit=_is_logit,
            segment_rows=_seg_rows, segment_dim=_seg_dim)
        st.download_button("⬇ Download full report (Excel, all tables)",
                           _wb_bytes,
                           file_name=f"driver_regression_{sel_brand}_{_dv_kind}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dr_dl_xlsx")
    except Exception as _wb_err:
        st.caption(f"⚠️ Excel export failed ({_wb_err}) — falling back to CSV.")
        _dl_df = pd.DataFrame([_prow(d) for d in drivers])
        st.download_button("⬇ Download regression table (CSV)",
                           _dl_df.to_csv(index=False).encode("utf-8"),
                           file_name=f"driver_regression_{sel_brand}_{_dv_kind}.csv",
                           mime="text/csv", key="dr_dl")

    # ── 2. Cross-brand importance heatmap ────────────────────────────────────
    st.markdown(f"**Cross-brand driver importance** — how strongly each top driver affects each brand")
    _top_brands = [b["brand_name"] for b in sorted(
        brands_list, key=lambda x: x.get("nps_base", 0) or 0, reverse=True)][:_xbrands]
    if any(b["brand_name"] == sel_brand for b in brands_list) and sel_brand not in _top_brands:
        _top_brands = ([sel_brand] + _top_brands)[:_xbrands]
    _xb = [b["brand_name"] for b in brands_list if b["brand_name"] in _top_brands]
    if not _xb:
        _xb = _top_brands

    _imp_by_brand = {}
    _signed_by_brand = {}
    with st.spinner(f"Running regression across {len(_xb)} brands…"):
        for _bn in _xb:
            _r = _driver_regression_for_brand(_bn, _dv_kind, _topbox, zone, gender, age_band, city, _model, _dv_attr, project_id=project_id)
            if _r and "error" not in _r:
                _imp_by_brand[_bn] = {d["attribute"]: d.get("importance", 0)
                                      for d in _r.get("significant_drivers", [])}
                _signed_by_brand[_bn] = {d["attribute"]: (d.get("std_coef") or 0)
                                         for d in _r.get("significant_drivers", [])}

    if len(_imp_by_brand) < 2:
        st.info("Not enough brands have sufficient data for the cross-brand comparison under these filters.")
        return

    # rank drivers by mean importance across brands, take top 12
    # prioritise sig drivers of the selected brand so they always appear
    _all_attrs = {}
    for bn, dd in _imp_by_brand.items():
        for a, v in dd.items():
            _all_attrs[a] = _all_attrs.get(a, 0) + v
    _sel_sig = list(_imp_by_brand.get(sel_brand, {}).keys())  # sig attrs for focus brand
    _rest = [a for a, _ in sorted(_all_attrs.items(), key=lambda x: -x[1]) if a not in _sel_sig]
    _order = (_sel_sig + _rest)[:12]
    _brand_cols = list(_imp_by_brand.keys())
    # z = signed importance (importance % carrying the sign of std β) for direction-aware reading
    _z = []
    for a in _order:
        row = []
        for bn in _brand_cols:
            imp = _imp_by_brand[bn].get(a, 0)
            sgn = 1 if _signed_by_brand[bn].get(a, 0) >= 0 else -1
            row.append(round(imp * sgn, 1))
        _z.append(row)
    _labels = [_clean_attr(a)[:42] for a in _order]
    _fig_hm = go.Figure(go.Heatmap(
        z=_z, x=_brand_cols, y=_labels, colorscale="RdYlGn", zmid=0,
        colorbar=dict(title="Signed<br>import. %", thickness=14, len=0.7),
        xgap=1, ygap=1,
        hovertemplate="<b>%{y}</b><br>%{x}<br>signed importance: %{z:.1f}%<extra></extra>",
    ))
    _fig_hm.update_layout(
        paper_bgcolor="white", plot_bgcolor="white",
        font=dict(family="Inter, sans-serif", size=11, color="#334155"),
        title=dict(text=f"Driver importance × brand — outcome: {_out_lbl}<br>"
                        f"<sup>Drivers ordered by overall importance · green raises the outcome, red lowers it</sup>",
                   font=dict(size=13), x=0.01, xanchor="left"),
        xaxis=dict(tickangle=-30, tickfont=dict(size=10)),
        yaxis=dict(autorange="reversed", tickfont=dict(size=10)),
        height=max(380, len(_order) * 34 + 150), margin=dict(l=200, t=70, b=90, r=20),
    )
    st.plotly_chart(_fig_hm, use_container_width=True)
    _hm_note = (" **Note:** When DV = imagery attribute, top rows are the selected brand's positive drivers. "
                "Red cells = competing attributes (associating a brand with those LOWERS association with this attribute — "
                "cross-category interference). Green = attributes that co-occur with this one."
                if _dv_kind == "ATTR" else "")
    st.caption("Each cell = that driver's signed relative importance for that brand. Compare columns to see "
               "how the SAME driver matters differently across brands; read down a column for a brand's own priorities."
               + _hm_note)



@st.cache_data(ttl=3600)
def _get_segment_filter_options():
    """Zone/Gender/Age-band/City options for the Segment Filters bar — read from THIS project's
    own active database, never hardcoded to any one project's real values. Different clients have
    different zones/cities/age-bands (a dairy client's cities are not project_1's electrical-
    appliance-survey cities) — hardcoding one project's list here silently filtered every other
    project's data to nothing whenever a respondent's real city/zone wasn't in the hardcoded list.
    Cached per active DB path so switching projects picks up fresh values.
    """
    import sqlite3
    from oxdata.db_loader import get_db_path
    opts = {"zones": [], "genders": [], "age_bands": [], "cities": []}
    try:
        conn = sqlite3.connect(str(get_db_path()))
        try:
            for key, col in (("zones", "zone_name"), ("genders", "gender"),
                              ("age_bands", "age_band"), ("cities", "city_name")):
                try:
                    rows = conn.execute(
                        f"SELECT DISTINCT {col} FROM v_respondents "
                        f"WHERE {col} IS NOT NULL AND TRIM({col}) != '' ORDER BY {col}"
                    ).fetchall()
                    opts[key] = [r[0] for r in rows]
                except Exception:
                    opts[key] = []
        finally:
            conn.close()
    except Exception:
        pass
    return opts


def render_brand_health_dashboard():
    inject_pulse_styles()
    # ── Brand Health page–specific visual overrides ───────────────────────────
    st.markdown("""
    <style>
    /* ── Tab bar: bold and spaced ─────────────────────────── */
    .stTabs [data-baseweb="tab-list"] {
        gap: 2px;
        background: #f8fafc;
        border-radius: 12px;
        padding: 4px;
        border: 1px solid #e2e8f0;
        margin-bottom: 18px;
    }
    .stTabs [data-baseweb="tab"] {
        border-radius: 8px;
        font-weight: 700;
        font-size: 0.82rem;
        letter-spacing: 0.02em;
        padding: 8px 18px;
        color: #6b7280;
        background: transparent;
        border: none;
        transition: all 0.15s ease;
    }
    .stTabs [aria-selected="true"] {
        background: #1a5d4d !important;
        color: #e5e7eb !important;
        box-shadow: 0 2px 8px rgba(26,93,77,0.25);
    }
    .stTabs [data-baseweb="tab"]:hover:not([aria-selected="true"]) {
        background: #f0fdf4;
        color: #1a5d4d;
    }
    /* ── Container borders: lighter, rounder ─────────────── */
    [data-testid="stVerticalBlockBorderWrapper"] > div {
        border-radius: 12px !important;
        border-color: #e2e8f0 !important;
    }
    /* ── Metric cards: consistent sizing ─────────────────── */
    [data-testid="stMetric"] {
        background: #f8fafc;
        border-radius: 10px;
        padding: 12px;
        border: 1px solid #e2e8f0;
    }
    [data-testid="stMetricValue"] { font-weight: 900; }
    [data-testid="stMetricDelta"] { font-size: 0.72rem; }
    /* ── Selectbox: tighter ───────────────────────────────── */
    .stSelectbox > div > div { border-radius: 8px; }
    /* selected option in dropdown list → white text on dark bg */
    [data-testid="stSelectbox"] [aria-selected="true"],
    .stSelectbox li[aria-selected="true"] { color: #ffffff !important; }
    /* selected value shown in the input field → keep readable */
    .stSelectbox [data-baseweb="select"] [data-testid="stMarkdownContainer"] p,
    .stSelectbox [data-baseweb="select"] > div > div { color: #111827; }
    /* ── Dataframe: clean header ──────────────────────────── */
    [data-testid="stDataFrame"] thead th {
        background: #f0fdf4 !important;
        color: #1a5d4d !important;
        font-weight: 700 !important;
    }
    /* ── Caption: softer ─────────────────────────────────── */
    .stCaption { color: #9ca3af; font-size: 0.68rem; }
    /* ── Expander: subtle ─────────────────────────────────── */
    [data-testid="stExpander"] summary {
        font-weight: 600;
        font-size: 0.8rem;
        color: #374151;
    }
    </style>
    """, unsafe_allow_html=True)

    active_project_id = st.session_state.get("active_project_id", "project_1")
    engine = BrandImageryEngine(project_id=active_project_id)
    if not os.path.exists(engine.db_path):
        show_error_card("Database Not Found",
                        f"Cannot reach survey database at <code>{engine.db_path}</code>.")
        return

    # Override NPS benchmark from project_meta so AK dairy gets correct colour-coding (not appliance avg)
    global NPS_INDUSTRY_AVG
    _CSAT_SCALE = _get_csat_scale()
    try:
        from oxdata.db_loader import get_project_meta as _bh_get_meta
        NPS_INDUSTRY_AVG = _bh_get_meta(active_project_id).get("nps_industry_avg", NPS_INDUSTRY_AVG)
    except Exception:
        pass

    _seg_opts = _get_segment_filter_options()

    _p_meta = {}
    try:
        from oxdata.db_loader import get_project_meta as _bh_get_meta_top
        _p_meta = _bh_get_meta_top(active_project_id)
    except Exception:
        pass
    _p_disp = _p_meta.get("display_name") or active_project_id.replace("_", " ").title()
    _p_ind = _p_meta.get("industry") or "Brand Intelligence"
    _p_has_excel = (Path(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))) / "data" / active_project_id / "master_mapping.xlsx").exists()
    _p_src = "📊 master_mapping.xlsx" if _p_has_excel else "🗄️ SQLite Engine"

    st.markdown(
        f'<div style="display:flex;justify-content:space-between;align-items:center;padding:8px 16px;'
        f'background:linear-gradient(135deg, rgba(26,93,77,0.08) 0%, rgba(14,165,233,0.05) 100%);'
        f'border:1px solid rgba(26,93,77,0.2);border-radius:10px;margin-bottom:12px;">'
        f'<div><span style="font-size:0.65rem;font-weight:800;letter-spacing:0.08em;text-transform:uppercase;color:#1a5d4d;">Active Project</span>'
        f'<div style="font-size:1.05rem;font-weight:900;color:#0f172a;letter-spacing:-0.01em;">{_p_disp} <span style="font-size:0.78rem;font-weight:500;color:#64748b;">({_p_ind})</span></div></div>'
        f'<div style="display:flex;gap:10px;align-items:center;">'
        f'<span style="background:rgba(26,93,77,0.1);border:1px solid rgba(26,93,77,0.25);border-radius:20px;padding:3px 12px;font-size:0.75rem;font-weight:700;color:#1a5d4d;">{_p_src}</span>'
        f'</div></div>',
        unsafe_allow_html=True,
    )

    # ── TOP-OF-PAGE FILTER BAR ────────────────────────────────────────────────
    # Zone/Gender/Age/City apply across funnel, imagery, portfolio, price, journey,
    # demographics & attitudes (via _resp_filter_cte). BEI/IPA/Attribute Ownership
    # remain All-India by design (composite stability) — labelled as such in-tab.
    with st.container(border=True):
        _fc0, _fc1, _fc2, _fc3, _fc4 = st.columns([2, 1, 1, 1, 2])
        with _fc0:
            st.markdown(
                '<div style="display:flex;align-items:center;gap:7px;padding-top:4px;">'
                '<span style="font-size:0.95rem;">🎛️</span>'
                '<div><div style="font-size:0.78rem;font-weight:800;letter-spacing:0.02em;'
                'color:#1a5d4d;">Segment Filters</div>'
                '<div style="font-size:0.62rem;color:#9ca3af;margin-top:1px;">'
                'Applies across the whole report</div></div></div>',
                unsafe_allow_html=True,
            )
        with _fc1:
            sel_zone = st.selectbox("Zone", ["All"] + _seg_opts["zones"],
                                    key="bh_zone", label_visibility="collapsed",
                                    placeholder="Zone")
        with _fc2:
            sel_gender = st.selectbox("Gender", ["All"] + _seg_opts["genders"], key="bh_gender",
                                      label_visibility="collapsed", placeholder="Gender")
        with _fc3:
            sel_age_band = st.selectbox("Age Band", ["All"] + _seg_opts["age_bands"], key="bh_age_band",
                                        label_visibility="collapsed", placeholder="Age Band")
        with _fc4:
            sel_city = st.selectbox(
                "City", ["All"] + _seg_opts["cities"],
                key="bh_city", label_visibility="collapsed", placeholder="City",
            )
    sel_months = None  # Single-wave study — wave window removed

    zone_arg     = "all" if sel_zone     == "All" else sel_zone
    gender_arg   = "all" if sel_gender   == "All" else sel_gender
    age_band_arg = "all" if sel_age_band == "All" else sel_age_band
    city_arg     = "all" if sel_city     == "All" else sel_city
    cat_arg      = "all"  # Category dimension not available in Wave 1

    # ── Active-filter chips + one-click reset ─────────────────────────────────
    _active_segs = [(lbl, val) for lbl, val in
                    [("Zone", sel_zone), ("Gender", sel_gender),
                     ("Age", sel_age_band), ("City", sel_city)] if val != "All"]
    if _active_segs:
        def _reset_bh_filters():
            for _k in ("bh_zone", "bh_gender", "bh_age_band", "bh_city"):
                st.session_state[_k] = "All"
        _chip_col, _reset_col = st.columns([6, 1])
        with _chip_col:
            _chips = "".join(
                f"<span style='display:inline-block;background:#ecfdf5;border:1px solid #86efac;"
                f"border-radius:20px;padding:2px 11px;margin:2px 6px 2px 0;color:#166534;"
                f"font-size:0.72rem;font-weight:600;'>{lbl}: {val}</span>"
                for lbl, val in _active_segs
            )
            st.markdown(
                f"<div style='padding-top:4px;'><span style='font-size:0.68rem;color:#9ca3af;"
                f"text-transform:uppercase;letter-spacing:0.06em;font-weight:700;margin-right:6px;'>"
                f"Active filters</span>{_chips}</div>",
                unsafe_allow_html=True,
            )
        with _reset_col:
            st.button("✕ Reset", on_click=_reset_bh_filters, use_container_width=True,
                      key="bh_reset_filters", help="Clear all segment filters")

    # ── Fetch core data ───────────────────────────────────────────────────────
    try:
        data = _get_cached_brand_health_data(
            cat_arg, zone_arg, city_arg, sel_months, gender_arg, age_band_arg, project_id=active_project_id
        )

    except Exception as e:
        show_error_card("Analytics Error", f"<code>{e}</code>")
        return

    base_n = data["base_n"]

    if data["status"] == "insufficient_data" or base_n < 1:
        st.warning(f"Insufficient data (Base N = {base_n:,}). Expand filters.")
        return

    brands_list = data.get("brands_list") or data.get("brands") or []

    # ── Brand selector + display controls ────────────────────────────────────
    brand_names = [b["brand_name"] for b in brands_list]
    _active_filters = [f for f in [
        sel_zone if sel_zone != "All" else None,
        sel_gender if sel_gender != "All" else None,
        sel_age_band if sel_age_band != "All" else None,
        sel_city if sel_city != "All" else None,
    ] if f]
    _filter_label = "  ·  ".join(_active_filters) if _active_filters else "All India"
    with st.container(border=True):
        fb1, fb2, fb_info = st.columns([3, 1, 3])
        _default_brand_idx = 0
        if "Akshayakalpa Organic" in brand_names:
            _default_brand_idx = brand_names.index("Akshayakalpa Organic")
        with fb1:
            sel_brand = st.selectbox(
                "Focus Brand",
                brand_names if brand_names else ["—"],
                index=_default_brand_idx,
                key=f"bh_brand_{active_project_id}",
                help="Select brand for deep-dive analysis across all tabs",
            )

        with fb2:
            min_base_n = st.select_slider(
                "Min Base N", options=[10, 20, 30, 50, 100], value=30,
                key="bh_min_base_n",
                help="Minimum respondents to show a cell in Competitive Benchmarking (suppresses low-N cells)",
            )
        with fb_info:
            st.markdown(
                f"<div style='padding-top:6px;'>"
                f"<span style='background:#f0fdf4;border:1px solid #86efac;border-radius:20px;"
                f"padding:3px 10px;color:#166534;font-weight:700;font-size:0.78rem;'>n = {base_n:,}</span>"
                f"&nbsp;<span style='font-size:0.75rem;color:#9ca3af;'>{_filter_label}</span>"
                f"</div>",
                unsafe_allow_html=True,
            )
    sel_theme = "All"
    focus_n   = 10  # reserved for future use — not yet wired to any chart

    # ── Sidebar: advanced chart settings only ────────────────────────────────
    # Apply persisted theme from session_state BEFORE widgets render
    # (CHART_THEME resets to defaults on every rerun as a module-level dict)
    _palette_keys = list(CHART_THEME["palettes"].keys())
    _saved_palette = st.session_state.get("bh_palette", "pulse_green")
    if _saved_palette in _palette_keys:
        CHART_THEME["palette"] = _saved_palette
    _font_opts = [
        "Inter, Arial, sans-serif",
        "Montserrat, sans-serif",
        "Playfair Display, Georgia, serif",
        "IBM Plex Sans, Arial, sans-serif",
        "Poppins, Arial, sans-serif",
    ]
    CHART_THEME["font_family"] = st.session_state.get("bh_font", _font_opts[0])
    CHART_THEME["font_size"]   = st.session_state.get("bh_font_size", 12)
    CHART_THEME["ppt_scale"]   = st.session_state.get("bh_ppt_scale", 1.0)

    with st.sidebar:
        with st.expander("Chart Appearance", expanded=False):
            palette_choice = st.selectbox(
                "Colour Palette", _palette_keys,
                index=_palette_keys.index(CHART_THEME["palette"]),
                key="bh_palette",
            )
            CHART_THEME["palette"] = palette_choice
            font_choice = st.selectbox("Chart Font", _font_opts, index=0, key="bh_font")
            CHART_THEME["font_family"] = font_choice
            font_size_choice = st.slider("Font Size", 9, 18, CHART_THEME["font_size"], 1,
                                         key="bh_font_size")
            CHART_THEME["font_size"] = font_size_choice
            ppt_scale = st.slider("Chart Scale (PPT)", 0.8, 1.8, 1.0, 0.1, key="bh_ppt_scale",
                                   help="Scale > 1.0 → larger charts for PPT export")
            CHART_THEME["ppt_scale"] = ppt_scale
        st.divider()
        if st.button("🔄 Clear Cache", use_container_width=True,
                     help="Refreshes AI narratives and all cached data."):
            st.cache_data.clear()
            st.rerun()

    brand_data = next((b for b in brands_list if b["brand_name"] == sel_brand), None)
    sidebar_context_block(brand=sel_brand, respondents=base_n, show_category=False)

    if not brand_data:
        st.info("Select a brand from the controls above to begin.")
        return

    # ── Deep-dive data (zone, city, rivals) ───────────────────────────────────
    with st.spinner("Loading geographic breakdown..."):
        try:
            zone_data = engine.get_zone_breakdown(sel_brand, base_n)
        except Exception as _e:
            st.warning(f"Zone breakdown unavailable: {_e}")
            zone_data = {}
        try:
            city_nps = engine.get_city_nps(sel_brand)
        except Exception as _e:
            city_nps = []
        try:
            rivals = engine.get_rival_metrics(sel_brand, base_n)
        except Exception as _e:
            rivals = []

    # ── Page header ───────────────────────────────────────────────────────────
    seg_label = sel_zone if sel_zone != "All" else "All India"

    _hero_banner(sel_brand, brand_data, base_n, seg_label, brands_list)

    # ── Fetch AI Narrative (Cached) ──────────────────────────────────────────
    import json as _json
    b_all_str = _json.dumps(brands_list, default=str)
    insights = _get_cached_narrative(sel_brand, brand_data, base_n, zone_data, city_nps, rivals, b_all_str, sel_theme, focus_n)
    
    # Ensure insights is a dict with all expected keys
    if not isinstance(insights, dict):
        insights = {"overview": str(insights), "geographic": "", "competitive": "", "nps": ""}
    for k in ["overview", "geographic", "competitive", "nps",
              "funnel", "radar", "nps_league", "city_story", "positioning",
              "salience_finding", "loyalty_finding", "dynamics_finding", "imagery_finding"]:
        if k not in insights:
            insights[k] = "Analysis pending..."

    # ── Secondary KPI row — 4 distinct cards ────────────────────────────────
    tom_leader   = max(brands_list, key=lambda x: x["tom_pct"])
    total_awa_pct = float(brand_data.get("total_awareness_pct", brand_data.get("aided_pct", 0)) or 0)
    aided_pct    = brand_data.get("aided_pct", 0)
    tom_pct      = brand_data.get("tom_pct", 0)
    spont_pct    = brand_data.get("spont_pct", 0)
    nps_v        = brand_data.get("nps")
    nps_base_n   = brand_data.get("nps_base", 0)
    aided_n      = brand_data.get("aided", 0)
    total_awa_n  = int(brand_data.get("total_awareness", brand_data.get("aided", 0)) or 0)

    # TOM rank among all brands
    tom_sorted = sorted(brands_list, key=lambda x: x["tom_pct"], reverse=True)
    tom_rank   = next((i + 1 for i, b in enumerate(tom_sorted) if b["brand_name"] == sel_brand), "—")
    tom_rank_label = f"#{tom_rank} of {len(tom_sorted)}" if isinstance(tom_rank, int) else "—"

    # NPS display
    nps_str     = f"{nps_v:+.0f}" if nps_v is not None else "N/A"
    nps_color   = "#22c55e" if (nps_v or 0) >= NPS_INDUSTRY_AVG else \
                  "#f59e0b" if (nps_v or 0) >= 0 else "#ef4444"
    nps_bg      = "#f0fdf4" if (nps_v or 0) >= NPS_INDUSTRY_AVG else \
                  "#fffbeb" if (nps_v or 0) >= 0 else "#fef2f2"
    nps_border  = "#dcfce7" if (nps_v or 0) >= NPS_INDUSTRY_AVG else \
                  "#fef3c7" if (nps_v or 0) >= 0 else "#fee2e2"
    nps_verdict = ("Above industry avg ✓" if (nps_v or 0) >= NPS_INDUSTRY_AVG
                   else "Below industry avg" if nps_v is not None else "Insufficient data")

    kpi_c1, kpi_c2, kpi_c3, kpi_c4 = st.columns(4)
    with kpi_c1:
        _metric_card(
            "🌏 Total Awareness", f"{total_awa_pct:.1f}%", "",
            subtext=f"TOM {tom_pct}% · Spont {round(spont_pct-tom_pct,1)}% · Aided {round(total_awa_pct-spont_pct,1)}%",
        )
    with kpi_c2:
        _metric_card(
            "🧠 Top of Mind", f"{tom_pct}%", "",
            subtext=f"{tom_rank_label} · Leader: {tom_leader['brand_name']} ({tom_leader['tom_pct']}%)",
        )
    with kpi_c3:
        _metric_card(
            "💬 Net Promoter Score", nps_str, "",
            subtext=nps_verdict,
            color=nps_color,
        )
    with kpi_c4:
        _consid_quick = float(brand_data.get("consideration_pct", 0) or 0)
        _ever_quick   = float(brand_data.get("ever_used_pct", 0) or 0)
        _metric_card(
            "🎯 Consideration Rate", f"{_consid_quick}%", "",
            subtext=f"Ever Used {_ever_quick}% · Aided {aided_n:,} respondents",
        )

    # ══════════════════════════════════════════════════════════════════════════
    # MAIN NAVIGATION TABS
    # st.tabs() would execute EVERY tab's body on every script rerun (Streamlit
    # only hides inactive tabs client-side, it doesn't skip their Python code) —
    # each tab here runs multiple R subprocesses + LLM calls, so a plain st.tabs
    # meant the page paid for Executive+Funnel+BrandEquity+Imagery+Loyalty+Advanced
    # on every single load regardless of which tab the user wanted. Use a radio
    # acting as the tab selector instead so only the active section computes.
    # See .regression_reference/AUDIT.md §5.
    # ══════════════════════════════════════════════════════════════════════════
    _BH_PAGE_LABELS = [
        "🏠 Executive",
        "📊 Funnel & Awareness",
        "🧪 Imagery & Analytics",
        "💎 Loyalty & Market",
    ]
    _bh_page = st.radio("Section", _BH_PAGE_LABELS, horizontal=True,
                         label_visibility="collapsed", key="bh_page_selector")

    # ── TAB 0: Executive Summary ──────────────────────────────────────────────
    if _bh_page == _BH_PAGE_LABELS[0]:
        # ── Section 0: Executive Command Briefing (C-suite role lenses) ──────
        try:
            _brief = _get_cached_briefing(
                sel_brand, brand_data, base_n, zone_data, city_nps, rivals,
                b_all_str, f"{sel_zone}|{sel_gender}|{sel_age_band}|{sel_city}",
            )
            _render_executive_briefing(_brief, sel_brand, _filter_label)
        except Exception as _be:
            print(f"[BH] exec briefing render skipped: {_be}")

        # ── Section 1: 6-metric KPI strip ────────────────────────────────────
        _bei_df = _get_brand_equity_scores()
        _bei_row = _bei_df[_bei_df["brand_name"] == sel_brand] if _bei_df is not None and not _bei_df.empty else None
        _raw_bei   = _bei_row["brand_equity_index"].iloc[0] if _bei_row is not None and not _bei_row.empty else None
        _bei_score = float(_raw_bei) if _raw_bei is not None and pd.notna(_raw_bei) else None
        _bei_rank  = int(_bei_df[_bei_df["brand_equity_index"] > (_bei_score or 0)].shape[0]) + 1 if _bei_score is not None else None
        _bei_n     = len(_bei_df) if _bei_df is not None else 0
        _consid_pct  = float(brand_data.get("consideration_pct", 0) or 0)
        _total_aware = float(brand_data.get("total_awareness_pct", brand_data.get("aided_pct", 0)) or 0)
        # CSAT comes from competitive benchmarking df (not in engine brand_data)
        _csat_v = None
        if _bei_row is not None and not _bei_row.empty and "CSAT" in _bei_row.columns:
            _csat_raw = _bei_row["CSAT"].iloc[0]
            _csat_v = float(_csat_raw) if pd.notna(_csat_raw) else None

        with st.container(border=True):
            _section_header("📊 Executive Scorecard",
                            f"{sel_brand} — key metrics at a glance · {_filter_label}")
            _kc = st.columns(6)
            def _exec_kpi(col, label, value, rank_str="", color="#1a5d4d", note=""):
                with col:
                    st.markdown(
                        f"<div style='text-align:center;padding:14px 6px 10px 6px;"
                        f"border-radius:10px;border:1px solid #e5e7eb;background:#fafafa;"
                        f"border-top:3px solid {color};'>"
                        f"<div style='font-size:0.55rem;font-weight:700;text-transform:uppercase;"
                        f"letter-spacing:0.09em;color:#9ca3af;margin-bottom:5px;'>{label}</div>"
                        f"<div style='font-size:1.8rem;font-weight:900;color:{color};line-height:1.1;'>"
                        f"{value}</div>"
                        f"<div style='font-size:0.63rem;color:#6b7280;margin-top:5px;"
                        f"white-space:nowrap;overflow:hidden;text-overflow:ellipsis;'>{rank_str}</div>"
                        f"<div style='font-size:0.55rem;color:#9ca3af;margin-top:2px;'>{note}</div>"
                        f"</div>",
                        unsafe_allow_html=True,
                    )

            # BEI score
            _exec_kpi(_kc[0], "Brand Equity",
                      f"{_bei_score:.0f}/100" if _bei_score is not None else "N/A",
                      rank_str=f"#{_bei_rank} of {_bei_n}" if _bei_rank else "—",
                      color="#1a5d4d" if (_bei_score or 0) >= 50 else "#d97706",
                      note="Composite score — internal formula")
            # Total Awareness
            _exec_kpi(_kc[1], "Total Awareness", f"{_total_aware:.1f}%",
                      rank_str=f"TOM {brand_data.get('tom_pct',0):.1f}%  Spont {round(brand_data.get('spont_pct',0)-brand_data.get('tom_pct',0),1)}%",
                      color="#0369a1", note="TOM+Spont+Inc.Aided")
            # Top of Mind
            _exec_kpi(_kc[2], "Top of Mind", f"{brand_data.get('tom_pct',0):.1f}%",
                      rank_str=f"{tom_rank_label}",
                      color="#7c3aed", note="Unprompted first recall")
            # NPS
            nps_val = brand_data.get("nps")
            _exec_kpi(_kc[3], "NPS",
                      f"{nps_val:+.0f}" if nps_val is not None else "N/A",
                      rank_str=nps_verdict,
                      color="#15803d" if (nps_val or 0) >= 45 else ("#d97706" if (nps_val or 0) >= 0 else "#dc2626"),
                      note="Promoters − Detractors")
            # Consideration
            _exec_kpi(_kc[4], "Consideration", f"{_consid_pct:.1f}%",
                      rank_str="Would consider brand",
                      color="#0891b2", note="Independent survey Q")
            # CSAT — scale-aware display (5-pt or 10-pt per project_meta["csat_scale"])
            _csat_suffix = f"/{_CSAT_SCALE}"
            _csat_green  = _CSAT_SCALE * 0.70   # ≥70% of scale = good
            _csat_amber  = _CSAT_SCALE * 0.50   # ≥50% of scale = ok
            _exec_kpi(_kc[5], f"CSAT (0–{_CSAT_SCALE})",
                      f"{_csat_v:.1f}{_csat_suffix}" if _csat_v is not None else "N/A",
                      rank_str="Recent buyers" if _csat_v is not None else "No data",
                      color="#059669" if (_csat_v or 0) >= _csat_green else ("#d97706" if (_csat_v or 0) >= _csat_amber else "#dc2626"),
                      note="Satisfaction score")

        # ── Section 2: Funnel quick view + AI Narrative (side by side) ───────
        _s2l, _s2r = st.columns([3, 2])
        with _s2l:
            with st.container(border=True):
                _section_header("🔽 Brand Funnel Snapshot",
                                "Absolute penetration at each consumer journey stage (% all respondents)")
                _fu_stages = [
                    ("Total Awareness", _total_aware, "#1a5d4d"),
                    ("Ever Used",       float(brand_data.get("ever_used_pct", 0) or 0), "#0369a1"),
                    ("Consideration",   _consid_pct,  "#7c3aed"),
                    ("Last Purchased",  float(brand_data.get("last_purchased_pct", 0) or 0), "#059669"),
                    ("Preferred",       float(brand_data.get("preferred_pct", 0) or 0), "#0891b2"),
                ]
                _max_v = max(v for _, v, _ in _fu_stages) or 1
                for _i, (_flabel, _fval, _fcol) in enumerate(_fu_stages):
                    _fw = max(4, round(_fval / _max_v * 100))
                    # Conversion rate from previous stage
                    _prev_val = _fu_stages[_i - 1][1] if _i > 0 else None
                    _conv_html = ""
                    if _prev_val and _prev_val > 0:
                        _conv = round(_fval / _prev_val * 100, 1)
                        _conv_col = "#15803d" if _conv >= 60 else ("#d97706" if _conv >= 35 else "#dc2626")
                        _conv_html = (
                            f"<div style='font-size:0.58rem;color:{_conv_col};font-weight:700;"
                            f"min-width:42px;text-align:center;'>"
                            f"↓{_conv:.0f}%</div>"
                        )
                    st.markdown(
                        f"<div style='display:flex;align-items:center;gap:8px;margin-bottom:6px;'>"
                        f"<div style='width:118px;font-size:0.72rem;font-weight:600;color:#6b7280;"
                        f"text-align:right;flex-shrink:0;'>{_flabel}</div>"
                        f"<div style='flex:1;background:#f3f4f6;border-radius:4px;height:22px;'>"
                        f"<div style='width:{_fw}%;background:{_fcol};border-radius:4px;height:22px;"
                        f"display:flex;align-items:center;padding-left:8px;'>"
                        f"<span style='font-size:0.72rem;font-weight:800;color: #e5e7eb;white-space:nowrap;'>"
                        f"{_fval:.1f}%</span></div></div>"
                        f"{_conv_html}</div>",
                        unsafe_allow_html=True,
                    )
                st.caption("↓X% = stage conversion from prior stage. ⚠️ Awareness & usage stages are independent survey questions.")
                _funnel_summary = (
                    f"Total Awareness {_total_aware:.1f}%, Ever Used {float(brand_data.get('ever_used_pct',0) or 0):.1f}%, "
                    f"Consideration {_consid_pct:.1f}%, Last Purchased {float(brand_data.get('last_purchased_pct',0) or 0):.1f}%, "
                    f"Preferred {float(brand_data.get('preferred_pct',0) or 0):.1f}%"
                )
                _render_ai_headline("funnel", _funnel_summary, brand=sel_brand,
                                    project_id=st.session_state.get("active_project_id", "project_1"))

        with _s2r:
            with st.container(border=True):
                _section_header("🤖 AI Strategic Overview", "LLM-generated brand intelligence")
                _ai_card(insights["overview"], "", "#30a76a")

        # ── Section 3: Competitive Position Table ────────────────────────────
        with st.container(border=True):
            _section_header("🏆 Competitive Position",
                            f"{sel_brand} rank vs. all tracked brands across 5 dimensions")
            _comp_df = _bei_df if (_bei_df is not None and not _bei_df.empty) else None
            if _comp_df is not None:
                _comp_df2 = _get_competitive_benchmarking()
                _pos_data = []
                _metrics_pos = [
                    ("Brand Equity Index", "brand_equity_index", _bei_df, None),
                ]
                # Build rank table from bei_df + competitive df
                _rank_specs = [
                    ("Brand Equity Index", "brand_equity_index", _bei_df),
                    ("Total Awareness",    "TOTAL_AIDED",         _comp_df2 if not _comp_df2.empty else None),
                    ("Top of Mind",        "TOM",                 _comp_df2 if not _comp_df2.empty else None),
                    ("NPS",                "nps",                 _comp_df2 if not _comp_df2.empty else None),
                    ("Consideration",      "CONSIDERATION",       _comp_df2 if not _comp_df2.empty else None),
                ]
                _tbl_rows = []
                for _mn, _col, _src in _rank_specs:
                    if _src is None or _col not in _src.columns:
                        continue
                    _src2 = _src.dropna(subset=[_col]).sort_values(_col, ascending=False).reset_index(drop=True)
                    _brand_pos = _src2[_src2["brand_name"] == sel_brand]
                    if _brand_pos.empty:
                        continue
                    _val = float(_brand_pos[_col].iloc[0])
                    _rank = int(_src2[_src2[_col] > _val].shape[0]) + 1
                    _n_brands = len(_src2)
                    _leader = _src2.iloc[0]["brand_name"] if not _src2.empty else "—"
                    _leader_val = float(_src2.iloc[0][_col]) if not _src2.empty else 0
                    _suffix = "/100" if _mn == "Brand Equity Index" else ("pts" if _mn == "NPS" else "%")
                    _tbl_rows.append({
                        "Metric": _mn,
                        f"{sel_brand}": f"{_val:.1f}{_suffix}",
                        "Rank": f"#{_rank} of {_n_brands}",
                        "Leader": f"{_leader} ({_leader_val:.1f}{_suffix})",
                    })
                if _tbl_rows:
                    _pos_tbl = pd.DataFrame(_tbl_rows)
                    st.dataframe(_pos_tbl, hide_index=True, use_container_width=True)
                else:
                    st.info("Competitive rank data not available.")
            else:
                st.info("BEI data required for competitive position view.")

        # ── Section 4: Awareness Landscape (full market view) ────────────────
        with st.container(border=True):
            _section_header("🌏 Market Awareness Landscape",
                            f"Stacked TOM · Spontaneous · Aided · respondent-weighted ALL benchmark · {_filter_label}")
            # ── Chart filters ───────────────────────────────────────────────
            _SORT_KEYS = {"Total Awareness": "total_awareness_pct", "Top of Mind": "tom_pct",
                          "Spontaneous": "spont_pct", "NPS": "nps"}
            _lc1, _lc2, _lc3 = st.columns([1.2, 1, 2])
            with _lc1:
                _awa_sort_lbl = st.selectbox("Rank by", list(_SORT_KEYS), index=0, key="awa_sort")
            _awa_key = _SORT_KEYS[_awa_sort_lbl]
            with _lc2:
                _awa_n = st.slider("Top N brands", 5, min(25, len(brands_list)),
                                   min(15, len(brands_list)), 1, key="awa_topn")
            _awa_pool = [b for b in brands_list if b.get(_awa_key) is not None]
            _awa_names = [b["brand_name"] for b in sorted(_awa_pool, key=lambda x: x.get(_awa_key, 0) or 0, reverse=True)]
            with _lc3:
                _awa_pick = st.multiselect("Or pick specific brands (overrides Top N)", _awa_names,
                                           default=[], key="awa_pick",
                                           help="Leave empty to show the Top N. Selected brands always include the focus brand.")
            _awa_sorted = sorted(brands_list, key=lambda x: x.get(_awa_key, 0) or 0, reverse=True)
            if _awa_pick:
                _sel = set(_awa_pick) | {sel_brand}
                _awa_show = [b for b in _awa_sorted if b["brand_name"] in _sel]
            else:
                _awa_show = _awa_sorted[:_awa_n]
                if sel_brand not in [b["brand_name"] for b in _awa_show]:
                    _focus = next((b for b in _awa_sorted if b["brand_name"] == sel_brand), None)
                    if _focus:
                        _awa_show = _awa_show + [_focus]
            if len(_awa_show) < 1:
                st.info("No brands match this filter.")
            else:
                # AI headline for landscape: top brand + focus brand position
                _top_brand = _awa_show[0]["brand_name"] if _awa_show else "—"
                _top_aware = _awa_show[0].get("total_awareness_pct", _awa_show[0].get("aided_pct", 0)) or 0
                _focus_rank = next((i+1 for i, b in enumerate(_awa_sorted) if b["brand_name"] == sel_brand), None)
                _awa_hl_summary = (
                    f"Market leader by total awareness: {_top_brand} ({_top_aware:.1f}%). "
                    f"{sel_brand} ranks #{_focus_rank} of {len(_awa_sorted)} brands. "
                    f"{sel_brand} TOM={float(brand_data.get('tom_pct',0) or 0):.1f}%, "
                    f"Total Aware={float(brand_data.get('total_awareness_pct', brand_data.get('aided_pct',0)) or 0):.1f}%."
                )
                _render_ai_headline("awareness_landscape", _awa_hl_summary, brand=sel_brand,
                                    project_id=st.session_state.get("active_project_id", "project_1"))
                st.plotly_chart(
                    _theme_fig(_awareness_landscape_chart(_awa_show, base_n=base_n,
                                                          top_n=len(_awa_show), all_brands_ref=_awa_sorted)),
                    use_container_width=True,
                )
                st.caption("Salience hierarchy: TOM (darkest) → Spontaneous → Aided (lightest). "
                           "▲▼ = significantly above/below the respondent-weighted ALL benchmark (95%). "
                           "Use the controls above to rank by a different metric, change Top N, or pick brands.")

    # ── Local section closures (capture outer scope, render into tabs) ───────────────

    def _tab_salience():
        with st.container(border=True):
            _section_header(
                "📊 Strategic Salience & Health Decomposition",
                f"{sel_brand} — Aided → Spontaneous → Top of Mind conversion"
            )
            with st.expander("ℹ️ Analysis Guide", expanded=False):
                st.markdown(f"**Critical Finding:** {insights['salience_finding']}")
                st.markdown("""
            **How to interpret:**
            - **Funnel Depth:** Shows the drop-off from general awareness (Aided) to active recall (Spontaneous) and ultimate preference (Top of Mind).
            - **Health Score:** A weighted index combining awareness and NPS to measure overall brand 'strength' compared to peers.
            """)
            _SEG_OPTIONS = {
                "Overall":  [],
                "Zone":     ["North", "South", "East", "West"],
                "Gender":   ["Male", "Female"],
                "Age Band": ["25-35", "36-50"],
                "City":     ["Delhi", "Mumbai", "Bangalore", "Chennai", "Hyderabad", "Kolkata",
                             "Ahmedabad", "Lucknow", "Patna", "Bhubaneshwar", "Nagaon", "Bikaner",
                             "Patiala", "Cochin", "Guntur", "Hassan", "Kolhapur", "Ujjain"],
            }
            other_brands = [b["brand_name"] for b in brands_list if b["brand_name"] != sel_brand]
            cmp_c1, cmp_c2 = st.columns([1, 3])
            with cmp_c1:
                st.markdown(
                    f"<div style='font-size:0.65rem;color:#9ca3af;text-transform:uppercase;"
                    f"font-weight:700;letter-spacing:0.05em;margin-bottom:4px;'>Base Brand</div>"
                    f"<div style='font-size:0.95rem;font-weight:800;color: #e5e7eb;"
                    f"padding:6px 12px;background:#1a5d4d;border-radius:8px;border:1px solid #30a76a;'>"
                    f"{sel_brand}</div>",
                    unsafe_allow_html=True,
                )
            with cmp_c2:
                cmp_brands_extra = st.multiselect(
                    "Add brands to compare (max 6)",
                    other_brands, default=[], max_selections=6, key="bh_cmp_brands",
                    placeholder="Select brands to compare side-by-side…",
                )
            seg_type = "Overall"
            seg_vals = []
            _cmp_conf = 0.95
            if cmp_brands_extra:
                cs1, cs2, cs3 = st.columns([1, 2, 1])
                with cs1:
                    seg_type = st.selectbox(
                        "Split comparison by",
                        ["Overall", "Zone", "Gender", "Age Band", "City"],
                        key="bh_seg_type",
                        help="Break down comparison funnels by this demographic dimension",
                    )
                with cs2:
                    seg_options_list = _SEG_OPTIONS.get(seg_type, [])
                    if seg_type != "Overall" and seg_options_list:
                        seg_vals = st.multiselect(
                            f"{seg_type} segments to show",
                            seg_options_list, default=seg_options_list[:2], key="bh_seg_vals",
                        )
                    else:
                        st.caption("Showing overall totals for each comparison brand.")
                with cs3:
                    _cmp_conf = st.selectbox("Flag differences at", [0.95, 0.90, 0.80], index=0,
                                             format_func=lambda a: f"{a:.0%}", key="bh_cmp_conf")
            _fc_col1, _fc_col2 = st.columns(2)
            with _fc_col1:
                sel_aware_stages = st.multiselect(  # noqa: E501
                    "Awareness Funnel Stages to display",
                    ["Total Awareness", "Spontaneous Recall", "Consideration", "Top of Mind"],
                    default=["Total Awareness", "Spontaneous Recall", "Consideration", "Top of Mind"],
                    key="bh_aware_stages_sel"
                )
            with _fc_col2:
                sel_conv_stages = st.multiselect(
                    "Conversion Funnel Stages to display",
                    ["Total Awareness", "Ever Tried (Trial)", "Current Usage", "Most Often Used Brand (MOUB)"],
                    default=["Total Awareness", "Ever Tried (Trial)", "Current Usage", "Most Often Used Brand (MOUB)"],
                    key="bh_conv_stages_sel"
                )

            # Pre-compute comparison data if brands are selected
            cmp_data = {}
            if cmp_brands_extra:
                cmp_brands_all = [sel_brand] + cmp_brands_extra
                seg_type_key = seg_type.lower().replace(" ", "_")
                with st.spinner("Computing comparison funnels…"):
                    try:
                        cmp_data = engine.get_funnel_comparison(
                            brands=cmp_brands_all,
                            segment_type=seg_type_key,
                            segment_values=seg_vals if seg_vals else None,
                        )
                    except Exception as _e:
                        st.error(f"Comparison error: {_e}")
                        cmp_data = {}

            _ftab_aware, _ftab_conv = st.tabs(["Awareness Funnel", "Conversion Funnel"])
            with _ftab_aware:
                _render_funnel_with_arrows(brand_data, visible_stages=sel_aware_stages)
                try:
                    _funnel_png = _funnel_png_bytes(brand_data)
                    st.download_button(
                        "📥 Download Awareness Funnel PNG", _funnel_png,
                        f"{sel_brand}_awareness_funnel.png", "image/png", key="dl_funnel_primary"
                    )
                except Exception as _e:
                    st.caption(f"⚠️ PNG export failed: {_e}")

                if cmp_data:
                    cmp_brands_all = [sel_brand] + cmp_brands_extra
                    st.divider()
                    st.markdown("**⚖ Brand Awareness Comparison**")
                    _render_comparison_funnels_html(cmp_data, sel_brand, alpha=round(1 - _cmp_conf, 2), funnel_type="awareness")
                    st.caption(
                        f"▲/▼ = significantly higher/lower than {sel_brand} at that stage "
                        f"(pooled two-proportion z-test, {_cmp_conf:.0%} confidence)."
                    )
                    n_cmp = len(list(cmp_data.keys()))
                    try:
                        import io as _io
                        _cmp_fig = _side_by_side_funnels(cmp_data, sel_brand, funnel_type="awareness")
                        _cmp_buf = _io.BytesIO()
                        _cmp_fig.write_image(_cmp_buf, format="png", scale=2,
                                             width=max(1000, n_cmp * 280), height=max(540, len(seg_vals) * 220 if seg_vals else 540))
                        _cmp_buf.seek(0)
                        st.download_button(
                            "📥 Download Awareness Comparison Funnel PNG", _cmp_buf,
                            f"{sel_brand}_awareness_comparison_funnel.png", "image/png", key="dl_funnel_cmp_aware"
                        )
                    except Exception as _png_err:
                        st.caption(f"⚠️ Comparison PNG export note: {_png_err}")

                    cmp_rows = []
                    for brand in cmp_brands_all:
                        for seg, d in cmp_data.get(brand, {}).items():
                            cmp_rows.append({
                                "Brand": brand, "Segment": seg,
                                "Total Awareness %": d.get("total_awareness_pct", d.get("aided_pct", 0)),
                                "Aided Only %":      d.get("aided_pct", 0),
                                "Spont %":           d.get("spont_pct", 0),
                                "Consideration %":   d.get("consideration_pct", 0),
                                "TOM %":             d.get("tom_pct", 0),
                                "Base N":            d.get("base_n", 0),
                            })
                    if cmp_rows:
                        with st.expander("📋 View Awareness Comparison Data Table", expanded=False):
                            st.dataframe(pd.DataFrame(cmp_rows), hide_index=True, use_container_width=True)

            with _ftab_conv:
                _render_conversion_funnel(brand_data, visible_stages=sel_conv_stages)
                try:
                    _conv_png = _conversion_funnel_png_bytes(brand_data)
                    st.download_button(
                        "📥 Download Conversion Funnel PNG", _conv_png,
                        f"{sel_brand}_conversion_funnel.png", "image/png", key="dl_funnel_conversion"
                    )
                except Exception as _e:
                    st.caption(f"⚠️ PNG export failed: {_e}")

                if cmp_data:
                    cmp_brands_all = [sel_brand] + cmp_brands_extra
                    st.divider()
                    st.markdown("**⚖ Brand Conversion Comparison**")
                    _render_comparison_funnels_html(cmp_data, sel_brand, alpha=round(1 - _cmp_conf, 2), funnel_type="conversion")
                    st.caption(
                        f"▲/▼ = significantly higher/lower than {sel_brand} at that stage "
                        f"(pooled two-proportion z-test, {_cmp_conf:.0%} confidence)."
                    )
                    n_cmp = len(list(cmp_data.keys()))
                    try:
                        import io as _io
                        _cmp_fig = _side_by_side_funnels(cmp_data, sel_brand, funnel_type="conversion")
                        _cmp_buf = _io.BytesIO()
                        _cmp_fig.write_image(_cmp_buf, format="png", scale=2,
                                             width=max(1000, n_cmp * 280), height=max(540, len(seg_vals) * 220 if seg_vals else 540))
                        _cmp_buf.seek(0)
                        st.download_button(
                            "📥 Download Conversion Comparison Funnel PNG", _cmp_buf,
                            f"{sel_brand}_conversion_comparison_funnel.png", "image/png", key="dl_funnel_cmp_conv"
                        )
                    except Exception as _png_err:
                        st.caption(f"⚠️ Comparison PNG export note: {_png_err}")

                    cmp_rows = []
                    for brand in cmp_brands_all:
                        for seg, d in cmp_data.get(brand, {}).items():
                            cmp_rows.append({
                                "Brand": brand, "Segment": seg,
                                "Total Awareness %": d.get("total_awareness_pct", d.get("aided_pct", 0)),
                                "Ever Tried %":      d.get("ever_used_pct", 0),
                                "Current Usage %":   d.get("current_pct", 0),
                                "Preferred %":       d.get("preferred_pct", 0),
                                "Base N":            d.get("base_n", 0),
                            })
                    if cmp_rows:
                        with st.expander("📋 View Conversion Comparison Data Table", expanded=False):
                            st.dataframe(pd.DataFrame(cmp_rows), hide_index=True, use_container_width=True)

            st.divider()
            _ai_card(insights['funnel'], "AI Funnel & Health Analysis", "#f59e0b")

    def _tab_imagery():
        with st.container(border=True):
            _section_header(
                "🧪 Imagery & Driver Analysis",
                "Key Driver Regression → BIP → CAN MAP: regression output filters the downstream tabs"
            )

            # ── multi-project guard (2026-08-04: H3 fix) ─────────────────────
            _active_pid = st.session_state.get("active_project_id", "project_1")
            _has_imagery = False
            try:
                from oxdata.db_loader import get_db_path as _gdbp
                import sqlite3 as _sql
                _img_db = _gdbp(required_table="fact_brand_imagery", project_id=_active_pid)
                if _img_db:
                    _c = _sql.connect(str(_img_db))
                    _img_cnt = _c.execute("SELECT COUNT(*) FROM fact_brand_imagery").fetchone()[0]
                    _c.close()
                    _has_imagery = _img_cnt > 0
            except Exception:
                _has_imagery = False

            if not _has_imagery:
                st.warning(
                    f"**Project `{_active_pid}` does not have Brand Imagery data (BQ3 battery).**  "
                    "Sections below are only available for studies with a brand imagery question battery.  "
                    "Switch to a project that has imagery data, or use **Add Project** to ingest a new study.",
                    icon="⚠️",
                )
                return
            # ─────────────────────────────────────────────────────────────────

            with st.expander("ℹ️ Analysis Guide", expanded=False):
                st.markdown(f"**Critical Finding:** {insights['imagery_finding']}")
                st.markdown("""
**How to use:** Run **Key Driver Regression** first. Significant driver attributes are stored and
automatically used to filter **BIP** and **CAN MAP** — so those tabs show only the attributes that
actually drive your outcome. Switch tabs in order: Regression → BIP → CAN MAP.
""")

            # ── Section 1: Key Driver Regression (always rendered → no widget state loss) ─
            with st.expander("🎯 Key Driver Regression", expanded=True):
                st.caption(
                    f"{sel_brand} — pick outcome → rank driver importance → feeds BIP & CAN MAP below."
                )
                try:
                    from oxdata.skills.r_bridge import r_available
                    if r_available():
                        _render_driver_regression(sel_brand, brands_list,
                                                  zone_arg, gender_arg, age_band_arg, city_arg,
                                                  project_id=_active_pid)
                    else:
                        st.info("R is not available on this host — Key Driver Regression requires Rscript.")
                except ImportError as _kdr_imp_err:
                    st.error(f"Key Driver Regression failed to load: {_kdr_imp_err}")
                _cur_dids  = st.session_state.get("driver_flow_attr_ids", [])
                _cur_brand = st.session_state.get("driver_flow_brand")
                if _cur_dids and _cur_brand == sel_brand:
                    st.success(f"✓ {len(_cur_dids)} driver attributes captured — "
                               "expand **BIP** or **CAN MAP** below to explore them filtered.")

            # ── Section 2: BIP (driver-filtered when regression has run, full otherwise) ─
            _bip_dids    = st.session_state.get("driver_flow_attr_ids", [])
            _bip_outcome = st.session_state.get("driver_flow_outcome", "")
            _bip_brand   = st.session_state.get("driver_flow_brand", "")
            _bip_use_filter = bool(_bip_dids) and _bip_brand == sel_brand
            _bip_label = (
                f"📊 Image Profiling (BIP)  —  🔗 filtered to {len(_bip_dids)} driver attrs"
                if _bip_use_filter else "📊 Image Profiling (BIP)"
            )
            with st.expander(_bip_label, expanded=True):
                if _bip_use_filter:
                    st.caption(
                        f"Showing only the **{len(_bip_dids)} significant driver attributes** "
                        f"from the {_bip_outcome} regression for {sel_brand}. "
                        "Re-run regression above to update the filter."
                    )
                with st.spinner("Running BIP Normalization…"):
                    _render_section_14_bip(
                        "All", zone_arg, gender_arg, age_band_arg, city_arg,
                        sel_brand=sel_brand, project_id=_active_pid,
                        attr_ids=_bip_dids if _bip_use_filter else None,
                    )

            # ── Section 3: CAN MAP (driver-filtered when regression has run, full otherwise) ─
            _ca_dids    = st.session_state.get("driver_flow_attr_ids", [])
            _ca_outcome = st.session_state.get("driver_flow_outcome", "")
            _ca_brand   = st.session_state.get("driver_flow_brand", "")
            _ca_aware   = st.session_state.get("driver_flow_awareness_stages")
            _ca_use_filter = bool(_ca_dids) and _ca_brand == sel_brand
            _ca_label = (
                f"🗺️ Perceptual Mapping (CAN MAP)  —  🔗 filtered to {len(_ca_dids)} driver attrs"
                if _ca_use_filter else "🗺️ Perceptual Mapping (CAN MAP)"
            )
            with st.expander(_ca_label, expanded=False):
                if _ca_use_filter:
                    st.caption(
                        f"Showing only the **{len(_ca_dids)} significant driver attributes** "
                        f"from the {_ca_outcome} regression for {sel_brand}. "
                        "Re-run regression above to update the filter."
                    )
                with st.spinner("Running Correspondence Analysis…"):
                    _render_section_13_can_map(
                        "All", zone_arg, gender_arg, age_band_arg, city_arg,
                        sel_brand, project_id=_active_pid,
                        attr_ids=_ca_dids if _ca_use_filter else None,
                        awareness_stages=_ca_aware if _ca_use_filter else None,
                    )

    def _tab_competitive():
        with st.container(border=True):
            _section_header(
                "⚔️ Competitive Benchmarking Panel",
                "Side-by-side comparison of all brands across funnel stages, NPS, and CSAT"
            )
            with st.expander("ℹ️ How to interpret", expanded=False):
                st.markdown("""
**What this shows:** Every brand measured on the same yardstick — funnel penetration %, NPS, and CSAT.

- **Heatmap Table:** Darker green = higher. Quickly spot who leads and who lags each metric.
- **Metric Rankings:** Pick any single metric and rank all brands. Selected brand highlighted dark green.
- **Funnel Comparison:** Line chart showing each brand's absolute penetration at each funnel stage.

*Base: min 30 respondents per cell. NPS base: min 20 NPS respondents. CSAT linked via last-purchased brand.*
                """)
            _render_competitive_benchmarking(sel_brand, min_base_n=min_base_n,
                                             zone=zone_arg, gender=gender_arg,
                                             age_band=age_band_arg, city=city_arg)

    def _tab_bei():
        with st.container(border=True):
            _section_header(
                "🏅 Brand Equity Index (BEI)",
                "Composite health score — TOM + Consideration + NPS + CSAT combined into single 0–100 index"
            )
            st.caption(
                "⚠ **Internal working formula, not an industry-standard index.** The 25/30/25/20 "
                "weights below are this team's own judgment call on how to combine these four "
                "metrics — not a cited or externally validated methodology. Useful as a relative "
                "ranking across brands in this dataset; treat the absolute score as directional, "
                "not authoritative."
            )
            with st.expander("ℹ️ How to interpret", expanded=False):
                st.markdown("""
**Formula:** BEI = (TOM% norm × 0.25) + (Consideration% norm × 0.30) + (NPS norm × 0.25) + (CSAT norm × 0.20)

Each metric is normalized 0–100 within the dataset before weighting, so all components are on equal footing.

- **Score > 66:** Strong brand equity (green zone)
- **Score 33–66:** Average equity (amber zone)
- **Score < 33:** Weak equity (red zone — needs intervention)
                """)
            _render_brand_equity_index(sel_brand)

    def _tab_ipa():
        with st.container(border=True):
            _section_header(
                "📊 Importance-Performance Grid",
                f"{sel_brand} — which attributes are strengths to maintain vs. opportunities to improve"
            )
            with st.expander("ℹ️ How to interpret", expanded=False):
                st.markdown("""
**Quadrants:**
- **Top-Right (Strengths — Maintain):** High importance + high brand association. Core assets, defend these.
- **Top-Left (Opportunities — Improve):** High importance + low brand association. Priority fix areas.
- **Bottom-Right (Monitor):** Low importance + high association. Not urgent but don't over-invest.
- **Bottom-Left (Low Priority):** Low importance + low association. De-prioritise.
                """)
            _render_ipa_grid(sel_brand)

    def _tab_funnel_leakage():
        with st.container(border=True):
            _section_header(
                "📉 Funnel Conversion & Leakage",
                f"{sel_brand} — where consumers drop off, vs. category average at each stage"
            )
            with st.expander("ℹ️ How to interpret", expanded=False):
                st.markdown("""
**What this shows:** At each funnel transition (e.g. Aided Aware → Ever Used), what % of the previous-stage audience moves forward?

- **Above category average:** Brand converts better than peers at that transition.
- **Below category average:** A leakage point — awareness not translating into action.
- **Biggest leakage = most actionable:** Low EVER_USED→CONSIDERATION signals product experience gaps.
                """)
            _render_funnel_leakage(sel_brand, zone=zone_arg, gender=gender_arg,
                                   age_band=age_band_arg, city=city_arg)

    def _tab_attribute_ownership():
        with st.container(border=True):
            _section_header(
                "🏆 Attribute Ownership Map",
                "Which brand 'owns' each product attribute in the consumer mind?"
            )
            with st.expander("ℹ️ How to interpret", expanded=False):
                st.markdown("""
**Owning an attribute** = highest association % among all brands — competitive moat.
**Lead Margin** = gap over 2nd-place brand. Small margin = vulnerable ownership.
**"Brand Position" tab:** Shows which attributes the selected brand owns and which are most winnable.

*Based on BQ3b brand-attribute association (442K rows).*
                """)
            _render_attribute_ownership(sel_brand)

    def _tab_key_driver():
        with st.container(border=True):
            _section_header(
                "🔬 Key Driver Impact Simulator",
                f"{sel_brand} — simulate NPS impact of improving specific brand attributes"
            )
            # Key Driver Regression now lives in its own tab: Imagery & Analytics →
            # "🎯 Key Driver Regression" (moved out of this expander to avoid duplicate
            # widget IDs and give it first-class placement).
            st.caption("📐 Driver regression moved to its own tab → **Imagery & Analytics · 🎯 Key Driver Regression** "
                       "(selectable outcome, top-box recode, importance ranking, cross-brand heatmap).")

            # ── R Factor Analysis (XLSTAT-style EFA) ─────────────────────────
            try:
                from oxdata.skills.r_bridge import r_available, run_r_stat
                if r_available():
                    with st.expander("🔬 R Factor Analysis — Attribute Dimensions (EFA)", expanded=False):
                        with st.spinner("Running R factor analysis…"):
                            import sqlite3
                            from oxdata.db_loader import get_db_path as _get_db_for_r
                            _fconn = sqlite3.connect(_get_db_for_r())
                            _fa_df = None
                            _fa_err = None
                            try:
                                _bi_fa = pd.read_sql(
                                    "SELECT bi.respondent_id, da.attr_label, bi.value "
                                    "FROM fact_brand_imagery bi "
                                    "JOIN dim_brand db ON bi.brand_id = db.brand_id "
                                    "JOIN dim_bq3_attribute da ON bi.attr_id = da.attr_id "
                                    "WHERE db.brand_name = ?",
                                    _fconn, params=[sel_brand]
                                )
                                if not _bi_fa.empty:
                                    _fa_df = _bi_fa.pivot_table(
                                        index="respondent_id", columns="attr_label",
                                        values="value", aggfunc="max", fill_value=0
                                    ).reset_index(drop=True)
                            except Exception as _fe:
                                _fa_err = str(_fe)
                            finally:
                                _fconn.close()

                        if _fa_df is not None and len(_fa_df) >= 30:
                            _fa_result = run_r_stat("factor_analysis", _fa_df)
                            if "error" in _fa_result:
                                st.warning(f"R factor analysis: {_fa_result['error']}")
                            else:
                                _fa_factors = _fa_result.get("factors", [])
                                _fa_var_exp = _fa_result.get("variance_explained", [])
                                _fa_total   = _fa_result.get("total_var_explained", 0)

                                # Variance explained bar chart
                                if _fa_var_exp:
                                    _fmet1, _fmet2 = st.columns(2)
                                    with _fmet1:
                                        st.metric("Factors extracted", _fa_result.get("n_factors", 0))
                                    with _fmet2:
                                        st.metric("Total variance explained", f"{_fa_total:.1%}")

                                    _fig_fa = go.Figure(go.Bar(
                                        x=[f"F{i+1}" for i in range(len(_fa_var_exp))],
                                        y=[v * 100 for v in _fa_var_exp],
                                        text=[f"{v:.1%}" for v in _fa_var_exp],
                                        textposition="outside",
                                        marker_color="#1a5d4d",
                                    ))
                                    _base_fa = {k: v for k, v in _chart_layout_base(280).items()
                                                if k not in ("xaxis", "yaxis")}
                                    _fig_fa.update_layout(
                                        **_base_fa,
                                        xaxis=dict(title="Factor"),
                                        yaxis=dict(title="Variance Explained %", range=[0, max(_fa_var_exp) * 120]),
                                        title=dict(text="EFA — Variance Explained per Factor (varimax rotation)", font=dict(size=12)),
                                        showlegend=False,
                                    )
                                    st.plotly_chart(_theme_fig(_fig_fa), use_container_width=True)

                                # Top attributes per factor
                                if _fa_factors:
                                    _fcols = st.columns(min(len(_fa_factors), 4))
                                    for _fi, (_fc, _fac) in enumerate(zip(_fcols, _fa_factors)):
                                        with _fc:
                                            top_a = _fac.get("top_attrs", [])
                                            st.markdown(
                                                f"<div style='font-weight:700;font-size:0.78rem;color:#1a5d4d;"
                                                f"padding:4px 0 6px 0;'>F{_fac['factor']} "
                                                f"({_fac.get('variance', 0):.1%})</div>",
                                                unsafe_allow_html=True,
                                            )
                                            for _attr in top_a:
                                                _load = _attr["loading"]
                                                _bar_w = int(abs(_load) * 100)
                                                _col   = "#16a34a" if _load > 0 else "#dc2626"
                                                st.markdown(
                                                    f"<div style='font-size:0.72rem;margin-bottom:3px;'>"
                                                    f"<span style='color:#374151'>{_attr['attr']}</span> "
                                                    f"<span style='color:{_col};font-weight:700'>{_load:+.2f}</span>"
                                                    f"<div style='height:4px;width:{_bar_w}%;background:{_col};"
                                                    f"border-radius:2px;margin-top:1px;opacity:0.6'></div></div>",
                                                    unsafe_allow_html=True,
                                                )
                                    st.caption("Loadings ≥ |0.40| are typically considered meaningful. Green = positive, Red = negative.")
                        else:
                            _nfa = len(_fa_df) if _fa_df is not None else 0
                            if _fa_err:
                                st.warning(f"R factor analysis data error: {_fa_err}")
                            else:
                                st.info(f"Need ≥30 respondents with imagery data for factor analysis "
                                        f"(found {_nfa} for {sel_brand}).")
            except ImportError:
                pass

            # Statistical tests section removed — undecided scope, to be redesigned

    def _tab_advocacy_loyalty():
        with st.container(border=True):
            _section_header(
                "💎 Advocacy & Loyalty",
                "Net Promoter Score decomposition, league ranking, and city-level performance"
            )
            with st.expander("ℹ️ How to interpret", expanded=False):
                st.markdown(f"**Critical Finding:** {insights['loyalty_finding']}")
                st.markdown(f"""
**NPS Formula:** NPS = % Promoters − % Detractors (range: −100 to +100). Industry avg: **+{NPS_INDUSTRY_AVG}**.
- **League Table:** Ranked among all tracked brands with ≥30 raters.
- **City Performance:** Zone colour-coded — each city bar reflects local brand advocacy.
- **Zone bands:** Red = detractor zone (<0) · Amber = developing (0–{NPS_INDUSTRY_AVG}) · Green = champion (>{NPS_INDUSTRY_AVG}).
                """)
            l_tabs = st.tabs(["NPS Deep Dive", "Performance League", "City Performance"])
            with l_tabs[0]:
                p_pct  = brand_data.get("nps_promoters_pct")
                pa_pct = brand_data.get("nps_passives_pct", 0)
                d_pct  = brand_data.get("nps_detractors_pct", 0)
                nps_v  = brand_data.get("nps")
                if p_pct is not None:
                    _nps_hl_summary = (
                        f"{sel_brand} NPS={nps_v:+.0f}, promoters={p_pct:.0f}%, "
                        f"passives={pa_pct:.0f}%, detractors={d_pct:.0f}%, "
                        f"industry avg=+{NPS_INDUSTRY_AVG}."
                    )
                    _render_ai_headline("nps", _nps_hl_summary, brand=sel_brand,
                                        project_id=st.session_state.get("active_project_id", "project_1"))
                if p_pct is not None:
                    # Compute league rank for this brand
                    _nps_elig = sorted(
                        [b for b in brands_list if b.get("nps") is not None and b.get("nps_base", 0) >= 30],
                        key=lambda b: b["nps"], reverse=True,
                    )
                    _nps_lg_rank = next((i + 1 for i, b in enumerate(_nps_elig) if b["brand_name"] == sel_brand), None)
                    _nps_lg_n    = len(_nps_elig)
                    _nps_pctile  = round((1 - (_nps_lg_rank - 1) / _nps_lg_n) * 100) if _nps_lg_rank else None

                    nps_col_score, nps_col_bar, nps_col_ai = st.columns([0.7, 1.8, 1])
                    with nps_col_score:
                        nps_color_hex = "#22c55e" if (nps_v or 0) >= NPS_INDUSTRY_AVG else "#f59e0b" if (nps_v or 0) >= 0 else "#ef4444"
                        delta_nps  = round((nps_v or 0) - NPS_INDUSTRY_AVG, 1)
                        delta_sign = "+" if delta_nps >= 0 else ""
                        _rank_line = (
                            f"#{_nps_lg_rank} of {_nps_lg_n} · Top {100-_nps_pctile+1:.0f}%"
                            if _nps_lg_rank and _nps_pctile else "—"
                        )
                        nps_html = (
                            f"<div style='background:#f9fafb; border:1px solid #e5e7eb; border-radius:14px;"
                            f" padding:20px 16px; text-align:center;'>"
                            f"<div style='font-size:0.58rem; text-transform:uppercase; letter-spacing:0.1em;"
                            f" color:#6b7280; margin-bottom:8px;'>Net Promoter Score</div>"
                            f"<div style='font-size:3.5rem; font-weight:900; color:{nps_color_hex}; line-height:1;'>{nps_v:+.0f}</div>"
                            f"<div style='font-size:0.72rem; color:#9ca3af; margin-top:6px;'>"
                            f"vs Industry +{NPS_INDUSTRY_AVG} &nbsp; <b style='color:{nps_color_hex}'>{delta_sign}{delta_nps}</b></div>"
                            f"<div style='font-size:0.68rem; color:{nps_color_hex}; font-weight:700; margin-top:4px;'>"
                            f"League: {_rank_line}</div>"
                            f"<hr style='margin:12px 0; border-color:#e5e7eb;'>"
                            f"<div style='display:grid; grid-template-columns:1fr 1fr; gap:8px;'>"
                            f"<div style='background:#f0fdf4; border:1px solid #dcfce7; border-radius:8px; padding:10px;'>"
                            f"<div style='font-size:1.25rem; font-weight:800; color:#16a34a;'>{p_pct:.0f}%</div>"
                            f"<div style='font-size:0.55rem; color:#166534; text-transform:uppercase; font-weight:700;'>Promoters</div></div>"
                            f"<div style='background:#fef2f2; border:1px solid #fee2e2; border-radius:8px; padding:10px;'>"
                            f"<div style='font-size:1.25rem; font-weight:800; color:#dc2626;'>{d_pct:.0f}%</div>"
                            f"<div style='font-size:0.55rem; color:#991b1b; text-transform:uppercase; font-weight:700;'>Detractors</div></div></div>"
                            f"<div style='margin-top:10px; background:#fffbeb; border:1px solid #fef3c7; border-radius:8px; padding:8px;'>"
                            f"<div style='font-size:1.1rem; font-weight:700; color:#92400e;'>{pa_pct:.0f}%</div>"
                            f"<div style='font-size:0.55rem; color:#78350f; text-transform:uppercase; font-weight:700;'>Passives</div></div>"
                            f"<div style='font-size:0.6rem; color:#9ca3af; margin-top:8px;'>n={brand_data.get('nps_base',0):,} raters</div>"
                            f"</div>"
                        )
                        st.markdown(nps_html, unsafe_allow_html=True)
                    with nps_col_bar:
                        st.plotly_chart(_theme_fig(_nps_stacked_bar(brand_data)), use_container_width=True)
                    with nps_col_ai:
                        _ai_card(insights['nps'], "AI Advocacy Insights", "#22c55e")
                else:
                    st.info(f"NPS data unavailable for {sel_brand}.")
            with l_tabs[1]:
                _eligible_count = sum(1 for b in brands_list if b.get("nps") is not None and b.get("nps_base", 0) >= 30)
                fig_nps_rank = None
                if _eligible_count < 5:
                    st.info(
                        f"NPS league unavailable — only {_eligible_count} brand(s) have ≥30 raters. "
                        "This project's NPS question is asked once overall, not per-brand, so most "
                        "brands have no qualifying rater base."
                    )
                else:
                    lg_ctl_l, lg_ctl_r = st.columns([2, 3])
                    with lg_ctl_l:
                        _l_min = 2 if _eligible_count < 5 else 5
                        _l_max = max(_l_min, min(30, _eligible_count))
                        _l_val = min(15, _l_max)
                        league_top_n = st.slider(
                            "Brands to show in league",
                            min_value=_l_min, max_value=_l_max,
                            value=_l_val, step=1, key="league_top_n",
                        )

                    with lg_ctl_r:
                        st.caption(
                            f"Showing top {league_top_n} of {_eligible_count} eligible brands (≥30 raters). "
                            f"{sel_brand} always included."
                        )
                    fig_nps_rank = _nps_rankings_chart(brands_list, min_raters=30, top_n=league_top_n, highlight=sel_brand)
                    if fig_nps_rank:
                        l_col, r_col = st.columns([2, 1])
                        with l_col:
                            st.plotly_chart(_theme_fig(fig_nps_rank), use_container_width=True)
                        with r_col:
                            _ai_card(insights['nps_league'], "AI League Analysis", "#f97316")
                    else:
                        st.info("Insufficient data for NPS league (min 30 raters per brand).")
            with l_tabs[2]:
                if city_nps:
                    city_count = len(city_nps)
                    if city_count < 2:
                        city_top_n = city_count
                    else:
                        city_ctl_l, city_ctl_r = st.columns([2, 3])
                        with city_ctl_l:
                            _min_val = 2 if city_count < 5 else 5
                            _max_val = max(_min_val, min(20, city_count))
                            _def_val = min(12, _max_val)
                            city_top_n = st.slider("Cities to show", _min_val, _max_val, _def_val, key="city_top_n")
                        with city_ctl_r:
                            st.caption("Sorted best → worst NPS. Bar colour = Zone. Min 15 raters per city.")
                    cl_col, cr_col = st.columns([2, 1])
                    with cl_col:
                        st.plotly_chart(_theme_fig(_city_nps_chart(city_nps, sel_brand, top_n=city_top_n)), use_container_width=True)
                    with cr_col:
                        _ai_card(insights['city_story'], "AI City Story", "#06b6d4")
                else:
                    st.info("City-level NPS data unavailable.")


    def _tab_csat():
        with st.container(border=True):
            _section_header(
                "⭐ Customer Satisfaction (CSAT)",
                "Satisfaction score (0–10) from recent buyers — distinct from NPS (would recommend)"
            )
            with st.expander("ℹ️ NPS vs CSAT", expanded=False):
                st.markdown("""
**NPS** = "Would you recommend?" (bq2b) — measures brand advocacy and word-of-mouth.
**CSAT** = "How satisfied are you?" (bq5) — measures fulfilment of expectations post-purchase.
A brand can have high CSAT but low NPS (satisfied but not enthusiastic enough to recommend).
Base = recent buyers only (4,704 respondents across categories).
                """)
            csat_data   = _get_brand_csat(sel_brand)
            mkt_csat    = _get_csat_overall()
            if csat_data:
                avg_csat   = csat_data["avg_csat"]
                n_csat     = csat_data["n"]
                mkt_avg    = mkt_csat["avg_csat"] if mkt_csat else avg_csat
                nps_v      = brand_data.get("nps", 0) or 0
                csat_color = "#22c55e" if avg_csat >= 7.5 else "#f59e0b" if avg_csat >= 6.0 else "#ef4444"
                gap_vs_mkt = round(avg_csat - mkt_avg, 2)
                gap_label  = f"+{gap_vs_mkt:.2f} vs market" if gap_vs_mkt >= 0 else f"{gap_vs_mkt:.2f} vs market"
                gap_color  = "#15803d" if gap_vs_mkt >= 0 else "#dc2626"
                _csat_summary = (
                    f"{sel_brand} CSAT={avg_csat:.1f}/10 (n={n_csat}), market avg={mkt_avg:.1f}/10, "
                    f"gap={gap_vs_mkt:+.2f}. NPS={nps_v:+.0f}."
                )
                _render_ai_headline("csat", _csat_summary, brand=sel_brand,
                                    project_id=st.session_state.get("active_project_id", "project_1"))
                c1, c2, c3 = st.columns([1, 2, 2])
                with c1:
                    st.markdown(
                        f"<div style='background:#f9fafb;border:1px solid #e5e7eb;border-radius:14px;"
                        f"padding:24px 16px;text-align:center;'>"
                        f"<div style='font-size:0.58rem;text-transform:uppercase;letter-spacing:0.1em;"
                        f"color:#6b7280;margin-bottom:8px;'>{sel_brand} CSAT</div>"
                        f"<div style='font-size:3.5rem;font-weight:900;color:{csat_color};line-height:1;'>"
                        f"{avg_csat:.1f}</div>"
                        f"<div style='font-size:0.75rem;color:#9ca3af;margin-top:6px;'>out of 10</div>"
                        f"<div style='font-size:0.65rem;color:{gap_color};margin-top:4px;font-weight:700;'>{gap_label}</div>"
                        f"<div style='font-size:0.65rem;color:#6b7280;margin-top:2px;'>n={n_csat:,} last buyers</div>"
                        f"<hr style='margin:12px 0;border-color:#e5e7eb;'>"
                        f"<div style='font-size:0.58rem;text-transform:uppercase;color:#6b7280;'>NPS ({sel_brand})</div>"
                        f"<div style='font-size:1.6rem;font-weight:800;color:#1d4ed8;'>{nps_v:+.0f}</div>"
                        f"</div>",
                        unsafe_allow_html=True,
                    )
                with c2:
                    csat_dist = _get_brand_csat_distribution(sel_brand)
                    try:
                        if not csat_dist.empty:
                            # Keep x numeric so the mean reference line lands exactly
                            # (add_vline on a categorical/string axis raises a TypeError).
                            _csat_score = pd.to_numeric(csat_dist["score"], errors="coerce")
                            _colors_csat = ["#ef4444" if s <= 3 else "#f59e0b" if s <= 6 else "#22c55e"
                                            for s in _csat_score]
                            _csat_nmax = float(pd.to_numeric(csat_dist["n"], errors="coerce").max() or 0)
                            fig_csat = go.Figure(go.Bar(
                                x=_csat_score, y=csat_dist["n"],
                                marker_color=_colors_csat,
                                text=csat_dist["n"],
                                textposition="outside",
                                hovertemplate="Score %{x}: %{y} respondents<extra></extra>",
                            ))
                            # Mean score vertical reference line (numeric axis → exact position)
                            fig_csat.add_vline(
                                x=float(avg_csat),
                                line_dash="dash", line_color="#374151", line_width=1.5,
                                annotation_text=f"Mean {avg_csat:.1f}",
                                annotation_position="top left",
                                annotation_font_size=10,
                            )
                            # Zone labels (numeric x positions)
                            for _zone_x, _zone_lbl, _zone_col in [
                                (1.5, "Detractors (0–3)", "#ef4444"),
                                (5.0, "Neutral (4–6)", "#f59e0b"),
                                (8.5, "Satisfied (7–10)", "#22c55e"),
                            ]:
                                fig_csat.add_annotation(
                                    x=_zone_x, y=_csat_nmax * 0.92,
                                    text=_zone_lbl, showarrow=False,
                                    font=dict(size=8, color=_zone_col),
                                    bgcolor="rgba(255,255,255,0.7)",
                                )
                            layout_base = {k: v for k, v in _chart_layout_base(290).items()
                                           if k not in ("xaxis", "yaxis")}
                            fig_csat.update_layout(
                                **layout_base,
                                xaxis=dict(title="Satisfaction Score (0–10)", gridcolor="#f1f5f9",
                                           dtick=1, range=[-0.6, 10.6]),
                                yaxis=dict(title="Respondents", gridcolor="#f1f5f9"),
                                title=dict(text=f"{sel_brand} CSAT Distribution (Recent Buyers)", font=dict(size=12)),
                            )
                            st.plotly_chart(_theme_fig(fig_csat), use_container_width=True)
                            st.caption("Red = Dissatisfied (0–3) · Amber = Neutral (4–6) · Green = Satisfied (7–10). Dashed = mean score.")
                    except Exception as _e:
                        st.warning(f"CSAT distribution unavailable: {_e}")
                with c3:
                    st.markdown(
                        f"<div style='background:#f0f9ff;border:1px solid #bae6fd;border-radius:12px;"
                        f"padding:20px 16px;'>"
                        f"<div style='font-size:0.65rem;text-transform:uppercase;letter-spacing:0.1em;"
                        f"color:#0369a1;font-weight:700;margin-bottom:12px;'>Satisfaction vs Advocacy Gap</div>"
                        f"<p style='font-size:0.78rem;color:#1e40af;line-height:1.5;'>"
                        f"CSAT = <b>satisfied</b>. NPS = would <b>recommend</b>. "
                        f"High CSAT + lower NPS = satisfied but not enthusiastic. "
                        f"Opportunity to convert satisfaction into active advocacy.</p>"
                        f"<div style='margin-top:10px;display:flex;gap:8px;'>"
                        f"<div style='flex:1;padding:10px;background:white;border-radius:8px;text-align:center;border:1px solid #e5e7eb;'>"
                        f"<div style='font-size:0.58rem;color:#6b7280;text-transform:uppercase;font-weight:700;'>{sel_brand}</div>"
                        f"<div style='font-size:1.6rem;font-weight:900;color:{csat_color};'>{avg_csat:.1f}/10</div>"
                        f"<div style='font-size:0.6rem;color:#9ca3af;'>brand CSAT</div>"
                        f"</div>"
                        f"<div style='flex:1;padding:10px;background:white;border-radius:8px;text-align:center;border:1px solid #e5e7eb;'>"
                        f"<div style='font-size:0.58rem;color:#6b7280;text-transform:uppercase;font-weight:700;'>Market Avg</div>"
                        f"<div style='font-size:1.6rem;font-weight:900;color:#6b7280;'>{mkt_avg:.1f}/10</div>"
                        f"<div style='font-size:0.6rem;color:{gap_color};font-weight:700;'>{gap_label}</div>"
                        f"</div>"
                        f"</div></div>",
                        unsafe_allow_html=True,
                    )
            else:
                st.info("CSAT data not available.")

    def _tab_portfolio():
        with st.container(border=True):
            _section_header(
                "🗂️ Portfolio Awareness",
                "Which product categories do consumers associate with each brand?"
            )
            with st.expander("ℹ️ How to interpret", expanded=False):
                st.markdown("""
Portfolio awareness (bq6) measures **category-brand linkage** — beyond just "aware of brand", do consumers know what products they make?
A brand with high ceiling fan awareness but low water heater association has a portfolio perception gap.
Base = respondents who are aware of the brand.
                """)
            _render_portfolio_awareness(sel_brand, zone_arg, gender_arg, age_band_arg, city_arg)

    def _tab_price_tier():
        with st.container(border=True):
            _section_header(
                "💰 Price Tier Distribution",
                "What price points did buyers actually pay, by category?"
            )
            with st.expander("ℹ️ How to interpret", expanded=False):
                st.markdown("""
Price tier data (bq0b) captures the **actual price range paid** by recent purchasers for each category.
Useful for: price segment concentration, premium vs. economy splits, comparing across zones.
Base = respondents who recently purchased in that category.
                """)
            _render_price_tier_distribution("All", zone_arg, gender_arg, age_band_arg, city_arg)

    def _tab_demographics():
        with st.container(border=True):
            _section_header(
                "👥 Consumer Demographics Profile",
                f"{sel_brand} — who considers this brand? Gender, age, zone breakdown with index vs. average"
            )
            with st.expander("ℹ️ How to interpret", expanded=False):
                st.markdown("""
**Index = 100:** This segment considers the brand at exactly the average rate.
**Index > 110 (green):** Over-index — this segment is disproportionately likely to consider the brand.
**Index < 90 (red):** Under-index — the brand is weak in this segment. Growth opportunity.
                """)
            _render_demographic_profile(sel_brand, zone_arg, gender_arg, age_band_arg, city_arg)

    def _tab_purchase_journey():
        with st.container(border=True):
            _section_header(
                "🛒 Purchase Journey",
                "Why buyers chose this category, where they researched, and how they decided"
            )
            with st.expander("ℹ️ About this data", expanded=False):
                st.markdown(
                    "Purchase journey data covers **recent buyers** across all tracked categories. "
                    "Responses are coded (numerical) — labels are standardised from the survey codebook."
                )
            _render_purchase_journey(zone_arg, gender_arg, age_band_arg, city_arg)

    def _tab_consumer_attitudes():
        with st.container(border=True):
            _section_header(
                "💭 Consumer Attitudes",
                "Category belief statements (AQ4) — agree/not sure/disagree. Base = category buyers/owners."
            )
            with st.expander("ℹ️ How to interpret", expanded=False):
                st.markdown("""
**AQ4** = Category attitude statements rated as Agree / Not Sure / Disagree.
Each category has 5–6 statements about usage habits, purchase behaviour, and product beliefs.
Use to identify consumer misconceptions, category barriers, and opportunity messaging angles.
Base = respondents who own or recently purchased in the category.
                """)
            try:
                import sqlite3
                from oxdata.db_loader import get_db_path
                _at_conn = sqlite3.connect(get_db_path())
                # Check dim_aq4_attribute exists
                has_dim = _at_conn.execute(
                    "SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name='dim_aq4_attribute'"
                ).fetchone()[0]
                if not has_dim:
                    st.info("Attitude dimension table not yet created. Run the DB migration script.")
                    _at_conn.close()
                    return

                # Category code → suffix mapping
                CAT_CODE_TO_SUFFIX = {1: 'cf', 4: 'led', 5: 'wh', 3: 'mg', 2: 'ac', 6: 'wp'}
                CAT_SUFFIX_TO_NAME = {'cf': 'Ceiling Fans', 'led': 'LED Batten', 'wh': 'Water Heater',
                                      'mg': 'Mixer Grinder', 'ac': 'Air Cooler', 'wp': 'Water Pumps'}
                SCORE_LABELS = {1: 'Agree', 2: 'Not Sure', 3: 'Disagree'}
                SCORE_COLORS = {1: '#15803d', 2: '#ca8a04', 3: '#dc2626'}

                available_cats = [r[0] for r in _at_conn.execute(
                    "SELECT DISTINCT category_code FROM fact_attitudes ORDER BY category_code"
                ).fetchall()]
                _at_conn.close()

                if not available_cats:
                    st.info("No attitude data available.")
                    return

                cat_names = [CAT_SUFFIX_TO_NAME.get(CAT_CODE_TO_SUFFIX.get(c, ''), f'Cat {c}') for c in available_cats]
                at_tab_labels = [f"{n}" for n in cat_names]
                at_tabs = st.tabs(at_tab_labels)

                for _at_tab, _cat_code in zip(at_tabs, available_cats):
                    with _at_tab:
                        _at_conn2 = sqlite3.connect(get_db_path())
                        _cat_suf = CAT_CODE_TO_SUFFIX.get(_cat_code, '')
                        _at_cte, _at_fp, _at_filt = _resp_filter_cte(zone_arg, gender_arg, age_band_arg, city_arg)
                        _at_base = " AND a.respondent_id IN (SELECT respondent_id FROM _base)" if _at_filt else ""
                        _at_prefix = f"WITH {_at_cte} " if _at_filt else ""
                        try:
                            _at_df = pd.read_sql(
                                f"""
                                {_at_prefix}
                                SELECT a.attr_num, d.attr_label,
                                       a.score, COUNT(*) as n
                                FROM fact_attitudes a
                                JOIN dim_aq4_attribute d
                                    ON d.category_suffix=? AND d.attr_num=a.attr_num
                                WHERE a.category_code=?{_at_base}
                                GROUP BY a.attr_num, d.attr_label, a.score
                                ORDER BY a.attr_num, a.score
                                """,
                                _at_conn2, params=_at_fp + [_cat_suf, _cat_code]
                            )
                        finally:
                            _at_conn2.close()

                        if _at_df.empty:
                            st.info("No data.")
                            continue

                        # Pivot to pct per statement
                        total_per_stmt = _at_df.groupby("attr_num")["n"].sum()
                        _at_df["pct"] = _at_df.apply(
                            lambda r: r["n"] / total_per_stmt[r["attr_num"]] * 100, axis=1
                        )

                        # For each statement: diverging Likert chart (Agree right, Disagree left)
                        labels = _at_df.drop_duplicates("attr_num").set_index("attr_num")["attr_label"].to_dict()
                        stmt_ids = sorted(labels.keys())

                        # Build pct lookup {(attr_num, score): pct}
                        _pct_lk = {}
                        for sid in stmt_ids:
                            for sc in (1, 2, 3):
                                _r = _at_df[(_at_df["attr_num"] == sid) & (_at_df["score"] == sc)]
                                _pct_lk[(sid, sc)] = float(_r["pct"].iloc[0]) if not _r.empty else 0.0

                        # Sort by Agree% descending (most agreed at top)
                        sorted_ids = sorted(stmt_ids, key=lambda s: _pct_lk.get((s, 1), 0), reverse=True)

                        def _trunc_lbl(txt, n=68):
                            return txt[:n - 1] + "…" if len(txt) > n else txt

                        y_labels     = [_trunc_lbl(labels[sid]) for sid in sorted_ids]
                        agree_vals   = [_pct_lk.get((sid, 1), 0) for sid in sorted_ids]
                        not_sure_v   = [_pct_lk.get((sid, 2), 0) for sid in sorted_ids]
                        disagree_neg = [-_pct_lk.get((sid, 3), 0) for sid in sorted_ids]

                        # Insight strip: strongest agree / most contested
                        _top_agreed   = labels[sorted_ids[0]]
                        _top_ag_pct   = agree_vals[0]
                        _most_dis_sid = sorted_ids[-1]
                        _most_dis_pct = abs(disagree_neg[-1])
                        _at_ic1, _at_ic2 = st.columns(2)
                        with _at_ic1:
                            st.markdown(
                                f"<div style='background:#f0fdf4;border:1px solid #bbf7d0;"
                                f"border-radius:10px;padding:10px 14px;margin-bottom:8px;'>"
                                f"<div style='font-size:0.58rem;font-weight:700;text-transform:uppercase;"
                                f"letter-spacing:0.08em;color:#15803d;margin-bottom:3px;'>Strongest Agreement</div>"
                                f"<div style='font-size:0.82rem;font-weight:700;color:#14532d;line-height:1.3;'>"
                                f"{_trunc_lbl(_top_agreed, 72)}</div>"
                                f"<div style='font-size:0.7rem;color:#166534;margin-top:3px;'>"
                                f"{_top_ag_pct:.0f}% agree</div>"
                                f"</div>",
                                unsafe_allow_html=True,
                            )
                        with _at_ic2:
                            st.markdown(
                                f"<div style='background:#fef2f2;border:1px solid #fecaca;"
                                f"border-radius:10px;padding:10px 14px;margin-bottom:8px;'>"
                                f"<div style='font-size:0.58rem;font-weight:700;text-transform:uppercase;"
                                f"letter-spacing:0.08em;color:#dc2626;margin-bottom:3px;'>Highest Resistance</div>"
                                f"<div style='font-size:0.82rem;font-weight:700;color:#7f1d1d;line-height:1.3;'>"
                                f"{_trunc_lbl(labels[_most_dis_sid], 72)}</div>"
                                f"<div style='font-size:0.7rem;color:#991b1b;margin-top:3px;'>"
                                f"{_most_dis_pct:.0f}% disagree</div>"
                                f"</div>",
                                unsafe_allow_html=True,
                            )

                        # Diverging bar: Disagree leftward (neg), Agree+NotSure rightward (pos)
                        fig_at = go.Figure()

                        fig_at.add_trace(go.Bar(
                            name="Disagree",
                            y=y_labels, x=disagree_neg,
                            orientation="h",
                            marker_color="#dc2626",
                            text=[f"{abs(v):.0f}%" if abs(v) >= 7 else "" for v in disagree_neg],
                            textposition="inside", insidetextanchor="middle",
                            textfont=dict(color="white", size=10),
                            customdata=[abs(v) for v in disagree_neg],
                            hovertemplate="<b>%{y}</b><br>Disagree: %{customdata:.1f}%<extra></extra>",
                        ))
                        fig_at.add_trace(go.Bar(
                            name="Agree",
                            y=y_labels, x=agree_vals,
                            orientation="h",
                            marker_color="#15803d",
                            text=[f"{v:.0f}%" if v >= 7 else "" for v in agree_vals],
                            textposition="inside", insidetextanchor="middle",
                            textfont=dict(color="white", size=10),
                            hovertemplate="<b>%{y}</b><br>Agree: %{x:.1f}%<extra></extra>",
                        ))
                        fig_at.add_trace(go.Bar(
                            name="Not Sure",
                            y=y_labels, x=not_sure_v,
                            orientation="h",
                            marker_color="#ca8a04", opacity=0.75,
                            text=[f"{v:.0f}%" if v >= 7 else "" for v in not_sure_v],
                            textposition="inside", insidetextanchor="middle",
                            textfont=dict(color="white", size=10),
                            hovertemplate="<b>%{y}</b><br>Not Sure: %{x:.1f}%<extra></extra>",
                        ))

                        fig_at.add_vline(x=0, line_color="#9ca3af", line_width=1.5)

                        _x_bound = max(
                            max(agree_vals) + max(not_sure_v) if agree_vals else 0,
                            max(abs(v) for v in disagree_neg) if disagree_neg else 0,
                        ) + 10
                        _base_at = {k: v for k, v in _chart_layout_base(max(380, len(sorted_ids) * 52)).items()
                                    if k not in ("xaxis", "yaxis", "legend", "margin")}
                        fig_at.update_layout(
                            **_base_at,
                            barmode="relative",
                            xaxis=dict(
                                title="← Disagree (%) · Agree (%) →",
                                range=[-_x_bound, _x_bound],
                                zeroline=True, zerolinecolor="#9ca3af",
                                tickformat=".0f",
                                gridcolor="#f1f5f9",
                            ),
                            yaxis=dict(automargin=True, tickfont=dict(size=11)),
                            legend=dict(orientation="h", yanchor="top", y=-0.06,
                                        x=0.5, xanchor="center",
                                        traceorder="normal"),
                            margin=dict(l=10, r=60, t=52, b=52),
                            title=dict(
                                text=f"{CAT_SUFFIX_TO_NAME.get(_cat_suf, '')} — Consumer Attitude Statements (Sorted by Agreement)",
                                font=dict(size=12),
                            ),
                        )
                        st.plotly_chart(_theme_fig(fig_at), use_container_width=True)
                        st.caption(
                            "Sorted highest→lowest Agreement. "
                            "Green = Agree · Red = Disagree (extending left) · Amber = Not Sure. "
                            "AQ4 base = category owners/buyers."
                        )
            except Exception as _e:
                st.warning(f"Attitudes section unavailable: {_e}")

    def _tab_market_dynamics():
        with st.container(border=True):
            _section_header(
                "🗺️ Market Dynamics & Positioning",
                "Geographic footprint, market landscape clusters, and competitive positioning"
            )
            with st.expander("ℹ️ How to interpret", expanded=False):
                st.markdown(f"**Critical Finding:** {insights['dynamics_finding']}")
                st.markdown("""
**Geographic Footprint:** Map bubble size = respondent base (market size). Colour intensity = TOM%.
Zone bars: grouped by TOM / Spont / Aided per zone; dashed lines = all-zone average.

**Market Landscape:** Heatmap = brand co-awareness (% aware of both). Cluster map uses PCA on zone TOM+NPS vectors.

**Competitive & Positioning:** Strategic Map — NPS vs TOM positioning, bubble size = Aided awareness.
                """)
            fig_pos = _brand_positioning_chart(brands_list, sel_brand)
            m_tabs = st.tabs(["Geographic Footprint", "Market Landscape", "Competitive & Positioning"])
            with m_tabs[0]:
                if zone_data:
                    india_map_fig = _india_zone_map(zone_data, sel_brand)
                    if india_map_fig:
                        st.plotly_chart(_theme_fig(india_map_fig), use_container_width=True)
                    col_geo_left, col_geo_right = st.columns([2, 1])
                    with col_geo_left:
                        st.plotly_chart(_theme_fig(_zone_comparison_chart(zone_data, sel_brand)), use_container_width=True)
                        zone_nps_fig = _zone_nps_chart(zone_data, sel_brand)
                        if zone_nps_fig:
                            st.plotly_chart(_theme_fig(zone_nps_fig), use_container_width=True)
                    with col_geo_right:
                        _ai_card(insights['geographic'], "AI Geographic Story", "#0ea5e9")
                else:
                    st.info("Geographic breakdown unavailable.")
            with m_tabs[1]:
                ml_col_left, ml_col_right = st.columns([1, 1])
                with ml_col_left:
                    st.markdown("**Brand Co-Awareness (Correlation)**")
                    with st.spinner("Computing brand correlations..."):
                        corr_data = _get_cached_correlation()
                    if corr_data and corr_data.get("brands"):
                        st.plotly_chart(_theme_fig(_correlation_heatmap(corr_data)), use_container_width=True)
                        st.caption("Each cell = % of respondents aware of both brands.")
                    else:
                        st.info("Correlation data unavailable.")
                with ml_col_right:
                    st.markdown("**Geographic Strength Clusters**")
                    with st.spinner("Computing positioning map..."):
                        zone_matrix = _get_cached_zone_matrix()
                    corr_map_fig = _correspondence_map(zone_matrix, highlight_brand=sel_brand)
                    if corr_map_fig:
                        st.plotly_chart(_theme_fig(corr_map_fig), use_container_width=True)
                    else:
                        st.info("Cluster map unavailable.")
                    st.caption("PCA on zone TOM+NPS vectors. Closer brands = similar regional strength profile.")
            with m_tabs[2]:
                # Rival-vs-metric ranking used to live here as a second "Rival Benchmarks" view
                # mode, duplicating the Funnel & Awareness tab's Competitive Benchmarking Panel
                # → Metric Rankings sub-tab (same "rank brands by metric" bar chart, just a
                # narrower rivals-only subset with fewer metrics). Removed — Strategic Map below
                # is the only chart here with no equivalent elsewhere on the page.
                if fig_pos:
                    p_left, p_right = st.columns([2, 1])
                    with p_left:
                        st.plotly_chart(_theme_fig(fig_pos), use_container_width=True)
                    with p_right:
                        _ai_card(insights['positioning'], "AI Positioning Insight", "#1a5d4d")
                    st.caption("Bubble size = Aided awareness %. ★ = selected brand.")
                else:
                    st.info("Positioning map unavailable.")

    # ── TAB 5: Driver Flow — Regression → BIP → CAN MAP in one connected pipeline ─────────────────
    def _tab_driver_flow():
        """Unified Driver Analysis Flow.

        Step 1: Choose outcome (NPS / CSAT / imagery attribute) → run regression.
        Step 2: BIP Normalization filtered to the significant driver attributes only.
        Step 3: CAN MAP filtered to the same driver attributes only.

        This is the industry-standard approach ('Rahul process'):
        regression tells you WHAT matters → BIP tells you WHO is strong on it → CAN MAP shows WHERE brands sit.
        """
        _active_pid = st.session_state.get("active_project_id", "project_1")

        with st.container(border=True):
            _section_header(
                "🔗 Driver-Linked Analysis Flow",
                "Regression → BIP → CAN MAP in one connected pipeline. "
                "Drivers from Step 1 auto-filter Steps 2 and 3."
            )
            st.markdown(
                "<div style='background:#f0fdf4;border:1px solid #86efac;border-radius:8px;"
                "padding:10px 14px;margin-bottom:12px;font-size:0.8rem;color:#166534;'>"
                "<b>How to use:</b> Pick your outcome (Step 1), run the regression. "
                "The attributes that significantly drive that outcome become the lens for BIP (Step 2) and CAN MAP (Step 3). "
                "This shows which attributes matter, which brand owns them, and how brands are positioned on what counts."
                "</div>",
                unsafe_allow_html=True,
            )

        # ── Step 1: Regression ───────────────────────────────────────────────
        with st.container(border=True):
            _section_header("Step 1 — Key Driver Regression",
                            f"{sel_brand} · pick outcome → find which attributes drive it")
            try:
                from oxdata.skills.r_bridge import r_available
                if not r_available():
                    st.info("R is not available — Key Driver Regression requires Rscript.")
                    return
            except ImportError as _e:
                st.error(f"R bridge not available: {_e}")
                return

            _fc1, _fc2, _fc3, _fc4 = st.columns(4)
            with _fc1:
                _df_dv_type = st.selectbox("Outcome", ["NPS", "CSAT", "Imagery attribute"],
                                           key="df_dv_type",
                                           help="What to explain. NPS/CSAT = satisfaction outcomes. "
                                                "Imagery attribute = which other attributes predict this one.")
            _df_dv_attr = None
            if _df_dv_type == "Imagery attribute":
                _df_attr_opts = _brand_attr_options(sel_brand, project_id=_active_pid)
                with _fc2:
                    if not _df_attr_opts:
                        st.info("No attributes available for this brand.")
                        return
                    _df_dv_attr = st.selectbox("Attribute (outcome)", _df_attr_opts, key="df_dvattr")
                _df_dv_kind = "ATTR"
                _df_topbox = 0
            else:
                _df_dv_kind = _df_dv_type
                # Scale-aware topbox options: CSAT uses project csat_scale, NPS always 0-10
                _df_scale = _CSAT_SCALE if _df_dv_type == "CSAT" else 10
                _df_max   = int(_df_scale)
                _df_t2    = _df_max - 1  # e.g. 9 for 10-pt, 4 for 5-pt
                _df_t3    = _df_max - 2  # e.g. 8 for 10-pt, 3 for 5-pt
                _DF_TOPBOX = {
                    f"Top-2 Box ({_df_t2}–{_df_max} → 1)": _df_t2,
                    f"Top-3 Box ({_df_t3}–{_df_max} → 1)": _df_t3,
                    f"Top-1 Box ({_df_max} → 1)":           _df_max,
                    f"Raw 0–{_df_max} score":                0,
                }
                with _fc2:
                    _df_tb_lbl = st.selectbox("Scale recode", list(_DF_TOPBOX), index=0,
                                              key="df_topbox")
                _df_topbox = _DF_TOPBOX[_df_tb_lbl]
            with _fc3:
                _df_model_lbl = st.selectbox("Model", ["Logistic", "Linear (LPM)"],
                                             key="df_model",
                                             help="Logistic for binary outcomes (top-box). Linear for raw 0–10.")
            _df_model = "logistic" if _df_model_lbl.startswith("Logistic") else "linear"
            with _fc4:
                _df_sig_thresh = st.select_slider("Significance", options=[0.01, 0.05, 0.10],
                                                  value=0.10, format_func=lambda v: f"p < {v}",
                                                  key="df_sig_thresh",
                                                  help="Attributes with p-value below this are treated as significant drivers.")

            _df_dv_is_binary = (_df_dv_kind == "ATTR") or (_df_topbox > 0)
            if _df_model == "logistic" and not _df_dv_is_binary:
                _df_model = "linear"
                st.caption("Auto-switched to Linear (raw outcome is not binary).")

            with st.spinner("Running regression…"):
                _df_res = _driver_regression_for_brand(
                    sel_brand, _df_dv_kind, _df_topbox,
                    zone_arg, gender_arg, age_band_arg, city_arg,
                    _df_model, _df_dv_attr,
                )

            if not _df_res:
                st.info(f"No {_df_dv_kind} + imagery data for {sel_brand}.")
                return
            if "error" in _df_res:
                if _df_res.get("error") == "insufficient":
                    st.info(f"Need ≥30 respondents with {_df_dv_kind} + imagery data.")
                else:
                    st.warning(f"Regression: {_df_res['error']}")
                return

            _df_all_drivers = _df_res.get("significant_drivers", [])
            if not _df_all_drivers:
                st.info("Regression returned no drivers.")
                return

            # Filter to significant at chosen threshold
            _df_sig_drivers = [d for d in _df_all_drivers
                               if d.get("p_value") is None or d.get("p_value", 1.0) <= _df_sig_thresh]
            if not _df_sig_drivers:
                _df_sig_drivers = sorted(_df_all_drivers, key=lambda d: d.get("importance", 0), reverse=True)[:8]
                st.caption(f"No drivers met p < {_df_sig_thresh} — showing top 8 by importance instead.")

            # AI headline from top driver results
            if _df_sig_drivers:
                _top3 = [_clean_attr(d["attribute"]) for d in _df_sig_drivers[:3]]
                _r2v  = _df_res.get("mcfadden_r2" if _df_model == "logistic" else "r_squared", 0)
                _img_summary = (
                    f"Top imagery drivers of {_df_dv_kind} for {sel_brand}: "
                    f"{', '.join(_top3)}. Model R²={_r2v:.3f}, n={_df_res.get('n',0)} respondent-brand pairs."
                )
                _render_ai_headline("imagery_drivers", _img_summary, brand=sel_brand,
                                    project_id=_active_pid)

            # Show quick driver ranking
            _df_metric_cols = st.columns(4)
            _df_metric_cols[0].metric("Significant drivers", len(_df_sig_drivers))
            _r2_key = "mcfadden_r2" if _df_model == "logistic" else "r_squared"
            _df_metric_cols[1].metric(
                "McFadden R²" if _df_model == "logistic" else "R²",
                f"{_df_res.get(_r2_key, 0):.3f}"
            )
            _df_metric_cols[2].metric("Respondents", _df_res.get("n", 0))
            _df_metric_cols[3].metric("Model", "Logistic" if _df_model == "logistic" else "Linear")

            # Driver bar chart
            _df_import_sorted = sorted(_df_sig_drivers, key=lambda d: d.get("importance", 0), reverse=True)
            _df_bar_fig = go.Figure(go.Bar(
                x=[d.get("importance", 0) for d in _df_import_sorted],
                y=[_clean_attr(d["attribute"]) for d in _df_import_sorted],
                orientation="h",
                marker=dict(
                    color=[("#22c55e" if (d.get("coef", d.get("std_coef", 0)) or 0) > 0 else "#ef4444")
                           for d in _df_import_sorted],
                    opacity=0.85,
                ),
                text=[f"{d.get('importance', 0):.1f}%" for d in _df_import_sorted],
                textposition="outside",
                hovertemplate="<b>%{y}</b><br>Importance: %{x:.2f}%<extra></extra>",
            ))
            _df_bar_fig.update_layout(
                **{k: v for k, v in _chart_layout_base(max(280, len(_df_import_sorted) * 28)).items()
                   if k not in ("xaxis", "yaxis")},
                xaxis=dict(title="Relative Importance %", range=[0, max(d.get("importance", 0) for d in _df_import_sorted) * 1.25]),
                yaxis=dict(autorange="reversed"),
                title=dict(text=f"Driver Importance — {sel_brand} ({_df_dv_kind} outcome)", font=dict(size=12)),
                showlegend=False,
            )
            st.plotly_chart(_theme_fig(_df_bar_fig), use_container_width=True)
            st.caption("Green = positive driver (more association → better outcome). Red = negative driver.")

            # Resolve attr_ids for significant drivers using DB lookup
            _df_driver_labels = [d["attribute"] for d in _df_sig_drivers]
            _df_attr_ids: list[int] = []
            try:
                import sqlite3 as _dfsql
                from oxdata.db_loader import get_db_path as _df_gdb
                _df_conn = _dfsql.connect(f"file:{_df_gdb(project_id=_active_pid)}?mode=ro", uri=True)
                for _lbl in _df_driver_labels:
                    _row = _df_conn.execute(
                        "SELECT attr_id FROM dim_bq3_attribute WHERE TRIM(attr_label) = ? COLLATE NOCASE",
                        (_lbl,)
                    ).fetchone()
                    if _row:
                        _df_attr_ids.append(_row[0])
                _df_conn.close()
            except Exception as _dfe:
                st.warning(f"Could not resolve driver attr_ids: {_dfe}")

            # Store in session state so downstream steps can reference them
            st.session_state["driver_flow_attr_ids"] = _df_attr_ids
            st.session_state["driver_flow_labels"]   = _df_driver_labels
            st.session_state["driver_flow_brand"]    = sel_brand
            st.session_state["driver_flow_outcome"]  = _df_dv_kind
            _DF_FUNNEL = {"TOM", "SPONT", "AIDED", "CONSIDERATION", "EVER_USED", "CURRENT_USER", "PREFERRED"}
            if _df_dv_kind in _DF_FUNNEL:
                _df_aware_stages = ["TOM", "SPONT", "AIDED"]
            elif _df_dv_kind == "CSAT":
                _df_aware_stages = ["LAST_PURCHASED"]
            else:
                _df_aware_stages = ["TOM", "SPONT", "AIDED"]
            st.session_state["driver_flow_awareness_stages"] = _df_aware_stages

        # ── Step 2: BIP filtered to driver attributes ────────────────────────
        with st.container(border=True):
            _section_header("Step 2 — Brand Image Profiling (Driver Attributes Only)",
                            "Which brands are STRONG on the attributes that actually drive your outcome?")

            _df_attr_ids_for_bip = st.session_state.get("driver_flow_attr_ids", [])
            if not _df_attr_ids_for_bip:
                st.info("Run Step 1 first to identify driver attributes.")
            else:
                with st.spinner("Running BIP on driver attributes…"):
                    try:
                        from lens.analytics.bip_engine import BIPNormalizationEngine as _BIPEng
                        _df_bip_eng = _BIPEng(project_id=_active_pid)
                        _df_bip_res = _df_bip_eng.compute_bip(
                            category="all",
                            zone=zone_arg, gender=gender_arg,
                            age_band=age_band_arg, city=city_arg,
                            attr_ids=_df_attr_ids_for_bip,
                        )
                    except Exception as _bipe:
                        st.warning(f"BIP error: {_bipe}")
                        _df_bip_res = None

                if _df_bip_res is None or _df_bip_res.get("status") != "ok":
                    msg = _df_bip_res.get("message", "No BIP data") if _df_bip_res else "BIP not available"
                    st.info(f"BIP: {msg}")
                else:
                    _df_raw_m   = _df_bip_res.get("raw_matrix")
                    _df_norm_m  = _df_bip_res.get("normalized_matrix")
                    _df_filt_m  = _df_bip_res.get("filtered_scores")
                    _df_sig_m   = _df_bip_res.get("significance_matrix")

                    # Transpose toggle
                    _df_bip_T = st.checkbox("Drivers as rows (brands as columns)", value=True,
                                            key="df_bip_transpose")

                    _bip_dtabs = st.tabs(["Raw Scores", "Normalised", "Filtered", "Significance"])

                    def _df_show_matrix(mat, label):
                        if mat is None or (hasattr(mat, "empty") and mat.empty):
                            st.info(f"No {label} data.")
                            return
                        df_show = mat.T if _df_bip_T else mat
                        _brand_col = sel_brand if sel_brand in df_show.columns else None
                        def _bip_highlight(col):
                            return ["background:#dcfce7;font-weight:700" if (col.name == _brand_col) else "" for _ in col]
                        st.dataframe(
                            df_show.style.apply(_bip_highlight, axis=0).format(precision=1),
                            use_container_width=True,
                        )

                    with _bip_dtabs[0]:
                        st.caption("Association % (raw) — share of each brand's aware respondents who associate this attribute.")
                        _df_show_matrix(_df_raw_m, "raw")
                    with _bip_dtabs[1]:
                        st.caption("Normalised vs market average (norm dev). Positive = brand over-indexes.")
                        _df_show_matrix(_df_norm_m, "normalised")
                    with _bip_dtabs[2]:
                        st.caption("Filtered: only cells where brand significantly differs from market norm.")
                        _df_show_matrix(_df_filt_m, "filtered")
                    with _bip_dtabs[3]:
                        st.caption("Significance: YES = brand's association % is statistically above market average.")
                        _df_show_matrix(_df_sig_m, "significance")

                    # Visual: stacked bar — brand strength per driver attribute
                    if _df_raw_m is not None and not _df_raw_m.empty:
                        _df_brands_bip = list(_df_raw_m.columns)
                        _df_attrs_bip  = list(_df_raw_m.index)
                        _bip_vis_fig = go.Figure()
                        _bip_pal = [
                            "#1a5d4d","#22c55e","#f59e0b","#3b82f6","#a855f7",
                            "#ef4444","#06b6d4","#84cc16","#f97316","#ec4899",
                        ]
                        for _bi_idx, _b in enumerate(_df_brands_bip):
                            _hi = _b == sel_brand
                            _bip_vis_fig.add_trace(go.Bar(
                                name=_b,
                                x=_df_attrs_bip,
                                y=_df_raw_m[_b].tolist(),
                                marker=dict(
                                    color=_bip_pal[_bi_idx % len(_bip_pal)],
                                    opacity=1.0 if _hi else 0.6,
                                    line=dict(width=2 if _hi else 0, color="#000000"),
                                ),
                                hovertemplate="<b>%{x}</b><br>%{fullData.name}: %{y:.1f}%<extra></extra>",
                            ))
                        _bip_vis_fig.update_layout(
                            **{k: v for k, v in _chart_layout_base(360).items()
                               if k not in ("xaxis", "yaxis", "legend", "barmode")},
                            barmode="group",
                            xaxis=dict(title="Driver Attribute", tickangle=-35),
                            yaxis=dict(title="Association %", range=[0, 100]),
                            legend=dict(orientation="h", yanchor="bottom", y=1.01, xanchor="right", x=1),
                            title=dict(text=f"Brand Strength on Driver Attributes · {sel_brand} highlighted",
                                       font=dict(size=12)),
                        )
                        st.plotly_chart(_theme_fig(_bip_vis_fig), use_container_width=True)

        # ── Step 3: CAN MAP filtered to driver attributes ────────────────────
        with st.container(border=True):
            _section_header("Step 3 — Brand Perceptual Map (Driver Attributes Only)",
                            "Where do brands sit relative to each other on the dimensions that matter?")

            _df_attr_ids_for_ca = st.session_state.get("driver_flow_attr_ids", [])
            _df_aware_for_ca    = st.session_state.get("driver_flow_awareness_stages")
            if not _df_attr_ids_for_ca:
                st.info("Run Step 1 first.")
            else:
                _dfc1, _dfc2 = st.columns(2)
                with _dfc1:
                    _df_ca_brands = st.slider("Max brands on map", 3, 15, 8, key="df_ca_brands")
                with _dfc2:
                    _df_ca_min_n = st.slider("Min respondents per brand", 5, 50, 10, key="df_ca_min_n")

                with st.spinner("Running Correspondence Analysis on driver attributes…"):
                    try:
                        from lens.analytics.can_map_engine import run_ca_pipeline as _df_run_ca
                        _df_ca_res = _df_run_ca(
                            project_id=_active_pid,
                            category="All",
                            zone=zone_arg, gender=gender_arg,
                            age_band=age_band_arg, city=city_arg,
                            attr_ids=_df_attr_ids_for_ca,
                            top_brands=_df_ca_brands,
                            top_attrs=len(_df_attr_ids_for_ca),
                            min_respondents=_df_ca_min_n,
                            awareness_stages=_df_aware_for_ca or None,
                        )
                    except Exception as _cae:
                        st.warning(f"CAN MAP error: {_cae}")
                        _df_ca_res = None

                if _df_ca_res is None or _df_ca_res.get("status") != "ok":
                    msg = _df_ca_res.get("message", "No data") if _df_ca_res else "CAN MAP not available"
                    st.info(f"CAN MAP: {msg}")
                else:
                    # Render CA scatter from brand_coords/attr_coords
                    try:
                        _bc = _df_ca_res.get("brand_coords", {})
                        _ac = _df_ca_res.get("attr_coords", {})
                        if _bc:
                            _df_ca_simple = go.Figure()
                            _bc_names = list(_bc.keys())
                            _bc_x = [_bc[n][0] for n in _bc_names]
                            _bc_y = [_bc[n][1] for n in _bc_names]
                            _bc_cols = ["#1a5d4d" if n == sel_brand else "#94a3b8" for n in _bc_names]
                            _bc_sizes= [16 if n == sel_brand else 10 for n in _bc_names]
                            _bc_syms = ["star" if n == sel_brand else "circle" for n in _bc_names]
                            _df_ca_simple.add_trace(go.Scatter(
                                x=_bc_x, y=_bc_y, mode="markers+text",
                                text=[f"★ {n}" if n == sel_brand else n for n in _bc_names],
                                textposition="top center",
                                marker=dict(size=_bc_sizes, color=_bc_cols, symbol=_bc_syms),
                                name="Brands",
                                hovertemplate="<b>%{text}</b><extra></extra>",
                            ))
                            if _ac:
                                _ac_names = list(_ac.keys())
                                _df_ca_simple.add_trace(go.Scatter(
                                    x=[_ac[n][0] for n in _ac_names],
                                    y=[_ac[n][1] for n in _ac_names],
                                    mode="markers+text",
                                    text=[_clean_attr(n) for n in _ac_names],
                                    textposition="bottom center",
                                    marker=dict(size=7, color="#f59e0b", symbol="diamond"),
                                    name="Attributes",
                                    hovertemplate="<b>%{text}</b><extra></extra>",
                                ))
                            _df_ca_simple.add_hline(y=0, line_dash="dot", line_color="#d1d5db")
                            _df_ca_simple.add_vline(x=0, line_dash="dot", line_color="#d1d5db")
                            _df_ca_simple.update_layout(
                                **{k: v for k, v in _chart_layout_base(520).items()
                                   if k not in ("xaxis", "yaxis", "legend")},
                                xaxis=dict(title=f"F1 ({_df_ca_res.get('explained_inertia', [0,0])[0]*100:.1f}%)"),
                                yaxis=dict(title=f"F2 ({_df_ca_res.get('explained_inertia', [0,0])[1]*100:.1f}%)"),
                                legend=dict(orientation="h", yanchor="bottom", y=1.01),
                                title=dict(text=f"Perceptual Map — Driver Attributes · {sel_brand} ★",
                                           font=dict(size=12)),
                            )
                            st.plotly_chart(_theme_fig(_df_ca_simple), use_container_width=True)

                    except Exception as _ca_render_err:
                        st.warning(f"Could not render CAN MAP: {_ca_render_err}")

                    _df_ei = _df_ca_res.get("explained_inertia", [])
                    if len(_df_ei) >= 2:
                        st.caption(
                            f"F1 explains {_df_ei[0]*100:.1f}%, F2 explains {_df_ei[1]*100:.1f}% of variation "
                            f"among driver attributes. ★ = {sel_brand}. Diamonds = driver attributes. "
                            "Brands close to an attribute have a stronger-than-average association with it."
                        )

    # ── Assemble tabs ─────────────────────────────────────────────────────────────────────────────
    if _bh_page == _BH_PAGE_LABELS[1]:
        _tab_salience()
        _tab_competitive()
        _tab_funnel_leakage()

    if _bh_page == _BH_PAGE_LABELS[2]:
        _tab_imagery()
        _tab_ipa()
        _tab_attribute_ownership()
        _tab_key_driver()

    if _bh_page == _BH_PAGE_LABELS[3]:
        _tab_advocacy_loyalty()
        _tab_csat()
        _tab_demographics()
        _tab_purchase_journey()
        _tab_consumer_attitudes()
        _tab_market_dynamics()
        _tab_portfolio()
        _tab_price_tier()

    # _tab_driver_flow() merged into _tab_imagery() (Step 1→2 connected pipeline)

    with st.container(border=True):
        st.subheader("🛠️ Data Workbench")
        with st.expander("Full Segment Data & Export", expanded=False):
            df_all = pd.DataFrame(brands_list)
            show_cols = ["brand_name", "aided_pct", "spont_pct", "tom_pct", "nps", "nps_base", "strat_score"]
            st.dataframe(df_all[show_cols], hide_index=True, use_container_width=True)

        with st.expander("🤖 Ask AI about this data", expanded=False):
            _wb_ctx = st.session_state.get("_wb_active_section_data")
            if _wb_ctx:
                st.caption(f"Scope: **{_wb_ctx['section']}** · captured {_wb_ctx['captured_at']}")
            else:
                st.caption("Scope: no section captured yet — open a Driver Regression, Kano, "
                          "MaxDiff, or TURF panel above first, then come back here.")
            _wb_c1, _wb_c2 = st.columns([1, 2.2])
            with _wb_c1:
                _wb_model_choice = st.selectbox("Model", list(_WB_MODEL_CHOICES), index=0, key="wb_model_choice",
                                                help="Free model is the default (rate-limited); paid models "
                                                     "are more reliable for larger payloads. Auto-falls back "
                                                     "to a paid model if the selected one errors.")
            with _wb_c2:
                _wb_question = st.text_input(
                    "Question (leave blank for a general analysis of this section)",
                    key="wb_question", placeholder="e.g. Which attributes should we prioritize and why?")
            if st.button("Ask AI", key="wb_ask_btn", type="primary", disabled=(_wb_ctx is None)):
                with st.spinner(f"Asking {_wb_model_choice}…"):
                    _wb_text, _wb_used_model, _wb_err = _wb_ask(_wb_question, _wb_model_choice)
                if _wb_text:
                    if _wb_err:  # fallback happened but still got an answer
                        st.warning(_wb_err)
                    st.markdown(f"**Answer** *(via {_wb_used_model})*:")
                    st.markdown(_wb_text)
                else:
                    st.error(f"AI request failed: {_wb_err or 'unknown error'}")


render_brand_health_dashboard()
